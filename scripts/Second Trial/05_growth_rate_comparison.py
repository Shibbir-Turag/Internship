from __future__ import annotations

import html
import json
import re
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# PATHS
# ============================================================

PROJECT_ROOT = Path(r"C:\Users\tshib\OneDrive\Desktop\Internship")

TRAY_STATUS_XLSX = (
    PROJECT_ROOT
    / "data"
    / "Second Trial"
    / "Tray Status.xlsx"
)

SCRIPT04_REPORTS = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "04_Visible_Emergence"
    / "_reports"
)

TRAY_SUMMARY_CSV = (
    SCRIPT04_REPORTS
    / "visible_emergence_tray_summary.csv"
)

CELL_RESULTS_CSV = (
    SCRIPT04_REPORTS
    / "visible_emergence_cell_results.csv"
)

OUTPUT_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "05_Treatment_Growth_Visuals"
)

CHARTS_ROOT = OUTPUT_ROOT / "charts"
REPORTS_ROOT = OUTPUT_ROOT / "_reports"
CONFIG_ROOT = OUTPUT_ROOT / "_config"


# ============================================================
# SETTINGS
# ============================================================

EXPECTED_CELLS = 70

DAY_LABELS = {
    1: "Day 1",
    2: "Day 2",
    3: "Day 3",
    4: "Day 4",
    5: "Day 5",
    9: "Day 9",
}

TREATMENT_ORDER = [
    "No Microbes",
    "Microbes",
]

INTERACTION_ORDER = [
    "No Microbes | Inside",
    "No Microbes | Outside",
    "Microbes | Inside",
    "Microbes | Outside",
]


# ============================================================
# HELPERS
# ============================================================

def normalise(value: object) -> str:
    return re.sub(
        r"[^a-z0-9]+",
        "",
        str(value).casefold(),
    )


def is_present(value: object) -> bool:
    return str(value).strip().casefold() in {
        "p",
        "yes",
        "y",
        "true",
        "1",
    }


def require_columns(
    dataframe: pd.DataFrame,
    required_columns: list[str],
    source_name: str,
) -> None:
    missing = [
        column
        for column in required_columns
        if column not in dataframe.columns
    ]

    if missing:
        raise ValueError(
            f"{source_name} is missing required columns: "
            + ", ".join(missing)
        )


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 34

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(
                color="FFFFFF",
                bold=True,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in worksheet.columns:
            letter = column_cells[0].column_letter

            longest = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )

            worksheet.column_dimensions[letter].width = min(
                max(longest + 2, 12),
                55,
            )

    workbook.save(path)


def group_sort_order(group_type: str, group_name: str) -> int:
    if group_type == "Treatment":
        return TREATMENT_ORDER.index(group_name)

    return INTERACTION_ORDER.index(group_name)


# ============================================================
# LOAD TRAY DESIGN
# ============================================================

def load_tray_design() -> pd.DataFrame:
    if not TRAY_STATUS_XLSX.exists():
        raise FileNotFoundError(
            f"Missing Tray Status workbook:\n{TRAY_STATUS_XLSX}"
        )

    raw = pd.read_excel(TRAY_STATUS_XLSX)

    header_lookup = {
        normalise(column): column
        for column in raw.columns
    }

    required_headers = {
        "trayno": "Tray No",
        "microbes": "Microbes",
        "nomicrobes": "No Microbes",
        "inside": "Inside",
        "outside": "Outside",
    }

    missing = [
        label
        for key, label in required_headers.items()
        if key not in header_lookup
    ]

    if missing:
        raise ValueError(
            "Tray Status.xlsx is missing: "
            + ", ".join(missing)
        )

    tray_col = header_lookup["trayno"]
    microbes_col = header_lookup["microbes"]
    no_microbes_col = header_lookup["nomicrobes"]
    inside_col = header_lookup["inside"]
    outside_col = header_lookup["outside"]

    records = []

    for _, row in raw.iterrows():
        tray_no = pd.to_numeric(
            row[tray_col],
            errors="coerce",
        )

        if pd.isna(tray_no):
            continue

        tray_no = int(tray_no)

        microbes = is_present(row[microbes_col])
        no_microbes = is_present(row[no_microbes_col])
        inside = is_present(row[inside_col])
        outside = is_present(row[outside_col])

        if microbes == no_microbes:
            raise ValueError(
                f"Tray {tray_no}: exactly one of Microbes / "
                "No Microbes must be marked."
            )

        if inside == outside:
            raise ValueError(
                f"Tray {tray_no}: exactly one of Inside / "
                "Outside must be marked."
            )

        treatment = "Microbes" if microbes else "No Microbes"
        environment = "Inside" if inside else "Outside"

        records.append(
            {
                "tray_no": tray_no,
                "treatment": treatment,
                "environment": environment,
                "interaction": f"{treatment} | {environment}",
            }
        )

    design = (
        pd.DataFrame(records)
        .drop_duplicates(subset=["tray_no"])
        .sort_values("tray_no")
        .reset_index(drop=True)
    )

    if design.empty:
        raise ValueError(
            "No valid tray design records were found."
        )

    return design


# ============================================================
# LOAD SCRIPT 04 RESULTS
# ============================================================

def load_script04_results() -> pd.DataFrame:
    if not TRAY_SUMMARY_CSV.exists():
        raise FileNotFoundError(
            f"Missing Script 04 tray summary:\n{TRAY_SUMMARY_CSV}"
        )

    if not CELL_RESULTS_CSV.exists():
        raise FileNotFoundError(
            f"Missing Script 04 cell results:\n{CELL_RESULTS_CSV}"
        )

    tray_summary = pd.read_csv(TRAY_SUMMARY_CSV)
    cell_results = pd.read_csv(CELL_RESULTS_CSV)

    require_columns(
        tray_summary,
        [
            "day_order",
            "tray_no",
            "capture_id",
            "tracked_emergence_percent",
            "raw_green_percent",
            "newly_emerged_today",
            "status",
        ],
        "visible_emergence_tray_summary.csv",
    )

    require_columns(
        cell_results,
        [
            "day_order",
            "tray_no",
            "capture_id",
            "cell_id",
            "green_area_percent",
        ],
        "visible_emergence_cell_results.csv",
    )

    tray_summary = tray_summary.loc[
        tray_summary["status"]
        .astype(str)
        .str.upper()
        .eq("PASS")
    ].copy()

    for column in [
        "day_order",
        "tray_no",
        "tracked_emergence_percent",
        "raw_green_percent",
        "newly_emerged_today",
    ]:
        tray_summary[column] = pd.to_numeric(
            tray_summary[column],
            errors="coerce",
        )

    tray_summary = tray_summary.dropna(
        subset=["day_order", "tray_no"]
    ).copy()

    tray_summary["day_order"] = (
        tray_summary["day_order"]
        .astype(int)
    )

    tray_summary["tray_no"] = (
        tray_summary["tray_no"]
        .astype(int)
    )

    for column in [
        "day_order",
        "tray_no",
        "cell_id",
        "green_area_percent",
    ]:
        cell_results[column] = pd.to_numeric(
            cell_results[column],
            errors="coerce",
        )

    cell_results = cell_results.dropna(
        subset=[
            "day_order",
            "tray_no",
            "cell_id",
            "green_area_percent",
        ]
    ).copy()

    cell_results["day_order"] = (
        cell_results["day_order"]
        .astype(int)
    )

    cell_results["tray_no"] = (
        cell_results["tray_no"]
        .astype(int)
    )

    green_cover = (
        cell_results.groupby(
            ["day_order", "tray_no", "capture_id"],
            as_index=False,
        )
        .agg(
            mean_green_cover_percent=(
                "green_area_percent",
                "mean",
            ),
            green_cell_record_count=(
                "cell_id",
                "count",
            ),
        )
    )

    invalid_counts = green_cover.loc[
        green_cover["green_cell_record_count"] != EXPECTED_CELLS
    ]

    if not invalid_counts.empty:
        raise ValueError(
            f"Every tray/day must have {EXPECTED_CELLS} cell records.\n"
            + invalid_counts.to_string(index=False)
        )

    data = tray_summary.merge(
        green_cover,
        on=["day_order", "tray_no", "capture_id"],
        how="left",
        validate="one_to_one",
    )

    if data["mean_green_cover_percent"].isna().any():
        missing = data.loc[
            data["mean_green_cover_percent"].isna(),
            ["day_order", "tray_no", "capture_id"],
        ]

        raise ValueError(
            "Missing green-cover results for:\n"
            + missing.to_string(index=False)
        )

    if "tray" not in data.columns:
        data["tray"] = (
            "Tray "
            + data["tray_no"].astype(str)
        )

    data["day"] = data["day_order"].map(DAY_LABELS)

    return data


# ============================================================
# TRAY METRICS
# ============================================================

def calculate_tray_metrics(
    data: pd.DataFrame,
) -> pd.DataFrame:
    required_days = {1, 5, 9}

    existing_days = set(data["day_order"].unique())

    missing_days = required_days - existing_days

    if missing_days:
        raise ValueError(
            "Required observation days are missing: "
            + ", ".join(
                f"Day {day}"
                for day in sorted(missing_days)
            )
        )

    pivot = data.pivot_table(
        index=[
            "tray_no",
            "tray",
            "treatment",
            "environment",
            "interaction",
        ],
        columns="day_order",
        values=[
            "tracked_emergence_percent",
            "mean_green_cover_percent",
        ],
        aggfunc="first",
    )

    pivot.columns = [
        f"{metric}_day_{day}"
        for metric, day in pivot.columns
    ]

    metrics = pivot.reset_index()

    required_metric_columns = [
        "tracked_emergence_percent_day_1",
        "tracked_emergence_percent_day_5",
        "tracked_emergence_percent_day_9",
        "mean_green_cover_percent_day_1",
        "mean_green_cover_percent_day_5",
        "mean_green_cover_percent_day_9",
    ]

    require_columns(
        metrics,
        required_metric_columns,
        "Tray metrics table",
    )

    metrics["day1_tracked_emergence_percent"] = (
        metrics["tracked_emergence_percent_day_1"]
    )

    metrics["day5_tracked_emergence_percent"] = (
        metrics["tracked_emergence_percent_day_5"]
    )

    metrics["day9_tracked_emergence_percent"] = (
        metrics["tracked_emergence_percent_day_9"]
    )

    metrics["day1_green_cover_percent"] = (
        metrics["mean_green_cover_percent_day_1"]
    )

    metrics["day5_green_cover_percent"] = (
        metrics["mean_green_cover_percent_day_5"]
    )

    metrics["day9_green_cover_percent"] = (
        metrics["mean_green_cover_percent_day_9"]
    )

    metrics["emergence_change_day1_to_day5_pp"] = (
        metrics["day5_tracked_emergence_percent"]
        - metrics["day1_tracked_emergence_percent"]
    )

    metrics["emergence_rate_day1_to_day5_pp_per_day"] = (
        metrics["emergence_change_day1_to_day5_pp"] / 4.0
    )

    metrics["green_cover_change_day1_to_day5_pp"] = (
        metrics["day5_green_cover_percent"]
        - metrics["day1_green_cover_percent"]
    )

    metrics["green_cover_rate_day1_to_day5_pp_per_day"] = (
        metrics["green_cover_change_day1_to_day5_pp"] / 4.0
    )

    metrics["green_cover_change_day5_to_day9_pp"] = (
        metrics["day9_green_cover_percent"]
        - metrics["day5_green_cover_percent"]
    )

    output_columns = [
        "tray_no",
        "tray",
        "treatment",
        "environment",
        "interaction",
        "day1_tracked_emergence_percent",
        "day5_tracked_emergence_percent",
        "day9_tracked_emergence_percent",
        "day1_green_cover_percent",
        "day5_green_cover_percent",
        "day9_green_cover_percent",
        "emergence_change_day1_to_day5_pp",
        "emergence_rate_day1_to_day5_pp_per_day",
        "green_cover_change_day1_to_day5_pp",
        "green_cover_rate_day1_to_day5_pp_per_day",
        "green_cover_change_day5_to_day9_pp",
    ]

    return (
        metrics[output_columns]
        .sort_values("tray_no")
        .reset_index(drop=True)
    )


# ============================================================
# GROUP METRICS
# ============================================================

def calculate_group_daily(
    data: pd.DataFrame,
    group_column: str,
    group_type: str,
) -> pd.DataFrame:
    result = (
        data.groupby(
            [group_column, "day_order", "day"],
            as_index=False,
        )
        .agg(
            tray_count=("tray_no", "nunique"),
            mean_tracked_emergence_percent=(
                "tracked_emergence_percent",
                "mean",
            ),
            sd_tracked_emergence_percent=(
                "tracked_emergence_percent",
                "std",
            ),
            mean_green_cover_percent=(
                "mean_green_cover_percent",
                "mean",
            ),
            sd_green_cover_percent=(
                "mean_green_cover_percent",
                "std",
            ),
            mean_newly_emerged_today=(
                "newly_emerged_today",
                "mean",
            ),
        )
        .rename(columns={group_column: "group"})
    )

    result["group_type"] = group_type

    result[
        "sd_tracked_emergence_percent"
    ] = result[
        "sd_tracked_emergence_percent"
    ].fillna(0.0)

    result["sd_green_cover_percent"] = result[
        "sd_green_cover_percent"
    ].fillna(0.0)

    return result


def calculate_group_rates(
    tray_metrics: pd.DataFrame,
    group_column: str,
    group_type: str,
) -> pd.DataFrame:
    metric_columns = [
        "day1_tracked_emergence_percent",
        "day5_tracked_emergence_percent",
        "day9_tracked_emergence_percent",
        "day1_green_cover_percent",
        "day5_green_cover_percent",
        "day9_green_cover_percent",
        "emergence_change_day1_to_day5_pp",
        "emergence_rate_day1_to_day5_pp_per_day",
        "green_cover_change_day1_to_day5_pp",
        "green_cover_rate_day1_to_day5_pp_per_day",
        "green_cover_change_day5_to_day9_pp",
    ]

    aggregations = {
        "tray_count": (
            "tray_no",
            "nunique",
        )
    }

    for metric in metric_columns:
        aggregations[f"mean_{metric}"] = (
            metric,
            "mean",
        )

        aggregations[f"sd_{metric}"] = (
            metric,
            "std",
        )

    output = (
        tray_metrics.groupby(
            group_column,
            as_index=False,
        )
        .agg(**aggregations)
        .rename(columns={group_column: "group"})
    )

    for column in output.columns:
        if column.startswith("sd_"):
            output[column] = output[column].fillna(0.0)

    output["group_type"] = group_type

    ordered_columns = [
        "group_type",
        "group",
        "tray_count",
    ] + [
        column
        for column in output.columns
        if column not in {
            "group_type",
            "group",
            "tray_count",
        }
    ]

    return output[ordered_columns]


# ============================================================
# CHARTS
# ============================================================

def save_trend_chart(
    daily: pd.DataFrame,
    group_type: str,
    groups: list[str],
    value_column: str,
    sd_column: str,
    title: str,
    y_label: str,
    output_path: Path,
) -> None:
    figure, axis = plt.subplots(
        figsize=(11, 6.5)
    )

    subset = daily.loc[
        daily["group_type"].eq(group_type)
    ].copy()

    for group in groups:
        series = subset.loc[
            subset["group"].eq(group)
        ].sort_values("day_order")

        if series.empty:
            continue

        axis.errorbar(
            series["day_order"],
            series[value_column],
            yerr=series[sd_column],
            marker="o",
            linewidth=2,
            capsize=4,
            label=(
                f"{group} "
                f"(n={series['tray_count'].iloc[0]})"
            ),
        )

    days = sorted(
        subset["day_order"].unique()
    )

    axis.set_title(title)
    axis.set_xlabel("Observation day")
    axis.set_ylabel(y_label)
    axis.set_xticks(days)
    axis.set_xticklabels(
        [
            DAY_LABELS[day]
            for day in days
        ]
    )
    axis.set_ylim(bottom=0)
    axis.grid(
        True,
        axis="y",
        alpha=0.30,
    )

    if 9 in days:
        axis.axvline(
            7,
            linestyle="--",
            linewidth=1,
        )

        axis.text(
            7.08,
            axis.get_ylim()[1] * 0.96,
            "Day 9 follow-up",
            va="top",
            fontsize=9,
        )

    axis.legend(loc="best")

    figure.tight_layout()

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    figure.savefig(
        output_path,
        dpi=220,
    )

    plt.close(figure)


def save_tray_ranking(
    tray_metrics: pd.DataFrame,
    output_path: Path,
) -> None:
    ranked = tray_metrics.sort_values(
        "green_cover_rate_day1_to_day5_pp_per_day"
    ).copy()

    ranked["label"] = (
        ranked["tray"].astype(str)
        + " — "
        + ranked["interaction"].astype(str)
    )

    figure, axis = plt.subplots(
        figsize=(12, 7)
    )

    axis.barh(
        ranked["label"],
        ranked[
            "green_cover_rate_day1_to_day5_pp_per_day"
        ],
    )

    axis.set_title(
        "Tray ranking: Day 1–Day 5 RGB green-cover growth rate"
    )

    axis.set_xlabel(
        "RGB green-cover change per day "
        "(percentage points)"
    )

    axis.set_ylabel(
        "Tray and treatment group"
    )

    axis.grid(
        True,
        axis="x",
        alpha=0.30,
    )

    figure.tight_layout()

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    figure.savefig(
        output_path,
        dpi=220,
    )

    plt.close(figure)


def save_day9_chart(
    interaction_rates: pd.DataFrame,
    output_path: Path,
) -> None:
    frame = interaction_rates.copy()

    frame["sort_order"] = frame["group"].map(
        {
            group: index
            for index, group in enumerate(
                INTERACTION_ORDER
            )
        }
    )

    frame = frame.sort_values("sort_order")

    figure, axis = plt.subplots(
        figsize=(11, 6.5)
    )

    axis.bar(
        frame["group"],
        frame[
            "mean_day9_green_cover_percent"
        ],
    )

    axis.set_title(
        "Day 9 RGB green-cover proxy by treatment and environment"
    )

    axis.set_xlabel(
        "Treatment and environment"
    )

    axis.set_ylabel(
        "Mean RGB green-cover proxy (%)"
    )

    axis.set_ylim(bottom=0)

    axis.grid(
        True,
        axis="y",
        alpha=0.30,
    )

    axis.tick_params(
        axis="x",
        rotation=18,
    )

    figure.tight_layout()

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    figure.savefig(
        output_path,
        dpi=220,
    )

    plt.close(figure)


# ============================================================
# HTML REPORT
# ============================================================

def create_html_report(
    output_path: Path,
    treatment_rates: pd.DataFrame,
    interaction_rates: pd.DataFrame,
    tray_metrics: pd.DataFrame,
    chart_files: list[Path],
) -> None:
    best_treatment = treatment_rates.loc[
        treatment_rates[
            "mean_day5_green_cover_percent"
        ].idxmax()
    ]

    best_interaction = interaction_rates.loc[
        interaction_rates[
            "mean_day5_green_cover_percent"
        ].idxmax()
    ]

    best_rate = interaction_rates.loc[
        interaction_rates[
            "mean_green_cover_rate_day1_to_day5_pp_per_day"
        ].idxmax()
    ]

    treatment_table = treatment_rates[
        [
            "group",
            "tray_count",
            "mean_day5_tracked_emergence_percent",
            "mean_day5_green_cover_percent",
            "mean_green_cover_rate_day1_to_day5_pp_per_day",
            "mean_day9_green_cover_percent",
        ]
    ].copy()

    interaction_table = interaction_rates[
        [
            "group",
            "tray_count",
            "mean_day5_tracked_emergence_percent",
            "mean_day5_green_cover_percent",
            "mean_green_cover_rate_day1_to_day5_pp_per_day",
            "mean_day9_green_cover_percent",
        ]
    ].copy()

    ranking_table = tray_metrics[
        [
            "tray",
            "treatment",
            "environment",
            "green_cover_rate_day1_to_day5_pp_per_day",
            "day5_tracked_emergence_percent",
            "day9_green_cover_percent",
        ]
    ].sort_values(
        "green_cover_rate_day1_to_day5_pp_per_day",
        ascending=False,
    )

    def dataframe_html(
        dataframe: pd.DataFrame,
    ) -> str:
        return dataframe.round(3).to_html(
            index=False,
            border=0,
            classes="data-table",
        )

    charts_html = "\n".join(
        [
            (
                "<figure>"
                f'<img src="../charts/{html.escape(chart.name)}" '
                f'alt="{html.escape(chart.stem)}">'
                f"<figcaption>{html.escape(chart.stem.replace('_', ' '))}</figcaption>"
                "</figure>"
            )
            for chart in chart_files
        ]
    )

    document = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Second Trial treatment and growth visuals</title>
<style>
body {{
    font-family: Arial, sans-serif;
    margin: 28px;
    max-width: 1250px;
    color: #1f2937;
}}
h1, h2 {{
    color: #111827;
}}
p {{
    line-height: 1.55;
}}
.notice {{
    background: #f3f4f6;
    border-left: 4px solid #6b7280;
    padding: 14px 16px;
}}
.data-table {{
    border-collapse: collapse;
    width: 100%;
    margin: 16px 0 30px;
}}
.data-table th, .data-table td {{
    border: 1px solid #d1d5db;
    padding: 8px;
    text-align: left;
}}
.data-table th {{
    background: #e5e7eb;
}}
figure {{
    margin: 28px 0;
}}
img {{
    width: 100%;
    max-width: 1100px;
    border: 1px solid #d1d5db;
}}
figcaption {{
    margin-top: 7px;
    color: #4b5563;
}}
</style>
</head>
<body>

<h1>Second Trial: Treatment and environment growth visuals</h1>

<p class="notice">
These comparisons are descriptive. RGB green cover is an image-based growth
proxy, not direct biomass. Each treatment × environment group has two trays,
so inside/outside patterns should not be treated as formal proof.
</p>

<h2>Observed results</h2>

<p>
The higher Day 5 mean RGB green-cover proxy between the two primary treatments
was observed for <strong>{html.escape(str(best_treatment["group"]))}</strong>.
The highest Day 5 green-cover proxy among all treatment × environment groups
was observed for
<strong>{html.escape(str(best_interaction["group"]))}</strong>.
The highest Day 1–Day 5 green-cover growth rate was observed for
<strong>{html.escape(str(best_rate["group"]))}</strong>.
</p>

<h2>Microbes versus No Microbes</h2>
{dataframe_html(treatment_table)}

<h2>Treatment × environment comparison</h2>
{dataframe_html(interaction_table)}

<h2>Individual tray ranking</h2>
{dataframe_html(ranking_table)}

<h2>Charts</h2>
{charts_html}

</body>
</html>"""

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    output_path.write_text(
        document,
        encoding="utf-8",
    )


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    CHARTS_ROOT.mkdir(
        parents=True,
        exist_ok=True,
    )

    REPORTS_ROOT.mkdir(
        parents=True,
        exist_ok=True,
    )

    CONFIG_ROOT.mkdir(
        parents=True,
        exist_ok=True,
    )

    tray_design = load_tray_design()
    data = load_script04_results()

    missing_trays = sorted(
        set(data["tray_no"])
        - set(tray_design["tray_no"])
    )

    if missing_trays:
        raise ValueError(
            "Tray Status.xlsx has no mapping for trays: "
            + ", ".join(
                str(value)
                for value in missing_trays
            )
        )

    data = data.merge(
        tray_design,
        on="tray_no",
        how="left",
        validate="many_to_one",
    )

    pair_counts = (
        data.groupby(
            ["day_order", "tray_no"]
        )
        .size()
        .reset_index(name="count")
    )

    invalid_pairs = pair_counts.loc[
        pair_counts["count"] != 1
    ]

    if not invalid_pairs.empty:
        raise ValueError(
            "Expected exactly one row for each Day/Tray:\n"
            + invalid_pairs.to_string(index=False)
        )

    tray_metrics = calculate_tray_metrics(data)

    treatment_daily = calculate_group_daily(
        data,
        "treatment",
        "Treatment",
    )

    interaction_daily = calculate_group_daily(
        data,
        "interaction",
        "Treatment x Environment",
    )

    group_daily = pd.concat(
        [
            treatment_daily,
            interaction_daily,
        ],
        ignore_index=True,
    )

    group_daily["sort_order"] = group_daily.apply(
        lambda row: group_sort_order(
            row["group_type"],
            row["group"],
        ),
        axis=1,
    )

    group_daily = (
        group_daily.sort_values(
            [
                "group_type",
                "sort_order",
                "day_order",
            ]
        )
        .drop(columns=["sort_order"])
        .reset_index(drop=True)
    )

    treatment_rates = calculate_group_rates(
        tray_metrics,
        "treatment",
        "Treatment",
    )

    interaction_rates = calculate_group_rates(
        tray_metrics,
        "interaction",
        "Treatment x Environment",
    )

    group_rates = pd.concat(
        [
            treatment_rates,
            interaction_rates,
        ],
        ignore_index=True,
    )

    group_rates["sort_order"] = group_rates.apply(
        lambda row: group_sort_order(
            row["group_type"],
            row["group"],
        ),
        axis=1,
    )

    group_rates = (
        group_rates.sort_values(
            [
                "group_type",
                "sort_order",
            ]
        )
        .drop(columns=["sort_order"])
        .reset_index(drop=True)
    )

    chart_files = [
        CHARTS_ROOT
        / "01_microbes_vs_no_microbes_emergence.png",

        CHARTS_ROOT
        / "02_microbes_vs_no_microbes_green_cover.png",

        CHARTS_ROOT
        / "03_interaction_emergence.png",

        CHARTS_ROOT
        / "04_interaction_green_cover.png",

        CHARTS_ROOT
        / "05_tray_green_cover_growth_rate_ranking.png",

        CHARTS_ROOT
        / "06_day9_green_cover_by_group.png",
    ]

    save_trend_chart(
        treatment_daily,
        "Treatment",
        TREATMENT_ORDER,
        "mean_tracked_emergence_percent",
        "sd_tracked_emergence_percent",
        "Tracked visible emergence: Microbes vs No Microbes",
        "Mean tracked visible emergence (%)",
        chart_files[0],
    )

    save_trend_chart(
        treatment_daily,
        "Treatment",
        TREATMENT_ORDER,
        "mean_green_cover_percent",
        "sd_green_cover_percent",
        "RGB green-cover proxy: Microbes vs No Microbes",
        "Mean RGB green-cover proxy (%)",
        chart_files[1],
    )

    save_trend_chart(
        interaction_daily,
        "Treatment x Environment",
        INTERACTION_ORDER,
        "mean_tracked_emergence_percent",
        "sd_tracked_emergence_percent",
        "Tracked visible emergence by treatment and environment",
        "Mean tracked visible emergence (%)",
        chart_files[2],
    )

    save_trend_chart(
        interaction_daily,
        "Treatment x Environment",
        INTERACTION_ORDER,
        "mean_green_cover_percent",
        "sd_green_cover_percent",
        "RGB green-cover proxy by treatment and environment",
        "Mean RGB green-cover proxy (%)",
        chart_files[3],
    )

    save_tray_ranking(
        tray_metrics,
        chart_files[4],
    )

    save_day9_chart(
        interaction_rates,
        chart_files[5],
    )

    tray_metrics.to_csv(
        REPORTS_ROOT
        / "tray_growth_metrics.csv",
        index=False,
    )

    group_daily.to_csv(
        REPORTS_ROOT
        / "group_daily_metrics.csv",
        index=False,
    )

    group_rates.to_csv(
        REPORTS_ROOT
        / "group_growth_rates.csv",
        index=False,
    )

    excel_path = (
        REPORTS_ROOT
        / "treatment_growth_report.xlsx"
    )

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
    ) as writer:
        tray_design.to_excel(
            writer,
            sheet_name="Tray Design",
            index=False,
        )

        tray_metrics.to_excel(
            writer,
            sheet_name="Tray Growth Metrics",
            index=False,
        )

        group_daily.to_excel(
            writer,
            sheet_name="Group Daily Metrics",
            index=False,
        )

        group_rates.to_excel(
            writer,
            sheet_name="Group Growth Rates",
            index=False,
        )

    style_workbook(excel_path)

    settings = {
        "growth_rate_window": "Day 1 to Day 5",
        "growth_rate_formula": "(Day 5 - Day 1) / 4",
        "day_9_handling": "Later follow-up only",
        "expected_cells_per_tray": EXPECTED_CELLS,
        "tray_design": tray_design.to_dict(
            orient="records"
        ),
    }

    (
        CONFIG_ROOT
        / "analysis_settings.json"
    ).write_text(
        json.dumps(
            settings,
            indent=2,
        ),
        encoding="utf-8",
    )

    create_html_report(
        REPORTS_ROOT
        / "treatment_visual_summary.html",
        treatment_rates,
        interaction_rates,
        tray_metrics,
        chart_files,
    )

    best_day5 = interaction_rates.loc[
        interaction_rates[
            "mean_day5_green_cover_percent"
        ].idxmax()
    ]

    best_rate = interaction_rates.loc[
        interaction_rates[
            "mean_green_cover_rate_day1_to_day5_pp_per_day"
        ].idxmax()
    ]

    print("\n" + "=" * 70)
    print("SCRIPT 05 FINISHED")
    print("=" * 70)

    print(
        "Highest observed Day 5 RGB green-cover group: "
        f"{best_day5['group']} "
        f"({best_day5['mean_day5_green_cover_percent']:.3f}%)."
    )

    print(
        "Highest observed Day 1–Day 5 RGB green-cover rate group: "
        f"{best_rate['group']} "
        f"({best_rate['mean_green_cover_rate_day1_to_day5_pp_per_day']:.3f} "
        "percentage points per day)."
    )

    print(f"\nCharts:\n{CHARTS_ROOT}")
    print(f"\nReports:\n{REPORTS_ROOT}")


if __name__ == "__main__":
    main()