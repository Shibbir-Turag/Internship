"""
SECOND TRIAL — FOUR-CORNER D + MULTISPECTRAL CROPPER

RULES
-----
• D.JPG uses its own four-corner crop.
• MS_NIR.TIF uses a separate four-corner crop.
• The MS_NIR crop transform is applied to:
      MS_G.TIF
      MS_R.TIF
      MS_RE.TIF
      MS_NIR.TIF
• F.JPG is ignored completely.
• No confirmation is requested. The crop is accepted immediately after
  the fourth corner click.

CLICK ORDER FOR EVERY CROP
--------------------------
1. Top-left
2. Top-right
3. Bottom-right
4. Bottom-left

INPUT STRUCTURE
---------------
OneDrive/Desktop/Internship/data/Second Trial/
    Tray Status.xlsx
    First Day/
        Tray 1/
        Tray 2/
    Second Day/
    Third Day/
    ...

OUTPUT STRUCTURE
----------------
OneDrive/Desktop/Internship/outputs/Second Trial/
    First Day/
        Tray 1/
        Tray 2/
    ...

The script also creates:
• second_trial_four_corner_crop_config.json
• second_trial_four_corner_crop_manifest.csv
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path

import cv2
import matplotlib.pyplot as plt
import numpy as np
import tifffile
from openpyxl import load_workbook
from PIL import Image


# ============================================================
# SETTINGS
# ============================================================

# Leave this as None if your OneDrive location is normal.
# If automatic detection fails, enter your actual path here.
#
# Example:
# INTERNSHIP_ROOT = Path(r"C:\Users\Shibbir\OneDrive\Desktop\Internship")

INTERNSHIP_ROOT = None

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".tif", ".tiff"}

D_GROUP = "D"
MS_GROUP = "MS"

CONFIG_FILENAME = "second_trial_four_corner_crop_config.json"
MANIFEST_FILENAME = "second_trial_four_corner_crop_manifest.csv"


# ============================================================
# PATHS
# ============================================================

def get_internship_root() -> Path:
    """Find OneDrive/Desktop/Internship automatically."""

    if INTERNSHIP_ROOT is not None:
        return Path(INTERNSHIP_ROOT)

    candidates = []

    if os.environ.get("OneDrive"):
        candidates.append(
            Path(os.environ["OneDrive"]) / "Desktop" / "Internship"
        )

    candidates.append(Path.home() / "OneDrive" / "Desktop" / "Internship")
    candidates.append(Path.home() / "Desktop" / "Internship")

    for candidate in candidates:
        if (candidate / "data" / "Second Trial").exists():
            return candidate

    return candidates[0]


# ============================================================
# GENERAL HELPERS
# ============================================================

def natural_sort_key(text: str):
    """Sort Tray 2 before Tray 10."""
    return [
        int(part) if part.isdigit() else part.lower()
        for part in re.split(r"(\d+)", text)
    ]


def get_tray_number(tray_name: str):
    """Extract tray number from labels such as Tray 1."""
    match = re.search(r"(\d+)", tray_name)
    return int(match.group(1)) if match else None


def parse_filter_list(value):
    """Convert comma-separated command-line filters into lowercase sets."""

    if not value:
        return None

    return {
        item.strip().lower()
        for item in value.split(",")
        if item.strip()
    }


def has_filename_token(stem: str, token: str) -> bool:
    """
    Check a DJI filename token safely.

    Example:
    DJI_0001_MS_NIR.TIF -> token MS_NIR
    DJI_0001_D.JPG      -> token D
    """

    pattern = rf"(?:^|_){re.escape(token)}(?:_|$)"
    return re.search(pattern, stem.upper()) is not None


# ============================================================
# DJI FILE CLASSIFICATION
# ============================================================

def classify_image(filename: str):
    """
    D.JPG:
        RGB / visual tray image

    MS_G, MS_R, MS_RE, MS_NIR:
        Raw multispectral bands

    F.JPG:
        False-colour preview only — intentionally ignored
    """

    stem = Path(filename).stem.upper()

    # Ignore F preview files completely.
    if has_filename_token(stem, "F"):
        return None, "F_IGNORED"

    # Multispectral bands.
    if has_filename_token(stem, "MS_NIR"):
        return MS_GROUP, "MS_NIR"

    if has_filename_token(stem, "MS_RE"):
        return MS_GROUP, "MS_RE"

    if has_filename_token(stem, "MS_R"):
        return MS_GROUP, "MS_R"

    if has_filename_token(stem, "MS_G"):
        return MS_GROUP, "MS_G"

    # RGB D file.
    if has_filename_token(stem, "D"):
        return D_GROUP, "D"

    return None, "UNKNOWN"


# ============================================================
# TRAY STATUS EXCEL
# ============================================================

def is_present(value) -> bool:
    """Treat P as a present marker."""
    return str(value).strip().upper() == "P"


def read_tray_status(xlsx_path: Path):
    """
    Read Tray Status.xlsx.

    Expected columns:
    Tray No | Microbes | No Microbes | Inside | Outside
    """

    if not xlsx_path.exists():
        print(f"\nWARNING: Tray Status.xlsx was not found:\n{xlsx_path}")
        return {}

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        return {}

    headers = {
        str(value).strip().lower(): index
        for index, value in enumerate(rows[0])
        if value is not None
    }

    required_columns = [
        "tray no",
        "microbes",
        "no microbes",
        "inside",
        "outside",
    ]

    for column in required_columns:
        if column not in headers:
            print(f"WARNING: Missing Tray Status column: {column}")
            return {}

    tray_metadata = {}

    for row in rows[1:]:
        if not row:
            continue

        tray_value = row[headers["tray no"]]

        if tray_value is None:
            continue

        try:
            tray_number = int(tray_value)
        except (TypeError, ValueError):
            continue

        microbes = is_present(row[headers["microbes"]])
        no_microbes = is_present(row[headers["no microbes"]])
        inside = is_present(row[headers["inside"]])
        outside = is_present(row[headers["outside"]])

        if microbes and not no_microbes:
            treatment = "Microbes"
        elif no_microbes and not microbes:
            treatment = "No Microbes"
        else:
            treatment = "Unclear"

        if inside and not outside:
            environment = "Inside"
        elif outside and not inside:
            environment = "Outside"
        else:
            environment = "Unclear"

        tray_metadata[tray_number] = {
            "tray_no": tray_number,
            "treatment": treatment,
            "environment": environment,
        }

    return tray_metadata


# ============================================================
# IMAGE READ / SAVE
# ============================================================

def read_image(image_path: Path):
    """Read original image pixels without changing values."""

    if image_path.suffix.lower() in {".tif", ".tiff"}:
        image_array = tifffile.imread(image_path)
    else:
        with Image.open(image_path) as image:
            image_array = np.asarray(image.copy())

    if image_array.ndim not in (2, 3):
        raise ValueError(
            f"Unsupported image dimensions: {image_array.shape} "
            f"for {image_path.name}"
        )

    return image_array


def make_display_image(image_array: np.ndarray):
    """
    Create a visible contrast-stretched preview only for mouse clicking.
    It does not alter the image data used for cropping.
    """

    if image_array.ndim == 2:
        low, high = np.percentile(image_array, [1, 99])

        if high <= low:
            high = low + 1

        return np.clip(
            (image_array - low) * 255 / (high - low),
            0,
            255,
        ).astype(np.uint8)

    if image_array.shape[2] > 3:
        image_array = image_array[:, :, :3]

    preview = np.zeros_like(image_array, dtype=np.uint8)

    for channel in range(image_array.shape[2]):
        band = image_array[:, :, channel]

        low, high = np.percentile(band, [1, 99])

        if high <= low:
            high = low + 1

        preview[:, :, channel] = np.clip(
            (band - low) * 255 / (high - low),
            0,
            255,
        ).astype(np.uint8)

    return preview


def save_image(output_path: Path, image_array: np.ndarray):
    """Save crop while retaining original pixel values."""

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.suffix.lower() in {".tif", ".tiff"}:
        tifffile.imwrite(output_path, image_array)
    else:
        Image.fromarray(image_array).save(
            output_path,
            quality=95,
            subsampling=0,
            optimize=True,
        )


# ============================================================
# FOUR-CORNER CROP FUNCTIONS
# ============================================================

def output_size_from_corners(corners: np.ndarray):
    """
    Work out the output width and height from the selected tray edges.
    """

    top_width = np.linalg.norm(corners[1] - corners[0])
    bottom_width = np.linalg.norm(corners[2] - corners[3])

    left_height = np.linalg.norm(corners[3] - corners[0])
    right_height = np.linalg.norm(corners[2] - corners[1])

    output_width = max(2, int(round(max(top_width, bottom_width))))
    output_height = max(2, int(round(max(left_height, right_height))))

    return output_width, output_height


def points_are_valid(corners: np.ndarray) -> bool:
    """Check that four points make a valid tray quadrilateral."""

    if corners.shape != (4, 2):
        return False

    contour = corners.reshape((-1, 1, 2)).astype(np.float32)

    if not cv2.isContourConvex(contour):
        return False

    area = abs(cv2.contourArea(contour))

    return area > 500


def select_four_corners(reference_image: Path, crop_type: str, label: str):
    """
    Click exactly four points in this strict order:

    1. Top-left
    2. Top-right
    3. Bottom-right
    4. Bottom-left

    No Enter key is required.
    No confirmation is requested.
    The window closes automatically after the fourth click.
    """

    image_array = read_image(reference_image)

    height, width = image_array.shape[:2]

    preview = make_display_image(image_array)

    print("\n" + "=" * 75)
    print(label)
    print(f"Crop type: {crop_type}")
    print(f"Reference image: {reference_image.name}")
    print("\nClick in this exact order:")
    print("1. TOP-LEFT")
    print("2. TOP-RIGHT")
    print("3. BOTTOM-RIGHT")
    print("4. BOTTOM-LEFT")
    print("\nCropping starts automatically after click number 4.")
    print("=" * 75)

    figure, axis = plt.subplots(figsize=(14, 9))

    if preview.ndim == 2:
        axis.imshow(preview, cmap="gray")
    else:
        axis.imshow(preview)

    axis.set_title(
        f"{label}\n{crop_type}\n"
        "Click: TOP-LEFT → TOP-RIGHT → BOTTOM-RIGHT → BOTTOM-LEFT",
        fontsize=12,
    )

    axis.axis("off")
    plt.tight_layout()

    clicked_points = []

    def on_click(event):
        """Collect and display four clicked corners."""

        if event.inaxes != axis:
            return

        if event.xdata is None or event.ydata is None:
            return

        if len(clicked_points) >= 4:
            return

        clicked_points.append((event.xdata, event.ydata))

        point_number = len(clicked_points)

        axis.plot(
            event.xdata,
            event.ydata,
            marker="o",
            markersize=7,
        )

        axis.text(
            event.xdata,
            event.ydata,
            f" {point_number}",
            fontsize=12,
            fontweight="bold",
        )

        figure.canvas.draw_idle()

        # Close instantly after the fourth click.
        if len(clicked_points) == 4:
            outline = np.array(
                clicked_points + [clicked_points[0]],
                dtype=np.float32,
            )

            axis.plot(
                outline[:, 0],
                outline[:, 1],
                linewidth=2,
            )

            figure.canvas.draw_idle()

            # Tiny visual delay so the fourth point is visible,
            # then processing continues automatically.
            plt.pause(0.15)
            plt.close(figure)

    connection_id = figure.canvas.mpl_connect(
        "button_press_event",
        on_click,
    )

    plt.show(block=True)

    figure.canvas.mpl_disconnect(connection_id)

    if len(clicked_points) != 4:
        raise RuntimeError(
            "Four corners were not selected. "
            "Run again and click all four tray corners."
        )

    corners = np.array(clicked_points, dtype=np.float32)

    corners[:, 0] = np.clip(corners[:, 0], 0, width - 1)
    corners[:, 1] = np.clip(corners[:, 1], 0, height - 1)

    if not points_are_valid(corners):
        raise RuntimeError(
            "The four points did not form a valid crop quadrilateral. "
            "Click in clockwise tray-corner order."
        )

    output_width, output_height = output_size_from_corners(corners)

    normalised_corners = [
        [
            round(float(x / width), 8),
            round(float(y / height), 8),
        ]
        for x, y in corners
    ]

    return {
        "reference_file": reference_image.name,
        "reference_width": width,
        "reference_height": height,
        "corners_pixels": corners.tolist(),
        "corners_normalised": normalised_corners,
        "output_width": output_width,
        "output_height": output_height,
        "selected_utc": datetime.now(timezone.utc).isoformat(),
    }


def normalised_corners_to_pixels(corners_normalised, width, height):
    """Convert saved normalised corners into the actual image dimensions."""

    corners = np.array(corners_normalised, dtype=np.float32)

    corners[:, 0] *= width
    corners[:, 1] *= height

    corners[:, 0] = np.clip(corners[:, 0], 0, width - 1)
    corners[:, 1] = np.clip(corners[:, 1], 0, height - 1)

    return corners.astype(np.float32)


def perspective_crop(
    image_array: np.ndarray,
    corners_normalised,
    output_width: int,
    output_height: int,
    interpolation,
):
    """
    Rectify the four selected tray corners into a straight rectangular output.
    """

    source_height, source_width = image_array.shape[:2]

    source_corners = normalised_corners_to_pixels(
        corners_normalised,
        source_width,
        source_height,
    )

    destination_corners = np.array(
        [
            [0, 0],
            [output_width - 1, 0],
            [output_width - 1, output_height - 1],
            [0, output_height - 1],
        ],
        dtype=np.float32,
    )

    transform_matrix = cv2.getPerspectiveTransform(
        source_corners,
        destination_corners,
    )

    cropped_image = cv2.warpPerspective(
        image_array,
        transform_matrix,
        (output_width, output_height),
        flags=interpolation,
        borderMode=cv2.BORDER_CONSTANT,
        borderValue=0,
    )

    return cropped_image, source_corners


# ============================================================
# CONFIGURATION FILE
# ============================================================

def load_config(config_path: Path):
    """Load saved four-corner selections."""

    if not config_path.exists():
        return {
            "version": 1,
            "created_utc": datetime.now(timezone.utc).isoformat(),
            "entries": {},
        }

    with open(config_path, "r", encoding="utf-8") as file:
        return json.load(file)


def save_config(config_path: Path, config_data):
    """Save selected crop corners."""

    config_path.parent.mkdir(parents=True, exist_ok=True)

    config_data["updated_utc"] = datetime.now(timezone.utc).isoformat()

    with open(config_path, "w", encoding="utf-8") as file:
        json.dump(config_data, file, indent=2)


# ============================================================
# FOLDER DISCOVERY
# ============================================================

def find_day_tray_folders(input_root, selected_days=None, selected_trays=None):
    """Find all Day > Tray folders."""

    results = []

    day_folders = sorted(
        [folder for folder in input_root.iterdir() if folder.is_dir()],
        key=lambda folder: natural_sort_key(folder.name),
    )

    for day_folder in day_folders:
        if selected_days and day_folder.name.lower() not in selected_days:
            continue

        tray_folders = sorted(
            [folder for folder in day_folder.iterdir() if folder.is_dir()],
            key=lambda folder: natural_sort_key(folder.name),
        )

        for tray_folder in tray_folders:
            if selected_trays and tray_folder.name.lower() not in selected_trays:
                continue

            results.append((day_folder, tray_folder))

    return results


def get_images_in_tray(tray_folder: Path):
    """List JPG and TIFF files inside one tray folder."""

    return sorted(
        [
            file
            for file in tray_folder.iterdir()
            if file.is_file() and file.suffix.lower() in IMAGE_EXTENSIONS
        ],
        key=lambda file: natural_sort_key(file.name),
    )


def find_reference_image(images, target_group):
    """
    D crop reference: D.JPG only.
    MS crop reference: MS_NIR.TIF only.
    """

    for image_path in images:
        group, band = classify_image(image_path.name)

        if target_group == D_GROUP:
            if group == D_GROUP and band == "D":
                return image_path

        if target_group == MS_GROUP:
            if group == MS_GROUP and band == "MS_NIR":
                return image_path

    return None


# ============================================================
# MANIFEST
# ============================================================

def write_manifest(manifest_path: Path, rows):
    """Write crop information and tray metadata into CSV."""

    manifest_path.parent.mkdir(parents=True, exist_ok=True)

    columns = [
        "day",
        "tray_folder",
        "tray_no",
        "treatment",
        "environment",
        "source_relative_path",
        "output_relative_path",
        "filename",
        "band",
        "crop_group",
        "reference_file",
        "source_width",
        "source_height",
        "corner_1_x",
        "corner_1_y",
        "corner_2_x",
        "corner_2_y",
        "corner_3_x",
        "corner_3_y",
        "corner_4_x",
        "corner_4_y",
        "output_width",
        "output_height",
        "status",
        "note",
    ]

    with open(manifest_path, "w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


# ============================================================
# MAIN PROCESS
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Four-corner cropper for Second Trial D and MS images."
    )

    parser.add_argument(
        "--redo",
        action="store_true",
        help="Select new four-corner crops for all selected trays.",
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing cropped output files.",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Check folders and file names without cropping.",
    )

    parser.add_argument(
        "--days",
        help='Example: --days "First Day,Second Day"',
    )

    parser.add_argument(
        "--trays",
        help='Example: --trays "Tray 1,Tray 2"',
    )

    args = parser.parse_args()

    internship_root = get_internship_root()

    input_root = internship_root / "data" / "Second Trial"
    output_root = internship_root / "outputs" / "Second Trial"
    script_root = internship_root / "scripts" / "Second Trial"

    tray_status_path = input_root / "Tray Status.xlsx"

    config_path = script_root / CONFIG_FILENAME
    manifest_path = output_root / MANIFEST_FILENAME

    print("\nSECOND TRIAL FOUR-CORNER CROPPER")
    print("=" * 75)
    print(f"Input folder:  {input_root}")
    print(f"Output folder: {output_root}")
    print(f"Script folder: {script_root}")
    print(f"Tray Status:   {tray_status_path}")

    if not input_root.exists():
        print("\nERROR: Input folder was not found.")
        print(f"Expected path:\n{input_root}")
        print("\nEdit INTERNSHIP_ROOT near the top if required.")
        return 1

    selected_days = parse_filter_list(args.days)
    selected_trays = parse_filter_list(args.trays)

    day_tray_folders = find_day_tray_folders(
        input_root,
        selected_days,
        selected_trays,
    )

    if not day_tray_folders:
        print("\nNo Day/Tray folders were found.")
        return 1

    tray_metadata = read_tray_status(tray_status_path)
    config_data = load_config(config_path)

    manifest_rows = []

    cropped_count = 0
    skipped_count = 0
    ignored_count = 0
    failed_count = 0

    print("\nDetected folders:")

    for day_folder, tray_folder in day_tray_folders:
        images = get_images_in_tray(tray_folder)

        usable_images = sum(
            1
            for image in images
            if classify_image(image.name)[0] is not None
        )

        print(
            f"  {day_folder.name} > {tray_folder.name}: "
            f"{len(images)} image(s), {usable_images} usable"
        )

    if args.dry_run:
        print("\nDry run completed. Nothing was cropped.")
        return 0

    output_root.mkdir(parents=True, exist_ok=True)
    script_root.mkdir(parents=True, exist_ok=True)

    # --redo automatically overwrites old cropped results.
    overwrite_outputs = args.overwrite or args.redo

    for day_folder, tray_folder in day_tray_folders:
        day_name = day_folder.name
        tray_name = tray_folder.name

        tray_number = get_tray_number(tray_name)

        metadata = tray_metadata.get(
            tray_number,
            {
                "tray_no": tray_number if tray_number else "",
                "treatment": "Unknown",
                "environment": "Unknown",
            },
        )

        entry_key = f"{day_name}/{tray_name}"

        images = get_images_in_tray(tray_folder)

        if not images:
            print(f"\nWARNING: No images found in {entry_key}")
            continue

        if entry_key not in config_data["entries"]:
            config_data["entries"][entry_key] = {}

        crop_settings = config_data["entries"][entry_key]

        # ====================================================
        # D.JPG: FOUR CLICKS
        # ====================================================

        d_reference = find_reference_image(images, D_GROUP)

        if d_reference is not None:
            if args.redo or D_GROUP not in crop_settings:
                try:
                    crop_settings[D_GROUP] = select_four_corners(
                        d_reference,
                        "D/RGB four-corner crop",
                        f"{day_name} > {tray_name}",
                    )

                    save_config(config_path, config_data)

                except Exception as error:
                    print(f"\nERROR selecting D crop for {entry_key}: {error}")
                    failed_count += 1
                    continue

        # ====================================================
        # MS_NIR.TIF: FOUR CLICKS
        # ====================================================

        ms_reference = find_reference_image(images, MS_GROUP)

        if ms_reference is not None:
            if args.redo or MS_GROUP not in crop_settings:
                try:
                    crop_settings[MS_GROUP] = select_four_corners(
                        ms_reference,
                        "MS four-corner crop using MS_NIR reference",
                        f"{day_name} > {tray_name}",
                    )

                    save_config(config_path, config_data)

                except Exception as error:
                    print(f"\nERROR selecting MS crop for {entry_key}: {error}")
                    failed_count += 1
                    continue

        # ====================================================
        # CROP ALL VALID FILES
        # ====================================================

        for source_file in images:
            crop_group, band = classify_image(source_file.name)

            source_relative_path = source_file.relative_to(input_root)
            output_file = output_root / source_relative_path

            manifest_row = {
                "day": day_name,
                "tray_folder": tray_name,
                "tray_no": metadata["tray_no"],
                "treatment": metadata["treatment"],
                "environment": metadata["environment"],
                "source_relative_path": str(source_relative_path),
                "output_relative_path": "",
                "filename": source_file.name,
                "band": band,
                "crop_group": crop_group if crop_group else "",
                "reference_file": "",
                "source_width": "",
                "source_height": "",
                "corner_1_x": "",
                "corner_1_y": "",
                "corner_2_x": "",
                "corner_2_y": "",
                "corner_3_x": "",
                "corner_3_y": "",
                "corner_4_x": "",
                "corner_4_y": "",
                "output_width": "",
                "output_height": "",
                "status": "",
                "note": "",
            }

            # F preview is deliberately excluded.
            if band == "F_IGNORED":
                manifest_row["status"] = "IGNORED"
                manifest_row["note"] = (
                    "F false-colour preview intentionally excluded."
                )

                manifest_rows.append(manifest_row)
                ignored_count += 1
                continue

            # Ignore unrelated files.
            if crop_group is None:
                manifest_row["status"] = "IGNORED"
                manifest_row["note"] = "Unrecognised filename pattern."

                manifest_rows.append(manifest_row)
                ignored_count += 1
                continue

            # No crop reference found.
            if crop_group not in crop_settings:
                manifest_row["status"] = "FAILED"
                manifest_row["note"] = (
                    f"No crop settings found for group: {crop_group}"
                )

                manifest_rows.append(manifest_row)
                failed_count += 1
                continue

            manifest_row["output_relative_path"] = str(
                output_file.relative_to(output_root)
            )

            # Keep existing file unless --overwrite or --redo was used.
            if output_file.exists() and not overwrite_outputs:
                manifest_row["status"] = "SKIPPED"
                manifest_row["reference_file"] = (
                    crop_settings[crop_group]["reference_file"]
                )
                manifest_row["note"] = (
                    "Output already exists. Use --overwrite or --redo to replace."
                )

                manifest_rows.append(manifest_row)
                skipped_count += 1
                continue

            try:
                image_array = read_image(source_file)

                source_height, source_width = image_array.shape[:2]

                crop_info = crop_settings[crop_group]

                # Preserve raw multispectral values: nearest-neighbour only.
                # RGB D image may use normal linear interpolation.
                if crop_group == MS_GROUP:
                    interpolation = cv2.INTER_NEAREST
                else:
                    interpolation = cv2.INTER_LINEAR

                cropped_image, actual_corners = perspective_crop(
                    image_array=image_array,
                    corners_normalised=crop_info["corners_normalised"],
                    output_width=crop_info["output_width"],
                    output_height=crop_info["output_height"],
                    interpolation=interpolation,
                )

                save_image(output_file, cropped_image)

                manifest_row["reference_file"] = crop_info["reference_file"]
                manifest_row["source_width"] = source_width
                manifest_row["source_height"] = source_height

                for index, point in enumerate(actual_corners, start=1):
                    manifest_row[f"corner_{index}_x"] = round(float(point[0]), 2)
                    manifest_row[f"corner_{index}_y"] = round(float(point[1]), 2)

                manifest_row["output_width"] = cropped_image.shape[1]
                manifest_row["output_height"] = cropped_image.shape[0]
                manifest_row["status"] = "CROPPED"

                if crop_group == D_GROUP:
                    manifest_row["note"] = (
                        "Perspective-cropped using D/RGB four-corner reference."
                    )
                else:
                    manifest_row["note"] = (
                        "Perspective-cropped using MS_NIR four-corner reference."
                    )

                manifest_rows.append(manifest_row)

                cropped_count += 1

                print(
                    f"CROPPED: {day_name} > {tray_name} > {source_file.name}"
                )

            except Exception as error:
                manifest_row["status"] = "FAILED"
                manifest_row["note"] = str(error)

                manifest_rows.append(manifest_row)

                failed_count += 1

                print(
                    f"FAILED: {day_name} > {tray_name} > "
                    f"{source_file.name} | {error}"
                )

    write_manifest(manifest_path, manifest_rows)
    save_config(config_path, config_data)

    print("\n" + "=" * 75)
    print("CROPPING FINISHED")
    print("=" * 75)
    print(f"Cropped: {cropped_count}")
    print(f"Skipped: {skipped_count}")
    print(f"Ignored: {ignored_count}")
    print(f"Failed:  {failed_count}")
    print(f"\nOutputs saved in:\n{output_root}")
    print(f"\nManifest CSV:\n{manifest_path}")
    print(f"\nCrop settings:\n{config_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())