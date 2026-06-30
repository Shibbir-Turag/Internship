"""
SCRIPT 04 — VISIBLE EMERGENCE TRACKING

Purpose
-------
Detect visible green seedling evidence inside the 70 square cell zones created
by Script 03, then track emergence across continuous observation days.

Important temporal rule
-----------------------
Every cell is analysed independently on every available day.

The script keeps two separate values:

1. Raw Current Green Evidence
   - Is green seedling evidence visible in this exact image/day?

2. Tracked Visible Emergence
   - Has the cell shown visible emergence on this day or any earlier day?

A previously emerged cell is still checked again on later days.
It is never skipped.

This measures visible emergence only. It does not claim underground germination.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageOps


# ============================================================
# 1) PATHS — CHANGE PROJECT_ROOT ONLY WHEN REUSING
# ============================================================

PROJECT_ROOT = Path(
    r"C:\Users\tshib\OneDrive\Desktop\Internship"
)

CROP_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "01_Crop_Dual_Reference"
)

GRID_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "03_Cell_Grid_Detection"
)

OUTPUT_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "04_Visible_Emergence"
)


# ============================================================
# 2) PROJECT SETTINGS
# ============================================================

ROWS = 7
COLS = 10
EXPECTED_CELLS = ROWS * COLS

DAY_NAME_TO_ORDER = {
    "day 1": 1,
    "day 2": 2,
    "day 3": 3,
    "day 4": 4,
    "day 5": 5,
    "day 9": 9,
    "first day": 1,
    "second day": 2,
    "third day": 3,
    "fourth day": 4,
    "fifth day": 5,
    "ninth day": 9,
}

ACCEPTED_GRID_STATUSES = {
    "PASS_AUTO",
    "PASS_MANUAL",
}

# Green-seedling evidence settings.
# OpenCV HSV hue range is 0–179.
GREEN_HUE_MIN = 20
GREEN_HUE_MAX = 95
GREEN_SATURATION_MIN = 28
GREEN_VALUE_MIN = 35

# Excess Green = 2G - R - B
EXCESS_GREEN_MIN = 12

# Helps reject brown mulch and soil.
GREEN_TO_RED_RATIO = 1.03
GREEN_TO_BLUE_RATIO = 1.00

# Minimum evidence required within one square zone.
MIN_GREEN_PIXELS_PER_CELL = 20
MIN_GREEN_COMPONENT_PIXELS = 12

MORPH_KERNEL_SIZE = 3


# ============================================================
# 3) GENERAL HELPERS
# ============================================================

def natural_key(text: str):
    return [
        int(part) if part.isdigit() else part.casefold()
        for part in re.split(r"(\d+)", text)
    ]


def day_sort_key(folder: Path):
    return (
        DAY_NAME_TO_ORDER.get(folder.name.casefold(), 999),
        natural_key(folder.name),
    )


def parse_filter_list(value: str | None):
    if not value:
        return None

    return {
        item.strip().casefold()
        for item in value.split(",")
        if item.strip()
    }


def tray_number_from_name(tray_name: str):
    match = re.search(r"(\d+)", tray_name)
    return int(match.group(1)) if match else ""


def safe_name(text: str):
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", text)


def relative_path(path: Path | None, root: Path):
    if path is None:
        return ""

    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


def record_key(day: str, tray: str, capture_id: str):
    return f"{day}|{tray}|{capture_id}"


def sortable_number(value, default=999):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def yes_no(value: bool):
    return "Yes" if value else "No"


# ============================================================
# 4) FIND CROPPED D/RGB FILES
# ============================================================

def parse_d_image(path: Path):
    """
    Example:
        DJI_20260618124632_0008_D.JPG
    """

    match = re.match(
        r"^(?P<capture>.+)_D$",
        path.stem.upper(),
    )

    return match.group("capture") if match else None


def find_d_images(tray_folder: Path):
    images = []

    for file in tray_folder.iterdir():
        if not file.is_file():
            continue

        if file.suffix.casefold() not in {".jpg", ".jpeg"}:
            continue

        capture_id = parse_d_image(file)

        if capture_id:
            images.append(
                {
                    "capture_id": capture_id,
                    "path": file,
                }
            )

    return sorted(
        images,
        key=lambda item: natural_key(item["capture_id"]),
    )


def read_rgb(path: Path):
    """Read cropped D/RGB image without changing the source file."""

    with Image.open(path) as image:
        image = ImageOps.exif_transpose(image)
        image = image.convert("RGB")

        return np.asarray(image)


# ============================================================
# 5) LOAD SCRIPT 03 GRID OUTPUTS
# ============================================================

def load_csv_rows(path: Path):
    if not path.exists():
        return []

    with path.open(
        "r",
        newline="",
        encoding="utf-8",
    ) as file:
        return list(csv.DictReader(file))


def load_grid_manifest():
    manifest_path = (
        GRID_ROOT
        / "_reports"
        / "cell_grid_manifest.csv"
    )

    manifest = {}

    for row in load_csv_rows(manifest_path):
        key = record_key(
            row.get("day", ""),
            row.get("tray", ""),
            row.get("capture_id", ""),
        )

        manifest[key] = row

    return manifest


def load_cell_coordinates():
    """
    Read square zones from Script 03.

    Returns:
        {
            "Day 1|Tray 1|CAPTURE_ID": [70 cell dictionaries],
            ...
        }
    """

    coordinates_path = (
        GRID_ROOT
        / "_reports"
        / "cell_coordinates.csv"
    )

    grouped = defaultdict(list)

    for row in load_csv_rows(coordinates_path):
        key = record_key(
            row.get("day", ""),
            row.get("tray", ""),
            row.get("capture_id", ""),
        )

        try:
            grouped[key].append(
                {
                    "cell_id": int(row["cell_id"]),
                    "row": int(row["row"]),
                    "column": int(row["column"]),
                    "x": float(row["x"]),
                    "y": float(row["y"]),
                    "square_x0": float(row["square_x0"]),
                    "square_y0": float(row["square_y0"]),
                    "square_x1": float(row["square_x1"]),
                    "square_y1": float(row["square_y1"]),
                    "square_side": float(row["square_side"]),
                    "coordinate_source": row.get(
                        "coordinate_source",
                        "",
                    ),
                    "needs_review": row.get(
                        "needs_review",
                        "",
                    ),
                }
            )
        except (KeyError, TypeError, ValueError):
            continue

    for key in grouped:
        grouped[key].sort(
            key=lambda item: item["cell_id"]
        )

    return grouped


# ============================================================
# 6) GREEN EVIDENCE DETECTION
# ============================================================

def build_green_mask(rgb: np.ndarray):
    """
    Build a conservative vegetation-like green mask.

    It combines:
    - HSV green range
    - saturation/value filters
    - excess green
    - green stronger than red and blue

    This is designed to reject most brown mulch, soil and white foam.
    """

    hsv = cv2.cvtColor(
        rgb,
        cv2.COLOR_RGB2HSV,
    )

    red = rgb[:, :, 0].astype(np.float32)
    green = rgb[:, :, 1].astype(np.float32)
    blue = rgb[:, :, 2].astype(np.float32)

    hue = hsv[:, :, 0]
    saturation = hsv[:, :, 1]
    value = hsv[:, :, 2]

    excess_green = (
        2.0 * green
        - red
        - blue
    )

    hsv_green = (
        (hue >= GREEN_HUE_MIN)
        & (hue <= GREEN_HUE_MAX)
        & (saturation >= GREEN_SATURATION_MIN)
        & (value >= GREEN_VALUE_MIN)
    )

    green_balance = (
        (green >= red * GREEN_TO_RED_RATIO)
        & (green >= blue * GREEN_TO_BLUE_RATIO)
    )

    mask = (
        hsv_green
        & green_balance
        & (excess_green >= EXCESS_GREEN_MIN)
    ).astype(np.uint8) * 255

    kernel = cv2.getStructuringElement(
        cv2.MORPH_ELLIPSE,
        (MORPH_KERNEL_SIZE, MORPH_KERNEL_SIZE),
    )

    mask = cv2.morphologyEx(
        mask,
        cv2.MORPH_OPEN,
        kernel,
        iterations=1,
    )

    mask = cv2.morphologyEx(
        mask,
        cv2.MORPH_CLOSE,
        kernel,
        iterations=1,
    )

    return mask


def clip_square_to_image(
    cell: dict,
    image_width: int,
    image_height: int,
):
    x0 = max(
        0,
        int(round(cell["square_x0"])),
    )

    y0 = max(
        0,
        int(round(cell["square_y0"])),
    )

    x1 = min(
        image_width,
        int(round(cell["square_x1"])),
    )

    y1 = min(
        image_height,
        int(round(cell["square_y1"])),
    )

    return x0, y0, x1, y1


def analyse_cell_green_evidence(
    green_mask: np.ndarray,
    cell: dict,
):
    """
    Analyse one Script 03 square ownership zone.

    A raw-positive result requires:
    - enough green pixels
    - a sufficiently large connected green component
    """

    image_height, image_width = green_mask.shape[:2]

    x0, y0, x1, y1 = clip_square_to_image(
        cell,
        image_width,
        image_height,
    )

    if x1 <= x0 or y1 <= y0:
        return {
            "raw_green_detected": False,
            "green_pixels": 0,
            "largest_green_component": 0,
            "green_area_percent": 0.0,
            "zone_pixels": 0,
        }

    zone = green_mask[y0:y1, x0:x1]

    zone_pixels = int(zone.size)
    green_pixels = int(np.count_nonzero(zone))

    if green_pixels == 0:
        return {
            "raw_green_detected": False,
            "green_pixels": 0,
            "largest_green_component": 0,
            "green_area_percent": 0.0,
            "zone_pixels": zone_pixels,
        }

    labels_count, _labels, stats, _centres = (
        cv2.connectedComponentsWithStats(
            zone,
            connectivity=8,
        )
    )

    component_areas = [
        int(stats[index, cv2.CC_STAT_AREA])
        for index in range(1, labels_count)
    ]

    largest_component = max(
        component_areas,
        default=0,
    )

    green_area_percent = (
        green_pixels
        / max(zone_pixels, 1)
        * 100.0
    )

    raw_green_detected = (
        green_pixels >= MIN_GREEN_PIXELS_PER_CELL
        and largest_component >= MIN_GREEN_COMPONENT_PIXELS
    )

    return {
        "raw_green_detected": raw_green_detected,
        "green_pixels": green_pixels,
        "largest_green_component": largest_component,
        "green_area_percent": green_area_percent,
        "zone_pixels": zone_pixels,
    }


# ============================================================
# 7) VISUAL QA OUTPUTS
# ============================================================

def save_emergence_overlay(
    rgb: np.ndarray,
    green_mask: np.ndarray,
    cell_results: list[dict],
    output_path: Path,
    title: str,
):
    """
    Overlay colours:
    Green = raw green evidence detected today
    Blue  = no raw evidence today, but emerged earlier
    Grey  = no emergence detected yet
    """

    overlay = cv2.cvtColor(
        rgb,
        cv2.COLOR_RGB2BGR,
    )

    green_layer = np.zeros_like(overlay)
    green_layer[:, :] = (0, 255, 0)

    blended = cv2.addWeighted(
        overlay,
        0.60,
        green_layer,
        0.40,
        0,
    )

    evidence_pixels = green_mask > 0
    overlay[evidence_pixels] = blended[evidence_pixels]

    image_height, image_width = overlay.shape[:2]

    median_side = float(
        np.median(
            [row["square_side"] for row in cell_results]
        )
    )

    line_width = max(
        2,
        int(round(median_side / 32)),
    )

    font_scale = max(
        0.35,
        min(1.0, median_side / 88),
    )

    font_thickness = max(
        1,
        int(round(font_scale * 2)),
    )

    for row in cell_results:
        x0 = max(
            0,
            int(round(row["square_x0"])),
        )

        y0 = max(
            0,
            int(round(row["square_y0"])),
        )

        x1 = min(
            image_width - 1,
            int(round(row["square_x1"])),
        )

        y1 = min(
            image_height - 1,
            int(round(row["square_y1"])),
        )

        if row["raw_green_detected"]:
            colour = (0, 200, 0)  # Green
        elif row["tracked_visible_emerged"]:
            colour = (255, 150, 0)  # Blue
        else:
            colour = (135, 135, 135)  # Grey

        cv2.rectangle(
            overlay,
            (x0, y0),
            (x1, y1),
            colour,
            line_width,
        )

        label = str(row["cell_id"])

        # FIX: Script 03 supplies x and y, not centre_x / centre_y.
        text_x = max(
            1,
            int(round(row["x"] - median_side * 0.10)),
        )

        text_y = min(
            image_height - 5,
            int(round(row["y"] + median_side * 0.08)),
        )

        cv2.putText(
            overlay,
            label,
            (text_x, text_y),
            cv2.FONT_HERSHEY_SIMPLEX,
            font_scale,
            colour,
            font_thickness,
            cv2.LINE_AA,
        )

    raw_count = sum(
        bool(row["raw_green_detected"])
        for row in cell_results
    )

    tracked_count = sum(
        bool(row["tracked_visible_emerged"])
        for row in cell_results
    )

    header_height = max(
        46,
        int(round(median_side * 0.34)),
    )

    cv2.rectangle(
        overlay,
        (0, 0),
        (image_width, header_height),
        (255, 255, 255),
        thickness=-1,
    )

    header_text = (
        f"{title} | "
        f"Raw green: {raw_count}/70 | "
        f"Tracked emerged: {tracked_count}/70"
    )

    cv2.putText(
        overlay,
        header_text,
        (12, int(header_height * 0.68)),
        cv2.FONT_HERSHEY_SIMPLEX,
        max(0.42, min(0.82, font_scale)),
        (0, 0, 0),
        2,
        cv2.LINE_AA,
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    cv2.imwrite(
        str(output_path),
        overlay,
        [
            cv2.IMWRITE_JPEG_QUALITY,
            95,
        ],
    )


def save_green_mask(
    green_mask: np.ndarray,
    output_path: Path,
):
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    cv2.imwrite(
        str(output_path),
        green_mask,
    )


# ============================================================
# 8) CSV / SETTINGS OUTPUTS
# ============================================================

def write_csv(
    path: Path,
    fieldnames: list[str],
    rows: list[dict],
):
    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with path.open(
        "w",
        newline="",
        encoding="utf-8",
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=fieldnames,
        )

        writer.writeheader()
        writer.writerows(rows)


def save_settings(path: Path):
    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    settings = {
        "purpose": (
            "Visible green seedling evidence detection "
            "with continuous-day tracking."
        ),
        "green_hue_min": GREEN_HUE_MIN,
        "green_hue_max": GREEN_HUE_MAX,
        "green_saturation_min": GREEN_SATURATION_MIN,
        "green_value_min": GREEN_VALUE_MIN,
        "excess_green_min": EXCESS_GREEN_MIN,
        "green_to_red_ratio": GREEN_TO_RED_RATIO,
        "green_to_blue_ratio": GREEN_TO_BLUE_RATIO,
        "min_green_pixels_per_cell": MIN_GREEN_PIXELS_PER_CELL,
        "min_green_component_pixels": MIN_GREEN_COMPONENT_PIXELS,
        "temporal_rule": (
            "Raw evidence is detected every day. "
            "Tracked emergence remains positive after first raw-positive day."
        ),
    }

    with path.open(
        "w",
        encoding="utf-8",
    ) as file:
        json.dump(
            settings,
            file,
            indent=2,
        )


# ============================================================
# 9) EXCEL REPORT
# ============================================================

def style_sheet(
    worksheet,
    headers: list[str],
    rows: list[list[Any]],
):
    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    thin = Side(
        style="thin",
        color="D9E2F3",
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    for column_number, header in enumerate(headers, start=1):
        cell = worksheet.cell(
            row=1,
            column=column_number,
            value=header,
        )

        cell.fill = header_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = border

    for row_number, row in enumerate(rows, start=2):
        for column_number, value in enumerate(row, start=1):
            cell = worksheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

            cell.border = border

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 34

    for column in worksheet.columns:
        letter = column[0].column_letter

        longest = max(
            len(str(cell.value))
            if cell.value is not None
            else 0
            for cell in column
        )

        worksheet.column_dimensions[letter].width = min(
            max(12, longest + 2),
            58,
        )


def create_excel_report(
    output_path: Path,
    tray_rows: list[dict],
    cell_rows: list[dict],
):
    workbook = Workbook()

    tray_sheet = workbook.active
    tray_sheet.title = "Tray Summary"

    tray_headers = [
        "Day Order",
        "Day",
        "Tray",
        "Tray No",
        "Capture ID",
        "Raw Green Cells",
        "Raw Green %",
        "Tracked Emerged Cells",
        "Tracked Emergence %",
        "Newly Emerged Today",
        "Carried Forward Cells",
        "Overlay Path",
        "Mask Path",
        "Status",
        "Notes",
    ]

    tray_values = [
        [
            row["day_order"],
            row["day"],
            row["tray"],
            row["tray_no"],
            row["capture_id"],
            row["raw_green_cells"],
            row["raw_green_percent"],
            row["tracked_emerged_cells"],
            row["tracked_emergence_percent"],
            row["newly_emerged_today"],
            row["carried_forward_cells"],
            row["overlay_path"],
            row["mask_path"],
            row["status"],
            row["notes"],
        ]
        for row in tray_rows
    ]

    style_sheet(
        tray_sheet,
        tray_headers,
        tray_values,
    )

    for row_number in range(2, len(tray_values) + 2):
        status_cell = tray_sheet.cell(
            row=row_number,
            column=14,
        )

        if status_cell.value == "PASS":
            status_cell.fill = PatternFill(
                "solid",
                fgColor="C6EFCE",
            )
        elif status_cell.value == "SKIPPED":
            status_cell.fill = PatternFill(
                "solid",
                fgColor="F4CCCC",
            )

    cell_sheet = workbook.create_sheet(
        "Cell Results"
    )

    cell_headers = [
        "Day Order",
        "Day",
        "Tray",
        "Tray No",
        "Capture ID",
        "Cell ID",
        "Row",
        "Column",
        "Raw Current Green Evidence",
        "Tracked Visible Emerged",
        "Carried Forward Only",
        "First Visible Emergence Day",
        "First Visible Emergence Day Order",
        "Green Pixels",
        "Largest Green Component",
        "Green Area %",
        "Square Side",
        "Coordinate Source",
    ]

    cell_values = [
        [
            row["day_order"],
            row["day"],
            row["tray"],
            row["tray_no"],
            row["capture_id"],
            row["cell_id"],
            row["row"],
            row["column"],
            row["raw_current_green_evidence"],
            row["tracked_visible_emerged"],
            row["carried_forward_only"],
            row["first_visible_emergence_day"],
            row["first_visible_emergence_day_order"],
            row["green_pixels"],
            row["largest_green_component"],
            row["green_area_percent"],
            row["square_side"],
            row["coordinate_source"],
        ]
        for row in cell_rows
    ]

    style_sheet(
        cell_sheet,
        cell_headers,
        cell_values,
    )

    readme_sheet = workbook.create_sheet("Read Me")

    readme_sheet["A1"] = "Script 04 — Visible Emergence Tracking"
    readme_sheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    notes = [
        "This report measures visible emergence, not underground germination.",
        "Every cell is assessed independently on every available image day.",
        "Raw Current Green Evidence means green evidence was detected in that exact image.",
        "Tracked Visible Emerged means the cell was positive on the current day or on an earlier day.",
        "A cell remains tracked as emerged after its first positive day, even when raw evidence is weak on a later day.",
        "Day 9 is treated as a later follow-up observation, not as a continuous one-day interval after Day 5.",
        "The square zones come from Script 03 and allow seedlings outside the cup rim to remain associated with their original planting cell.",
        "Use overlays and masks to validate the green-evidence threshold before interpreting biological results.",
    ]

    for row_number, note in enumerate(notes, start=3):
        readme_sheet.cell(
            row=row_number,
            column=1,
            value=note,
        )

    readme_sheet.column_dimensions["A"].width = 120

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(output_path)


# ============================================================
# 10) MAIN WORKFLOW
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Script 04: Detect visible green seedling evidence "
            "and apply cumulative emergence tracking."
        )
    )

    parser.add_argument(
        "--days",
        help='Optional example: --days "Day 1,Day 2,Day 3"',
    )

    parser.add_argument(
        "--trays",
        help='Optional example: --trays "Tray 1,Tray 2"',
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="List valid D/RGB and grid combinations only.",
    )

    args = parser.parse_args()

    if not CROP_ROOT.exists():
        print(
            "ERROR: Script 01 crop folder not found:\n"
            f"{CROP_ROOT}"
        )
        return 1

    if not GRID_ROOT.exists():
        print(
            "ERROR: Script 03 output folder not found:\n"
            f"{GRID_ROOT}"
        )
        return 1

    grid_manifest = load_grid_manifest()
    cell_coordinates = load_cell_coordinates()

    if not cell_coordinates:
        print(
            "ERROR: No Script 03 cell coordinate records found."
        )
        return 1

    selected_days = parse_filter_list(args.days)
    selected_trays = parse_filter_list(args.trays)

    day_folders = sorted(
        [
            folder
            for folder in CROP_ROOT.iterdir()
            if folder.is_dir()
            and folder.name.casefold() in DAY_NAME_TO_ORDER
        ],
        key=day_sort_key,
    )

    jobs = []

    for day_folder in day_folders:
        if (
            selected_days
            and day_folder.name.casefold()
            not in selected_days
        ):
            continue

        day_name = day_folder.name
        day_order = DAY_NAME_TO_ORDER.get(
            day_name.casefold(),
            999,
        )

        tray_folders = sorted(
            [
                folder
                for folder in day_folder.iterdir()
                if folder.is_dir()
            ],
            key=lambda folder: natural_key(folder.name),
        )

        for tray_folder in tray_folders:
            if (
                selected_trays
                and tray_folder.name.casefold()
                not in selected_trays
            ):
                continue

            tray_name = tray_folder.name

            for d_image in find_d_images(tray_folder):
                capture_id = d_image["capture_id"]

                key = record_key(
                    day_name,
                    tray_name,
                    capture_id,
                )

                jobs.append(
                    {
                        "day_name": day_name,
                        "day_order": day_order,
                        "tray_name": tray_name,
                        "tray_no": tray_number_from_name(
                            tray_name
                        ),
                        "capture_id": capture_id,
                        "d_path": d_image["path"],
                        "grid_status": grid_manifest.get(
                            key,
                            {},
                        ).get("status", ""),
                        "grid_cells": cell_coordinates.get(
                            key,
                            [],
                        ),
                    }
                )

    if not jobs:
        print(
            "No cropped D/RGB and Script 03 grid combinations found."
        )
        return 1

    jobs.sort(
        key=lambda row: (
            row["day_order"],
            natural_key(row["tray_name"]),
            natural_key(row["capture_id"]),
        )
    )

    print("\nSCRIPT 04 — VISIBLE EMERGENCE TRACKING")
    print("=" * 70)
    print(f"D/RGB input:\n{CROP_ROOT}")
    print(f"\nSquare-grid input:\n{GRID_ROOT}")
    print(f"\nOutput folder:\n{OUTPUT_ROOT}")

    for job in jobs:
        valid_grid = (
            job["grid_status"] in ACCEPTED_GRID_STATUSES
            and len(job["grid_cells"]) == EXPECTED_CELLS
        )

        state = "READY" if valid_grid else "SKIPPED_INVALID_GRID"

        print(
            f"{state}: {job['day_name']} > "
            f"{job['tray_name']} > {job['capture_id']}"
        )

    if args.dry_run:
        print("\nDry run complete. No outputs created.")
        return 0

    settings_path = (
        OUTPUT_ROOT
        / "_config"
        / "emergence_detection_settings.json"
    )

    save_settings(settings_path)

    raw_payloads = {}
    skipped_jobs = []

    # Raw detection is performed for every cell on every day.
    for job in jobs:
        valid_grid = (
            job["grid_status"] in ACCEPTED_GRID_STATUSES
            and len(job["grid_cells"]) == EXPECTED_CELLS
        )

        key = record_key(
            job["day_name"],
            job["tray_name"],
            job["capture_id"],
        )

        if not valid_grid:
            skipped_jobs.append(
                {
                    **job,
                    "reason": (
                        "Grid status was not accepted or did not contain "
                        "exactly 70 square cell zones."
                    ),
                }
            )
            continue

        try:
            rgb = read_rgb(job["d_path"])
            green_mask = build_green_mask(rgb)

            cell_results = []

            for cell in job["grid_cells"]:
                evidence = analyse_cell_green_evidence(
                    green_mask,
                    cell,
                )

                cell_results.append(
                    {
                        **cell,
                        **evidence,
                    }
                )

            raw_payloads[key] = {
                "job": job,
                "rgb": rgb,
                "green_mask": green_mask,
                "cell_results": cell_results,
            }

            print(
                f"RAW DETECTED: {job['day_name']} > "
                f"{job['tray_name']} > {job['capture_id']}"
            )

        except Exception as error:
            skipped_jobs.append(
                {
                    **job,
                    "reason": str(error),
                }
            )

            print(
                f"FAIL: {job['day_name']} > "
                f"{job['tray_name']} > {job['capture_id']} | {error}"
            )

    # Group all valid days by tray to apply cumulative tracking.
    payloads_by_tray = defaultdict(list)

    for payload in raw_payloads.values():
        tray_name = payload["job"]["tray_name"]
        payloads_by_tray[tray_name].append(payload)

    tray_report_rows = []
    cell_report_rows = []

    for tray_name, payloads in payloads_by_tray.items():
        payloads.sort(
            key=lambda payload: (
                payload["job"]["day_order"],
                natural_key(payload["job"]["capture_id"]),
            )
        )

        first_visible_day_by_cell = {}

        for payload in payloads:
            job = payload["job"]
            cell_results = payload["cell_results"]

            raw_count = 0
            tracked_count = 0
            newly_emerged_count = 0
            carried_forward_count = 0

            for cell in cell_results:
                cell_id = cell["cell_id"]

                raw_positive = bool(
                    cell["raw_green_detected"]
                )

                if raw_positive:
                    raw_count += 1

                first_day = first_visible_day_by_cell.get(
                    cell_id
                )

                if raw_positive and first_day is None:
                    first_day = {
                        "day_name": job["day_name"],
                        "day_order": job["day_order"],
                    }

                    first_visible_day_by_cell[cell_id] = first_day
                    newly_emerged_count += 1

                tracked_positive = first_day is not None

                carried_forward_only = (
                    tracked_positive
                    and not raw_positive
                )

                if tracked_positive:
                    tracked_count += 1

                if carried_forward_only:
                    carried_forward_count += 1

                cell["tracked_visible_emerged"] = tracked_positive
                cell["carried_forward_only"] = carried_forward_only

                cell["first_visible_emergence_day"] = (
                    first_day["day_name"]
                    if first_day
                    else ""
                )

                cell[
                    "first_visible_emergence_day_order"
                ] = (
                    first_day["day_order"]
                    if first_day
                    else ""
                )

                cell_report_rows.append(
                    {
                        "day_order": job["day_order"],
                        "day": job["day_name"],
                        "tray": job["tray_name"],
                        "tray_no": job["tray_no"],
                        "capture_id": job["capture_id"],
                        "cell_id": cell["cell_id"],
                        "row": cell["row"],
                        "column": cell["column"],
                        "raw_current_green_evidence": yes_no(
                            raw_positive
                        ),
                        "tracked_visible_emerged": yes_no(
                            tracked_positive
                        ),
                        "carried_forward_only": yes_no(
                            carried_forward_only
                        ),
                        "first_visible_emergence_day": cell[
                            "first_visible_emergence_day"
                        ],
                        "first_visible_emergence_day_order": cell[
                            "first_visible_emergence_day_order"
                        ],
                        "green_pixels": cell["green_pixels"],
                        "largest_green_component": cell[
                            "largest_green_component"
                        ],
                        "green_area_percent": round(
                            float(cell["green_area_percent"]),
                            5,
                        ),
                        "square_side": round(
                            float(cell["square_side"]),
                            3,
                        ),
                        "coordinate_source": cell[
                            "coordinate_source"
                        ],
                        "square_x0": round(
                            float(cell["square_x0"]),
                            3,
                        ),
                        "square_y0": round(
                            float(cell["square_y0"]),
                            3,
                        ),
                        "square_x1": round(
                            float(cell["square_x1"]),
                            3,
                        ),
                        "square_y1": round(
                            float(cell["square_y1"]),
                            3,
                        ),
                        "centre_x": round(
                            float(cell["x"]),
                            3,
                        ),
                        "centre_y": round(
                            float(cell["y"]),
                            3,
                        ),
                    }
                )

            output_folder = (
                OUTPUT_ROOT
                / job["day_name"]
                / job["tray_name"]
            )

            safe_capture_id = safe_name(
                job["capture_id"]
            )

            overlay_path = (
                output_folder
                / f"{safe_capture_id}_visible_emergence_overlay.jpg"
            )

            mask_path = (
                output_folder
                / f"{safe_capture_id}_green_evidence_mask.png"
            )

            overlay_title = (
                f"{job['day_name']} | "
                f"{job['tray_name']} | "
                f"{job['capture_id']}"
            )

            save_emergence_overlay(
                payload["rgb"],
                payload["green_mask"],
                cell_results,
                overlay_path,
                overlay_title,
            )

            save_green_mask(
                payload["green_mask"],
                mask_path,
            )

            tray_report_rows.append(
                {
                    "day_order": job["day_order"],
                    "day": job["day_name"],
                    "tray": job["tray_name"],
                    "tray_no": job["tray_no"],
                    "capture_id": job["capture_id"],
                    "raw_green_cells": raw_count,
                    "raw_green_percent": round(
                        raw_count
                        / EXPECTED_CELLS
                        * 100.0,
                        3,
                    ),
                    "tracked_emerged_cells": tracked_count,
                    "tracked_emergence_percent": round(
                        tracked_count
                        / EXPECTED_CELLS
                        * 100.0,
                        3,
                    ),
                    "newly_emerged_today": newly_emerged_count,
                    "carried_forward_cells": carried_forward_count,
                    "overlay_path": relative_path(
                        overlay_path,
                        OUTPUT_ROOT,
                    ),
                    "mask_path": relative_path(
                        mask_path,
                        OUTPUT_ROOT,
                    ),
                    "status": "PASS",
                    "notes": (
                        "Raw detection completed for this image. "
                        "Tracked emergence carries previous emergence "
                        "forward without skipping current-day detection."
                    ),
                }
            )

            print(
                f"TRACKED: {job['day_name']} > "
                f"{job['tray_name']} | "
                f"raw={raw_count}/70 | "
                f"tracked={tracked_count}/70 | "
                f"new={newly_emerged_count}"
            )

    # Add invalid or failed jobs to tray-level report.
    for skipped in skipped_jobs:
        tray_report_rows.append(
            {
                "day_order": skipped["day_order"],
                "day": skipped["day_name"],
                "tray": skipped["tray_name"],
                "tray_no": skipped["tray_no"],
                "capture_id": skipped["capture_id"],
                "raw_green_cells": "",
                "raw_green_percent": "",
                "tracked_emerged_cells": "",
                "tracked_emergence_percent": "",
                "newly_emerged_today": "",
                "carried_forward_cells": "",
                "overlay_path": "",
                "mask_path": "",
                "status": "SKIPPED",
                "notes": skipped["reason"],
            }
        )

    tray_report_rows.sort(
        key=lambda row: (
            sortable_number(row["day_order"]),
            natural_key(row["tray"]),
            natural_key(row["capture_id"]),
        )
    )

    cell_report_rows.sort(
        key=lambda row: (
            sortable_number(row["day_order"]),
            natural_key(row["tray"]),
            sortable_number(row["cell_id"]),
        )
    )

    report_folder = OUTPUT_ROOT / "_reports"

    tray_csv_path = (
        report_folder
        / "visible_emergence_tray_summary.csv"
    )

    cell_csv_path = (
        report_folder
        / "visible_emergence_cell_results.csv"
    )

    report_xlsx_path = (
        report_folder
        / "visible_emergence_report.xlsx"
    )

    tray_fields = [
        "day_order",
        "day",
        "tray",
        "tray_no",
        "capture_id",
        "raw_green_cells",
        "raw_green_percent",
        "tracked_emerged_cells",
        "tracked_emergence_percent",
        "newly_emerged_today",
        "carried_forward_cells",
        "overlay_path",
        "mask_path",
        "status",
        "notes",
    ]

    cell_fields = [
        "day_order",
        "day",
        "tray",
        "tray_no",
        "capture_id",
        "cell_id",
        "row",
        "column",
        "raw_current_green_evidence",
        "tracked_visible_emerged",
        "carried_forward_only",
        "first_visible_emergence_day",
        "first_visible_emergence_day_order",
        "green_pixels",
        "largest_green_component",
        "green_area_percent",
        "square_side",
        "coordinate_source",
        "square_x0",
        "square_y0",
        "square_x1",
        "square_y1",
        "centre_x",
        "centre_y",
    ]

    write_csv(
        tray_csv_path,
        tray_fields,
        tray_report_rows,
    )

    write_csv(
        cell_csv_path,
        cell_fields,
        cell_report_rows,
    )

    create_excel_report(
        report_xlsx_path,
        tray_report_rows,
        cell_report_rows,
    )

    pass_count = sum(
        row["status"] == "PASS"
        for row in tray_report_rows
    )

    skipped_count = sum(
        row["status"] == "SKIPPED"
        for row in tray_report_rows
    )

    print("\n" + "=" * 70)
    print("SCRIPT 04 FINISHED")
    print("=" * 70)
    print(f"PASS: {pass_count}")
    print(f"SKIPPED: {skipped_count}")
    print(f"\nExcel report:\n{report_xlsx_path}")
    print(f"\nTray summary:\n{tray_csv_path}")
    print(f"\nCell results:\n{cell_csv_path}")
    print(f"\nOverlays and masks:\n{OUTPUT_ROOT}")

    return 0 if skipped_count == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())