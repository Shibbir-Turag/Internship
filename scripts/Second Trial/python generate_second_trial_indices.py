from __future__ import annotations

import argparse
import csv
import os
import re
from pathlib import Path

import numpy as np
import tifffile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont


# ============================================================
# SETTINGS
# ============================================================

# Leave this as None if your OneDrive path is detected normally.
# If the script cannot find your Internship folder, set it manually:
#
# INTERNSHIP_ROOT = Path(
#     r"C:\Users\tshib\OneDrive\Desktop\Internship"
# )

INTERNSHIP_ROOT = None

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".tif", ".tiff"}

ANALYSIS_FOLDER = "Analysis"

NODATA_VALUE = -9999.0
EPSILON = 1e-10


# ============================================================
# PATH HELPERS
# ============================================================

def get_internship_root() -> Path:
    """Find OneDrive/Desktop/Internship automatically."""

    if INTERNSHIP_ROOT is not None:
        return Path(INTERNSHIP_ROOT)

    candidates = []

    if os.environ.get("OneDrive"):
        candidates.append(
            Path(os.environ["OneDrive"]) / "Desktop" / "Internship"
        )

    candidates.append(Path.home() / "OneDrive" / "Desktop" / "Internship")
    candidates.append(Path.home() / "Desktop" / "Internship")

    for candidate in candidates:
        if (candidate / "outputs" / "Second Trial").exists():
            return candidate

    return candidates[0]


def natural_sort_key(text: str):
    """Sort Tray 2 before Tray 10."""
    return [
        int(part) if part.isdigit() else part.casefold()
        for part in re.split(r"(\d+)", text)
    ]


def parse_filter_list(value: str | None):
    """Convert comma-separated command-line filters into lowercase sets."""

    if not value:
        return None

    return {
        item.strip().lower()
        for item in value.split(",")
        if item.strip()
    }


def get_tray_number(tray_name: str):
    """Extract number from labels such as Tray 1."""
    match = re.search(r"(\d+)", tray_name)
    return int(match.group(1)) if match else None


# ============================================================
# FILE CLASSIFICATION
# ============================================================

def has_filename_token(stem: str, token: str) -> bool:
    """Match DJI-style underscore-delimited filename tokens."""

    pattern = rf"(?:^|_){re.escape(token)}(?:_|$)"
    return re.search(pattern, stem.upper()) is not None


def classify_image(filename: str):
    """
    Return capture ID and band name.

    Examples:
        DJI_0001_MS_NIR.TIF -> DJI_0001, MS_NIR
        DJI_0001_D.JPG      -> DJI_0001, D

    F preview files are ignored.
    """

    stem = Path(filename).stem.upper()

    for token, band in [
        ("MS_NIR", "MS_NIR"),
        ("MS_RE", "MS_RE"),
        ("MS_R", "MS_R"),
        ("MS_G", "MS_G"),
        ("D", "D"),
        ("F", "F_IGNORED"),
    ]:
        if has_filename_token(stem, token):
            capture_id = re.sub(
                rf"_{re.escape(token)}$",
                "",
                stem,
            )

            return capture_id if capture_id else stem, band

    return stem, "UNKNOWN"


# ============================================================
# TRAY STATUS EXCEL
# ============================================================

def is_present(value) -> bool:
    """Treat P as a positive tray-status marker."""
    return str(value).strip().upper() == "P"


def read_tray_status(xlsx_path: Path):
    """
    Expected columns:
    Tray No | Microbes | No Microbes | Inside | Outside
    """

    if not xlsx_path.exists():
        print(f"WARNING: Tray Status.xlsx not found: {xlsx_path}")
        return {}

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        return {}

    headers = {
        str(value).strip().lower(): index
        for index, value in enumerate(rows[0])
        if value is not None
    }

    required = [
        "tray no",
        "microbes",
        "no microbes",
        "inside",
        "outside",
    ]

    if any(name not in headers for name in required):
        print("WARNING: Tray Status.xlsx does not have expected columns.")
        return {}

    tray_metadata = {}

    for row in rows[1:]:
        if not row or row[headers["tray no"]] is None:
            continue

        try:
            tray_number = int(row[headers["tray no"]])
        except (TypeError, ValueError):
            continue

        microbes = is_present(row[headers["microbes"]])
        no_microbes = is_present(row[headers["no microbes"]])
        inside = is_present(row[headers["inside"]])
        outside = is_present(row[headers["outside"]])

        treatment = (
            "Microbes"
            if microbes and not no_microbes
            else "No Microbes"
            if no_microbes and not microbes
            else "Unclear"
        )

        environment = (
            "Inside"
            if inside and not outside
            else "Outside"
            if outside and not inside
            else "Unclear"
        )

        tray_metadata[tray_number] = {
            "tray_no": tray_number,
            "treatment": treatment,
            "environment": environment,
        }

    return tray_metadata


# ============================================================
# IMAGE READING
# ============================================================

def read_single_band_tiff(image_path: Path) -> np.ndarray:
    """Read a single-band TIFF as float32."""

    array = tifffile.imread(image_path)

    if array.ndim == 2:
        return array.astype(np.float32)

    if array.ndim == 3 and array.shape[-1] == 1:
        return array[..., 0].astype(np.float32)

    if array.ndim == 3 and array.shape[0] == 1:
        return array[0, ...].astype(np.float32)

    raise ValueError(
        f"{image_path.name} must be a single-band TIFF. "
        f"Found shape: {array.shape}"
    )


def read_d_preview(image_path: Path | None):
    """
    Read D.JPG only for visual reference in preview PNGs.
    It is not used for NDVI or NDRE.
    """

    if image_path is None or not image_path.exists():
        return None

    try:
        with Image.open(image_path) as image:
            return image.convert("RGB")
    except Exception:
        return None


# ============================================================
# NDVI / NDRE CALCULATIONS
# ============================================================

def calculate_indices(
    red: np.ndarray,
    red_edge: np.ndarray,
    nir: np.ndarray,
):
    """
    NDVI = (NIR - Red) / (NIR + Red)
    NDRE = (NIR - Red Edge) / (NIR + Red Edge)
    """

    if red.shape != red_edge.shape or red.shape != nir.shape:
        raise ValueError(
            "Multispectral band dimensions do not match. "
            f"R={red.shape}, RE={red_edge.shape}, NIR={nir.shape}"
        )

    valid = (
        np.isfinite(red)
        & np.isfinite(red_edge)
        & np.isfinite(nir)
        & (red >= 0)
        & (red_edge >= 0)
        & (nir >= 0)
        & ((nir + red) > EPSILON)
        & ((nir + red_edge) > EPSILON)
    )

    ndvi = np.full(nir.shape, NODATA_VALUE, dtype=np.float32)
    ndre = np.full(nir.shape, NODATA_VALUE, dtype=np.float32)

    ndvi[valid] = (
        (nir[valid] - red[valid])
        / (nir[valid] + red[valid])
    )

    ndre[valid] = (
        (nir[valid] - red_edge[valid])
        / (nir[valid] + red_edge[valid])
    )

    return ndvi, ndre, valid


def calculate_stats(index_array: np.ndarray):
    """Calculate stats excluding nodata values."""

    values = index_array[
        np.isfinite(index_array)
        & (index_array != NODATA_VALUE)
    ]

    if values.size == 0:
        return {
            "n": "",
            "mean": "",
            "median": "",
            "min": "",
            "max": "",
            "sd": "",
            "p10": "",
            "p90": "",
        }

    return {
        "n": int(values.size),
        "mean": float(np.mean(values)),
        "median": float(np.median(values)),
        "min": float(np.min(values)),
        "max": float(np.max(values)),
        "sd": float(np.std(values)),
        "p10": float(np.percentile(values, 10)),
        "p90": float(np.percentile(values, 90)),
    }


def save_index_tiff(
    output_path: Path,
    index_array: np.ndarray,
    description: str,
):
    """Save float32 NDVI/NDRE raster."""

    output_path.parent.mkdir(parents=True, exist_ok=True)

    tifffile.imwrite(
        output_path,
        index_array.astype(np.float32),
        description=description,
        metadata={"axes": "YX"},
    )


# ============================================================
# PREVIEW PNG FUNCTIONS
# ============================================================

def make_index_colour_preview(index_array: np.ndarray) -> Image.Image:
    """
    Create visual-only colour map.

    -1.0 = blue
     0.0 = yellow
     1.0 = green
    nodata = black
    """

    valid = (
        np.isfinite(index_array)
        & (index_array != NODATA_VALUE)
    )

    scaled = np.clip((index_array + 1.0) / 2.0, 0, 1)

    output = np.zeros(
        (*index_array.shape, 3),
        dtype=np.uint8,
    )

    low_mask = valid & (scaled <= 0.5)
    high_mask = valid & (scaled > 0.5)

    low = np.zeros_like(scaled)
    low[low_mask] = scaled[low_mask] / 0.5

    output[..., 0][low_mask] = (
        245 * low[low_mask]
    ).astype(np.uint8)

    output[..., 1][low_mask] = (
        40 + 195 * low[low_mask]
    ).astype(np.uint8)

    output[..., 2][low_mask] = (
        150 - 30 * low[low_mask]
    ).astype(np.uint8)

    high = np.zeros_like(scaled)
    high[high_mask] = (
        (scaled[high_mask] - 0.5) / 0.5
    )

    output[..., 0][high_mask] = (
        245 * (1 - high[high_mask])
    ).astype(np.uint8)

    output[..., 1][high_mask] = (
        235 - 125 * high[high_mask]
    ).astype(np.uint8)

    output[..., 2][high_mask] = (
        120 - 85 * high[high_mask]
    ).astype(np.uint8)

    return Image.fromarray(output, "RGB")


def make_panel(
    image: Image.Image | None,
    width: int,
    height: int,
    title: str,
):
    """Fit image into a labelled preview panel."""

    panel = Image.new("RGB", (width, height), "white")

    draw = ImageDraw.Draw(panel)
    font = ImageFont.load_default()

    draw.text((10, 8), title, fill="black", font=font)

    if image is None:
        draw.text(
            (10, 35),
            "Not available",
            fill="black",
            font=font,
        )
        return panel

    image_copy = image.copy()

    image_copy.thumbnail((width - 12, height - 42))

    x = (width - image_copy.width) // 2
    y = 30 + (height - 30 - image_copy.height) // 2

    panel.paste(image_copy, (x, y))

    return panel


def create_preview(
    output_path: Path,
    day_name: str,
    tray_name: str,
    capture_id: str,
    d_image: Image.Image | None,
    ndvi: np.ndarray,
    ndre: np.ndarray,
    ndvi_stats: dict,
    ndre_stats: dict,
):
    """Create D/RGB + NDVI + NDRE preview PNG."""

    panel_width = 420
    panel_height = 330
    header_height = 48

    canvas = Image.new(
        "RGB",
        (panel_width * 3, panel_height + header_height),
        "#ECECEC",
    )

    draw = ImageDraw.Draw(canvas)
    font = ImageFont.load_default()

    draw.text(
        (10, 15),
        f"{day_name} | {tray_name} | {capture_id}",
        fill="black",
        font=font,
    )

    d_panel = make_panel(
        d_image,
        panel_width,
        panel_height,
        "D / RGB reference",
    )

    ndvi_title = (
        f"NDVI | mean={ndvi_stats['mean']:.4f}"
        if ndvi_stats["n"] != ""
        else "NDVI | no valid pixels"
    )

    ndre_title = (
        f"NDRE | mean={ndre_stats['mean']:.4f}"
        if ndre_stats["n"] != ""
        else "NDRE | no valid pixels"
    )

    ndvi_panel = make_panel(
        make_index_colour_preview(ndvi),
        panel_width,
        panel_height,
        ndvi_title,
    )

    ndre_panel = make_panel(
        make_index_colour_preview(ndre),
        panel_width,
        panel_height,
        ndre_title,
    )

    canvas.paste(d_panel, (0, header_height))
    canvas.paste(ndvi_panel, (panel_width, header_height))
    canvas.paste(ndre_panel, (panel_width * 2, header_height))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path)


# ============================================================
# EXCEL REPORT
# ============================================================

def create_excel_report(report_path: Path, rows):
    """Create the NDVI/NDRE summary workbook."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Index Summary"

    headers = [
        "Day",
        "Tray",
        "Tray No",
        "Treatment",
        "Environment",
        "Capture Set",
        "Total Pixels",
        "Valid Pixels",
        "Valid %",
        "NDVI Mean",
        "NDVI Median",
        "NDVI Min",
        "NDVI Max",
        "NDVI SD",
        "NDVI P10",
        "NDVI P90",
        "NDRE Mean",
        "NDRE Median",
        "NDRE Min",
        "NDRE Max",
        "NDRE SD",
        "NDRE P10",
        "NDRE P90",
        "NDVI TIFF",
        "NDRE TIFF",
        "Preview PNG",
        "Status",
        "Notes",
    ]

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column, header)

        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = border

    for row_index, row in enumerate(rows, start=2):
        values = [
            row["day"],
            row["tray"],
            row["tray_no"],
            row["treatment"],
            row["environment"],
            row["capture_id"],
            row["total_pixels"],
            row["valid_pixels"],
            row["valid_percent"],
            row["ndvi"]["mean"],
            row["ndvi"]["median"],
            row["ndvi"]["min"],
            row["ndvi"]["max"],
            row["ndvi"]["sd"],
            row["ndvi"]["p10"],
            row["ndvi"]["p90"],
            row["ndre"]["mean"],
            row["ndre"]["median"],
            row["ndre"]["min"],
            row["ndre"]["max"],
            row["ndre"]["sd"],
            row["ndre"]["p10"],
            row["ndre"]["p90"],
            row["ndvi_path"],
            row["ndre_path"],
            row["preview_path"],
            row["status"],
            row["notes"],
        ]

        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

            cell.border = border

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 34

    for column in sheet.columns:
        letter = column[0].column_letter

        longest = max(
            len(str(cell.value))
            if cell.value is not None
            else 0
            for cell in column
        )

        sheet.column_dimensions[letter].width = min(
            max(12, longest + 2),
            48,
        )

    notes_sheet = workbook.create_sheet("Read Me")

    notes_sheet["A1"] = "Second Trial NDVI / NDRE Notes"
    notes_sheet["A1"].font = Font(bold=True, size=14)

    notes = [
        "NDVI = (NIR - Red) / (NIR + Red).",
        "NDRE = (NIR - Red Edge) / (NIR + Red Edge).",
        "F preview files are not used.",
        "D/RGB images are visual references only.",
        "These are unmasked tray-wide index maps, including soil/background.",
        "Do not use tray-wide means as final plant-only treatment findings.",
        "The next step is vegetation masking, canopy cover, germination estimation, and plant-only index summaries.",
        "Compare days carefully because light and acquisition conditions can vary.",
    ]

    for row_index, note in enumerate(notes, start=3):
        notes_sheet.cell(row_index, 1, note)

    notes_sheet.column_dimensions["A"].width = 120

    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(report_path)


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generate NDVI and NDRE from Second Trial crops."
    )

    parser.add_argument(
        "--days",
        help='Optional example: --days "First Day,Second Day"',
    )

    parser.add_argument(
        "--trays",
        help='Optional example: --trays "Tray 1,Tray 2"',
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing NDVI/NDRE outputs.",
    )

    parser.add_argument(
        "--no-previews",
        action="store_true",
        help="Do not create preview PNG files.",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Only show recognised capture sets.",
    )

    args = parser.parse_args()

    internship_root = get_internship_root()

    cropped_root = (
        internship_root
        / "outputs"
        / "Second Trial"
    )

    analysis_root = cropped_root / ANALYSIS_FOLDER

    tray_status_path = (
        internship_root
        / "data"
        / "Second Trial"
        / "Tray Status.xlsx"
    )

    if not cropped_root.exists():
        print(f"ERROR: Cropped outputs not found:\n{cropped_root}")
        return 1

    selected_days = parse_filter_list(args.days)
    selected_trays = parse_filter_list(args.trays)

    metadata = read_tray_status(tray_status_path)

    excluded_folders = {"QA", ANALYSIS_FOLDER}

    work_items = []

    day_folders = sorted(
        [
            folder
            for folder in cropped_root.iterdir()
            if folder.is_dir()
            and folder.name not in excluded_folders
        ],
        key=lambda folder: natural_sort_key(folder.name),
    )

    for day_folder in day_folders:
        if selected_days and day_folder.name.lower() not in selected_days:
            continue

        tray_folders = sorted(
            [
                folder
                for folder in day_folder.iterdir()
                if folder.is_dir()
            ],
            key=lambda folder: natural_sort_key(folder.name),
        )

        for tray_folder in tray_folders:
            if selected_trays and tray_folder.name.lower() not in selected_trays:
                continue

            capture_sets = {}

            image_files = sorted(
                [
                    file
                    for file in tray_folder.iterdir()
                    if file.is_file()
                    and file.suffix.lower() in IMAGE_EXTENSIONS
                ],
                key=lambda file: natural_sort_key(file.name),
            )

            for image_file in image_files:
                capture_id, band = classify_image(image_file.name)

                if band in {"UNKNOWN", "F_IGNORED"}:
                    continue

                capture_sets.setdefault(capture_id, {})
                capture_sets[capture_id].setdefault(
                    band,
                    [],
                ).append(image_file)

            for capture_id, band_files in capture_sets.items():
                paths = {
                    band: files[0]
                    for band, files in band_files.items()
                }

                missing = [
                    band
                    for band in ["MS_R", "MS_RE", "MS_NIR"]
                    if band not in paths
                ]

                work_items.append(
                    {
                        "day_folder": day_folder,
                        "tray_folder": tray_folder,
                        "capture_id": capture_id,
                        "paths": paths,
                        "missing": missing,
                    }
                )

    if not work_items:
        print("No recognised capture sets found.")
        return 1

    print("\nDETECTED CAPTURE SETS")

    for item in work_items:
        state = (
            "READY"
            if not item["missing"]
            else "MISSING: " + ", ".join(item["missing"])
        )

        print(
            f"{item['day_folder'].name} > "
            f"{item['tray_folder'].name} > "
            f"{item['capture_id']} : {state}"
        )

    if args.dry_run:
        ready_count = sum(
            1
            for item in work_items
            if not item["missing"]
        )

        print(f"\nDry run complete. Ready sets: {ready_count}")
        return 0

    analysis_root.mkdir(parents=True, exist_ok=True)

    report_rows = []

    for item in work_items:
        day_name = item["day_folder"].name
        tray_name = item["tray_folder"].name
        capture_id = item["capture_id"]
        paths = item["paths"]

        number = get_tray_number(tray_name)

        tray_info = metadata.get(
            number,
            {
                "tray_no": number if number else "",
                "treatment": "Unknown",
                "environment": "Unknown",
            },
        )

        row = {
            "day": day_name,
            "tray": tray_name,
            "tray_no": tray_info["tray_no"],
            "treatment": tray_info["treatment"],
            "environment": tray_info["environment"],
            "capture_id": capture_id,
            "total_pixels": "",
            "valid_pixels": "",
            "valid_percent": "",
            "ndvi": {
                "n": "",
                "mean": "",
                "median": "",
                "min": "",
                "max": "",
                "sd": "",
                "p10": "",
                "p90": "",
            },
            "ndre": {
                "n": "",
                "mean": "",
                "median": "",
                "min": "",
                "max": "",
                "sd": "",
                "p10": "",
                "p90": "",
            },
            "ndvi_path": "",
            "ndre_path": "",
            "preview_path": "",
            "status": "",
            "notes": "",
        }

        if item["missing"]:
            row["status"] = "FAIL"
            row["notes"] = (
                "Missing required bands: "
                + ", ".join(item["missing"])
            )

            report_rows.append(row)
            continue

        try:
            red = read_single_band_tiff(paths["MS_R"])
            red_edge = read_single_band_tiff(paths["MS_RE"])
            nir = read_single_band_tiff(paths["MS_NIR"])

            ndvi, ndre, valid_mask = calculate_indices(
                red,
                red_edge,
                nir,
            )

            ndvi_stats = calculate_stats(ndvi)
            ndre_stats = calculate_stats(ndre)

            safe_capture_id = re.sub(
                r"[^A-Za-z0-9_.-]+",
                "_",
                capture_id,
            )

            output_folder = (
                analysis_root
                / day_name
                / tray_name
            )

            ndvi_path = (
                output_folder
                / f"{safe_capture_id}_NDVI.tif"
            )

            ndre_path = (
                output_folder
                / f"{safe_capture_id}_NDRE.tif"
            )

            preview_path = (
                output_folder
                / f"{safe_capture_id}_index_preview.png"
            )

            if args.overwrite or not ndvi_path.exists():
                save_index_tiff(
                    ndvi_path,
                    ndvi,
                    (
                        "NDVI generated from "
                        f"NIR={paths['MS_NIR'].name}; "
                        f"Red={paths['MS_R'].name}; "
                        f"Nodata={NODATA_VALUE}"
                    ),
                )

            if args.overwrite or not ndre_path.exists():
                save_index_tiff(
                    ndre_path,
                    ndre,
                    (
                        "NDRE generated from "
                        f"NIR={paths['MS_NIR'].name}; "
                        f"RedEdge={paths['MS_RE'].name}; "
                        f"Nodata={NODATA_VALUE}"
                    ),
                )

            if not args.no_previews:
                if args.overwrite or not preview_path.exists():
                    create_preview(
                        preview_path,
                        day_name,
                        tray_name,
                        capture_id,
                        read_d_preview(paths.get("D")),
                        ndvi,
                        ndre,
                        ndvi_stats,
                        ndre_stats,
                    )

            row.update(
                {
                    "total_pixels": int(ndvi.size),
                    "valid_pixels": int(valid_mask.sum()),
                    "valid_percent": float(
                        valid_mask.mean() * 100
                    ),
                    "ndvi": ndvi_stats,
                    "ndre": ndre_stats,
                    "ndvi_path": str(
                        ndvi_path.relative_to(cropped_root)
                    ),
                    "ndre_path": str(
                        ndre_path.relative_to(cropped_root)
                    ),
                    "preview_path": (
                        ""
                        if args.no_previews
                        else str(
                            preview_path.relative_to(cropped_root)
                        )
                    ),
                    "status": "PASS",
                    "notes": (
                        "Unmasked relative index maps created. "
                        "Use vegetation masking next for plant-only metrics."
                    ),
                }
            )

            print(
                f"CREATED: {day_name} > {tray_name} > {capture_id} | "
                f"NDVI={ndvi_stats['mean']:.4f}, "
                f"NDRE={ndre_stats['mean']:.4f}"
            )

        except Exception as error:
            row["status"] = "FAIL"
            row["notes"] = str(error)

            print(
                f"FAILED: {day_name} > {tray_name} > "
                f"{capture_id} | {error}"
            )

        report_rows.append(row)

    create_excel_report(
        analysis_root / "second_trial_index_summary.xlsx",
        report_rows,
    )

    csv_fields = [
        "day",
        "tray",
        "tray_no",
        "treatment",
        "environment",
        "capture_id",
        "total_pixels",
        "valid_pixels",
        "valid_percent",
        "ndvi_mean",
        "ndre_mean",
        "ndvi_path",
        "ndre_path",
        "preview_path",
        "status",
        "notes",
    ]

    csv_path = (
        analysis_root
        / "second_trial_index_manifest.csv"
    )

    with open(
        csv_path,
        "w",
        newline="",
        encoding="utf-8",
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=csv_fields,
        )

        writer.writeheader()

        for row in report_rows:
            writer.writerow(
                {
                    "day": row["day"],
                    "tray": row["tray"],
                    "tray_no": row["tray_no"],
                    "treatment": row["treatment"],
                    "environment": row["environment"],
                    "capture_id": row["capture_id"],
                    "total_pixels": row["total_pixels"],
                    "valid_pixels": row["valid_pixels"],
                    "valid_percent": row["valid_percent"],
                    "ndvi_mean": row["ndvi"]["mean"],
                    "ndre_mean": row["ndre"]["mean"],
                    "ndvi_path": row["ndvi_path"],
                    "ndre_path": row["ndre_path"],
                    "preview_path": row["preview_path"],
                    "status": row["status"],
                    "notes": row["notes"],
                }
            )

    passed = sum(
        1
        for row in report_rows
        if row["status"] == "PASS"
    )

    failed = len(report_rows) - passed

    print("\nINDEX GENERATION FINISHED")
    print(f"PASS: {passed}")
    print(f"FAIL: {failed}")

    print(
        "\nExcel summary:\n"
        f"{analysis_root / 'second_trial_index_summary.xlsx'}"
    )

    print(f"\nAnalysis folder:\n{analysis_root}")

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())