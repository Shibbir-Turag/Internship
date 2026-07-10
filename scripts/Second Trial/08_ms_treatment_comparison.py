from __future__ import annotations

"""
SCRIPT 08 — MULTISPECTRAL TREATMENT COMPARISON

Purpose
-------
Compare Script 07 relative NDVI / NDRE results by:
- Microbes vs No Microbes
- No Microbes | Inside vs No Microbes | Outside
- Microbes | Inside vs Microbes | Outside
- Individual tray-level NDVI / NDRE growth-rate ranking

Day 1 to Day 5 is treated as the continuous growth-rate window:
    rate = (Day 5 - Day 1) / 4

Day 9 is treated as a later follow-up observation, not as one day after Day 5.

This script creates a PDF visual report, not an HTML report.
"""

import argparse
import json
import math
import re
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image as PDFImage,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


# ============================================================
# 1) PATHS — CHANGE PROJECT_ROOT ONLY WHEN REUSING
# ============================================================

PROJECT_ROOT = Path(r"C:\Users\tshib\OneDrive\Desktop\Internship")

TRAY_STATUS_XLSX = (
    PROJECT_ROOT
    / "data"
    / "Second Trial"
    / "Tray Status.xlsx"
)

SCRIPT07_REPORTS = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "07_MS_Vegetation_Indices"
    / "_reports"
)

MS_TRAY_SUMMARY_CSV = (
    SCRIPT07_REPORTS
    / "ms_index_tray_summary.csv"
)

OUTPUT_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "08_MS_Treatment_Comparison"
)

CHARTS_ROOT = OUTPUT_ROOT / "charts"
REPORTS_ROOT = OUTPUT_ROOT / "_reports"
CONFIG_ROOT = OUTPUT_ROOT / "_config"


# ============================================================
# 2) SETTINGS
# ============================================================

EXPECTED_TRAYS = 8
EXPECTED_DAYS = [1, 2, 3, 4, 5, 9]

DAY_LABELS = {
    1: "Day 1",
    2: "Day 2",
    3: "Day 3",
    4: "Day 4",
    5: "Day 5",
    9: "Day 9",
}

TREATMENT_ORDER = [
    "No Microbes",
    "Microbes",
]

INTERACTION_ORDER = [
    "No Microbes | Inside",
    "No Microbes | Outside",
    "Microbes | Inside",
    "Microbes | Outside",
]

ACCEPTED_SCRIPT07_STATUSES = {
    "PASS",
    "CHECK",
}


# ============================================================
# 3) BASIC HELPERS
# ============================================================

def normalise(value: object) -> str:
    return re.sub(
        r"[^a-z0-9]+",
        "",
        str(value).casefold(),
    )


def is_present(value: object) -> bool:
    return str(value).strip().casefold() in {
        "p",
        "yes",
        "y",
        "true",
        "1",
    }


def require_columns(
    dataframe: pd.DataFrame,
    required_columns: list[str],
    source_name: str,
) -> None:
    missing = [
        column
        for column in required_columns
        if column not in dataframe.columns
    ]

    if missing:
        raise ValueError(
            f"{source_name} is missing required column(s): "
            + ", ".join(missing)
        )


def group_sort_order(group_type: str, group_name: str) -> int:
    if group_type == "Treatment":
        if group_name in TREATMENT_ORDER:
            return TREATMENT_ORDER.index(group_name)
        return 999

    if group_name in INTERACTION_ORDER:
        return INTERACTION_ORDER.index(group_name)

    return 999


def safe_round_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = 4,
) -> pd.DataFrame:
    output = dataframe.copy()

    numeric_columns = output.select_dtypes(
        include=["number"]
    ).columns

    output[numeric_columns] = output[numeric_columns].round(decimals)

    return output


def safe_idxmax_row(
    dataframe: pd.DataFrame,
    column: str,
):
    valid = dataframe.dropna(subset=[column])

    if valid.empty:
        return None

    return valid.loc[valid[column].idxmax()]


def value_for_day(
    rows_by_day: dict[int, pd.Series],
    day_number: int,
    metric: str,
):
    row = rows_by_day.get(day_number)

    if row is None:
        return math.nan

    return row.get(metric, math.nan)


def format_leader(row, metric: str) -> str:
    if row is None:
        return "not available"

    value = row.get(metric, math.nan)

    if pd.isna(value):
        return str(row["group"])

    return f"{row['group']} ({value:.4f})"


# ============================================================
# 4) LOAD TRAY DESIGN
# ============================================================

def load_tray_design() -> pd.DataFrame:
    if not TRAY_STATUS_XLSX.exists():
        raise FileNotFoundError(
            f"Tray Status workbook not found:\n{TRAY_STATUS_XLSX}"
        )

    raw = pd.read_excel(TRAY_STATUS_XLSX)

    header_lookup = {
        normalise(column): column
        for column in raw.columns
    }

    required = {
        "trayno": "Tray No",
        "microbes": "Microbes",
        "nomicrobes": "No Microbes",
        "inside": "Inside",
        "outside": "Outside",
    }

    missing = [
        display
        for key, display in required.items()
        if key not in header_lookup
    ]

    if missing:
        raise ValueError(
            "Tray Status.xlsx is missing required column(s): "
            + ", ".join(missing)
        )

    tray_col = header_lookup["trayno"]
    microbes_col = header_lookup["microbes"]
    no_microbes_col = header_lookup["nomicrobes"]
    inside_col = header_lookup["inside"]
    outside_col = header_lookup["outside"]

    records = []

    for _, row in raw.iterrows():
        tray_no = pd.to_numeric(
            row[tray_col],
            errors="coerce",
        )

        if pd.isna(tray_no):
            continue

        tray_no = int(tray_no)

        microbes = is_present(row[microbes_col])
        no_microbes = is_present(row[no_microbes_col])
        inside = is_present(row[inside_col])
        outside = is_present(row[outside_col])

        if microbes == no_microbes:
            raise ValueError(
                f"Tray {tray_no}: exactly one of Microbes / "
                "No Microbes must be marked."
            )

        if inside == outside:
            raise ValueError(
                f"Tray {tray_no}: exactly one of Inside / "
                "Outside must be marked."
            )

        treatment = "Microbes" if microbes else "No Microbes"
        environment = "Inside" if inside else "Outside"

        records.append(
            {
                "tray_no": tray_no,
                "treatment": treatment,
                "environment": environment,
                "interaction": f"{treatment} | {environment}",
            }
        )

    design = (
        pd.DataFrame(records)
        .drop_duplicates(subset=["tray_no"])
        .sort_values("tray_no")
        .reset_index(drop=True)
    )

    if design.empty:
        raise ValueError(
            "No valid tray design records were found."
        )

    return design


# ============================================================
# 5) LOAD SCRIPT 07 RESULTS
# ============================================================

def load_script07_tray_summary() -> pd.DataFrame:
    if not MS_TRAY_SUMMARY_CSV.exists():
        raise FileNotFoundError(
            f"Script 07 tray summary not found:\n{MS_TRAY_SUMMARY_CSV}"
        )

    dataframe = pd.read_csv(MS_TRAY_SUMMARY_CSV)

    require_columns(
        dataframe,
        [
            "day_order",
            "day",
            "tray",
            "tray_no",
            "capture_id",
            "cells_processed",
            "cells_with_valid_ndvi",
            "cells_with_valid_ndre",
            "mean_cell_ndvi",
            "median_cell_ndvi",
            "mean_cell_ndre",
            "median_cell_ndre",
            "status",
        ],
        "ms_index_tray_summary.csv",
    )

    dataframe["status"] = dataframe["status"].astype(str).str.upper()

    dataframe = dataframe.loc[
        dataframe["status"].isin(ACCEPTED_SCRIPT07_STATUSES)
    ].copy()

    if dataframe.empty:
        raise ValueError(
            "No PASS or CHECK rows found in Script 07 tray summary."
        )

    numeric_columns = [
        "day_order",
        "tray_no",
        "cells_processed",
        "cells_with_valid_ndvi",
        "cells_with_valid_ndre",
        "mean_cell_ndvi",
        "median_cell_ndvi",
        "mean_cell_ndre",
        "median_cell_ndre",
    ]

    for column in numeric_columns:
        dataframe[column] = pd.to_numeric(
            dataframe[column],
            errors="coerce",
        )

    dataframe = dataframe.dropna(
        subset=[
            "day_order",
            "tray_no",
            "mean_cell_ndvi",
            "median_cell_ndvi",
            "mean_cell_ndre",
            "median_cell_ndre",
        ]
    ).copy()

    dataframe["day_order"] = dataframe["day_order"].astype(int)
    dataframe["tray_no"] = dataframe["tray_no"].astype(int)
    dataframe["day"] = dataframe["day_order"].map(DAY_LABELS)

    return dataframe


def create_missing_record_report(
    dataframe: pd.DataFrame,
    tray_design: pd.DataFrame,
) -> pd.DataFrame:
    expected_records = []

    for tray_no in sorted(tray_design["tray_no"].unique()):
        for day_order in EXPECTED_DAYS:
            expected_records.append(
                {
                    "tray_no": int(tray_no),
                    "day_order": int(day_order),
                    "day": DAY_LABELS.get(
                        int(day_order),
                        f"Day {day_order}",
                    ),
                }
            )

    expected = pd.DataFrame(expected_records)

    observed = dataframe[
        ["tray_no", "day_order"]
    ].drop_duplicates()

    missing = expected.merge(
        observed,
        on=["tray_no", "day_order"],
        how="left",
        indicator=True,
    )

    missing = missing.loc[
        missing["_merge"].eq("left_only")
    ].drop(columns=["_merge"])

    if missing.empty:
        return pd.DataFrame(
            columns=[
                "tray_no",
                "day_order",
                "day",
                "treatment",
                "environment",
                "interaction",
                "notes",
            ]
        )

    missing = missing.merge(
        tray_design,
        on="tray_no",
        how="left",
    )

    missing["notes"] = (
        "Missing accepted Script 07 PASS/CHECK record for this Day/Tray."
    )

    return missing[
        [
            "tray_no",
            "day_order",
            "day",
            "treatment",
            "environment",
            "interaction",
            "notes",
        ]
    ].sort_values(
        ["day_order", "tray_no"]
    ).reset_index(drop=True)


def validate_no_duplicate_day_tray(dataframe: pd.DataFrame) -> None:
    counts = (
        dataframe.groupby(["day_order", "tray_no"])
        .size()
        .reset_index(name="count")
    )

    duplicates = counts.loc[
        counts["count"] > 1
    ]

    if not duplicates.empty:
        raise ValueError(
            "More than one Script 07 tray summary row was found "
            "for the same Day/Tray:\n"
            + duplicates.to_string(index=False)
        )


# ============================================================
# 6) TRAY-LEVEL METRICS
# ============================================================

def calculate_tray_metrics(dataframe: pd.DataFrame) -> pd.DataFrame:
    records = []

    for tray_no, group in dataframe.groupby("tray_no"):
        group = group.sort_values("day_order")

        rows_by_day = {
            int(row["day_order"]): row
            for _, row in group.iterrows()
        }

        reference = group.iloc[0]

        day1_ndvi = value_for_day(
            rows_by_day,
            1,
            "mean_cell_ndvi",
        )

        day5_ndvi = value_for_day(
            rows_by_day,
            5,
            "mean_cell_ndvi",
        )

        day9_ndvi = value_for_day(
            rows_by_day,
            9,
            "mean_cell_ndvi",
        )

        day1_ndre = value_for_day(
            rows_by_day,
            1,
            "mean_cell_ndre",
        )

        day5_ndre = value_for_day(
            rows_by_day,
            5,
            "mean_cell_ndre",
        )

        day9_ndre = value_for_day(
            rows_by_day,
            9,
            "mean_cell_ndre",
        )

        ndvi_change = (
            day5_ndvi - day1_ndvi
            if pd.notna(day1_ndvi) and pd.notna(day5_ndvi)
            else math.nan
        )

        ndre_change = (
            day5_ndre - day1_ndre
            if pd.notna(day1_ndre) and pd.notna(day5_ndre)
            else math.nan
        )

        ndvi_change_day5_to_day9 = (
            day9_ndvi - day5_ndvi
            if pd.notna(day9_ndvi) and pd.notna(day5_ndvi)
            else math.nan
        )

        ndre_change_day5_to_day9 = (
            day9_ndre - day5_ndre
            if pd.notna(day9_ndre) and pd.notna(day5_ndre)
            else math.nan
        )

        records.append(
            {
                "tray_no": int(tray_no),
                "tray": str(reference["tray"]),
                "treatment": str(reference["treatment"]),
                "environment": str(reference["environment"]),
                "interaction": str(reference["interaction"]),
                "day1_mean_ndvi": day1_ndvi,
                "day5_mean_ndvi": day5_ndvi,
                "day9_mean_ndvi": day9_ndvi,
                "day1_mean_ndre": day1_ndre,
                "day5_mean_ndre": day5_ndre,
                "day9_mean_ndre": day9_ndre,
                "ndvi_change_day1_to_day5": ndvi_change,
                "ndvi_rate_day1_to_day5_per_day": (
                    ndvi_change / 4.0
                    if pd.notna(ndvi_change)
                    else math.nan
                ),
                "ndre_change_day1_to_day5": ndre_change,
                "ndre_rate_day1_to_day5_per_day": (
                    ndre_change / 4.0
                    if pd.notna(ndre_change)
                    else math.nan
                ),
                "ndvi_change_day5_to_day9": ndvi_change_day5_to_day9,
                "ndre_change_day5_to_day9": ndre_change_day5_to_day9,
                "available_day_count": int(group["day_order"].nunique()),
            }
        )

    return (
        pd.DataFrame(records)
        .sort_values("tray_no")
        .reset_index(drop=True)
    )


# ============================================================
# 7) GROUP-LEVEL METRICS
# ============================================================

def calculate_group_daily(
    dataframe: pd.DataFrame,
    group_column: str,
    group_type: str,
) -> pd.DataFrame:
    output = (
        dataframe.groupby(
            [group_column, "day_order", "day"],
            as_index=False,
        )
        .agg(
            tray_count=("tray_no", "nunique"),
            mean_ndvi=("mean_cell_ndvi", "mean"),
            sd_ndvi=("mean_cell_ndvi", "std"),
            median_ndvi=("median_cell_ndvi", "mean"),
            mean_ndre=("mean_cell_ndre", "mean"),
            sd_ndre=("mean_cell_ndre", "std"),
            median_ndre=("median_cell_ndre", "mean"),
            mean_valid_ndvi_cells=("cells_with_valid_ndvi", "mean"),
            mean_valid_ndre_cells=("cells_with_valid_ndre", "mean"),
        )
        .rename(columns={group_column: "group"})
    )

    output["group_type"] = group_type

    for column in ["sd_ndvi", "sd_ndre"]:
        output[column] = output[column].fillna(0.0)

    return output


def calculate_group_growth_rates(
    tray_metrics: pd.DataFrame,
    group_column: str,
    group_type: str,
) -> pd.DataFrame:
    metric_columns = [
        "day1_mean_ndvi",
        "day5_mean_ndvi",
        "day9_mean_ndvi",
        "day1_mean_ndre",
        "day5_mean_ndre",
        "day9_mean_ndre",
        "ndvi_change_day1_to_day5",
        "ndvi_rate_day1_to_day5_per_day",
        "ndre_change_day1_to_day5",
        "ndre_rate_day1_to_day5_per_day",
        "ndvi_change_day5_to_day9",
        "ndre_change_day5_to_day9",
    ]

    aggregations = {
        "tray_count": ("tray_no", "nunique"),
    }

    for metric in metric_columns:
        aggregations[f"mean_{metric}"] = (metric, "mean")
        aggregations[f"sd_{metric}"] = (metric, "std")

    output = (
        tray_metrics.groupby(
            group_column,
            as_index=False,
        )
        .agg(**aggregations)
        .rename(columns={group_column: "group"})
    )

    output["group_type"] = group_type

    for column in output.columns:
        if column.startswith("sd_"):
            output[column] = output[column].fillna(0.0)

    ordered = [
        "group_type",
        "group",
        "tray_count",
    ] + [
        column
        for column in output.columns
        if column not in {
            "group_type",
            "group",
            "tray_count",
        }
    ]

    return output[ordered]


def sort_grouped_frame(dataframe: pd.DataFrame) -> pd.DataFrame:
    dataframe = dataframe.copy()

    dataframe["sort_order"] = dataframe.apply(
        lambda row: group_sort_order(
            row["group_type"],
            row["group"],
        ),
        axis=1,
    )

    sort_columns = ["group_type", "sort_order"]

    if "day_order" in dataframe.columns:
        sort_columns.append("day_order")

    dataframe = (
        dataframe.sort_values(sort_columns)
        .drop(columns=["sort_order"])
        .reset_index(drop=True)
    )

    return dataframe


# ============================================================
# 8) CHARTS
# ============================================================

def save_trend_chart(
    daily: pd.DataFrame,
    group_type: str,
    groups: list[str],
    value_column: str,
    sd_column: str,
    title: str,
    y_label: str,
    output_path: Path,
) -> None:
    figure, axis = plt.subplots(figsize=(11, 6.5))

    subset = daily.loc[
        daily["group_type"].eq(group_type)
    ].copy()

    for group_name in groups:
        series = subset.loc[
            subset["group"].eq(group_name)
        ].sort_values("day_order")

        if series.empty:
            continue

        axis.errorbar(
            series["day_order"],
            series[value_column],
            yerr=series[sd_column],
            marker="o",
            linewidth=2,
            capsize=4,
            label=f"{group_name} (n={int(series['tray_count'].iloc[0])})",
        )

    observation_days = sorted(
        subset["day_order"].unique()
    )

    axis.set_title(title)
    axis.set_xlabel("Observation day")
    axis.set_ylabel(y_label)
    axis.set_xticks(observation_days)
    axis.set_xticklabels(
        [
            DAY_LABELS.get(day, f"Day {day}")
            for day in observation_days
        ]
    )
    axis.grid(True, axis="y", alpha=0.30)

    if 9 in observation_days:
        axis.axvline(7, linestyle="--", linewidth=1)
        axis.text(
            7.08,
            axis.get_ylim()[1] * 0.96,
            "Day 9 follow-up",
            va="top",
            fontsize=9,
        )

    axis.legend(loc="best")

    figure.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=220)
    plt.close(figure)


def save_rate_ranking_chart(
    tray_metrics: pd.DataFrame,
    metric: str,
    title: str,
    x_label: str,
    output_path: Path,
) -> None:
    ranked = tray_metrics.dropna(
        subset=[metric]
    ).sort_values(metric).copy()

    if ranked.empty:
        return

    ranked["label"] = (
        ranked["tray"].astype(str)
        + " — "
        + ranked["interaction"].astype(str)
    )

    figure, axis = plt.subplots(figsize=(12, 7))

    axis.barh(
        ranked["label"],
        ranked[metric],
    )

    axis.set_title(title)
    axis.set_xlabel(x_label)
    axis.set_ylabel("Tray and treatment group")
    axis.grid(True, axis="x", alpha=0.30)

    figure.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=220)
    plt.close(figure)


def save_day9_bar_chart(
    group_rates: pd.DataFrame,
    metric: str,
    title: str,
    y_label: str,
    output_path: Path,
) -> None:
    frame = group_rates.loc[
        group_rates["group_type"].eq("Treatment x Environment")
    ].copy()

    frame = frame.dropna(subset=[metric])

    if frame.empty:
        return

    frame["sort_order"] = frame["group"].map(
        {
            group: index
            for index, group in enumerate(INTERACTION_ORDER)
        }
    )

    frame = frame.sort_values("sort_order")

    figure, axis = plt.subplots(figsize=(11, 6.5))

    axis.bar(
        frame["group"],
        frame[metric],
    )

    axis.set_title(title)
    axis.set_xlabel("Treatment and environment")
    axis.set_ylabel(y_label)
    axis.grid(True, axis="y", alpha=0.30)
    axis.tick_params(axis="x", rotation=18)

    figure.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=220)
    plt.close(figure)


# ============================================================
# 9) EXCEL REPORT
# ============================================================

def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 34

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(
                color="FFFFFF",
                bold=True,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in worksheet.columns:
            letter = column_cells[0].column_letter

            longest = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )

            worksheet.column_dimensions[letter].width = min(
                max(longest + 2, 12),
                58,
            )

    workbook.save(path)


def create_excel_report(
    output_path: Path,
    tray_design: pd.DataFrame,
    tray_metrics: pd.DataFrame,
    group_daily: pd.DataFrame,
    group_rates: pd.DataFrame,
    missing_records: pd.DataFrame,
) -> None:
    readme = pd.DataFrame(
        {
            "Notes": [
                "This report compares relative NDVI and NDRE by treatment and environment.",
                "Day 1 to Day 5 is the continuous growth-rate window.",
                "Growth rate = (Day 5 - Day 1) / 4.",
                "Day 9 is a later follow-up observation.",
                "NDVI = (MS_NIR - MS_R) / (MS_NIR + MS_R).",
                "NDRE = (MS_NIR - MS_RE) / (MS_NIR + MS_RE).",
                "If MS bands are not calibrated reflectance products, results should be described as relative image-derived indices.",
                "Treatment x Environment groups contain two trays each, so results are descriptive rather than formal statistical proof.",
                "If any Day/Tray records were missing from Script 07, they are listed in the Missing Records sheet.",
                "The visual summary report is generated as PDF, not HTML.",
            ]
        }
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        tray_design.to_excel(
            writer,
            sheet_name="Tray Design",
            index=False,
        )

        tray_metrics.to_excel(
            writer,
            sheet_name="Tray MS Metrics",
            index=False,
        )

        group_daily.to_excel(
            writer,
            sheet_name="Group Daily MS",
            index=False,
        )

        group_rates.to_excel(
            writer,
            sheet_name="Group MS Growth Rates",
            index=False,
        )

        missing_records.to_excel(
            writer,
            sheet_name="Missing Records",
            index=False,
        )

        readme.to_excel(
            writer,
            sheet_name="Read Me",
            index=False,
        )

    style_workbook(output_path)


# ============================================================
# 10) PDF REPORT
# ============================================================

def make_pdf_table(
    dataframe: pd.DataFrame,
    max_rows: int | None = None,
    font_size: int = 6,
    available_width: float = 10.9 * inch,
):
    frame = safe_round_dataframe(dataframe, 4).copy()

    if max_rows is not None:
        frame = frame.head(max_rows)

    styles = getSampleStyleSheet()

    cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["BodyText"],
        fontSize=font_size,
        leading=font_size + 1,
        wordWrap="CJK",
    )

    header_style = ParagraphStyle(
        "TableHeader",
        parent=styles["BodyText"],
        fontSize=font_size,
        leading=font_size + 1,
        textColor=colors.white,
        alignment=1,
        wordWrap="CJK",
    )

    values = []

    values.append(
        [
            Paragraph(str(column), header_style)
            for column in frame.columns
        ]
    )

    for _, row in frame.iterrows():
        values.append(
            [
                Paragraph(
                    "" if pd.isna(value) else str(value),
                    cell_style,
                )
                for value in row.tolist()
            ]
        )

    column_count = max(1, len(frame.columns))

    column_widths = [
        available_width / column_count
        for _ in range(column_count)
    ]

    table = Table(
        values,
        repeatRows=1,
        colWidths=column_widths,
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    return table


def create_pdf_report(
    output_path: Path,
    group_rates: pd.DataFrame,
    tray_metrics: pd.DataFrame,
    missing_records: pd.DataFrame,
    chart_files: list[Path],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    document = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        rightMargin=0.35 * inch,
        leftMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch,
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]
    heading_style = styles["Heading2"]
    body_style = styles["BodyText"]

    story = []

    story.append(
        Paragraph(
            "Second Trial: Multispectral Treatment Comparison",
            title_style,
        )
    )

    story.append(Spacer(1, 10))

    story.append(
        Paragraph(
            "This report compares relative image-derived NDVI and NDRE values by treatment and environment. "
            "If the original multispectral bands are not calibrated reflectance products, interpret these as relative spectral indicators rather than absolute field-calibrated vegetation indices.",
            body_style,
        )
    )

    story.append(Spacer(1, 12))

    interaction_rows = group_rates.loc[
        group_rates["group_type"].eq("Treatment x Environment")
    ].copy()

    treatment_rows = group_rates.loc[
        group_rates["group_type"].eq("Treatment")
    ].copy()

    best_day5_ndvi = safe_idxmax_row(
        interaction_rows,
        "mean_day5_mean_ndvi",
    )

    best_day5_ndre = safe_idxmax_row(
        interaction_rows,
        "mean_day5_mean_ndre",
    )

    best_ndvi_rate = safe_idxmax_row(
        interaction_rows,
        "mean_ndvi_rate_day1_to_day5_per_day",
    )

    best_ndre_rate = safe_idxmax_row(
        interaction_rows,
        "mean_ndre_rate_day1_to_day5_per_day",
    )

    story.append(
        Paragraph(
            "Observed Spectral Leaders",
            heading_style,
        )
    )

    story.append(
        Paragraph(
            f"Highest Day 5 mean NDVI: {format_leader(best_day5_ndvi, 'mean_day5_mean_ndvi')}<br/>"
            f"Highest Day 5 mean NDRE: {format_leader(best_day5_ndre, 'mean_day5_mean_ndre')}<br/>"
            f"Highest Day 1-Day 5 NDVI rate: {format_leader(best_ndvi_rate, 'mean_ndvi_rate_day1_to_day5_per_day')}<br/>"
            f"Highest Day 1-Day 5 NDRE rate: {format_leader(best_ndre_rate, 'mean_ndre_rate_day1_to_day5_per_day')}",
            body_style,
        )
    )

    story.append(Spacer(1, 12))

    treatment_table = treatment_rows[
        [
            "group",
            "tray_count",
            "mean_day5_mean_ndvi",
            "mean_day5_mean_ndre",
            "mean_ndvi_rate_day1_to_day5_per_day",
            "mean_ndre_rate_day1_to_day5_per_day",
            "mean_day9_mean_ndvi",
            "mean_day9_mean_ndre",
        ]
    ]

    story.append(
        Paragraph(
            "Microbes versus No Microbes",
            heading_style,
        )
    )

    story.append(make_pdf_table(treatment_table, font_size=7))
    story.append(PageBreak())

    interaction_table = interaction_rows[
        [
            "group",
            "tray_count",
            "mean_day5_mean_ndvi",
            "mean_day5_mean_ndre",
            "mean_ndvi_rate_day1_to_day5_per_day",
            "mean_ndre_rate_day1_to_day5_per_day",
            "mean_day9_mean_ndvi",
            "mean_day9_mean_ndre",
        ]
    ]

    story.append(
        Paragraph(
            "Treatment x Environment Comparison",
            heading_style,
        )
    )

    story.append(make_pdf_table(interaction_table, font_size=7))
    story.append(Spacer(1, 14))

    tray_ranking = tray_metrics[
        [
            "tray",
            "treatment",
            "environment",
            "available_day_count",
            "day5_mean_ndvi",
            "day5_mean_ndre",
            "ndvi_rate_day1_to_day5_per_day",
            "ndre_rate_day1_to_day5_per_day",
            "day9_mean_ndvi",
            "day9_mean_ndre",
        ]
    ].sort_values(
        "ndvi_rate_day1_to_day5_per_day",
        ascending=False,
        na_position="last",
    )

    story.append(
        Paragraph(
            "Tray Ranking by NDVI Growth Rate",
            heading_style,
        )
    )

    story.append(make_pdf_table(tray_ranking, font_size=6))
    story.append(PageBreak())

    story.append(
        Paragraph(
            "Missing Day/Tray Records",
            heading_style,
        )
    )

    if missing_records.empty:
        story.append(
            Paragraph(
                "No missing Day/Tray records were found.",
                body_style,
            )
        )
    else:
        story.append(make_pdf_table(missing_records, font_size=7))

    story.append(PageBreak())

    story.append(
        Paragraph(
            "Charts",
            heading_style,
        )
    )

    for index, chart in enumerate(chart_files):
        if not chart.exists():
            continue

        if index > 0:
            story.append(PageBreak())

        story.append(
            Paragraph(
                chart.stem.replace("_", " "),
                styles["Heading3"],
            )
        )

        story.append(Spacer(1, 6))

        story.append(
            PDFImage(
                str(chart),
                width=9.8 * inch,
                height=5.8 * inch,
            )
        )

    document.build(story)


# ============================================================
# 11) MAIN
# ============================================================

def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Script 08: compare Script 07 NDVI/NDRE results by "
            "Microbes/No Microbes and Inside/Outside."
        )
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate inputs and print tray design only.",
    )

    args = parser.parse_args()

    tray_design = load_tray_design()
    ms_data = load_script07_tray_summary()

    missing_trays = sorted(
        set(ms_data["tray_no"]) - set(tray_design["tray_no"])
    )

    if missing_trays:
        raise ValueError(
            "Tray Status.xlsx has no mapping for tray(s): "
            + ", ".join(str(value) for value in missing_trays)
        )

    ms_data = ms_data.merge(
        tray_design,
        on="tray_no",
        how="left",
        validate="many_to_one",
    )

    validate_no_duplicate_day_tray(ms_data)

    missing_records = create_missing_record_report(
        ms_data,
        tray_design,
    )

    print("\nSCRIPT 08 — MS TREATMENT COMPARISON")
    print("=" * 70)
    print(f"Tray Status:\n{TRAY_STATUS_XLSX}")
    print(f"\nScript 07 tray summary:\n{MS_TRAY_SUMMARY_CSV}")
    print(f"\nOutput folder:\n{OUTPUT_ROOT}\n")

    print("Tray design:")

    for row in tray_design.itertuples():
        print(
            f"Tray {row.tray_no}: {row.treatment} | {row.environment}"
        )

    print(f"\nScript 07 records used: {len(ms_data)}")

    print(
        "Accepted Script 07 statuses: "
        + ", ".join(sorted(ACCEPTED_SCRIPT07_STATUSES))
    )

    if not missing_records.empty:
        print("\nMissing Day/Tray records detected:")
        print(missing_records.to_string(index=False))

    if args.dry_run:
        print("\nDry run complete. No outputs created.")
        return 0

    CHARTS_ROOT.mkdir(parents=True, exist_ok=True)
    REPORTS_ROOT.mkdir(parents=True, exist_ok=True)
    CONFIG_ROOT.mkdir(parents=True, exist_ok=True)

    tray_metrics = calculate_tray_metrics(ms_data)

    treatment_daily = calculate_group_daily(
        ms_data,
        "treatment",
        "Treatment",
    )

    interaction_daily = calculate_group_daily(
        ms_data,
        "interaction",
        "Treatment x Environment",
    )

    group_daily = pd.concat(
        [
            treatment_daily,
            interaction_daily,
        ],
        ignore_index=True,
    )

    group_daily = sort_grouped_frame(group_daily)

    treatment_rates = calculate_group_growth_rates(
        tray_metrics,
        "treatment",
        "Treatment",
    )

    interaction_rates = calculate_group_growth_rates(
        tray_metrics,
        "interaction",
        "Treatment x Environment",
    )

    group_rates = pd.concat(
        [
            treatment_rates,
            interaction_rates,
        ],
        ignore_index=True,
    )

    group_rates = sort_grouped_frame(group_rates)

    chart_files = [
        CHARTS_ROOT / "01_microbes_vs_no_microbes_ndvi.png",
        CHARTS_ROOT / "02_microbes_vs_no_microbes_ndre.png",
        CHARTS_ROOT / "03_interaction_ndvi.png",
        CHARTS_ROOT / "04_interaction_ndre.png",
        CHARTS_ROOT / "05_tray_ndvi_growth_rate_ranking.png",
        CHARTS_ROOT / "06_tray_ndre_growth_rate_ranking.png",
        CHARTS_ROOT / "07_day9_ndvi_by_group.png",
        CHARTS_ROOT / "08_day9_ndre_by_group.png",
    ]

    save_trend_chart(
        treatment_daily,
        "Treatment",
        TREATMENT_ORDER,
        "mean_ndvi",
        "sd_ndvi",
        "Relative NDVI: Microbes vs No Microbes",
        "Mean relative NDVI",
        chart_files[0],
    )

    save_trend_chart(
        treatment_daily,
        "Treatment",
        TREATMENT_ORDER,
        "mean_ndre",
        "sd_ndre",
        "Relative NDRE: Microbes vs No Microbes",
        "Mean relative NDRE",
        chart_files[1],
    )

    save_trend_chart(
        interaction_daily,
        "Treatment x Environment",
        INTERACTION_ORDER,
        "mean_ndvi",
        "sd_ndvi",
        "Relative NDVI by treatment and environment",
        "Mean relative NDVI",
        chart_files[2],
    )

    save_trend_chart(
        interaction_daily,
        "Treatment x Environment",
        INTERACTION_ORDER,
        "mean_ndre",
        "sd_ndre",
        "Relative NDRE by treatment and environment",
        "Mean relative NDRE",
        chart_files[3],
    )

    save_rate_ranking_chart(
        tray_metrics,
        "ndvi_rate_day1_to_day5_per_day",
        "Tray ranking: Day 1-Day 5 relative NDVI growth rate",
        "Relative NDVI change per day",
        chart_files[4],
    )

    save_rate_ranking_chart(
        tray_metrics,
        "ndre_rate_day1_to_day5_per_day",
        "Tray ranking: Day 1-Day 5 relative NDRE growth rate",
        "Relative NDRE change per day",
        chart_files[5],
    )

    save_day9_bar_chart(
        group_rates,
        "mean_day9_mean_ndvi",
        "Day 9 relative NDVI by treatment and environment",
        "Mean relative NDVI",
        chart_files[6],
    )

    save_day9_bar_chart(
        group_rates,
        "mean_day9_mean_ndre",
        "Day 9 relative NDRE by treatment and environment",
        "Mean relative NDRE",
        chart_files[7],
    )

    tray_metrics.to_csv(
        REPORTS_ROOT / "ms_tray_index_metrics.csv",
        index=False,
    )

    group_daily.to_csv(
        REPORTS_ROOT / "ms_group_daily_metrics.csv",
        index=False,
    )

    group_rates.to_csv(
        REPORTS_ROOT / "ms_group_growth_rates.csv",
        index=False,
    )

    missing_records.to_csv(
        REPORTS_ROOT / "missing_script07_day_tray_records.csv",
        index=False,
    )

    create_excel_report(
        REPORTS_ROOT / "ms_treatment_comparison_report.xlsx",
        tray_design,
        tray_metrics,
        group_daily,
        group_rates,
        missing_records,
    )

    pdf_report_path = REPORTS_ROOT / "ms_treatment_visual_summary.pdf"

    create_pdf_report(
        pdf_report_path,
        group_rates,
        tray_metrics,
        missing_records,
        chart_files,
    )

    settings = {
        "purpose": (
            "Treatment and environment comparison using Script 07 "
            "relative NDVI and NDRE outputs."
        ),
        "input_tray_status": str(TRAY_STATUS_XLSX),
        "input_script07_tray_summary": str(MS_TRAY_SUMMARY_CSV),
        "growth_rate_window": "Day 1 to Day 5",
        "growth_rate_formula": "(Day 5 - Day 1) / 4",
        "day_9_handling": "Later follow-up only",
        "accepted_script07_statuses": sorted(ACCEPTED_SCRIPT07_STATUSES),
        "missing_record_count": int(len(missing_records)),
        "report_format": "PDF",
        "report_file": str(pdf_report_path),
        "interpretation_note": (
            "If source MS bands are not calibrated reflectance products, "
            "NDVI and NDRE should be described as relative image-derived indices."
        ),
        "tray_design": tray_design.to_dict(orient="records"),
    }

    (CONFIG_ROOT / "analysis_settings.json").write_text(
        json.dumps(settings, indent=2),
        encoding="utf-8",
    )

    interaction_rows = group_rates.loc[
        group_rates["group_type"].eq("Treatment x Environment")
    ].copy()

    best_day5_ndvi = safe_idxmax_row(
        interaction_rows,
        "mean_day5_mean_ndvi",
    )

    best_day5_ndre = safe_idxmax_row(
        interaction_rows,
        "mean_day5_mean_ndre",
    )

    best_ndvi_rate = safe_idxmax_row(
        interaction_rows,
        "mean_ndvi_rate_day1_to_day5_per_day",
    )

    best_ndre_rate = safe_idxmax_row(
        interaction_rows,
        "mean_ndre_rate_day1_to_day5_per_day",
    )

    print("\n" + "=" * 70)
    print("SCRIPT 08 FINISHED")
    print("=" * 70)

    if best_day5_ndvi is not None:
        print(
            "Highest observed Day 5 mean NDVI group: "
            f"{best_day5_ndvi['group']} "
            f"({best_day5_ndvi['mean_day5_mean_ndvi']:.4f})."
        )

    if best_day5_ndre is not None:
        print(
            "Highest observed Day 5 mean NDRE group: "
            f"{best_day5_ndre['group']} "
            f"({best_day5_ndre['mean_day5_mean_ndre']:.4f})."
        )

    if best_ndvi_rate is not None:
        print(
            "Highest observed Day 1-Day 5 NDVI rate group: "
            f"{best_ndvi_rate['group']} "
            f"({best_ndvi_rate['mean_ndvi_rate_day1_to_day5_per_day']:.5f} per day)."
        )

    if best_ndre_rate is not None:
        print(
            "Highest observed Day 1-Day 5 NDRE rate group: "
            f"{best_ndre_rate['group']} "
            f"({best_ndre_rate['mean_ndre_rate_day1_to_day5_per_day']:.5f} per day)."
        )

    if not missing_records.empty:
        print(
            f"\nWARNING: {len(missing_records)} expected Day/Tray record(s) "
            "were missing from accepted Script 07 results. "
            "They were listed in the Missing Records report."
        )

    print(f"\nCharts:\n{CHARTS_ROOT}")
    print(f"\nReports:\n{REPORTS_ROOT}")

    print(
        "\nOpen this PDF report:\n"
        f"{pdf_report_path}"
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())