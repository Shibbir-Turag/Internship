from __future__ import annotations

"""
SCRIPT 09 - SECOND TRIAL SYNTHESIS

Purpose
-------
Combine the RGB and multispectral outputs into one final Second Trial synthesis.

Inputs
------
1. Tray Status.xlsx
2. Script 05 RGB treatment growth metrics
3. Script 07 MS NDVI / NDRE tray summary

Outputs
-------
outputs/Second Trial/09_Second_Trial_Synthesis/
    charts/
    _reports/
        second_trial_master_summary.csv
        second_trial_group_summary.csv
        second_trial_master_summary.xlsx
        second_trial_synthesis_report.pdf
    _config/
        synthesis_settings.json

Interpretation
--------------
The overall performance score is descriptive. It is not a formal statistical
test. It ranks trays and groups based on the available image-derived indicators.
"""

import argparse
import json
import math
import re
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image as PDFImage,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


# ============================================================
# 1) PATHS - CHANGE PROJECT_ROOT ONLY WHEN REUSING
# ============================================================

PROJECT_ROOT = Path(r"C:\Users\tshib\OneDrive\Desktop\Internship")

TRAY_STATUS_XLSX = (
    PROJECT_ROOT
    / "data"
    / "Second Trial"
    / "Tray Status.xlsx"
)

SCRIPT05_REPORTS = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "05_Treatment_Growth_Visuals"
    / "_reports"
)

RGB_TRAY_METRICS_CSV = (
    SCRIPT05_REPORTS
    / "tray_growth_metrics.csv"
)

SCRIPT07_REPORTS = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "07_MS_Vegetation_Indices"
    / "_reports"
)

MS_TRAY_SUMMARY_CSV = (
    SCRIPT07_REPORTS
    / "ms_index_tray_summary.csv"
)

OUTPUT_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "09_Second_Trial_Synthesis"
)

CHARTS_ROOT = OUTPUT_ROOT / "charts"
REPORTS_ROOT = OUTPUT_ROOT / "_reports"
CONFIG_ROOT = OUTPUT_ROOT / "_config"


# ============================================================
# 2) SETTINGS
# ============================================================

DAY_LABELS = {
    1: "Day 1",
    2: "Day 2",
    3: "Day 3",
    4: "Day 4",
    5: "Day 5",
    9: "Day 9",
}

EXPECTED_DAYS = [1, 2, 3, 4, 5, 9]

TREATMENT_ORDER = [
    "No Microbes",
    "Microbes",
]

INTERACTION_ORDER = [
    "No Microbes | Inside",
    "No Microbes | Outside",
    "Microbes | Inside",
    "Microbes | Outside",
]

ACCEPTED_SCRIPT07_STATUSES = {
    "PASS",
    "CHECK",
}

PERFORMANCE_COMPONENTS = [
    "day5_tracked_emergence_percent",
    "day5_green_cover_percent",
    "day5_mean_ndvi",
    "day5_mean_ndre",
    "green_cover_rate_day1_to_day5_pp_per_day",
    "ndvi_rate_day1_to_day5_per_day",
    "ndre_rate_day1_to_day5_per_day",
]


# ============================================================
# 3) GENERAL HELPERS
# ============================================================

def normalise(value: object) -> str:
    return re.sub(
        r"[^a-z0-9]+",
        "",
        str(value).casefold(),
    )


def is_present(value: object) -> bool:
    return str(value).strip().casefold() in {
        "p",
        "yes",
        "y",
        "true",
        "1",
    }


def require_columns(
    dataframe: pd.DataFrame,
    required_columns: list[str],
    source_name: str,
) -> None:
    missing = [
        column
        for column in required_columns
        if column not in dataframe.columns
    ]

    if missing:
        raise ValueError(
            f"{source_name} is missing required column(s): "
            + ", ".join(missing)
        )


def group_sort_order(group_name: str) -> int:
    if group_name in INTERACTION_ORDER:
        return INTERACTION_ORDER.index(group_name)

    return 999


def safe_round_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = 4,
) -> pd.DataFrame:
    output = dataframe.copy()

    numeric_columns = output.select_dtypes(
        include=["number"]
    ).columns

    output[numeric_columns] = output[numeric_columns].round(decimals)

    return output


def minmax_score(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(
        series,
        errors="coerce",
    )

    valid = values.dropna()

    if valid.empty:
        return pd.Series(
            [math.nan] * len(values),
            index=values.index,
        )

    minimum = valid.min()
    maximum = valid.max()

    if maximum == minimum:
        return pd.Series(
            [
                50.0 if pd.notna(value) else math.nan
                for value in values
            ],
            index=values.index,
        )

    return (values - minimum) / (maximum - minimum) * 100.0


def get_day_value(
    rows_by_day: dict[int, pd.Series],
    day_order: int,
    metric: str,
):
    row = rows_by_day.get(day_order)

    if row is None:
        return math.nan

    return row.get(metric, math.nan)


def safe_idxmax_row(
    dataframe: pd.DataFrame,
    column: str,
):
    valid = dataframe.dropna(subset=[column])

    if valid.empty:
        return None

    return valid.loc[valid[column].idxmax()]


# ============================================================
# 4) LOAD TRAY DESIGN
# ============================================================

def load_tray_design() -> pd.DataFrame:
    if not TRAY_STATUS_XLSX.exists():
        raise FileNotFoundError(
            f"Tray Status workbook not found:\n{TRAY_STATUS_XLSX}"
        )

    raw = pd.read_excel(TRAY_STATUS_XLSX)

    header_lookup = {
        normalise(column): column
        for column in raw.columns
    }

    required = {
        "trayno": "Tray No",
        "microbes": "Microbes",
        "nomicrobes": "No Microbes",
        "inside": "Inside",
        "outside": "Outside",
    }

    missing = [
        label
        for key, label in required.items()
        if key not in header_lookup
    ]

    if missing:
        raise ValueError(
            "Tray Status.xlsx is missing required column(s): "
            + ", ".join(missing)
        )

    tray_col = header_lookup["trayno"]
    microbes_col = header_lookup["microbes"]
    no_microbes_col = header_lookup["nomicrobes"]
    inside_col = header_lookup["inside"]
    outside_col = header_lookup["outside"]

    records = []

    for _, row in raw.iterrows():
        tray_no = pd.to_numeric(
            row[tray_col],
            errors="coerce",
        )

        if pd.isna(tray_no):
            continue

        tray_no = int(tray_no)

        microbes = is_present(row[microbes_col])
        no_microbes = is_present(row[no_microbes_col])
        inside = is_present(row[inside_col])
        outside = is_present(row[outside_col])

        if microbes == no_microbes:
            raise ValueError(
                f"Tray {tray_no}: exactly one of Microbes / No Microbes must be marked."
            )

        if inside == outside:
            raise ValueError(
                f"Tray {tray_no}: exactly one of Inside / Outside must be marked."
            )

        treatment = "Microbes" if microbes else "No Microbes"
        environment = "Inside" if inside else "Outside"

        records.append(
            {
                "tray_no": tray_no,
                "tray": f"Tray {tray_no}",
                "treatment": treatment,
                "environment": environment,
                "interaction": f"{treatment} | {environment}",
            }
        )

    design = (
        pd.DataFrame(records)
        .drop_duplicates(subset=["tray_no"])
        .sort_values("tray_no")
        .reset_index(drop=True)
    )

    if design.empty:
        raise ValueError(
            "No valid tray design records were found."
        )

    return design


# ============================================================
# 5) LOAD RGB METRICS FROM SCRIPT 05
# ============================================================

def load_rgb_metrics() -> pd.DataFrame:
    if not RGB_TRAY_METRICS_CSV.exists():
        raise FileNotFoundError(
            f"Script 05 RGB tray metrics not found:\n{RGB_TRAY_METRICS_CSV}"
        )

    dataframe = pd.read_csv(RGB_TRAY_METRICS_CSV)

    require_columns(
        dataframe,
        [
            "tray_no",
            "day1_tracked_emergence_percent",
            "day5_tracked_emergence_percent",
            "day9_tracked_emergence_percent",
            "day1_green_cover_percent",
            "day5_green_cover_percent",
            "day9_green_cover_percent",
            "emergence_rate_day1_to_day5_pp_per_day",
            "green_cover_rate_day1_to_day5_pp_per_day",
        ],
        "tray_growth_metrics.csv",
    )

    numeric_columns = [
        "tray_no",
        "day1_tracked_emergence_percent",
        "day5_tracked_emergence_percent",
        "day9_tracked_emergence_percent",
        "day1_green_cover_percent",
        "day5_green_cover_percent",
        "day9_green_cover_percent",
        "emergence_change_day1_to_day5_pp",
        "emergence_rate_day1_to_day5_pp_per_day",
        "green_cover_change_day1_to_day5_pp",
        "green_cover_rate_day1_to_day5_pp_per_day",
        "green_cover_change_day5_to_day9_pp",
    ]

    for column in numeric_columns:
        if column in dataframe.columns:
            dataframe[column] = pd.to_numeric(
                dataframe[column],
                errors="coerce",
            )

    dataframe = dataframe.dropna(
        subset=["tray_no"]
    ).copy()

    dataframe["tray_no"] = dataframe["tray_no"].astype(int)

    keep_columns = [
        "tray_no",
        "day1_tracked_emergence_percent",
        "day5_tracked_emergence_percent",
        "day9_tracked_emergence_percent",
        "day1_green_cover_percent",
        "day5_green_cover_percent",
        "day9_green_cover_percent",
        "emergence_change_day1_to_day5_pp",
        "emergence_rate_day1_to_day5_pp_per_day",
        "green_cover_change_day1_to_day5_pp",
        "green_cover_rate_day1_to_day5_pp_per_day",
        "green_cover_change_day5_to_day9_pp",
    ]

    keep_columns = [
        column
        for column in keep_columns
        if column in dataframe.columns
    ]

    return dataframe[keep_columns].drop_duplicates(
        subset=["tray_no"]
    )


# ============================================================
# 6) LOAD MS METRICS FROM SCRIPT 07
# ============================================================

def load_ms_data() -> pd.DataFrame:
    if not MS_TRAY_SUMMARY_CSV.exists():
        raise FileNotFoundError(
            f"Script 07 MS tray summary not found:\n{MS_TRAY_SUMMARY_CSV}"
        )

    dataframe = pd.read_csv(MS_TRAY_SUMMARY_CSV)

    require_columns(
        dataframe,
        [
            "day_order",
            "day",
            "tray",
            "tray_no",
            "mean_cell_ndvi",
            "mean_cell_ndre",
            "status",
        ],
        "ms_index_tray_summary.csv",
    )

    dataframe["status"] = dataframe["status"].astype(str).str.upper()

    dataframe = dataframe.loc[
        dataframe["status"].isin(ACCEPTED_SCRIPT07_STATUSES)
    ].copy()

    numeric_columns = [
        "day_order",
        "tray_no",
        "cells_processed",
        "cells_with_valid_ndvi",
        "cells_with_valid_ndre",
        "mean_cell_ndvi",
        "median_cell_ndvi",
        "mean_cell_ndre",
        "median_cell_ndre",
    ]

    for column in numeric_columns:
        if column in dataframe.columns:
            dataframe[column] = pd.to_numeric(
                dataframe[column],
                errors="coerce",
            )

    dataframe = dataframe.dropna(
        subset=[
            "day_order",
            "tray_no",
            "mean_cell_ndvi",
            "mean_cell_ndre",
        ]
    ).copy()

    dataframe["day_order"] = dataframe["day_order"].astype(int)
    dataframe["tray_no"] = dataframe["tray_no"].astype(int)
    dataframe["day"] = dataframe["day_order"].map(DAY_LABELS)

    return dataframe


def calculate_ms_tray_metrics(ms_data: pd.DataFrame) -> pd.DataFrame:
    records = []

    for tray_no, group in ms_data.groupby("tray_no"):
        group = group.sort_values("day_order")

        rows_by_day = {
            int(row["day_order"]): row
            for _, row in group.iterrows()
        }

        day1_ndvi = get_day_value(rows_by_day, 1, "mean_cell_ndvi")
        day5_ndvi = get_day_value(rows_by_day, 5, "mean_cell_ndvi")
        day9_ndvi = get_day_value(rows_by_day, 9, "mean_cell_ndvi")

        day1_ndre = get_day_value(rows_by_day, 1, "mean_cell_ndre")
        day5_ndre = get_day_value(rows_by_day, 5, "mean_cell_ndre")
        day9_ndre = get_day_value(rows_by_day, 9, "mean_cell_ndre")

        ndvi_change = (
            day5_ndvi - day1_ndvi
            if pd.notna(day1_ndvi) and pd.notna(day5_ndvi)
            else math.nan
        )

        ndre_change = (
            day5_ndre - day1_ndre
            if pd.notna(day1_ndre) and pd.notna(day5_ndre)
            else math.nan
        )

        records.append(
            {
                "tray_no": int(tray_no),
                "day1_mean_ndvi": day1_ndvi,
                "day5_mean_ndvi": day5_ndvi,
                "day9_mean_ndvi": day9_ndvi,
                "day1_mean_ndre": day1_ndre,
                "day5_mean_ndre": day5_ndre,
                "day9_mean_ndre": day9_ndre,
                "ndvi_change_day1_to_day5": ndvi_change,
                "ndvi_rate_day1_to_day5_per_day": (
                    ndvi_change / 4.0
                    if pd.notna(ndvi_change)
                    else math.nan
                ),
                "ndre_change_day1_to_day5": ndre_change,
                "ndre_rate_day1_to_day5_per_day": (
                    ndre_change / 4.0
                    if pd.notna(ndre_change)
                    else math.nan
                ),
                "ndvi_change_day5_to_day9": (
                    day9_ndvi - day5_ndvi
                    if pd.notna(day9_ndvi) and pd.notna(day5_ndvi)
                    else math.nan
                ),
                "ndre_change_day5_to_day9": (
                    day9_ndre - day5_ndre
                    if pd.notna(day9_ndre) and pd.notna(day5_ndre)
                    else math.nan
                ),
                "ms_available_day_count": int(group["day_order"].nunique()),
            }
        )

    return (
        pd.DataFrame(records)
        .sort_values("tray_no")
        .reset_index(drop=True)
    )


# ============================================================
# 7) SYNTHESIS TABLES
# ============================================================

def create_master_summary(
    tray_design: pd.DataFrame,
    rgb_metrics: pd.DataFrame,
    ms_metrics: pd.DataFrame,
) -> pd.DataFrame:
    master = tray_design.merge(
        rgb_metrics,
        on="tray_no",
        how="left",
        validate="one_to_one",
    )

    master = master.merge(
        ms_metrics,
        on="tray_no",
        how="left",
        validate="one_to_one",
    )

    for component in PERFORMANCE_COMPONENTS:
        score_column = f"{component}_score"

        if component in master.columns:
            master[score_column] = minmax_score(
                master[component]
            )
        else:
            master[score_column] = math.nan

    score_columns = [
        f"{component}_score"
        for component in PERFORMANCE_COMPONENTS
    ]

    master["overall_observed_performance_score"] = master[
        score_columns
    ].mean(
        axis=1,
        skipna=True,
    )

    master["overall_observed_rank"] = master[
        "overall_observed_performance_score"
    ].rank(
        ascending=False,
        method="min",
    )

    master["overall_observed_rank"] = master[
        "overall_observed_rank"
    ].astype("Int64")

    master["data_completeness_note"] = master.apply(
        create_completeness_note,
        axis=1,
    )

    ordered_columns = [
        "tray_no",
        "tray",
        "treatment",
        "environment",
        "interaction",
        "day1_tracked_emergence_percent",
        "day5_tracked_emergence_percent",
        "day9_tracked_emergence_percent",
        "day1_green_cover_percent",
        "day5_green_cover_percent",
        "day9_green_cover_percent",
        "day1_mean_ndvi",
        "day5_mean_ndvi",
        "day9_mean_ndvi",
        "day1_mean_ndre",
        "day5_mean_ndre",
        "day9_mean_ndre",
        "emergence_rate_day1_to_day5_pp_per_day",
        "green_cover_rate_day1_to_day5_pp_per_day",
        "ndvi_rate_day1_to_day5_per_day",
        "ndre_rate_day1_to_day5_per_day",
        "emergence_change_day1_to_day5_pp",
        "green_cover_change_day1_to_day5_pp",
        "ndvi_change_day1_to_day5",
        "ndre_change_day1_to_day5",
        "green_cover_change_day5_to_day9_pp",
        "ndvi_change_day5_to_day9",
        "ndre_change_day5_to_day9",
        "overall_observed_performance_score",
        "overall_observed_rank",
        "data_completeness_note",
    ]

    ordered_columns = [
        column
        for column in ordered_columns
        if column in master.columns
    ]

    return master[ordered_columns].sort_values(
        ["overall_observed_rank", "tray_no"],
        na_position="last",
    ).reset_index(drop=True)


def create_completeness_note(row: pd.Series) -> str:
    missing = []

    key_columns = {
        "RGB emergence": "day5_tracked_emergence_percent",
        "RGB green cover": "day5_green_cover_percent",
        "NDVI": "day5_mean_ndvi",
        "NDRE": "day5_mean_ndre",
    }

    for label, column in key_columns.items():
        if column not in row.index or pd.isna(row[column]):
            missing.append(label)

    if not missing:
        return "Complete key RGB and MS metrics available."

    return "Missing: " + ", ".join(missing)


def create_group_summary(master: pd.DataFrame) -> pd.DataFrame:
    group_rows = []

    for group_type, group_column in [
        ("Treatment", "treatment"),
        ("Treatment x Environment", "interaction"),
    ]:
        grouped = master.groupby(
            group_column,
            dropna=False,
        )

        for group_name, group in grouped:
            group_rows.append(
                {
                    "group_type": group_type,
                    "group": group_name,
                    "tray_count": int(group["tray_no"].nunique()),
                    "mean_day5_emergence_percent": group[
                        "day5_tracked_emergence_percent"
                    ].mean(),
                    "mean_day5_rgb_green_cover_percent": group[
                        "day5_green_cover_percent"
                    ].mean(),
                    "mean_day5_ndvi": group[
                        "day5_mean_ndvi"
                    ].mean(),
                    "mean_day5_ndre": group[
                        "day5_mean_ndre"
                    ].mean(),
                    "mean_rgb_green_cover_rate": group[
                        "green_cover_rate_day1_to_day5_pp_per_day"
                    ].mean(),
                    "mean_ndvi_rate": group[
                        "ndvi_rate_day1_to_day5_per_day"
                    ].mean(),
                    "mean_ndre_rate": group[
                        "ndre_rate_day1_to_day5_per_day"
                    ].mean(),
                    "mean_overall_score": group[
                        "overall_observed_performance_score"
                    ].mean(),
                    "best_rank_in_group": group[
                        "overall_observed_rank"
                    ].min(),
                }
            )

    summary = pd.DataFrame(group_rows)

    summary["sort_order"] = summary.apply(
        lambda row: (
            0
            if row["group_type"] == "Treatment"
            else 1,
            group_sort_order(str(row["group"])),
        ),
        axis=1,
    )

    summary = summary.sort_values(
        "sort_order"
    ).drop(
        columns=["sort_order"]
    ).reset_index(drop=True)

    return summary


# ============================================================
# 8) CHARTS
# ============================================================

def save_tray_score_ranking(
    master: pd.DataFrame,
    output_path: Path,
) -> None:
    frame = master.dropna(
        subset=["overall_observed_performance_score"]
    ).sort_values(
        "overall_observed_performance_score"
    )

    if frame.empty:
        return

    labels = (
        frame["tray"].astype(str)
        + " - "
        + frame["interaction"].astype(str)
    )

    figure, axis = plt.subplots(figsize=(12, 7))

    axis.barh(
        labels,
        frame["overall_observed_performance_score"],
    )

    axis.set_title(
        "Second Trial tray ranking by overall observed performance score"
    )
    axis.set_xlabel("Overall observed performance score")
    axis.set_ylabel("Tray")
    axis.grid(True, axis="x", alpha=0.30)

    figure.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=220)
    plt.close(figure)


def save_group_score_ranking(
    group_summary: pd.DataFrame,
    output_path: Path,
) -> None:
    frame = group_summary.loc[
        group_summary["group_type"].eq("Treatment x Environment")
    ].dropna(
        subset=["mean_overall_score"]
    ).sort_values(
        "mean_overall_score"
    )

    if frame.empty:
        return

    figure, axis = plt.subplots(figsize=(11, 6.5))

    axis.barh(
        frame["group"],
        frame["mean_overall_score"],
    )

    axis.set_title(
        "Second Trial group ranking by mean overall performance score"
    )
    axis.set_xlabel("Mean overall observed performance score")
    axis.set_ylabel("Treatment x Environment")
    axis.grid(True, axis="x", alpha=0.30)

    figure.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=220)
    plt.close(figure)


def save_scatter(
    master: pd.DataFrame,
    x_column: str,
    y_column: str,
    title: str,
    x_label: str,
    y_label: str,
    output_path: Path,
) -> None:
    frame = master.dropna(
        subset=[x_column, y_column]
    ).copy()

    if frame.empty:
        return

    figure, axis = plt.subplots(figsize=(8.5, 6.5))

    for interaction, group in frame.groupby("interaction"):
        axis.scatter(
            group[x_column],
            group[y_column],
            label=interaction,
            s=70,
        )

        for _, row in group.iterrows():
            axis.annotate(
                str(row["tray"]),
                (
                    row[x_column],
                    row[y_column],
                ),
                textcoords="offset points",
                xytext=(5, 5),
                fontsize=8,
            )

    axis.set_title(title)
    axis.set_xlabel(x_label)
    axis.set_ylabel(y_label)
    axis.grid(True, alpha=0.30)
    axis.legend(loc="best", fontsize=8)

    figure.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=220)
    plt.close(figure)


def save_group_heatmap(
    group_summary: pd.DataFrame,
    output_path: Path,
) -> None:
    frame = group_summary.loc[
        group_summary["group_type"].eq("Treatment x Environment")
    ].copy()

    metrics = [
        "mean_day5_emergence_percent",
        "mean_day5_rgb_green_cover_percent",
        "mean_day5_ndvi",
        "mean_day5_ndre",
        "mean_rgb_green_cover_rate",
        "mean_ndvi_rate",
        "mean_ndre_rate",
        "mean_overall_score",
    ]

    available_metrics = [
        metric
        for metric in metrics
        if metric in frame.columns
    ]

    if frame.empty or not available_metrics:
        return

    heatmap_data = frame.set_index("group")[available_metrics]

    normalised = heatmap_data.apply(minmax_score, axis=0)

    figure, axis = plt.subplots(figsize=(13, 5.8))

    image = axis.imshow(
        normalised.to_numpy(),
        aspect="auto",
    )

    axis.set_title(
        "Treatment x Environment summary heatmap - normalised 0 to 100"
    )

    axis.set_xticks(range(len(available_metrics)))
    axis.set_xticklabels(
        [
            metric.replace("_", " ")
            for metric in available_metrics
        ],
        rotation=35,
        ha="right",
    )

    axis.set_yticks(range(len(normalised.index)))
    axis.set_yticklabels(normalised.index)

    for row_index in range(normalised.shape[0]):
        for column_index in range(normalised.shape[1]):
            value = normalised.iloc[row_index, column_index]

            if pd.notna(value):
                axis.text(
                    column_index,
                    row_index,
                    f"{value:.0f}",
                    ha="center",
                    va="center",
                    fontsize=8,
                )

    figure.colorbar(
        image,
        ax=axis,
        label="Normalised score",
    )

    figure.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=220)
    plt.close(figure)


# ============================================================
# 9) EXCEL AND PDF REPORTS
# ============================================================

def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 34

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(
                color="FFFFFF",
                bold=True,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in worksheet.columns:
            letter = column_cells[0].column_letter

            longest = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )

            worksheet.column_dimensions[letter].width = min(
                max(longest + 2, 12),
                58,
            )

    workbook.save(path)


def create_excel_report(
    output_path: Path,
    master: pd.DataFrame,
    group_summary: pd.DataFrame,
    tray_design: pd.DataFrame,
) -> None:
    readme = pd.DataFrame(
        {
            "Notes": [
                "This workbook combines RGB and multispectral results for the Second Trial.",
                "RGB emergence and green-cover values come from Script 05.",
                "NDVI and NDRE values come from Script 07.",
                "Overall observed performance score is descriptive and based on normalised image-derived indicators.",
                "The score is not a formal statistical test.",
                "Day 1 to Day 5 is the continuous growth-rate window.",
                "Day 9 is treated as a later follow-up observation.",
                "PDF is the default visual report format for internship reports.",
            ]
        }
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        safe_round_dataframe(master, 5).to_excel(
            writer,
            sheet_name="Master Summary",
            index=False,
        )

        safe_round_dataframe(group_summary, 5).to_excel(
            writer,
            sheet_name="Group Summary",
            index=False,
        )

        tray_design.to_excel(
            writer,
            sheet_name="Tray Design",
            index=False,
        )

        readme.to_excel(
            writer,
            sheet_name="Read Me",
            index=False,
        )

    style_workbook(output_path)


def make_pdf_table(
    dataframe: pd.DataFrame,
    max_rows: int | None = None,
    font_size: int = 6,
    available_width: float = 10.9 * inch,
):
    frame = safe_round_dataframe(dataframe, 4).copy()

    if max_rows is not None:
        frame = frame.head(max_rows)

    styles = getSampleStyleSheet()

    cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["BodyText"],
        fontSize=font_size,
        leading=font_size + 1,
        wordWrap="CJK",
    )

    header_style = ParagraphStyle(
        "TableHeader",
        parent=styles["BodyText"],
        fontSize=font_size,
        leading=font_size + 1,
        textColor=colors.white,
        alignment=1,
        wordWrap="CJK",
    )

    values = [
        [
            Paragraph(str(column), header_style)
            for column in frame.columns
        ]
    ]

    for _, row in frame.iterrows():
        values.append(
            [
                Paragraph(
                    "" if pd.isna(value) else str(value),
                    cell_style,
                )
                for value in row.tolist()
            ]
        )

    column_count = max(1, len(frame.columns))
    column_widths = [
        available_width / column_count
        for _ in range(column_count)
    ]

    table = Table(
        values,
        repeatRows=1,
        colWidths=column_widths,
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F9FAFB")],
                ),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    return table


def create_pdf_report(
    output_path: Path,
    master: pd.DataFrame,
    group_summary: pd.DataFrame,
    chart_files: list[Path],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    document = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        rightMargin=0.35 * inch,
        leftMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch,
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(
        Paragraph(
            "Second Trial Synthesis Report",
            styles["Title"],
        )
    )

    story.append(Spacer(1, 10))

    story.append(
        Paragraph(
            "This report combines the RGB emergence and green-cover workflow with the multispectral NDVI and NDRE workflow. "
            "The overall score is a descriptive ranking based on normalised image-derived indicators. "
            "It should not be interpreted as a formal statistical test.",
            styles["BodyText"],
        )
    )

    story.append(Spacer(1, 12))

    best_tray = safe_idxmax_row(
        master,
        "overall_observed_performance_score",
    )

    best_group = safe_idxmax_row(
        group_summary.loc[
            group_summary["group_type"].eq("Treatment x Environment")
        ],
        "mean_overall_score",
    )

    story.append(
        Paragraph(
            "Observed Leaders",
            styles["Heading2"],
        )
    )

    if best_tray is not None:
        story.append(
            Paragraph(
                f"Highest ranked tray: {best_tray['tray']} - {best_tray['interaction']} "
                f"(score {best_tray['overall_observed_performance_score']:.2f}).",
                styles["BodyText"],
            )
        )

    if best_group is not None:
        story.append(
            Paragraph(
                f"Highest ranked treatment x environment group: {best_group['group']} "
                f"(mean score {best_group['mean_overall_score']:.2f}).",
                styles["BodyText"],
            )
        )

    story.append(Spacer(1, 12))

    group_table = group_summary[
        [
            "group_type",
            "group",
            "tray_count",
            "mean_day5_emergence_percent",
            "mean_day5_rgb_green_cover_percent",
            "mean_day5_ndvi",
            "mean_day5_ndre",
            "mean_overall_score",
        ]
    ]

    story.append(
        Paragraph(
            "Group Summary",
            styles["Heading2"],
        )
    )

    story.append(make_pdf_table(group_table, font_size=6))
    story.append(PageBreak())

    master_table = master[
        [
            "tray",
            "treatment",
            "environment",
            "day5_tracked_emergence_percent",
            "day5_green_cover_percent",
            "day5_mean_ndvi",
            "day5_mean_ndre",
            "overall_observed_performance_score",
            "overall_observed_rank",
        ]
    ]

    story.append(
        Paragraph(
            "Tray-Level Master Summary",
            styles["Heading2"],
        )
    )

    story.append(make_pdf_table(master_table, font_size=6))
    story.append(PageBreak())

    story.append(
        Paragraph(
            "Report Charts",
            styles["Heading2"],
        )
    )

    first_chart = True

    for chart in chart_files:
        if not chart.exists():
            continue

        if not first_chart:
            story.append(PageBreak())

        first_chart = False

        story.append(
            Paragraph(
                chart.stem.replace("_", " "),
                styles["Heading3"],
            )
        )

        story.append(Spacer(1, 6))

        story.append(
            PDFImage(
                str(chart),
                width=9.8 * inch,
                height=5.8 * inch,
            )
        )

    document.build(story)


# ============================================================
# 10) MAIN
# ============================================================

def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Script 09: combine RGB and multispectral outputs into "
            "one Second Trial synthesis report."
        )
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate inputs only.",
    )

    args = parser.parse_args()

    tray_design = load_tray_design()
    rgb_metrics = load_rgb_metrics()
    ms_data = load_ms_data()
    ms_metrics = calculate_ms_tray_metrics(ms_data)

    missing_rgb = sorted(
        set(tray_design["tray_no"])
        - set(rgb_metrics["tray_no"])
    )

    missing_ms = sorted(
        set(tray_design["tray_no"])
        - set(ms_metrics["tray_no"])
    )

    print("\nSCRIPT 09 - SECOND TRIAL SYNTHESIS")
    print("=" * 70)
    print(f"Tray Status:\n{TRAY_STATUS_XLSX}")
    print(f"\nScript 05 RGB metrics:\n{RGB_TRAY_METRICS_CSV}")
    print(f"\nScript 07 MS metrics:\n{MS_TRAY_SUMMARY_CSV}")
    print(f"\nOutput folder:\n{OUTPUT_ROOT}")

    if missing_rgb:
        print(
            "\nWARNING: missing RGB metric rows for trays: "
            + ", ".join(map(str, missing_rgb))
        )

    if missing_ms:
        print(
            "\nWARNING: missing MS metric rows for trays: "
            + ", ".join(map(str, missing_ms))
        )

    if args.dry_run:
        print("\nDry run complete. No outputs created.")
        return 0

    CHARTS_ROOT.mkdir(parents=True, exist_ok=True)
    REPORTS_ROOT.mkdir(parents=True, exist_ok=True)
    CONFIG_ROOT.mkdir(parents=True, exist_ok=True)

    master = create_master_summary(
        tray_design,
        rgb_metrics,
        ms_metrics,
    )

    group_summary = create_group_summary(master)

    chart_files = [
        CHARTS_ROOT / "01_final_tray_score_ranking.png",
        CHARTS_ROOT / "02_final_group_score_ranking.png",
        CHARTS_ROOT / "03_day5_rgb_green_cover_vs_ndvi.png",
        CHARTS_ROOT / "04_day5_rgb_green_cover_vs_ndre.png",
        CHARTS_ROOT / "05_treatment_summary_heatmap.png",
    ]

    save_tray_score_ranking(
        master,
        chart_files[0],
    )

    save_group_score_ranking(
        group_summary,
        chart_files[1],
    )

    save_scatter(
        master,
        "day5_green_cover_percent",
        "day5_mean_ndvi",
        "Day 5 RGB green cover vs relative NDVI",
        "Day 5 RGB green-cover proxy (%)",
        "Day 5 mean relative NDVI",
        chart_files[2],
    )

    save_scatter(
        master,
        "day5_green_cover_percent",
        "day5_mean_ndre",
        "Day 5 RGB green cover vs relative NDRE",
        "Day 5 RGB green-cover proxy (%)",
        "Day 5 mean relative NDRE",
        chart_files[3],
    )

    save_group_heatmap(
        group_summary,
        chart_files[4],
    )

    master_csv = (
        REPORTS_ROOT
        / "second_trial_master_summary.csv"
    )

    group_csv = (
        REPORTS_ROOT
        / "second_trial_group_summary.csv"
    )

    excel_report = (
        REPORTS_ROOT
        / "second_trial_master_summary.xlsx"
    )

    pdf_report = (
        REPORTS_ROOT
        / "second_trial_synthesis_report.pdf"
    )

    safe_round_dataframe(master, 6).to_csv(
        master_csv,
        index=False,
    )

    safe_round_dataframe(group_summary, 6).to_csv(
        group_csv,
        index=False,
    )

    create_excel_report(
        excel_report,
        master,
        group_summary,
        tray_design,
    )

    create_pdf_report(
        pdf_report,
        master,
        group_summary,
        chart_files,
    )

    settings = {
        "purpose": (
            "Combine Second Trial RGB and multispectral analysis results "
            "into one final synthesis dataset and PDF report."
        ),
        "input_tray_status": str(TRAY_STATUS_XLSX),
        "input_rgb_metrics": str(RGB_TRAY_METRICS_CSV),
        "input_ms_metrics": str(MS_TRAY_SUMMARY_CSV),
        "output_master_csv": str(master_csv),
        "output_group_csv": str(group_csv),
        "output_excel_report": str(excel_report),
        "output_pdf_report": str(pdf_report),
        "report_format": "PDF",
        "score_components": PERFORMANCE_COMPONENTS,
        "score_method": (
            "Each available component is min-max normalised from 0 to 100 "
            "across trays, then averaged to create the overall observed "
            "performance score."
        ),
        "interpretation_note": (
            "The score is descriptive and should not be treated as a formal "
            "statistical test."
        ),
        "day_1_to_day_5_rule": (
            "Day 1 to Day 5 is the continuous growth-rate window."
        ),
        "day_9_rule": (
            "Day 9 is treated as a later follow-up observation."
        ),
    }

    (
        CONFIG_ROOT
        / "synthesis_settings.json"
    ).write_text(
        json.dumps(settings, indent=2),
        encoding="utf-8",
    )

    best_tray = safe_idxmax_row(
        master,
        "overall_observed_performance_score",
    )

    best_group = safe_idxmax_row(
        group_summary.loc[
            group_summary["group_type"].eq("Treatment x Environment")
        ],
        "mean_overall_score",
    )

    print("\n" + "=" * 70)
    print("SCRIPT 09 FINISHED")
    print("=" * 70)

    if best_tray is not None:
        print(
            "Highest ranked tray: "
            f"{best_tray['tray']} | {best_tray['interaction']} | "
            f"score={best_tray['overall_observed_performance_score']:.2f}"
        )

    if best_group is not None:
        print(
            "Highest ranked group: "
            f"{best_group['group']} | "
            f"mean score={best_group['mean_overall_score']:.2f}"
        )

    print(f"\nMaster CSV:\n{master_csv}")
    print(f"\nGroup CSV:\n{group_csv}")
    print(f"\nExcel report:\n{excel_report}")
    print(f"\nPDF report:\n{pdf_report}")
    print(f"\nCharts:\n{CHARTS_ROOT}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())