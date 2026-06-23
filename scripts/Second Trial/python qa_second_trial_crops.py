"""
SECOND TRIAL CROPPED OUTPUT QA SCRIPT

This script does NOT modify, move, or crop any image.

It checks cropped outputs in:

OneDrive/Desktop/Internship/outputs/Second Trial/

It creates:

outputs/Second Trial/QA/
    second_trial_crop_qa_report.xlsx
    second_trial_crop_qa_summary.csv
    previews/
        <Day>/
            <Tray>/
                <capture_set>_preview.png

Expected useful files per capture set:
    D.JPG
    MS_G.TIF
    MS_R.TIF
    MS_RE.TIF
    MS_NIR.TIF

F.JPG is intentionally ignored.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
from datetime import datetime
from pathlib import Path

import numpy as np
import tifffile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont


# ============================================================
# SETTINGS
# ============================================================

# Leave as None if your OneDrive path is normal.
# If needed, enter your exact path here.
#
# Example:
# INTERNSHIP_ROOT = Path(r"C:\Users\tshib\OneDrive\Desktop\Internship")

INTERNSHIP_ROOT = None

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".tif", ".tiff"}

REQUIRED_BANDS = ["D", "MS_G", "MS_R", "MS_RE", "MS_NIR"]

OUTPUT_QA_FOLDER = "QA"
EXCEL_REPORT_NAME = "second_trial_crop_qa_report.xlsx"
CSV_SUMMARY_NAME = "second_trial_crop_qa_summary.csv"

PREVIEW_WIDTH = 360
PREVIEW_HEIGHT = 260


# ============================================================
# PATH FUNCTIONS
# ============================================================

def get_internship_root() -> Path:
    """Find OneDrive/Desktop/Internship automatically."""

    if INTERNSHIP_ROOT is not None:
        return Path(INTERNSHIP_ROOT)

    candidates = []

    if os.environ.get("OneDrive"):
        candidates.append(
            Path(os.environ["OneDrive"]) / "Desktop" / "Internship"
        )

    candidates.append(Path.home() / "OneDrive" / "Desktop" / "Internship")
    candidates.append(Path.home() / "Desktop" / "Internship")

    for candidate in candidates:
        if (candidate / "outputs" / "Second Trial").exists():
            return candidate

    return candidates[0]


# ============================================================
# GENERAL HELPERS
# ============================================================

def natural_sort_key(text: str):
    """Sort Tray 2 before Tray 10."""
    return [
        int(part) if part.isdigit() else part.lower()
        for part in re.split(r"(\d+)", text)
    ]


def parse_filter_list(value):
    """Convert comma-separated command-line values into lowercase sets."""

    if not value:
        return None

    return {
        item.strip().lower()
        for item in value.split(",")
        if item.strip()
    }


def get_tray_number(tray_name: str):
    """Extract number from names such as Tray 1."""
    match = re.search(r"(\d+)", tray_name)
    return int(match.group(1)) if match else None


def relative_or_blank(path: Path | None, root: Path) -> str:
    """Return a safe relative path."""
    if path is None:
        return ""

    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


# ============================================================
# DJI FILE CLASSIFICATION
# ============================================================

def has_filename_token(stem: str, token: str) -> bool:
    """
    Check file tokens safely.

    Example:
    DJI_0001_MS_NIR.TIF -> MS_NIR
    DJI_0001_D.JPG      -> D
    """

    pattern = rf"(?:^|_){re.escape(token)}(?:_|$)"
    return re.search(pattern, stem.upper()) is not None


def classify_image(filename: str):
    """
    Return:
        capture_id, band

    Examples:
        DJI_0001_D.JPG      -> DJI_0001, D
        DJI_0001_MS_NIR.TIF -> DJI_0001, MS_NIR

    F is intentionally ignored.
    """

    stem = Path(filename).stem.upper()

    band_patterns = [
        ("MS_NIR", "MS_NIR"),
        ("MS_RE", "MS_RE"),
        ("MS_R", "MS_R"),
        ("MS_G", "MS_G"),
        ("D", "D"),
        ("F", "F_IGNORED"),
    ]

    for token, label in band_patterns:
        if has_filename_token(stem, token):
            suffix_pattern = rf"_(?:{re.escape(token)})$"
            capture_id = re.sub(suffix_pattern, "", stem)

            if not capture_id:
                capture_id = stem

            return capture_id, label

    return stem, "UNKNOWN"


# ============================================================
# TRAY STATUS READING
# ============================================================

def is_present(value) -> bool:
    """Treat P as a positive tray-status marker."""
    return str(value).strip().upper() == "P"


def read_tray_status(xlsx_path: Path):
    """
    Expected columns:

    Tray No | Microbes | No Microbes | Inside | Outside
    """

    if not xlsx_path.exists():
        print(f"WARNING: Tray Status.xlsx not found: {xlsx_path}")
        return {}

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        return {}

    headers = {
        str(value).strip().lower(): index
        for index, value in enumerate(rows[0])
        if value is not None
    }

    expected_columns = [
        "tray no",
        "microbes",
        "no microbes",
        "inside",
        "outside",
    ]

    missing_columns = [
        name for name in expected_columns
        if name not in headers
    ]

    if missing_columns:
        print(
            "WARNING: Tray Status.xlsx is missing: "
            + ", ".join(missing_columns)
        )
        return {}

    tray_data = {}

    for row in rows[1:]:
        if not row:
            continue

        tray_value = row[headers["tray no"]]

        if tray_value is None:
            continue

        try:
            tray_number = int(tray_value)
        except (TypeError, ValueError):
            continue

        microbes = is_present(row[headers["microbes"]])
        no_microbes = is_present(row[headers["no microbes"]])
        inside = is_present(row[headers["inside"]])
        outside = is_present(row[headers["outside"]])

        if microbes and not no_microbes:
            treatment = "Microbes"
        elif no_microbes and not microbes:
            treatment = "No Microbes"
        else:
            treatment = "Unclear"

        if inside and not outside:
            environment = "Inside"
        elif outside and not inside:
            environment = "Outside"
        else:
            environment = "Unclear"

        tray_data[tray_number] = {
            "tray_no": tray_number,
            "treatment": treatment,
            "environment": environment,
        }

    return tray_data


# ============================================================
# FILE INFORMATION
# ============================================================

def get_image_info(image_path: Path):
    """
    Read basic image information without changing any image.
    """

    suffix = image_path.suffix.lower()

    try:
        if suffix in {".tif", ".tiff"}:
            with tifffile.TiffFile(image_path) as tif:
                page = tif.pages[0]
                shape = tuple(page.shape)
                dtype = str(page.dtype)

                if len(shape) >= 2:
                    height = shape[0]
                    width = shape[1]
                else:
                    height = ""
                    width = ""

                return {
                    "width": width,
                    "height": height,
                    "shape": str(shape),
                    "dtype": dtype,
                    "readable": True,
                    "error": "",
                }

        with Image.open(image_path) as image:
            width, height = image.size

            return {
                "width": width,
                "height": height,
                "shape": f"({height}, {width})",
                "dtype": image.mode,
                "readable": True,
                "error": "",
            }

    except Exception as error:
        return {
            "width": "",
            "height": "",
            "shape": "",
            "dtype": "",
            "readable": False,
            "error": str(error),
        }


# ============================================================
# PREVIEW IMAGE FUNCTIONS
# ============================================================

def normalise_for_preview(image_array: np.ndarray):
    """
    Create a display-only uint8 preview.

    Original source arrays are not changed.
    """

    if image_array.ndim == 2:
        low, high = np.percentile(image_array, [1, 99])

        if high <= low:
            high = low + 1

        preview = np.clip(
            (image_array - low) * 255 / (high - low),
            0,
            255,
        ).astype(np.uint8)

        return Image.fromarray(preview).convert("RGB")

    if image_array.shape[2] > 3:
        image_array = image_array[:, :, :3]

    result = np.zeros_like(image_array, dtype=np.uint8)

    for channel in range(image_array.shape[2]):
        band = image_array[:, :, channel]

        low, high = np.percentile(band, [1, 99])

        if high <= low:
            high = low + 1

        result[:, :, channel] = np.clip(
            (band - low) * 255 / (high - low),
            0,
            255,
        ).astype(np.uint8)

    return Image.fromarray(result).convert("RGB")


def load_preview_image(image_path: Path | None):
    """
    Load one image and convert it into a visible preview.
    Returns None if unavailable.
    """

    if image_path is None or not image_path.exists():
        return None

    try:
        suffix = image_path.suffix.lower()

        if suffix in {".tif", ".tiff"}:
            image_array = tifffile.imread(image_path)
        else:
            with Image.open(image_path) as image:
                image_array = np.asarray(image.convert("RGB"))

        return normalise_for_preview(image_array)

    except Exception:
        return None


def fit_image(image: Image.Image, width: int, height: int):
    """Fit image into preview panel while keeping aspect ratio."""

    background = Image.new("RGB", (width, height), "white")

    image_copy = image.copy()
    image_copy.thumbnail((width - 10, height - 35))

    x = (width - image_copy.width) // 2
    y = 25 + (height - 25 - image_copy.height) // 2

    background.paste(image_copy, (x, y))

    return background


def make_placeholder(label: str, width: int, height: int):
    """Create placeholder panel for missing or unreadable files."""

    image = Image.new("RGB", (width, height), "#F4CCCC")
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()

    draw.text((10, 10), label, fill="black", font=font)

    return image


def create_capture_preview(
    output_path: Path,
    day_name: str,
    tray_name: str,
    capture_id: str,
    files_by_band: dict,
):
    """
    Build one contact sheet:

    D | MS_G | MS_R | MS_RE | MS_NIR
    """

    ordered_bands = ["D", "MS_G", "MS_R", "MS_RE", "MS_NIR"]

    panel_width = PREVIEW_WIDTH
    panel_height = PREVIEW_HEIGHT

    title_height = 45
    total_width = panel_width * len(ordered_bands)
    total_height = panel_height + title_height

    canvas = Image.new("RGB", (total_width, total_height), "#EDEDED")
    draw = ImageDraw.Draw(canvas)
    font = ImageFont.load_default()

    title = f"{day_name} | {tray_name} | {capture_id}"
    draw.text((10, 12), title, fill="black", font=font)

    for index, band in enumerate(ordered_bands):
        source_path = files_by_band.get(band)

        preview = load_preview_image(source_path)

        if preview is None:
            panel = make_placeholder(
                f"{band}\nMissing or unreadable",
                panel_width,
                panel_height,
            )
        else:
            panel = fit_image(preview, panel_width, panel_height)

            panel_draw = ImageDraw.Draw(panel)
            panel_draw.text(
                (10, 8),
                band,
                fill="black",
                font=font,
            )

        x_position = index * panel_width
        canvas.paste(panel, (x_position, title_height))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path)


# ============================================================
# EXCEL REPORT
# ============================================================

def write_sheet(worksheet, headers, rows):
    """Write a styled worksheet."""

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="5B9BD5")

    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for column_number, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_number, value=header)

        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = thin_border

    for row_number, row_values in enumerate(rows, start=2):
        for column_number, value in enumerate(row_values, start=1):
            cell = worksheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            cell.border = thin_border

            if headers[column_number - 1] == "Status":
                if value == "PASS":
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif value == "CHECK":
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                elif value == "FAIL":
                    cell.fill = PatternFill("solid", fgColor="F4CCCC")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            text = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(text))

        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            45,
        )

    worksheet.row_dimensions[1].height = 32


def create_excel_report(
    report_path: Path,
    summary_rows,
    file_rows,
    tray_rows,
):
    """Create a multi-sheet Excel QA workbook."""

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Capture QA"

    summary_headers = [
        "Day",
        "Tray",
        "Tray No",
        "Treatment",
        "Environment",
        "Capture Set",
        "D",
        "MS_G",
        "MS_R",
        "MS_RE",
        "MS_NIR",
        "Missing Bands",
        "Duplicate Bands",
        "MS Dimensions",
        "Preview Path",
        "Status",
        "Notes",
    ]

    summary_values = [
        [
            row["day"],
            row["tray"],
            row["tray_no"],
            row["treatment"],
            row["environment"],
            row["capture_id"],
            row["D"],
            row["MS_G"],
            row["MS_R"],
            row["MS_RE"],
            row["MS_NIR"],
            row["missing_bands"],
            row["duplicate_bands"],
            row["ms_dimensions"],
            row["preview_path"],
            row["status"],
            row["notes"],
        ]
        for row in summary_rows
    ]

    write_sheet(summary_sheet, summary_headers, summary_values)

    file_sheet = workbook.create_sheet("File Inventory")

    file_headers = [
        "Day",
        "Tray",
        "Tray No",
        "Treatment",
        "Environment",
        "Capture Set",
        "Band",
        "Filename",
        "Relative Path",
        "Width",
        "Height",
        "Shape",
        "Data Type / Mode",
        "Readable",
        "Error",
    ]

    file_values = [
        [
            row["day"],
            row["tray"],
            row["tray_no"],
            row["treatment"],
            row["environment"],
            row["capture_id"],
            row["band"],
            row["filename"],
            row["relative_path"],
            row["width"],
            row["height"],
            row["shape"],
            row["dtype"],
            row["readable"],
            row["error"],
        ]
        for row in file_rows
    ]

    write_sheet(file_sheet, file_headers, file_values)

    tray_sheet = workbook.create_sheet("Tray Summary")

    tray_headers = [
        "Day",
        "Tray",
        "Tray No",
        "Treatment",
        "Environment",
        "Capture Sets",
        "Pass",
        "Check",
        "Fail",
        "Status",
        "Notes",
    ]

    tray_values = [
        [
            row["day"],
            row["tray"],
            row["tray_no"],
            row["treatment"],
            row["environment"],
            row["capture_sets"],
            row["pass_count"],
            row["check_count"],
            row["fail_count"],
            row["status"],
            row["notes"],
        ]
        for row in tray_rows
    ]

    write_sheet(tray_sheet, tray_headers, tray_values)

    note_sheet = workbook.create_sheet("Read Me")

    note_sheet["A1"] = "Second Trial Crop QA Notes"
    note_sheet["A1"].font = Font(bold=True, size=14)

    notes = [
        "This workbook is a non-destructive quality-control report.",
        "F preview images are intentionally ignored.",
        "D crop dimensions do not need to match the multispectral crop dimensions.",
        "MS_G, MS_R, MS_RE and MS_NIR should have matching crop dimensions within one capture set.",
        "PASS means all five required files were found, readable, and MS dimensions matched.",
        "CHECK means manual review is needed, such as duplicate files or incomplete information.",
        "FAIL means required bands are missing, unreadable, or MS dimensions are inconsistent.",
        "Preview sheets are saved as PNG files in the QA/previews folder.",
    ]

    for row_number, note in enumerate(notes, start=3):
        note_sheet.cell(row=row_number, column=1, value=note)

    note_sheet.column_dimensions["A"].width = 120

    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(report_path)


# ============================================================
# MAIN QA PROCESS
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Run QA checks on Second Trial cropped outputs."
    )

    parser.add_argument(
        "--days",
        help='Optional filter. Example: --days "First Day,Second Day"',
    )

    parser.add_argument(
        "--trays",
        help='Optional filter. Example: --trays "Tray 1,Tray 2"',
    )

    parser.add_argument(
        "--no-previews",
        action="store_true",
        help="Do not generate PNG preview sheets.",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show detected folders but do not create reports.",
    )

    args = parser.parse_args()

    internship_root = get_internship_root()

    input_root = internship_root / "outputs" / "Second Trial"
    source_data_root = internship_root / "data" / "Second Trial"

    tray_status_path = source_data_root / "Tray Status.xlsx"

    qa_root = input_root / OUTPUT_QA_FOLDER
    previews_root = qa_root / "previews"

    excel_report_path = qa_root / EXCEL_REPORT_NAME
    csv_summary_path = qa_root / CSV_SUMMARY_NAME

    selected_days = parse_filter_list(args.days)
    selected_trays = parse_filter_list(args.trays)

    print("\nSECOND TRIAL CROPPED OUTPUT QA")
    print("=" * 70)
    print(f"Crop output root: {input_root}")
    print(f"QA output root:   {qa_root}")
    print(f"Tray Status file: {tray_status_path}")

    if not input_root.exists():
        print("\nERROR: Cropped output folder not found.")
        print(f"Expected:\n{input_root}")
        return 1

    tray_metadata = read_tray_status(tray_status_path)

    day_folders = sorted(
        [
            folder
            for folder in input_root.iterdir()
            if folder.is_dir() and folder.name != OUTPUT_QA_FOLDER
        ],
        key=lambda folder: natural_sort_key(folder.name),
    )

    summary_rows = []
    file_rows = []
    tray_rows = []

    discovered_pairs = []

    for day_folder in day_folders:
        if selected_days and day_folder.name.lower() not in selected_days:
            continue

        tray_folders = sorted(
            [
                folder
                for folder in day_folder.iterdir()
                if folder.is_dir()
            ],
            key=lambda folder: natural_sort_key(folder.name),
        )

        for tray_folder in tray_folders:
            if selected_trays and tray_folder.name.lower() not in selected_trays:
                continue

            discovered_pairs.append((day_folder, tray_folder))

    if not discovered_pairs:
        print("\nNo Day/Tray folders matched your filters.")
        return 1

    print("\nDetected folders:")

    for day_folder, tray_folder in discovered_pairs:
        print(f"  {day_folder.name} > {tray_folder.name}")

    if args.dry_run:
        print("\nDry run complete. No report or previews created.")
        return 0

    qa_root.mkdir(parents=True, exist_ok=True)

    for day_folder, tray_folder in discovered_pairs:
        day_name = day_folder.name
        tray_name = tray_folder.name

        tray_number = get_tray_number(tray_name)

        metadata = tray_metadata.get(
            tray_number,
            {
                "tray_no": tray_number if tray_number is not None else "",
                "treatment": "Unknown",
                "environment": "Unknown",
            },
        )

        image_files = sorted(
            [
                file
                for file in tray_folder.iterdir()
                if file.is_file()
                and file.suffix.lower() in IMAGE_EXTENSIONS
            ],
            key=lambda file: natural_sort_key(file.name),
        )

        capture_sets = {}

        for image_path in image_files:
            capture_id, band = classify_image(image_path.name)

            # F is ignored entirely.
            if band == "F_IGNORED":
                continue

            # Unknown files are included in inventory but not analysis.
            if capture_id not in capture_sets:
                capture_sets[capture_id] = {}

            capture_sets[capture_id].setdefault(band, []).append(image_path)

        if not capture_sets:
            tray_rows.append(
                {
                    "day": day_name,
                    "tray": tray_name,
                    "tray_no": metadata["tray_no"],
                    "treatment": metadata["treatment"],
                    "environment": metadata["environment"],
                    "capture_sets": 0,
                    "pass_count": 0,
                    "check_count": 0,
                    "fail_count": 1,
                    "status": "FAIL",
                    "notes": "No recognised D or multispectral output files found.",
                }
            )
            continue

        tray_pass = 0
        tray_check = 0
        tray_fail = 0

        for capture_id in sorted(
            capture_sets.keys(),
            key=natural_sort_key,
        ):
            files_by_band_lists = capture_sets[capture_id]

            selected_files = {}
            duplicate_bands = []

            for band, paths in files_by_band_lists.items():
                if len(paths) > 1:
                    duplicate_bands.append(band)

                selected_files[band] = paths[0]

            missing_bands = [
                band
                for band in REQUIRED_BANDS
                if band not in selected_files
            ]

            infos = {}

            for band, image_path in selected_files.items():
                image_info = get_image_info(image_path)
                infos[band] = image_info

                file_rows.append(
                    {
                        "day": day_name,
                        "tray": tray_name,
                        "tray_no": metadata["tray_no"],
                        "treatment": metadata["treatment"],
                        "environment": metadata["environment"],
                        "capture_id": capture_id,
                        "band": band,
                        "filename": image_path.name,
                        "relative_path": relative_or_blank(
                            image_path,
                            input_root,
                        ),
                        "width": image_info["width"],
                        "height": image_info["height"],
                        "shape": image_info["shape"],
                        "dtype": image_info["dtype"],
                        "readable": "Yes" if image_info["readable"] else "No",
                        "error": image_info["error"],
                    }
                )

            unreadable_bands = [
                band
                for band, info in infos.items()
                if not info["readable"]
            ]

            ms_dimensions = []

            for band in ["MS_G", "MS_R", "MS_RE", "MS_NIR"]:
                if band in infos and infos[band]["readable"]:
                    ms_dimensions.append(
                        (
                            infos[band]["width"],
                            infos[band]["height"],
                        )
                    )

            unique_ms_dimensions = sorted(set(ms_dimensions))

            ms_dimension_match = (
                len(unique_ms_dimensions) == 1
                and len(ms_dimensions) == 4
            )

            notes = []

            if missing_bands:
                notes.append(
                    "Missing: " + ", ".join(missing_bands)
                )

            if duplicate_bands:
                notes.append(
                    "Duplicate files: " + ", ".join(duplicate_bands)
                )

            if unreadable_bands:
                notes.append(
                    "Unreadable: " + ", ".join(unreadable_bands)
                )

            if not ms_dimension_match:
                notes.append(
                    "MS crop dimensions are incomplete or inconsistent."
                )

            if missing_bands or unreadable_bands or not ms_dimension_match:
                status = "FAIL"
                tray_fail += 1
            elif duplicate_bands:
                status = "CHECK"
                tray_check += 1
            else:
                status = "PASS"
                tray_pass += 1

            preview_path = ""

            if not args.no_previews:
                safe_capture_name = re.sub(
                    r"[^A-Za-z0-9_.-]+",
                    "_",
                    capture_id,
                )

                preview_file = (
                    previews_root
                    / day_name
                    / tray_name
                    / f"{safe_capture_name}_preview.png"
                )

                create_capture_preview(
                    preview_file,
                    day_name,
                    tray_name,
                    capture_id,
                    selected_files,
                )

                preview_path = relative_or_blank(
                    preview_file,
                    input_root,
                )

            summary_rows.append(
                {
                    "day": day_name,
                    "tray": tray_name,
                    "tray_no": metadata["tray_no"],
                    "treatment": metadata["treatment"],
                    "environment": metadata["environment"],
                    "capture_id": capture_id,
                    "D": relative_or_blank(selected_files.get("D"), input_root),
                    "MS_G": relative_or_blank(selected_files.get("MS_G"), input_root),
                    "MS_R": relative_or_blank(selected_files.get("MS_R"), input_root),
                    "MS_RE": relative_or_blank(selected_files.get("MS_RE"), input_root),
                    "MS_NIR": relative_or_blank(selected_files.get("MS_NIR"), input_root),
                    "missing_bands": ", ".join(missing_bands),
                    "duplicate_bands": ", ".join(duplicate_bands),
                    "ms_dimensions": (
                        str(unique_ms_dimensions)
                        if unique_ms_dimensions
                        else ""
                    ),
                    "preview_path": preview_path,
                    "status": status,
                    "notes": " | ".join(notes) if notes else "All checks passed.",
                }
            )

        if tray_fail > 0:
            tray_status = "FAIL"
            tray_notes = "One or more capture sets failed QA."
        elif tray_check > 0:
            tray_status = "CHECK"
            tray_notes = "One or more capture sets need review."
        else:
            tray_status = "PASS"
            tray_notes = "All capture sets passed QA."

        tray_rows.append(
            {
                "day": day_name,
                "tray": tray_name,
                "tray_no": metadata["tray_no"],
                "treatment": metadata["treatment"],
                "environment": metadata["environment"],
                "capture_sets": len(capture_sets),
                "pass_count": tray_pass,
                "check_count": tray_check,
                "fail_count": tray_fail,
                "status": tray_status,
                "notes": tray_notes,
            }
        )

    create_excel_report(
        excel_report_path,
        summary_rows,
        file_rows,
        tray_rows,
    )

    csv_headers = [
        "day",
        "tray",
        "tray_no",
        "treatment",
        "environment",
        "capture_id",
        "D",
        "MS_G",
        "MS_R",
        "MS_RE",
        "MS_NIR",
        "missing_bands",
        "duplicate_bands",
        "ms_dimensions",
        "preview_path",
        "status",
        "notes",
    ]

    with open(
        csv_summary_path,
        "w",
        newline="",
        encoding="utf-8",
    ) as csv_file:
        writer = csv.DictWriter(
            csv_file,
            fieldnames=csv_headers,
        )

        writer.writeheader()
        writer.writerows(summary_rows)

    pass_count = sum(
        1 for row in summary_rows
        if row["status"] == "PASS"
    )

    check_count = sum(
        1 for row in summary_rows
        if row["status"] == "CHECK"
    )

    fail_count = sum(
        1 for row in summary_rows
        if row["status"] == "FAIL"
    )

    print("\n" + "=" * 70)
    print("QA FINISHED")
    print("=" * 70)
    print(f"Capture sets checked: {len(summary_rows)}")
    print(f"PASS: {pass_count}")
    print(f"CHECK: {check_count}")
    print(f"FAIL: {fail_count}")
    print(f"\nExcel report:\n{excel_report_path}")
    print(f"\nCSV summary:\n{csv_summary_path}")

    if not args.no_previews:
        print(f"\nPreview images:\n{previews_root}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())