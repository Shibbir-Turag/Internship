from __future__ import annotations

"""
SCRIPT 04 — THIRD TRIAL VISIBLE EMERGENCE AND RGB GREEN-COVER ANALYSIS

Purpose
-------
Analyse visible seedling emergence in Third Trial cropped D/RGB tray images
using the 70 square cell ownership zones created by Script 03.

This script DOES:
- read cropped D/RGB images from Script 01
- read D/RGB 70-cell square grid coordinates from Script 03
- read Trial 3 treatment and observation schedule from Trial_3_Tray_Status.xlsx
- detect current-day green seedling evidence inside each cell square
- track cumulative visible emergence over time
- calculate RGB green-cover percentage per cell and tray
- flag possible Day 7 bug-eaten/missing seedlings
- save overlays, masks, CSV reports, Excel report, and settings JSON

This script DOES NOT:
- estimate/impute Day 7 missing crop values
- calculate adjusted growth rates
- calculate NDVI or NDRE
- use multispectral images

Important terminology
---------------------
This measures visible emergence only. It does not prove underground germination.

Important Day 7 rule
--------------------
If a cell had visible emergence earlier but has no visible green evidence on
Day 7, this script flags it as possible Day 7 bug-eaten/missing crop.

The observed Day 7 result is still preserved. Adjustment/imputation must be
done later in Script 05, where observed and adjusted values can be kept separate.
"""

import argparse
import csv
import json
import re
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageOps


# ============================================================
# 1) PATHS — CHANGE PROJECT_ROOT ONLY IF NEEDED
# ============================================================

PROJECT_ROOT = Path(
    r"C:\Users\tshib\OneDrive\Desktop\Internship"
)

CROP_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Third trial"
    / "01_Crop_Dual_Reference"
)

GRID_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Third trial"
    / "03_Cell_Grid_Detection"
)

TRAY_STATUS_XLSX = (
    PROJECT_ROOT
    / "data"
    / "Third Trial"
    / "Trial_3_Tray_Status.xlsx"
)

OUTPUT_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Third trial"
    / "04_Visible_Emergence"
)


# ============================================================
# 2) TRIAL SETTINGS
# ============================================================

ROWS = 7
COLS = 10
EXPECTED_CELLS = ROWS * COLS

DAY_NAME_TO_ORDER = {
    "day 1": 1,
    "day 2": 2,
    "day 3": 3,
    "day 4": 4,
    "day 5": 5,
    "day 6": 6,
    "day 7": 7,
}

ACCEPTED_GRID_STATUSES = {
    "PASS_AUTO",
    "PASS_MANUAL",
}

# Green-seedling evidence settings.
# These are carried over from the successful Trial 2 Script 04 logic.
# OpenCV HSV hue range is 0–179.
GREEN_HUE_MIN = 20
GREEN_HUE_MAX = 95
GREEN_SATURATION_MIN = 28
GREEN_VALUE_MIN = 35

# Excess Green = 2G - R - B
EXCESS_GREEN_MIN = 12

# Helps reject brown mulch/soil.
GREEN_TO_RED_RATIO = 1.03
GREEN_TO_BLUE_RATIO = 1.00

# Minimum evidence required inside one square ownership zone.
MIN_GREEN_PIXELS_PER_CELL = 20
MIN_GREEN_COMPONENT_PIXELS = 12

MORPH_KERNEL_SIZE = 3


# ============================================================
# 3) FALLBACK TRAY DESIGN AND OBSERVATION SCHEDULE
# ============================================================

FALLBACK_TRAY_DESIGN = {
    1: {
        "tray_no": 1,
        "tray": "Tray 1",
        "label_code": "N/M",
        "microbe_status": "No Microbes",
        "treatment": "Ideal",
        "label_environment": "Inside",
        "environment_type": "Fixed",
        "watering_rule": "Daily watering",
    },
    2: {
        "tray_no": 2,
        "tray": "Tray 2",
        "label_code": "N/M",
        "microbe_status": "No Microbes",
        "treatment": "Ideal",
        "label_environment": "Outside",
        "environment_type": "Fixed",
        "watering_rule": "Daily watering",
    },
    3: {
        "tray_no": 3,
        "tray": "Tray 3",
        "label_code": "N/M",
        "microbe_status": "No Microbes",
        "treatment": "Moisture",
        "label_environment": "Outside",
        "environment_type": "Fixed",
        "watering_rule": "Scheduled watering",
    },
    4: {
        "tray_no": 4,
        "tray": "Tray 4",
        "label_code": "N/M",
        "microbe_status": "No Microbes",
        "treatment": "Heat",
        "label_environment": "Dynamic",
        "environment_type": "Dynamic",
        "watering_rule": "Daily watering",
    },
    5: {
        "tray_no": 5,
        "tray": "Tray 5",
        "label_code": "M",
        "microbe_status": "Microbes",
        "treatment": "Moisture",
        "label_environment": "Inside",
        "environment_type": "Fixed",
        "watering_rule": "Scheduled watering",
    },
    6: {
        "tray_no": 6,
        "tray": "Tray 6",
        "label_code": "M",
        "microbe_status": "Microbes",
        "treatment": "Heat",
        "label_environment": "Dynamic",
        "environment_type": "Dynamic",
        "watering_rule": "Daily watering",
    },
    7: {
        "tray_no": 7,
        "tray": "Tray 7",
        "label_code": "M",
        "microbe_status": "Microbes",
        "treatment": "Ideal",
        "label_environment": "Outside",
        "environment_type": "Fixed",
        "watering_rule": "Daily watering",
    },
    8: {
        "tray_no": 8,
        "tray": "Tray 8",
        "label_code": "M",
        "microbe_status": "Microbes",
        "treatment": "Heat",
        "label_environment": "Dynamic",
        "environment_type": "Dynamic",
        "watering_rule": "Daily watering",
    },
    9: {
        "tray_no": 9,
        "tray": "Tray 9",
        "label_code": "M",
        "microbe_status": "Microbes",
        "treatment": "Ideal",
        "label_environment": "Inside",
        "environment_type": "Fixed",
        "watering_rule": "Daily watering",
    },
    10: {
        "tray_no": 10,
        "tray": "Tray 10",
        "label_code": "M",
        "microbe_status": "Microbes",
        "treatment": "Moisture",
        "label_environment": "Outside",
        "environment_type": "Fixed",
        "watering_rule": "Scheduled watering",
    },
    11: {
        "tray_no": 11,
        "tray": "Tray 11",
        "label_code": "N/M",
        "microbe_status": "No Microbes",
        "treatment": "Heat",
        "label_environment": "Dynamic",
        "environment_type": "Dynamic",
        "watering_rule": "Daily watering",
    },
    12: {
        "tray_no": 12,
        "tray": "Tray 12",
        "label_code": "N/M",
        "microbe_status": "No Microbes",
        "treatment": "Moisture",
        "label_environment": "Inside",
        "environment_type": "Fixed",
        "watering_rule": "Scheduled watering",
    },
}

# Corrected Trial 3 schedule:
# Day 0 = 27/06/2026 planting.
# Day 1 photo = 29/06/2026.
# Day 7 final photo = 07/07/2026.
FALLBACK_OBSERVATION_SCHEDULE = {
    0: {
        "day_order": 0,
        "day": "Day 0",
        "calendar_date": "2026-06-27",
        "photo_taken": "No",
        "days_since_planting": 0,
        "days_since_day1": "",
        "days_since_previous_photo": "",
        "heat_tray_location": "Inside",
        "heat_phase": "Planting day; heat trays inside",
        "moisture_watered_today": "Yes",
        "moisture_phase": "Watered on planting day",
        "notes": "Seeds planted.",
    },
    1: {
        "day_order": 1,
        "day": "Day 1",
        "calendar_date": "2026-06-29",
        "photo_taken": "Yes",
        "days_since_planting": 2,
        "days_since_day1": 0,
        "days_since_previous_photo": "",
        "heat_tray_location": "Inside",
        "heat_phase": "Inside before heat exposure",
        "moisture_watered_today": "No",
        "moisture_phase": "After Day 0 watering; not watered",
        "notes": "First image set.",
    },
    2: {
        "day_order": 2,
        "day": "Day 2",
        "calendar_date": "2026-06-30",
        "photo_taken": "Yes",
        "days_since_planting": 3,
        "days_since_day1": 1,
        "days_since_previous_photo": 1,
        "heat_tray_location": "Inside",
        "heat_phase": "Inside before heat exposure",
        "moisture_watered_today": "No",
        "moisture_phase": "Dry period after Day 0 watering",
        "notes": "Second image set.",
    },
    3: {
        "day_order": 3,
        "day": "Day 3",
        "calendar_date": "2026-07-01",
        "photo_taken": "Yes",
        "days_since_planting": 4,
        "days_since_day1": 2,
        "days_since_previous_photo": 1,
        "heat_tray_location": "Outside",
        "heat_phase": "Moved outside; heat exposure begins",
        "moisture_watered_today": "Yes",
        "moisture_phase": "Watered on Day 3",
        "notes": "Heat trays moved outside; Moisture trays watered.",
    },
    4: {
        "day_order": 4,
        "day": "Day 4",
        "calendar_date": "2026-07-02",
        "photo_taken": "Yes",
        "days_since_planting": 5,
        "days_since_day1": 3,
        "days_since_previous_photo": 1,
        "heat_tray_location": "Outside",
        "heat_phase": "Outside heat exposure",
        "moisture_watered_today": "No",
        "moisture_phase": "Dry period after Day 3 watering",
        "notes": "Heat trays outside; Moisture trays not watered.",
    },
    5: {
        "day_order": 5,
        "day": "Day 5",
        "calendar_date": "2026-07-03",
        "photo_taken": "Yes",
        "days_since_planting": 6,
        "days_since_day1": 4,
        "days_since_previous_photo": 1,
        "heat_tray_location": "Outside",
        "heat_phase": "Outside heat exposure",
        "moisture_watered_today": "No",
        "moisture_phase": "Dry period after Day 3 watering",
        "notes": "Heat trays outside; Moisture trays not watered.",
    },
    6: {
        "day_order": 6,
        "day": "Day 6",
        "calendar_date": "2026-07-04",
        "photo_taken": "Yes",
        "days_since_planting": 7,
        "days_since_day1": 5,
        "days_since_previous_photo": 1,
        "heat_tray_location": "Inside",
        "heat_phase": "Moved inside; recovery begins",
        "moisture_watered_today": "Yes",
        "moisture_phase": "Watered on Day 6",
        "notes": "Heat trays moved inside; Moisture trays watered.",
    },
    7: {
        "day_order": 7,
        "day": "Day 7",
        "calendar_date": "2026-07-07",
        "photo_taken": "Yes",
        "days_since_planting": 10,
        "days_since_day1": 8,
        "days_since_previous_photo": 3,
        "heat_tray_location": "Inside",
        "heat_phase": "Inside recovery; final photo",
        "moisture_watered_today": "No",
        "moisture_phase": "Unwatered after Day 6; final photo",
        "notes": (
            "Final image set after skipped 05/07/2026 and 06/07/2026 photo days. "
            "Possible bug-eaten/missing crops should be flagged in analysis."
        ),
    },
}


# ============================================================
# 4) GENERAL HELPERS
# ============================================================

def natural_key(text: object):
    return [
        int(part) if part.isdigit() else part.casefold()
        for part in re.split(r"(\d+)", str(text))
    ]


def normalise(text: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(text).casefold())


def yes_no(value: bool) -> str:
    return "Yes" if value else "No"


def parse_filter_list(value: str | None):
    if not value:
        return None

    return {
        item.strip().casefold()
        for item in value.split(",")
        if item.strip()
    }


def tray_number_from_name(tray_name: str):
    match = re.search(r"(\d+)", tray_name)
    return int(match.group(1)) if match else ""


def relative_path(path: Path | None, root: Path) -> str:
    if path is None:
        return ""

    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


def record_key(day: str, tray: str, capture_id: str) -> str:
    return f"{day}|{tray}|{capture_id}"


def day_sort_key(folder: Path):
    return (
        DAY_NAME_TO_ORDER.get(folder.name.casefold(), 999),
        natural_key(folder.name),
    )


def safe_text(value: object, default: str = "") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def to_int_or_blank(value: object):
    if value is None or str(value).strip() == "":
        return ""

    try:
        return int(float(value))
    except (TypeError, ValueError):
        return ""


def excel_date_to_iso(value: object) -> str:
    """
    Handles:
    - Python datetime/date objects from openpyxl
    - Excel serial numbers such as 46202
    - already-formatted date strings
    """

    if value is None or str(value).strip() == "":
        return ""

    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
        except Exception:
            return str(value)

    text = str(value).strip()

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass

    return text


def header_lookup(sheet) -> dict[str, int]:
    return {
        normalise(cell.value): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }


def get_sheet_by_name(workbook, wanted_name: str):
    wanted = normalise(wanted_name)

    for sheet_name in workbook.sheetnames:
        if normalise(sheet_name) == wanted:
            return workbook[sheet_name]

    return None


def cell_value(row, headers: dict[str, int], *names: str):
    for name in names:
        key = normalise(name)
        if key in headers:
            return row[headers[key] - 1].value
    return None


# ============================================================
# 5) LOAD TRIAL 3 TRAY STATUS WORKBOOK
# ============================================================

def load_tray_design() -> tuple[dict[int, dict], str]:
    design = {
        tray_no: dict(values)
        for tray_no, values in FALLBACK_TRAY_DESIGN.items()
    }

    if not TRAY_STATUS_XLSX.exists():
        return design, (
            "WARNING: Trial_3_Tray_Status.xlsx was not found. "
            "Built-in fallback tray design was used."
        )

    workbook = load_workbook(TRAY_STATUS_XLSX, data_only=True)
    sheet = get_sheet_by_name(workbook, "Tray Status")

    if sheet is None:
        return design, (
            "WARNING: Tray Status sheet was not found. "
            "Built-in fallback tray design was used."
        )

    headers = header_lookup(sheet)
    loaded_count = 0

    for row in sheet.iter_rows(min_row=2):
        tray_no = to_int_or_blank(cell_value(row, headers, "Tray No", "TrayNo"))

        if tray_no == "":
            continue

        existing = design.get(tray_no, {}).copy()

        record = {
            "tray_no": tray_no,
            "tray": safe_text(
                cell_value(row, headers, "Tray"),
                f"Tray {tray_no}",
            ),
            "label_code": safe_text(
                cell_value(row, headers, "Label Code"),
                existing.get("label_code", ""),
            ),
            "microbe_status": safe_text(
                cell_value(row, headers, "Microbe Status"),
                existing.get("microbe_status", ""),
            ),
            "treatment": safe_text(
                cell_value(row, headers, "Treatment", "Treatment Type"),
                existing.get("treatment", ""),
            ),
            "label_environment": safe_text(
                cell_value(row, headers, "Label Environment", "Environment"),
                existing.get("label_environment", ""),
            ),
            "actual_environment_rule": safe_text(
                cell_value(row, headers, "Actual Environment Rule"),
                existing.get("actual_environment_rule", ""),
            ),
            "environment_type": safe_text(
                cell_value(row, headers, "Environment Type"),
                existing.get("environment_type", ""),
            ),
            "full_label_meaning": safe_text(
                cell_value(row, headers, "Full Label Meaning"),
                existing.get("full_label_meaning", ""),
            ),
            "watering_rule": safe_text(
                cell_value(row, headers, "Watering Rule"),
                existing.get("watering_rule", ""),
            ),
            "watering_notes": safe_text(
                cell_value(row, headers, "Watering Notes"),
                existing.get("watering_notes", ""),
            ),
        }

        design[tray_no] = record
        loaded_count += 1

    if loaded_count == 0:
        return design, (
            "WARNING: No valid tray rows were found in Tray Status sheet. "
            "Built-in fallback tray design was used."
        )

    return design, f"Tray design loaded from workbook: {loaded_count} tray rows."


def load_observation_schedule() -> tuple[dict[int, dict], str]:
    schedule = {
        day_order: dict(values)
        for day_order, values in FALLBACK_OBSERVATION_SCHEDULE.items()
    }

    if not TRAY_STATUS_XLSX.exists():
        return schedule, (
            "WARNING: Trial_3_Tray_Status.xlsx was not found. "
            "Built-in fallback schedule was used."
        )

    workbook = load_workbook(TRAY_STATUS_XLSX, data_only=True)
    sheet = get_sheet_by_name(workbook, "Observation Schedule")

    if sheet is None:
        return schedule, (
            "WARNING: Observation Schedule sheet was not found. "
            "Built-in fallback schedule was used."
        )

    headers = header_lookup(sheet)
    loaded_count = 0

    for row in sheet.iter_rows(min_row=2):
        day_order = to_int_or_blank(cell_value(row, headers, "Day Order", "DayOrder"))

        if day_order == "":
            continue

        existing = schedule.get(day_order, {}).copy()

        record = {
            "day_order": day_order,
            "day": safe_text(
                cell_value(row, headers, "Day"),
                existing.get("day", f"Day {day_order}"),
            ),
            "calendar_date": excel_date_to_iso(
                cell_value(row, headers, "Calendar Date", "Date")
            ),
            "photo_taken": safe_text(
                cell_value(row, headers, "Photo Taken"),
                existing.get("photo_taken", ""),
            ),
            "days_since_planting": to_int_or_blank(
                cell_value(row, headers, "Days Since Planting")
            ),
            "days_since_day1": to_int_or_blank(
                cell_value(row, headers, "Days Since Day 1")
            ),
            "days_since_previous_photo": to_int_or_blank(
                cell_value(row, headers, "Days Since Previous Photo")
            ),
            "heat_tray_location": safe_text(
                cell_value(row, headers, "Heat Tray Location"),
                existing.get("heat_tray_location", ""),
            ),
            "heat_phase": safe_text(
                cell_value(row, headers, "Heat Phase"),
                existing.get("heat_phase", ""),
            ),
            "moisture_watered_today": safe_text(
                cell_value(row, headers, "Moisture Watered Today"),
                existing.get("moisture_watered_today", ""),
            ),
            "moisture_phase": safe_text(
                cell_value(row, headers, "Moisture Phase"),
                existing.get("moisture_phase", ""),
            ),
            "notes": safe_text(
                cell_value(row, headers, "Notes"),
                existing.get("notes", ""),
            ),
        }

        schedule[day_order] = record
        loaded_count += 1

    if loaded_count == 0:
        return schedule, (
            "WARNING: No valid rows were found in Observation Schedule. "
            "Built-in fallback schedule was used."
        )

    return schedule, f"Observation schedule loaded from workbook: {loaded_count} day rows."


def tray_day_metadata(
    tray_no: int,
    day_order: int,
    tray_design: dict[int, dict],
    observation_schedule: dict[int, dict],
) -> dict:
    tray = tray_design.get(
        tray_no,
        {
            "tray_no": tray_no,
            "tray": f"Tray {tray_no}",
            "label_code": "",
            "microbe_status": "",
            "treatment": "",
            "label_environment": "",
            "environment_type": "",
            "watering_rule": "",
            "watering_notes": "",
        },
    )

    day = observation_schedule.get(day_order, {})

    treatment = safe_text(tray.get("treatment", ""))
    label_environment = safe_text(tray.get("label_environment", ""))

    if treatment.casefold() == "heat":
        observed_environment = safe_text(day.get("heat_tray_location", ""))
        heat_phase_for_tray = safe_text(day.get("heat_phase", ""))
    else:
        observed_environment = label_environment
        heat_phase_for_tray = "N/A"

    if treatment.casefold() == "moisture":
        moisture_watered_today_for_tray = safe_text(day.get("moisture_watered_today", ""))
        moisture_phase_for_tray = safe_text(day.get("moisture_phase", ""))
    else:
        moisture_watered_today_for_tray = "Daily watering"
        moisture_phase_for_tray = "Daily watering control"

    return {
        **tray,
        **day,
        "observed_environment": observed_environment,
        "heat_phase_for_tray": heat_phase_for_tray,
        "moisture_watered_today_for_tray": moisture_watered_today_for_tray,
        "moisture_phase_for_tray": moisture_phase_for_tray,
        "day7_bug_rule_applies": yes_no(day_order == 7),
    }


# ============================================================
# 6) FIND CROPPED D/RGB FILES
# ============================================================

def parse_d_image(path: Path):
    """
    Example:
        DJI_20260629153749_0001_D.JPG
    """

    match = re.match(
        r"^(?P<capture>.+)_D$",
        path.stem.upper(),
    )

    return match.group("capture") if match else None


def find_d_images(tray_folder: Path):
    images = []

    for file in tray_folder.iterdir():
        if not file.is_file():
            continue

        if file.suffix.casefold() not in {".jpg", ".jpeg"}:
            continue

        capture_id = parse_d_image(file)

        if capture_id:
            images.append(
                {
                    "capture_id": capture_id,
                    "path": file,
                }
            )

    return sorted(
        images,
        key=lambda item: natural_key(item["capture_id"]),
    )


def collect_jobs(days_filter=None, trays_filter=None):
    if not CROP_ROOT.exists():
        raise FileNotFoundError(
            f"Input crop folder not found:\n{CROP_ROOT}"
        )

    jobs = []

    day_folders = sorted(
        [
            folder
            for folder in CROP_ROOT.iterdir()
            if folder.is_dir()
            and folder.name.casefold() in DAY_NAME_TO_ORDER
        ],
        key=day_sort_key,
    )

    for day_folder in day_folders:
        if days_filter and day_folder.name.casefold() not in days_filter:
            continue

        tray_folders = sorted(
            [
                folder
                for folder in day_folder.iterdir()
                if folder.is_dir()
                and folder.name.casefold().startswith("tray")
            ],
            key=lambda folder: natural_key(folder.name),
        )

        for tray_folder in tray_folders:
            if trays_filter and tray_folder.name.casefold() not in trays_filter:
                continue

            for image in find_d_images(tray_folder):
                jobs.append(
                    {
                        "day": day_folder.name,
                        "day_order": DAY_NAME_TO_ORDER.get(
                            day_folder.name.casefold(),
                            999,
                        ),
                        "tray": tray_folder.name,
                        "tray_no": tray_number_from_name(tray_folder.name),
                        "capture_id": image["capture_id"],
                        "path": image["path"],
                    }
                )

    return sorted(
        jobs,
        key=lambda item: (
            item["day_order"],
            int(item["tray_no"]),
            natural_key(item["capture_id"]),
        ),
    )


def read_rgb(path: Path):
    with Image.open(path) as image:
        image = ImageOps.exif_transpose(image)
        image = image.convert("RGB")
        return np.asarray(image)


# ============================================================
# 7) LOAD SCRIPT 03 GRID OUTPUTS
# ============================================================

def load_csv_rows(path: Path):
    if not path.exists():
        return []

    with path.open(
        "r",
        newline="",
        encoding="utf-8",
    ) as file:
        return list(csv.DictReader(file))


def load_grid_manifest(accept_check_auto: bool):
    manifest_path = (
        GRID_ROOT
        / "_reports"
        / "cell_grid_manifest.csv"
    )

    accepted_statuses = set(ACCEPTED_GRID_STATUSES)

    if accept_check_auto:
        accepted_statuses.add("CHECK_AUTO")

    manifest = {}

    for row in load_csv_rows(manifest_path):
        key = record_key(
            row.get("day", ""),
            row.get("tray", ""),
            row.get("capture_id", ""),
        )

        status = safe_text(row.get("status", "")).upper()

        row["accepted_for_script04"] = yes_no(
            status in accepted_statuses
        )

        manifest[key] = row

    return manifest


def load_cell_coordinates():
    coordinates_path = (
        GRID_ROOT
        / "_reports"
        / "cell_coordinates.csv"
    )

    grouped = defaultdict(list)

    for row in load_csv_rows(coordinates_path):
        key = record_key(
            row.get("day", ""),
            row.get("tray", ""),
            row.get("capture_id", ""),
        )

        try:
            grouped[key].append(
                {
                    "cell_id": int(row["cell_id"]),
                    "row": int(row["row"]),
                    "column": int(row["column"]),
                    "x": float(row["x"]),
                    "y": float(row["y"]),
                    "square_x0": float(row["square_x0"]),
                    "square_y0": float(row["square_y0"]),
                    "square_x1": float(row["square_x1"]),
                    "square_y1": float(row["square_y1"]),
                    "square_side": float(row["square_side"]),
                    "grid_error": float(row.get("grid_error", 0.0) or 0.0),
                    "coordinate_source": row.get("coordinate_source", ""),
                    "needs_review": row.get("needs_review", ""),
                }
            )
        except (KeyError, TypeError, ValueError):
            continue

    for key in grouped:
        grouped[key].sort(
            key=lambda item: item["cell_id"]
        )

    return grouped


# ============================================================
# 8) GREEN EVIDENCE DETECTION
# ============================================================

def build_green_mask(rgb: np.ndarray):
    """
    Build a conservative vegetation-like green mask.

    It combines:
    - HSV green range
    - saturation/value filters
    - excess green
    - green stronger than red and blue
    - small morphological cleanup
    """

    hsv = cv2.cvtColor(
        rgb,
        cv2.COLOR_RGB2HSV,
    )

    red = rgb[:, :, 0].astype(np.float32)
    green = rgb[:, :, 1].astype(np.float32)
    blue = rgb[:, :, 2].astype(np.float32)

    hue = hsv[:, :, 0]
    saturation = hsv[:, :, 1]
    value = hsv[:, :, 2]

    excess_green = (
        2.0 * green
        - red
        - blue
    )

    mask = (
        (hue >= GREEN_HUE_MIN)
        & (hue <= GREEN_HUE_MAX)
        & (saturation >= GREEN_SATURATION_MIN)
        & (value >= GREEN_VALUE_MIN)
        & (green >= red * GREEN_TO_RED_RATIO)
        & (green >= blue * GREEN_TO_BLUE_RATIO)
        & (excess_green >= EXCESS_GREEN_MIN)
    ).astype(np.uint8) * 255

    kernel = cv2.getStructuringElement(
        cv2.MORPH_ELLIPSE,
        (MORPH_KERNEL_SIZE, MORPH_KERNEL_SIZE),
    )

    mask = cv2.morphologyEx(
        mask,
        cv2.MORPH_OPEN,
        kernel,
        iterations=1,
    )

    mask = cv2.morphologyEx(
        mask,
        cv2.MORPH_CLOSE,
        kernel,
        iterations=1,
    )

    return mask


def clip_square_to_image(
    cell: dict,
    image_width: int,
    image_height: int,
):
    x0 = max(0, int(round(cell["square_x0"])))
    y0 = max(0, int(round(cell["square_y0"])))
    x1 = min(image_width, int(round(cell["square_x1"])))
    y1 = min(image_height, int(round(cell["square_y1"])))

    return x0, y0, x1, y1


def analyse_cell_green_evidence(
    green_mask: np.ndarray,
    cell: dict,
    min_green_pixels: int,
    min_component_pixels: int,
):
    image_height, image_width = green_mask.shape[:2]

    x0, y0, x1, y1 = clip_square_to_image(
        cell,
        image_width,
        image_height,
    )

    if x1 <= x0 or y1 <= y0:
        return {
            "raw_green_detected": False,
            "green_pixels": 0,
            "largest_green_component": 0,
            "green_area_percent": 0.0,
            "zone_pixels": 0,
        }

    zone = green_mask[y0:y1, x0:x1]

    zone_pixels = int(zone.size)
    green_pixels = int(np.count_nonzero(zone))

    if green_pixels == 0:
        return {
            "raw_green_detected": False,
            "green_pixels": 0,
            "largest_green_component": 0,
            "green_area_percent": 0.0,
            "zone_pixels": zone_pixels,
        }

    labels_count, _labels, stats, _centres = cv2.connectedComponentsWithStats(
        zone,
        connectivity=8,
    )

    component_areas = [
        int(stats[index, cv2.CC_STAT_AREA])
        for index in range(1, labels_count)
    ]

    largest_component = max(
        component_areas,
        default=0,
    )

    green_area_percent = (
        green_pixels
        / max(zone_pixels, 1)
        * 100.0
    )

    raw_green_detected = (
        green_pixels >= min_green_pixels
        and largest_component >= min_component_pixels
    )

    return {
        "raw_green_detected": raw_green_detected,
        "green_pixels": green_pixels,
        "largest_green_component": largest_component,
        "green_area_percent": green_area_percent,
        "zone_pixels": zone_pixels,
    }


# ============================================================
# 9) VISUAL OUTPUTS
# ============================================================

def save_green_mask(
    green_mask: np.ndarray,
    output_path: Path,
):
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    cv2.imwrite(
        str(output_path),
        green_mask,
    )


def save_emergence_overlay(
    rgb: np.ndarray,
    green_mask: np.ndarray,
    cell_results: list[dict],
    output_path: Path,
    title: str,
):
    """
    Overlay colours:
    Green  = raw green evidence visible today
    Blue   = no raw evidence today, but emerged earlier
    Grey   = no visible emergence yet
    Yellow = Script 03 grid cell marked needs_review
    Red    = possible Day 7 bug-eaten/missing crop
    """

    overlay = cv2.cvtColor(
        rgb,
        cv2.COLOR_RGB2BGR,
    )

    green_layer = np.zeros_like(overlay)
    green_layer[:, :] = (0, 255, 0)

    blended = cv2.addWeighted(
        overlay,
        0.60,
        green_layer,
        0.40,
        0,
    )

    overlay[green_mask > 0] = blended[green_mask > 0]

    image_height, image_width = overlay.shape[:2]

    median_side = float(
        np.median(
            [row["square_side"] for row in cell_results]
        )
    )

    line_width = max(
        2,
        int(round(median_side / 32)),
    )

    font_scale = max(
        0.35,
        min(1.0, median_side / 88),
    )

    font_thickness = max(
        1,
        int(round(font_scale * 2)),
    )

    for row in cell_results:
        x0 = max(0, int(round(row["square_x0"])))
        y0 = max(0, int(round(row["square_y0"])))
        x1 = min(image_width - 1, int(round(row["square_x1"])))
        y1 = min(image_height - 1, int(round(row["square_y1"])))

        if row["possible_day7_bug_eaten_bool"]:
            colour = (0, 0, 220)          # Red
        elif str(row["grid_needs_review"]).casefold() == "yes":
            colour = (0, 220, 255)        # Yellow
        elif row["raw_green_bool"]:
            colour = (0, 200, 0)          # Green
        elif row["tracked_visible_emerged_bool"]:
            colour = (255, 150, 0)        # Blue
        else:
            colour = (135, 135, 135)      # Grey

        cv2.rectangle(
            overlay,
            (x0, y0),
            (x1, y1),
            colour,
            line_width,
        )

        text_x = max(
            1,
            int(round(row["x"] - median_side * 0.10)),
        )

        text_y = min(
            image_height - 5,
            int(round(row["y"] + median_side * 0.08)),
        )

        cv2.putText(
            overlay,
            str(row["cell_id"]),
            (text_x, text_y),
            cv2.FONT_HERSHEY_SIMPLEX,
            font_scale,
            colour,
            font_thickness,
            cv2.LINE_AA,
        )

    raw_count = sum(
        bool(row["raw_green_bool"])
        for row in cell_results
    )

    tracked_count = sum(
        bool(row["tracked_visible_emerged_bool"])
        for row in cell_results
    )

    bug_count = sum(
        bool(row["possible_day7_bug_eaten_bool"])
        for row in cell_results
    )

    header_height = max(
        48,
        int(round(median_side * 0.36)),
    )

    cv2.rectangle(
        overlay,
        (0, 0),
        (image_width, header_height),
        (255, 255, 255),
        thickness=-1,
    )

    header_text = (
        f"{title} | Raw green: {raw_count}/70 | "
        f"Tracked emerged: {tracked_count}/70 | "
        f"Possible Day 7 missing: {bug_count}"
    )

    cv2.putText(
        overlay,
        header_text,
        (12, int(header_height * 0.68)),
        cv2.FONT_HERSHEY_SIMPLEX,
        max(0.42, min(0.82, font_scale)),
        (0, 0, 0),
        2,
        cv2.LINE_AA,
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    cv2.imwrite(
        str(output_path),
        overlay,
        [
            cv2.IMWRITE_JPEG_QUALITY,
            95,
        ],
    )


# ============================================================
# 10) REPORT WRITING
# ============================================================

def style_excel_report(path: Path):
    workbook = load_workbook(path)

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 34

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(
                color="FFFFFF",
                bold=True,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            letter = column_cells[0].column_letter

            longest = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )

            worksheet.column_dimensions[letter].width = min(
                max(12, longest + 2),
                58,
            )

    workbook.save(path)


def write_reports(
    tray_rows: list[dict],
    cell_rows: list[dict],
    first_rows: list[dict],
    settings: dict,
):
    reports_root = OUTPUT_ROOT / "_reports"
    config_root = OUTPUT_ROOT / "_config"

    reports_root.mkdir(parents=True, exist_ok=True)
    config_root.mkdir(parents=True, exist_ok=True)

    tray_csv = reports_root / "visible_emergence_tray_summary.csv"
    cell_csv = reports_root / "visible_emergence_cell_results.csv"
    first_csv = reports_root / "first_emergence_summary.csv"
    excel_path = reports_root / "visible_emergence_report.xlsx"
    settings_path = config_root / "visible_emergence_settings.json"

    tray_frame = pd.DataFrame(tray_rows)
    cell_frame = pd.DataFrame(cell_rows)
    first_frame = pd.DataFrame(first_rows)

    tray_frame.to_csv(tray_csv, index=False)
    cell_frame.to_csv(cell_csv, index=False)
    first_frame.to_csv(first_csv, index=False)

    readme_frame = pd.DataFrame(
        {
            "Notes": [
                "Raw Current Green Evidence = visible green seedling evidence in that exact image/day.",
                "Tracked Visible Emerged = cumulative visible emergence after the first raw-positive observation.",
                "RGB green-cover percentage is calculated from green pixels inside each Script 03 square ownership zone.",
                "Script 04 preserves observed Day 7 values only.",
                "Possible Day 7 bug-eaten/missing crop cells are flagged only.",
                "Adjusted/imputed Day 7 values must be created later in Script 05, not here.",
                "Corrected Trial 3 schedule is used: Day 1 photo date = 29/06/2026 and Day 7 photo date = 07/07/2026.",
            ]
        }
    )

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
    ) as writer:
        tray_frame.to_excel(
            writer,
            sheet_name="Tray Summary",
            index=False,
        )

        cell_frame.to_excel(
            writer,
            sheet_name="Cell Results",
            index=False,
        )

        first_frame.to_excel(
            writer,
            sheet_name="First Emergence",
            index=False,
        )

        readme_frame.to_excel(
            writer,
            sheet_name="Read Me",
            index=False,
        )

    style_excel_report(excel_path)

    with settings_path.open("w", encoding="utf-8") as file:
        json.dump(
            settings,
            file,
            indent=2,
        )

    return {
        "tray_csv": tray_csv,
        "cell_csv": cell_csv,
        "first_csv": first_csv,
        "excel_path": excel_path,
        "settings_path": settings_path,
    }


# ============================================================
# 11) FIRST EMERGENCE SUMMARY
# ============================================================

def build_first_emergence_rows(
    tracking: dict,
    cell_identity: dict,
    possible_bug_keys: set,
):
    rows = []

    for key in sorted(cell_identity):
        state = tracking.get(
            key,
            {
                "emerged": False,
                "first_day": "",
                "first_day_order": "",
                "first_date": "",
            },
        )

        identity = cell_identity[key]

        possible_bug = key in possible_bug_keys

        rows.append(
            {
                **identity,
                "ever_visible_emerged": yes_no(bool(state["emerged"])),
                "first_visible_emergence_day": state["first_day"],
                "first_visible_emergence_day_order": state["first_day_order"],
                "first_visible_emergence_date": state["first_date"],
                "possible_day7_bug_eaten": yes_no(possible_bug),
                "notes": (
                    "Possible Day 7 disappearance after earlier emergence. "
                    "Use this flag only in later adjusted growth-rate analysis."
                    if possible_bug
                    else ""
                ),
            }
        )

    return rows


# ============================================================
# 12) MAIN ANALYSIS
# ============================================================

def make_skipped_tray_row(
    job: dict,
    meta: dict,
    status: str,
    notes: str,
):
    return {
        "day_order": job["day_order"],
        "day": job["day"],
        "calendar_date": meta.get("calendar_date", ""),
        "days_since_planting": meta.get("days_since_planting", ""),
        "days_since_day1": meta.get("days_since_day1", ""),
        "days_since_previous_photo": meta.get("days_since_previous_photo", ""),
        "tray": job["tray"],
        "tray_no": job["tray_no"],
        "label_code": meta.get("label_code", ""),
        "microbe_status": meta.get("microbe_status", ""),
        "treatment": meta.get("treatment", ""),
        "label_environment": meta.get("label_environment", ""),
        "observed_environment": meta.get("observed_environment", ""),
        "environment_type": meta.get("environment_type", ""),
        "watering_rule": meta.get("watering_rule", ""),
        "heat_phase": meta.get("heat_phase_for_tray", ""),
        "moisture_watered_today": meta.get("moisture_watered_today_for_tray", ""),
        "moisture_phase": meta.get("moisture_phase_for_tray", ""),
        "day7_bug_rule_applies": meta.get("day7_bug_rule_applies", ""),
        "capture_id": job["capture_id"],
        "raw_green_cells": "",
        "raw_green_percent": "",
        "tracked_emerged_cells": "",
        "tracked_emergence_percent": "",
        "newly_emerged_today": "",
        "newly_emerged_percent": "",
        "carried_forward_cells": "",
        "mean_green_area_percent": "",
        "mean_rgb_green_cover_percent": "",
        "possible_day7_bug_eaten_cells": "",
        "overlay_path": "",
        "mask_path": "",
        "status": status,
        "notes": notes,
    }


def run_analysis(args) -> int:
    tray_design, tray_design_note = load_tray_design()
    observation_schedule, schedule_note = load_observation_schedule()

    days_filter = parse_filter_list(args.days)
    trays_filter = parse_filter_list(args.trays)

    jobs = collect_jobs(
        days_filter=days_filter,
        trays_filter=trays_filter,
    )

    print("\nSCRIPT 04 — THIRD TRIAL VISIBLE EMERGENCE")
    print("=" * 72)
    print(f"Crop folder:\n{CROP_ROOT}")
    print(f"\nGrid folder:\n{GRID_ROOT}")
    print(f"\nTray status workbook:\n{TRAY_STATUS_XLSX}")
    print(f"\nOutput folder:\n{OUTPUT_ROOT}")
    print(f"\n{tray_design_note}")
    print(schedule_note)
    print(f"\nJobs found: {len(jobs)}\n")

    for job in jobs:
        print(
            f"{job['day']} > {job['tray']} > {job['capture_id']}"
        )

    if args.dry_run:
        print("\nDry run complete. No outputs created.")
        return 0

    if not jobs:
        print("\nNo cropped D/RGB images were found.")
        return 1

    grid_manifest = load_grid_manifest(
        accept_check_auto=args.accept_check_auto,
    )

    coordinates_by_key = load_cell_coordinates()

    tracking = {}
    cell_identity = {}
    possible_bug_keys = set()

    tray_rows = []
    cell_rows = []

    for job in jobs:
        day = job["day"]
        day_order = int(job["day_order"])
        tray = job["tray"]
        tray_no = int(job["tray_no"])
        capture_id = job["capture_id"]

        grid_key = record_key(
            day,
            tray,
            capture_id,
        )

        meta = tray_day_metadata(
            tray_no,
            day_order,
            tray_design,
            observation_schedule,
        )

        manifest_row = grid_manifest.get(grid_key)
        cells = coordinates_by_key.get(grid_key, [])

        print(f"\nProcessing: {day} > {tray} > {capture_id}")

        if manifest_row is None:
            notes = "No Script 03 grid manifest row was found for this image."
            tray_rows.append(
                make_skipped_tray_row(
                    job,
                    meta,
                    "SKIPPED_NO_GRID_MANIFEST",
                    notes,
                )
            )
            print(f"SKIPPED_NO_GRID_MANIFEST | {notes}")
            continue

        if str(manifest_row.get("accepted_for_script04", "No")).casefold() != "yes":
            notes = (
                f"Script 03 grid status is {manifest_row.get('status', '')}. "
                "Only PASS_AUTO and PASS_MANUAL are accepted by default. "
                "Use --accept-check-auto only after visual review."
            )

            tray_rows.append(
                make_skipped_tray_row(
                    job,
                    meta,
                    "SKIPPED_GRID_STATUS",
                    notes,
                )
            )

            print(f"SKIPPED_GRID_STATUS | {notes}")
            continue

        if len(cells) != EXPECTED_CELLS:
            notes = f"Expected 70 cell coordinate rows, found {len(cells)}."

            tray_rows.append(
                make_skipped_tray_row(
                    job,
                    meta,
                    "SKIPPED_COORDINATES",
                    notes,
                )
            )

            print(f"SKIPPED_COORDINATES | {notes}")
            continue

        rgb = read_rgb(job["path"])
        green_mask = build_green_mask(rgb)

        image_cell_rows = []

        for cell in cells:
            cell_id = int(cell["cell_id"])
            track_key = (tray_no, cell_id)

            state = tracking.setdefault(
                track_key,
                {
                    "emerged": False,
                    "first_day": "",
                    "first_day_order": "",
                    "first_date": "",
                },
            )

            cell_identity[track_key] = {
                "tray_no": tray_no,
                "tray": tray,
                "cell_id": cell_id,
                "row": cell["row"],
                "column": cell["column"],
                "label_code": meta.get("label_code", ""),
                "microbe_status": meta.get("microbe_status", ""),
                "treatment": meta.get("treatment", ""),
                "label_environment": meta.get("label_environment", ""),
                "environment_type": meta.get("environment_type", ""),
                "watering_rule": meta.get("watering_rule", ""),
            }

            previously_emerged = bool(state["emerged"])

            evidence = analyse_cell_green_evidence(
                green_mask,
                cell,
                min_green_pixels=args.min_green_pixels,
                min_component_pixels=args.min_component_pixels,
            )

            raw_green_bool = bool(evidence["raw_green_detected"])

            newly_emerged_bool = (
                raw_green_bool
                and not previously_emerged
            )

            if newly_emerged_bool:
                state["emerged"] = True
                state["first_day"] = day
                state["first_day_order"] = day_order
                state["first_date"] = meta.get("calendar_date", "")

            tracked_visible_emerged_bool = bool(state["emerged"])

            carried_forward_only_bool = (
                tracked_visible_emerged_bool
                and not raw_green_bool
            )

            disappeared_after_previous_emergence_bool = (
                previously_emerged
                and not raw_green_bool
            )

            possible_day7_bug_eaten_bool = (
                day_order == 7
                and disappeared_after_previous_emergence_bool
            )

            if possible_day7_bug_eaten_bool:
                possible_bug_keys.add(track_key)

            row = {
                "day_order": day_order,
                "day": day,
                "calendar_date": meta.get("calendar_date", ""),
                "days_since_planting": meta.get("days_since_planting", ""),
                "days_since_day1": meta.get("days_since_day1", ""),
                "days_since_previous_photo": meta.get("days_since_previous_photo", ""),
                "tray": tray,
                "tray_no": tray_no,
                "label_code": meta.get("label_code", ""),
                "microbe_status": meta.get("microbe_status", ""),
                "treatment": meta.get("treatment", ""),
                "label_environment": meta.get("label_environment", ""),
                "observed_environment": meta.get("observed_environment", ""),
                "environment_type": meta.get("environment_type", ""),
                "watering_rule": meta.get("watering_rule", ""),
                "heat_phase": meta.get("heat_phase_for_tray", ""),
                "moisture_watered_today": meta.get("moisture_watered_today_for_tray", ""),
                "moisture_phase": meta.get("moisture_phase_for_tray", ""),
                "day7_bug_rule_applies": meta.get("day7_bug_rule_applies", ""),
                "capture_id": capture_id,
                "cell_id": cell_id,
                "row": cell["row"],
                "column": cell["column"],
                "x": cell["x"],
                "y": cell["y"],
                "square_x0": cell["square_x0"],
                "square_y0": cell["square_y0"],
                "square_x1": cell["square_x1"],
                "square_y1": cell["square_y1"],
                "square_side": cell["square_side"],
                "raw_green_bool": raw_green_bool,
                "raw_current_green_evidence": yes_no(raw_green_bool),
                "tracked_visible_emerged_bool": tracked_visible_emerged_bool,
                "tracked_visible_emerged": yes_no(tracked_visible_emerged_bool),
                "newly_emerged_today": yes_no(newly_emerged_bool),
                "carried_forward_only": yes_no(carried_forward_only_bool),
                "first_visible_emergence_day": state["first_day"],
                "first_visible_emergence_day_order": state["first_day_order"],
                "first_visible_emergence_date": state["first_date"],
                "disappeared_after_previous_emergence": yes_no(
                    disappeared_after_previous_emergence_bool
                ),
                "possible_day7_bug_eaten_bool": possible_day7_bug_eaten_bool,
                "possible_day7_bug_eaten": yes_no(possible_day7_bug_eaten_bool),
                "green_pixels": evidence["green_pixels"],
                "largest_green_component": evidence["largest_green_component"],
                "green_area_percent": evidence["green_area_percent"],
                "rgb_green_cover_percent": evidence["green_area_percent"],
                "zone_pixels": evidence["zone_pixels"],
                "coordinate_source": cell.get("coordinate_source", ""),
                "grid_error": cell.get("grid_error", ""),
                "grid_needs_review": cell.get("needs_review", ""),
                "notes": (
                    "Possible Day 7 bug-eaten/missing crop. Observed value preserved; adjustment belongs in Script 05."
                    if possible_day7_bug_eaten_bool
                    else ""
                ),
            }

            image_cell_rows.append(row)
            cell_rows.append(row)

        output_folder = (
            OUTPUT_ROOT
            / day
            / tray
        )

        overlay_path = (
            output_folder
            / "overlays"
            / f"{capture_id}_visible_emergence_overlay.jpg"
        )

        mask_path = (
            output_folder
            / "masks"
            / f"{capture_id}_green_mask.png"
        )

        if args.overwrite or not overlay_path.exists():
            save_emergence_overlay(
                rgb,
                green_mask,
                image_cell_rows,
                overlay_path,
                title=f"{day} | {tray} | {capture_id}",
            )

        if args.overwrite or not mask_path.exists():
            save_green_mask(
                green_mask,
                mask_path,
            )

        raw_green_cells = sum(
            bool(row["raw_green_bool"])
            for row in image_cell_rows
        )

        tracked_emerged_cells = sum(
            bool(row["tracked_visible_emerged_bool"])
            for row in image_cell_rows
        )

        newly_emerged_cells = sum(
            row["newly_emerged_today"] == "Yes"
            for row in image_cell_rows
        )

        carried_forward_cells = sum(
            row["carried_forward_only"] == "Yes"
            for row in image_cell_rows
        )

        possible_day7_bug_cells = sum(
            bool(row["possible_day7_bug_eaten_bool"])
            for row in image_cell_rows
        )

        mean_green_area_percent = float(
            np.mean(
                [
                    row["green_area_percent"]
                    for row in image_cell_rows
                ]
            )
        )

        tray_rows.append(
            {
                "day_order": day_order,
                "day": day,
                "calendar_date": meta.get("calendar_date", ""),
                "days_since_planting": meta.get("days_since_planting", ""),
                "days_since_day1": meta.get("days_since_day1", ""),
                "days_since_previous_photo": meta.get("days_since_previous_photo", ""),
                "tray": tray,
                "tray_no": tray_no,
                "label_code": meta.get("label_code", ""),
                "microbe_status": meta.get("microbe_status", ""),
                "treatment": meta.get("treatment", ""),
                "label_environment": meta.get("label_environment", ""),
                "observed_environment": meta.get("observed_environment", ""),
                "environment_type": meta.get("environment_type", ""),
                "watering_rule": meta.get("watering_rule", ""),
                "heat_phase": meta.get("heat_phase_for_tray", ""),
                "moisture_watered_today": meta.get("moisture_watered_today_for_tray", ""),
                "moisture_phase": meta.get("moisture_phase_for_tray", ""),
                "day7_bug_rule_applies": meta.get("day7_bug_rule_applies", ""),
                "capture_id": capture_id,
                "raw_green_cells": raw_green_cells,
                "raw_green_percent": raw_green_cells / EXPECTED_CELLS * 100.0,
                "tracked_emerged_cells": tracked_emerged_cells,
                "tracked_emergence_percent": tracked_emerged_cells / EXPECTED_CELLS * 100.0,
                "newly_emerged_today": newly_emerged_cells,
                "newly_emerged_percent": newly_emerged_cells / EXPECTED_CELLS * 100.0,
                "carried_forward_cells": carried_forward_cells,
                "mean_green_area_percent": mean_green_area_percent,
                "mean_rgb_green_cover_percent": mean_green_area_percent,
                "possible_day7_bug_eaten_cells": possible_day7_bug_cells,
                "overlay_path": relative_path(overlay_path, OUTPUT_ROOT),
                "mask_path": relative_path(mask_path, OUTPUT_ROOT),
                "status": "PASS",
                "notes": (
                    "Observed values only; no Day 7 imputation performed."
                ),
            }
        )

        print(
            f"PASS | raw={raw_green_cells}/70 | "
            f"tracked={tracked_emerged_cells}/70 | "
            f"new={newly_emerged_cells} | "
            f"possible_day7_missing={possible_day7_bug_cells}"
        )

    first_rows = build_first_emergence_rows(
        tracking,
        cell_identity,
        possible_bug_keys,
    )

    settings = {
        "purpose": "Third Trial visible emergence and RGB green-cover analysis",
        "crop_root": str(CROP_ROOT),
        "grid_root": str(GRID_ROOT),
        "tray_status_xlsx": str(TRAY_STATUS_XLSX),
        "output_root": str(OUTPUT_ROOT),
        "rows": ROWS,
        "columns": COLS,
        "expected_cells": EXPECTED_CELLS,
        "accepted_grid_statuses": sorted(ACCEPTED_GRID_STATUSES),
        "accept_check_auto": bool(args.accept_check_auto),
        "corrected_day1_photo_date": "2026-06-29",
        "day7_photo_date": "2026-07-07",
        "green_hue_min": GREEN_HUE_MIN,
        "green_hue_max": GREEN_HUE_MAX,
        "green_saturation_min": GREEN_SATURATION_MIN,
        "green_value_min": GREEN_VALUE_MIN,
        "excess_green_min": EXCESS_GREEN_MIN,
        "green_to_red_ratio": GREEN_TO_RED_RATIO,
        "green_to_blue_ratio": GREEN_TO_BLUE_RATIO,
        "min_green_pixels_per_cell": args.min_green_pixels,
        "min_green_component_pixels": args.min_component_pixels,
        "morph_kernel_size": MORPH_KERNEL_SIZE,
        "tray_design_note": tray_design_note,
        "schedule_note": schedule_note,
        "day7_bug_rule": (
            "Observed Day 7 values are preserved. Cells that had earlier emergence "
            "but no raw green evidence on Day 7 are flagged as possible bug-eaten/missing. "
            "Adjusted Day 7 growth estimates must be created later in Script 05."
        ),
    }

    output_paths = write_reports(
        tray_rows,
        cell_rows,
        first_rows,
        settings,
    )

    status_counts = defaultdict(int)

    for row in tray_rows:
        status_counts[row.get("status", "")] += 1

    possible_day7_bug_total = sum(
        int(row.get("possible_day7_bug_eaten_cells") or 0)
        for row in tray_rows
        if row.get("status") == "PASS"
    )

    print("\n" + "=" * 72)
    print("SCRIPT 04 FINISHED")
    print("=" * 72)

    for status, count in sorted(status_counts.items()):
        print(f"{status}: {count}")

    print(f"\nPossible Day 7 bug-eaten/missing crop flags: {possible_day7_bug_total}")

    print(f"\nTray summary:\n{output_paths['tray_csv']}")
    print(f"\nCell results:\n{output_paths['cell_csv']}")
    print(f"\nFirst emergence summary:\n{output_paths['first_csv']}")
    print(f"\nExcel report:\n{output_paths['excel_path']}")
    print(f"\nSettings:\n{output_paths['settings_path']}")

    failed_or_skipped = [
        row
        for row in tray_rows
        if row.get("status") != "PASS"
    ]

    if failed_or_skipped:
        print(
            "\nSome images were skipped. Check visible_emergence_tray_summary.csv "
            "before moving to Script 05."
        )
        return 1

    return 0


# ============================================================
# 13) CLI
# ============================================================

def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Trial 3 Script 04: visible emergence and RGB green-cover analysis."
        )
    )

    parser.add_argument(
        "--days",
        help='Example: --days "Day 1,Day 7"',
    )

    parser.add_argument(
        "--trays",
        help='Example: --trays "Tray 1,Tray 12"',
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing overlay and mask images.",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="List eligible jobs without processing.",
    )

    parser.add_argument(
        "--accept-check-auto",
        action="store_true",
        help=(
            "Accept Script 03 CHECK_AUTO grid rows. Use only after visual review."
        ),
    )

    parser.add_argument(
        "--min-green-pixels",
        type=int,
        default=MIN_GREEN_PIXELS_PER_CELL,
        help="Minimum total green pixels required inside one cell square.",
    )

    parser.add_argument(
        "--min-component-pixels",
        type=int,
        default=MIN_GREEN_COMPONENT_PIXELS,
        help="Minimum largest connected green component size inside one cell square.",
    )

    args = parser.parse_args()

    return run_analysis(args)


if __name__ == "__main__":
    raise SystemExit(main())