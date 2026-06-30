"""
SCRIPT 03 — 70 SQUARE CELL GRID DETECTION

Purpose
-------
Detect the fixed 7 x 10 = 70 planting-cell layout from the cropped D/RGB
tray images created by Script 01.

This script DOES:
- use cropped D/RGB images only
- fit one global 7 x 10 tray lattice
- create 70 square cell ownership zones
- save square overlays, coordinates, CSV reports and Excel reports
- mark uncertain fitted cells in yellow for targeted review

This script DOES NOT:
- count seedlings
- calculate germination rate
- calculate NDVI or NDRE
- use F preview images
- transfer RGB coordinates to multispectral images

Why squares?
------------
The square cell zone is larger than the physical cup. This allows a seedling
that grows slightly outside the cup rim to remain assigned to its original
planting cell during Script 04 visible-emergence analysis.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any

import cv2
import matplotlib

matplotlib.use("TkAgg")

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageOps


# ============================================================
# 1) PATHS — CHANGE PROJECT_ROOT ONLY WHEN REUSING
# ============================================================

PROJECT_ROOT = Path(
    r"C:\Users\tshib\OneDrive\Desktop\Internship"
)

INPUT_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "01_Crop_Dual_Reference"
)

OUTPUT_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "03_Cell_Grid_Detection"
)


# ============================================================
# 2) TRAY SETTINGS
# ============================================================

ROWS = 7
COLS = 10
EXPECTED_CELLS = ROWS * COLS

DAY_NAME_TO_ORDER = {
    "day 1": 1,
    "day 2": 2,
    "day 3": 3,
    "day 4": 4,
    "day 5": 5,
    "day 9": 9,
    "first day": 1,
    "second day": 2,
    "third day": 3,
    "fourth day": 4,
    "fifth day": 5,
    "ninth day": 9,
}

# Images are temporarily resized only to speed up auto detection.
# Saved coordinates remain in the original cropped-image resolution.
DEFAULT_MAX_DIMENSION = 1800

# Square side length as a fraction of the smaller grid pitch.
# 0.90 keeps a narrow gap between neighbouring ownership zones.
DEFAULT_SQUARE_ZONE_RATIO = 0.90

# Automatic quality thresholds.
PASS_MIN_DIRECT_CELLS = 60
PASS_MAX_GRID_ERROR = 0.16

CHECK_MIN_DIRECT_CELLS = 45
CHECK_MAX_GRID_ERROR = 0.30


# ============================================================
# 3) GENERAL HELPERS
# ============================================================

def natural_key(text: str):
    return [
        int(part) if part.isdigit() else part.casefold()
        for part in re.split(r"(\d+)", text)
    ]


def day_sort_key(folder: Path):
    return (
        DAY_NAME_TO_ORDER.get(folder.name.casefold(), 999),
        natural_key(folder.name),
    )


def parse_filter_list(value: str | None):
    if not value:
        return None

    return {
        item.strip().casefold()
        for item in value.split(",")
        if item.strip()
    }


def tray_number_from_name(tray_name: str):
    match = re.search(r"(\d+)", tray_name)
    return int(match.group(1)) if match else ""


def safe_name(text: str):
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", text)


def relative_path(path: Path | None, root: Path):
    if path is None:
        return ""

    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


def record_key(day: str, tray: str, capture_id: str):
    return f"{day}|{tray}|{capture_id}"


def numeric_or_default(value, default=999):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


# ============================================================
# 4) FIND CROPPED D/RGB IMAGES FROM SCRIPT 01
# ============================================================

def parse_d_image(path: Path):
    """
    Example:
        DJI_20260618124632_0008_D.JPG
    """

    match = re.match(
        r"^(?P<capture>.+)_D$",
        path.stem.upper(),
    )

    return match.group("capture") if match else None


def find_d_images(tray_folder: Path):
    images = []

    for file in tray_folder.iterdir():
        if not file.is_file():
            continue

        if file.suffix.casefold() not in {".jpg", ".jpeg"}:
            continue

        capture_id = parse_d_image(file)

        if capture_id:
            images.append(
                {
                    "capture_id": capture_id,
                    "path": file,
                }
            )

    return sorted(
        images,
        key=lambda item: natural_key(item["capture_id"]),
    )


# ============================================================
# 5) IMAGE READING / TEMPORARY RESIZING
# ============================================================

def read_rgb(path: Path):
    """Read cropped D/RGB image without changing the source file."""

    with Image.open(path) as image:
        image = ImageOps.exif_transpose(image)
        image = image.convert("RGB")

        return np.asarray(image)


def resize_for_detection(
    rgb: np.ndarray,
    max_dimension: int,
):
    """
    Resize only for faster detection.
    Final cell coordinates are later converted back to full resolution.
    """

    height, width = rgb.shape[:2]
    largest_dimension = max(width, height)

    if max_dimension <= 0 or largest_dimension <= max_dimension:
        return rgb.copy(), 1.0

    scale = max_dimension / largest_dimension

    resized = cv2.resize(
        rgb,
        (
            int(round(width * scale)),
            int(round(height * scale)),
        ),
        interpolation=cv2.INTER_AREA,
    )

    return resized, scale


# ============================================================
# 6) Hough Candidate Evidence
# ============================================================

def candidate_evidence(
    grayscale: np.ndarray,
    x_value: float,
    y_value: float,
    radius: float,
    dark_threshold: float,
):
    """
    A real planting cup usually has:
    - dark soil/mulch in its core
    - a lighter white rim around the outside

    This rejects false Hough detections on blank foam background.
    """

    height, width = grayscale.shape[:2]

    outer_radius = max(8, int(round(radius * 1.15)))

    x0 = max(0, int(round(x_value)) - outer_radius)
    x1 = min(width, int(round(x_value)) + outer_radius + 1)

    y0 = max(0, int(round(y_value)) - outer_radius)
    y1 = min(height, int(round(y_value)) + outer_radius + 1)

    if x1 <= x0 or y1 <= y0:
        return {
            "core_dark_fraction": 0.0,
            "contrast": -999.0,
            "quality": -999.0,
        }

    crop = grayscale[y0:y1, x0:x1].astype(np.float32)

    yy, xx = np.ogrid[y0:y1, x0:x1]

    distance = np.sqrt(
        (xx - x_value) ** 2
        + (yy - y_value) ** 2
    )

    core_mask = distance <= radius * 0.48

    ring_mask = (
        (distance >= radius * 0.70)
        & (distance <= radius * 1.08)
    )

    if (
        np.count_nonzero(core_mask) < 20
        or np.count_nonzero(ring_mask) < 20
    ):
        return {
            "core_dark_fraction": 0.0,
            "contrast": -999.0,
            "quality": -999.0,
        }

    core_values = crop[core_mask]
    ring_values = crop[ring_mask]

    core_dark_fraction = float(
        np.mean(core_values <= dark_threshold)
    )

    core_mean = float(np.mean(core_values))
    ring_mean = float(np.mean(ring_values))

    contrast = ring_mean - core_mean

    quality = (
        core_dark_fraction * 100.0
        + max(0.0, contrast)
    )

    return {
        "core_dark_fraction": core_dark_fraction,
        "contrast": contrast,
        "quality": quality,
    }


def suppress_duplicates(
    candidates: list[dict],
    minimum_distance: float,
):
    """
    Hough detection can return multiple circles for one cup.
    Keep the strongest evidence candidate within each local neighbourhood.
    """

    selected = []

    for candidate in sorted(
        candidates,
        key=lambda item: item["quality"],
        reverse=True,
    ):
        close_to_existing = False

        for kept in selected:
            distance = np.hypot(
                candidate["x"] - kept["x"],
                candidate["y"] - kept["y"],
            )

            if distance < minimum_distance:
                close_to_existing = True
                break

        if not close_to_existing:
            selected.append(candidate)

    return selected


def detect_pot_candidates(rgb: np.ndarray):
    """
    Detect likely cup centres using Hough circles and reject background
    false positives using dark-soil / light-rim evidence.
    """

    grayscale = cv2.cvtColor(
        rgb,
        cv2.COLOR_RGB2GRAY,
    )

    grayscale = cv2.createCLAHE(
        clipLimit=2.0,
        tileGridSize=(8, 8),
    ).apply(grayscale)

    blurred = cv2.GaussianBlur(
        grayscale,
        (5, 5),
        0,
    )

    height, width = grayscale.shape[:2]

    estimated_pitch = min(
        width / COLS,
        height / ROWS,
    )

    min_distance = max(
        20,
        int(round(estimated_pitch * 0.42)),
    )

    min_radius = max(
        8,
        int(round(estimated_pitch * 0.16)),
    )

    max_radius = max(
        min_radius + 6,
        int(round(estimated_pitch * 0.52)),
    )

    dark_threshold = float(
        np.percentile(grayscale, 55)
    )

    raw_candidates = []

    # Multiple thresholds improve recovery under different lighting.
    for param2 in [58, 52, 46, 40, 34, 28]:
        circles = cv2.HoughCircles(
            blurred,
            cv2.HOUGH_GRADIENT,
            dp=1.2,
            minDist=min_distance,
            param1=100,
            param2=param2,
            minRadius=min_radius,
            maxRadius=max_radius,
        )

        if circles is None:
            continue

        for x_value, y_value, radius in np.round(
            circles[0]
        ).astype(np.float32):
            evidence = candidate_evidence(
                grayscale,
                float(x_value),
                float(y_value),
                float(radius),
                dark_threshold,
            )

            raw_candidates.append(
                {
                    "x": float(x_value),
                    "y": float(y_value),
                    "raw_radius": float(radius),
                    **evidence,
                }
            )

    if not raw_candidates:
        raise ValueError(
            "No circular planting-cup candidates were detected."
        )

    # Background false detections usually have little dark soil evidence.
    usable = [
        item
        for item in raw_candidates
        if item["core_dark_fraction"] >= 0.06
        and item["quality"] > 8.0
    ]

    # Preserve the strongest candidates if strict filtering was too severe.
    if len(usable) < 35:
        usable = sorted(
            raw_candidates,
            key=lambda item: item["quality"],
            reverse=True,
        )[:160]

    usable = suppress_duplicates(
        usable,
        minimum_distance=max(
            10.0,
            estimated_pitch * 0.22,
        ),
    )

    usable = sorted(
        usable,
        key=lambda item: item["quality"],
        reverse=True,
    )[:160]

    if len(usable) < 35:
        raise ValueError(
            f"Only {len(usable)} usable cup-centre candidates remained."
        )

    return usable, {
        "estimated_pitch": estimated_pitch,
        "candidate_count": len(usable),
    }


# ============================================================
# 7) GLOBAL 7 x 10 GRID FITTING
# ============================================================

def cluster_1d_centres(
    values: np.ndarray,
    cluster_count: int,
):
    """Use 1D k-means to estimate expected row/column positions."""

    values = values.astype(np.float32).reshape(-1, 1)

    criteria = (
        cv2.TERM_CRITERIA_EPS
        + cv2.TERM_CRITERIA_MAX_ITER,
        100,
        0.1,
    )

    cv2.setRNGSeed(42)

    _, _, centres = cv2.kmeans(
        values,
        cluster_count,
        None,
        criteria,
        20,
        cv2.KMEANS_PP_CENTERS,
    )

    return np.sort(centres.reshape(-1))


def build_grid_nodes(
    x_centres: np.ndarray,
    y_centres: np.ndarray,
):
    nodes = []

    for row_index, y_value in enumerate(y_centres, start=1):
        for column_index, x_value in enumerate(
            x_centres,
            start=1,
        ):
            nodes.append(
                {
                    "cell_id": (
                        (row_index - 1) * COLS
                        + column_index
                    ),
                    "row": row_index,
                    "column": column_index,
                    "x": float(x_value),
                    "y": float(y_value),
                }
            )

    return nodes


def greedy_match(
    nodes: list[dict],
    candidates: list[dict],
    pitch: float,
    maximum_normalised_distance: float,
):
    """
    Assign one high-quality candidate to at most one expected grid node.
    """

    pairs = []

    for node_index, node in enumerate(nodes):
        for candidate_index, candidate in enumerate(candidates):
            distance = np.hypot(
                node["x"] - candidate["x"],
                node["y"] - candidate["y"],
            )

            normalised_distance = distance / max(pitch, 1e-6)

            if normalised_distance <= maximum_normalised_distance:
                # Prefer candidates close to the node and with stronger evidence.
                cost = (
                    normalised_distance
                    - min(candidate["quality"], 100.0) / 4000.0
                )

                pairs.append(
                    (
                        cost,
                        normalised_distance,
                        node_index,
                        candidate_index,
                    )
                )

    pairs.sort(key=lambda item: item[0])

    used_nodes = set()
    used_candidates = set()
    matches = {}

    for _cost, normalised_distance, node_index, candidate_index in pairs:
        if node_index in used_nodes:
            continue

        if candidate_index in used_candidates:
            continue

        matches[node_index] = (
            candidate_index,
            normalised_distance,
        )

        used_nodes.add(node_index)
        used_candidates.add(candidate_index)

    return matches


def fit_affine_grid(
    nodes: list[dict],
    candidates: list[dict],
    matches: dict,
):
    """
    Fit a global affine model:

        x = a0 + a1 * column + a2 * row
        y = b0 + b1 * column + b2 * row

    This prevents one false candidate from shifting an individual cell
    into blank background.
    """

    if len(matches) < 20:
        raise ValueError(
            "Too few reliable cell candidates to fit the tray lattice."
        )

    features = []
    positions = []

    for node_index, (candidate_index, _error) in matches.items():
        node = nodes[node_index]
        candidate = candidates[candidate_index]

        features.append(
            [
                1.0,
                float(node["column"] - 1),
                float(node["row"] - 1),
            ]
        )

        positions.append(
            [
                candidate["x"],
                candidate["y"],
            ]
        )

    features = np.asarray(features, dtype=np.float64)
    positions = np.asarray(positions, dtype=np.float64)

    coefficients, _, _, _ = np.linalg.lstsq(
        features,
        positions,
        rcond=None,
    )

    refined_nodes = []

    for row_index in range(1, ROWS + 1):
        for column_index in range(1, COLS + 1):
            feature = np.asarray(
                [
                    1.0,
                    float(column_index - 1),
                    float(row_index - 1),
                ],
                dtype=np.float64,
            )

            x_value, y_value = feature @ coefficients

            refined_nodes.append(
                {
                    "cell_id": (
                        (row_index - 1) * COLS
                        + column_index
                    ),
                    "row": row_index,
                    "column": column_index,
                    "x": float(x_value),
                    "y": float(y_value),
                }
            )

    column_vector = coefficients[1]
    row_vector = coefficients[2]

    x_pitch = float(np.linalg.norm(column_vector))
    y_pitch = float(np.linalg.norm(row_vector))

    return refined_nodes, x_pitch, y_pitch


def auto_detect_square_grid(
    rgb: np.ndarray,
    max_dimension: int,
    square_zone_ratio: float,
):
    """
    Detect 70 cell centres from cropped D/RGB and create square zones.

    Every final square comes from the global 7 x 10 lattice.
    Direct candidates are only used to refine a lattice position when they
    fit the expected grid and show strong local cup evidence.
    """

    detection_rgb, scale = resize_for_detection(
        rgb,
        max_dimension,
    )

    candidates, candidate_metrics = detect_pot_candidates(
        detection_rgb
    )

    x_values = np.asarray(
        [candidate["x"] for candidate in candidates],
        dtype=np.float32,
    )

    y_values = np.asarray(
        [candidate["y"] for candidate in candidates],
        dtype=np.float32,
    )

    x_centres = cluster_1d_centres(
        x_values,
        COLS,
    )

    y_centres = cluster_1d_centres(
        y_values,
        ROWS,
    )

    initial_x_spacing = np.diff(x_centres)
    initial_y_spacing = np.diff(y_centres)

    x_pitch = float(np.median(initial_x_spacing))
    y_pitch = float(np.median(initial_y_spacing))

    initial_pitch = min(x_pitch, y_pitch)

    if initial_pitch <= 0:
        raise ValueError(
            "Invalid initial grid spacing."
        )

    initial_nodes = build_grid_nodes(
        x_centres,
        y_centres,
    )

    initial_matches = greedy_match(
        initial_nodes,
        candidates,
        initial_pitch,
        maximum_normalised_distance=0.65,
    )

    refined_nodes, x_pitch, y_pitch = fit_affine_grid(
        initial_nodes,
        candidates,
        initial_matches,
    )

    # Refine once more using closer matches only.
    refined_pitch = min(x_pitch, y_pitch)

    final_matches = greedy_match(
        refined_nodes,
        candidates,
        refined_pitch,
        maximum_normalised_distance=0.32,
    )

    inlier_matches = {
        node_index: match
        for node_index, match in final_matches.items()
        if match[1] <= 0.24
    }

    if len(inlier_matches) >= 20:
        refined_nodes, x_pitch, y_pitch = fit_affine_grid(
            refined_nodes,
            candidates,
            inlier_matches,
        )

        refined_pitch = min(x_pitch, y_pitch)

        final_matches = greedy_match(
            refined_nodes,
            candidates,
            refined_pitch,
            maximum_normalised_distance=0.30,
        )

    square_side_detection = (
        refined_pitch
        * square_zone_ratio
    )

    direct_cell_count = 0
    grid_errors = []
    cells_detection = []

    for node_index, node in enumerate(refined_nodes):
        if node_index in final_matches:
            candidate_index, normalised_error = final_matches[node_index]
            candidate = candidates[candidate_index]

            direct_supported = (
                normalised_error <= 0.24
                and candidate["core_dark_fraction"] >= 0.06
            )
        else:
            candidate = None
            normalised_error = 1.0
            direct_supported = False

        if direct_supported:
            centre_x = candidate["x"]
            centre_y = candidate["y"]
            source = "direct_cup_evidence"
            review = "No"
            direct_cell_count += 1
        else:
            centre_x = node["x"]
            centre_y = node["y"]
            source = "global_lattice_fitted"
            review = "Yes"

        grid_errors.append(float(normalised_error))

        half_side = square_side_detection / 2.0

        cells_detection.append(
            {
                "cell_id": node["cell_id"],
                "row": node["row"],
                "column": node["column"],
                "x": float(centre_x),
                "y": float(centre_y),
                "square_side": float(square_side_detection),
                "square_x0": float(centre_x - half_side),
                "square_y0": float(centre_y - half_side),
                "square_x1": float(centre_x + half_side),
                "square_y1": float(centre_y + half_side),
                "grid_error": float(normalised_error),
                "coordinate_source": source,
                "needs_review": review,
            }
        )

    # Convert to full cropped-image coordinates.
    full_resolution_cells = []

    for cell in cells_detection:
        full_resolution_cells.append(
            {
                **cell,
                "x": cell["x"] / scale,
                "y": cell["y"] / scale,
                "square_side": cell["square_side"] / scale,
                "square_x0": cell["square_x0"] / scale,
                "square_y0": cell["square_y0"] / scale,
                "square_x1": cell["square_x1"] / scale,
                "square_y1": cell["square_y1"] / scale,
            }
        )

    spacing_cv = max(
        float(
            np.std(initial_x_spacing)
            / max(np.mean(initial_x_spacing), 1e-6)
        ),
        float(
            np.std(initial_y_spacing)
            / max(np.mean(initial_y_spacing), 1e-6)
        ),
    )

    uncertain_cell_ids = [
        cell["cell_id"]
        for cell in full_resolution_cells
        if cell["needs_review"] == "Yes"
    ]

    metrics = {
        "candidate_count": candidate_metrics["candidate_count"],
        "direct_cell_count": direct_cell_count,
        "uncertain_cell_count": len(uncertain_cell_ids),
        "uncertain_cell_ids": uncertain_cell_ids,
        "square_side": square_side_detection / scale,
        "x_pitch": x_pitch / scale,
        "y_pitch": y_pitch / scale,
        "spacing_cv": spacing_cv,
        "median_grid_error": float(np.median(grid_errors)),
        "max_grid_error": float(max(grid_errors)),
        "detection_scale": scale,
    }

    return full_resolution_cells, metrics


# ============================================================
# 8) MANUAL FOUR-CORNER FALLBACK
# ============================================================

def get_manual_corner_centres(
    rgb: np.ndarray,
    title: str,
):
    """
    Click centres in this exact order:

    1. Cell 1  = top-left
    2. Cell 10 = top-right
    3. Cell 70 = bottom-right
    4. Cell 61 = bottom-left
    """

    figure, axis = plt.subplots(figsize=(15, 10))

    axis.imshow(rgb)

    axis.set_title(
        title
        + "\nClick: Cell 1 → Cell 10 → Cell 70 → Cell 61"
        + "\nRight click removes the latest point.",
        fontsize=12,
    )

    axis.axis("off")

    plt.tight_layout()
    plt.show(block=False)

    points = plt.ginput(
        n=4,
        timeout=0,
        show_clicks=True,
        mouse_add=1,
        mouse_pop=3,
        mouse_stop=2,
    )

    plt.close(figure)

    if len(points) != 4:
        return None

    return np.asarray(
        points,
        dtype=np.float32,
    )


def build_manual_square_grid(
    corner_points: np.ndarray,
    square_zone_ratio: float,
):
    """Generate all 70 square cell zones from four clicked corner cells."""

    source_grid = np.asarray(
        [
            [0, 0],
            [COLS - 1, 0],
            [COLS - 1, ROWS - 1],
            [0, ROWS - 1],
        ],
        dtype=np.float32,
    )

    transform = cv2.getPerspectiveTransform(
        source_grid,
        corner_points.astype(np.float32),
    )

    grid_points = []

    for row_index in range(ROWS):
        for column_index in range(COLS):
            grid_points.append(
                [[column_index, row_index]]
            )

    grid_points = np.asarray(
        grid_points,
        dtype=np.float32,
    )

    image_points = cv2.perspectiveTransform(
        grid_points,
        transform,
    ).reshape(-1, 2)

    horizontal_distances = []
    vertical_distances = []

    for row_index in range(ROWS):
        for column_index in range(COLS - 1):
            left = image_points[
                row_index * COLS + column_index
            ]
            right = image_points[
                row_index * COLS + column_index + 1
            ]

            horizontal_distances.append(
                float(np.linalg.norm(right - left))
            )

    for row_index in range(ROWS - 1):
        for column_index in range(COLS):
            top = image_points[
                row_index * COLS + column_index
            ]
            bottom = image_points[
                (row_index + 1) * COLS + column_index
            ]

            vertical_distances.append(
                float(np.linalg.norm(bottom - top))
            )

    pitch = float(
        np.median(
            horizontal_distances
            + vertical_distances
        )
    )

    square_side = pitch * square_zone_ratio
    half_side = square_side / 2.0

    cells = []

    for index, (x_value, y_value) in enumerate(
        image_points,
        start=1,
    ):
        row_index = ((index - 1) // COLS) + 1
        column_index = ((index - 1) % COLS) + 1

        cells.append(
            {
                "cell_id": index,
                "row": row_index,
                "column": column_index,
                "x": float(x_value),
                "y": float(y_value),
                "square_side": float(square_side),
                "square_x0": float(x_value - half_side),
                "square_y0": float(y_value - half_side),
                "square_x1": float(x_value + half_side),
                "square_y1": float(y_value + half_side),
                "grid_error": 0.0,
                "coordinate_source": "manual_four_corner_grid",
                "needs_review": "No",
            }
        )

    metrics = {
        "candidate_count": "",
        "direct_cell_count": EXPECTED_CELLS,
        "uncertain_cell_count": 0,
        "uncertain_cell_ids": [],
        "square_side": square_side,
        "x_pitch": pitch,
        "y_pitch": pitch,
        "spacing_cv": 0.0,
        "median_grid_error": 0.0,
        "max_grid_error": 0.0,
        "detection_scale": 1.0,
    }

    return cells, metrics


# ============================================================
# 9) GRID QUALITY STATUS
# ============================================================

def decide_auto_status(metrics: dict):
    if (
        metrics["direct_cell_count"] >= PASS_MIN_DIRECT_CELLS
        and metrics["median_grid_error"] <= PASS_MAX_GRID_ERROR
    ):
        return "PASS_AUTO"

    if (
        metrics["direct_cell_count"] >= CHECK_MIN_DIRECT_CELLS
        and metrics["median_grid_error"] <= CHECK_MAX_GRID_ERROR
    ):
        return "CHECK_AUTO"

    return "FAIL"


# ============================================================
# 10) OVERLAY CREATION
# ============================================================

def save_square_grid_overlay(
    rgb: np.ndarray,
    cells: list[dict],
    output_path: Path,
    title: str,
):
    """
    Green squares = directly supported by cup evidence.
    Yellow squares = fitted global lattice positions needing review.
    Blue squares = manually created grid.
    """

    overlay = cv2.cvtColor(
        rgb,
        cv2.COLOR_RGB2BGR,
    )

    image_height, image_width = overlay.shape[:2]

    median_side = float(
        np.median(
            [cell["square_side"] for cell in cells]
        )
    )

    line_width = max(
        2,
        int(round(median_side / 32)),
    )

    font_scale = max(
        0.35,
        min(1.05, median_side / 85),
    )

    font_thickness = max(
        1,
        int(round(font_scale * 2)),
    )

    for cell in cells:
        x0 = max(
            0,
            int(round(cell["square_x0"])),
        )
        y0 = max(
            0,
            int(round(cell["square_y0"])),
        )
        x1 = min(
            image_width - 1,
            int(round(cell["square_x1"])),
        )
        y1 = min(
            image_height - 1,
            int(round(cell["square_y1"])),
        )

        if cell["coordinate_source"] == "manual_four_corner_grid":
            colour = (255, 150, 0)  # Blue
        elif cell["needs_review"] == "Yes":
            colour = (0, 180, 255)  # Yellow
        else:
            colour = (0, 180, 0)  # Green

        cv2.rectangle(
            overlay,
            (x0, y0),
            (x1, y1),
            colour,
            line_width,
        )

        label = str(cell["cell_id"])

        text_x = max(
            1,
            int(round(cell["x"] - median_side * 0.10)),
        )
        text_y = min(
            image_height - 5,
            int(round(cell["y"] + median_side * 0.08)),
        )

        cv2.putText(
            overlay,
            label,
            (text_x, text_y),
            cv2.FONT_HERSHEY_SIMPLEX,
            font_scale,
            colour,
            font_thickness,
            cv2.LINE_AA,
        )

    header_height = max(
        42,
        int(round(median_side * 0.30)),
    )

    cv2.rectangle(
        overlay,
        (0, 0),
        (image_width, header_height),
        (255, 255, 255),
        thickness=-1,
    )

    cv2.putText(
        overlay,
        title,
        (12, int(header_height * 0.68)),
        cv2.FONT_HERSHEY_SIMPLEX,
        max(0.42, min(0.85, font_scale)),
        (0, 0, 0),
        2,
        cv2.LINE_AA,
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    cv2.imwrite(
        str(output_path),
        overlay,
        [
            cv2.IMWRITE_JPEG_QUALITY,
            95,
        ],
    )


# ============================================================
# 11) CSV / JSON HELPERS
# ============================================================

def load_json(path: Path):
    if not path.exists():
        return {"records": {}}

    try:
        with path.open("r", encoding="utf-8") as file:
            return json.load(file)
    except Exception:
        return {"records": {}}


def save_json(path: Path, data: dict):
    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with path.open("w", encoding="utf-8") as file:
        json.dump(
            data,
            file,
            indent=2,
        )


def load_existing_csv(path: Path):
    if not path.exists():
        return []

    try:
        with path.open(
            "r",
            newline="",
            encoding="utf-8",
        ) as file:
            return list(csv.DictReader(file))
    except Exception:
        return []


def write_csv(
    path: Path,
    fieldnames: list[str],
    rows: list[dict],
):
    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with path.open(
        "w",
        newline="",
        encoding="utf-8",
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=fieldnames,
        )

        writer.writeheader()
        writer.writerows(rows)


# ============================================================
# 12) EXCEL REPORT
# ============================================================

def style_sheet(
    worksheet,
    headers: list[str],
    rows: list[list[Any]],
):
    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    thin = Side(
        style="thin",
        color="D9E2F3",
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    for column_number, header in enumerate(headers, start=1):
        cell = worksheet.cell(
            row=1,
            column=column_number,
            value=header,
        )

        cell.fill = header_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = border

    for row_number, row in enumerate(rows, start=2):
        for column_number, value in enumerate(row, start=1):
            cell = worksheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

            cell.border = border

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 34

    for column in worksheet.columns:
        letter = column[0].column_letter

        longest = max(
            len(str(cell.value))
            if cell.value is not None
            else 0
            for cell in column
        )

        worksheet.column_dimensions[letter].width = min(
            max(12, longest + 2),
            58,
        )


def create_excel_report(
    output_path: Path,
    summary_rows: list[dict],
    coordinate_rows: list[dict],
):
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Cell Grid Summary"

    summary_headers = [
        "Day Order",
        "Day",
        "Tray",
        "Tray No",
        "Capture ID",
        "Source D Image",
        "Method",
        "Cells Generated",
        "Pot Candidates",
        "Direct Cells",
        "Uncertain Cells",
        "Uncertain Cell IDs",
        "Square Zone Side",
        "Median Grid Error",
        "Maximum Grid Error",
        "Review Required",
        "Overlay Path",
        "Status",
        "Notes",
    ]

    summary_values = [
        [
            row.get("day_order", ""),
            row.get("day", ""),
            row.get("tray", ""),
            row.get("tray_no", ""),
            row.get("capture_id", ""),
            row.get("source_d", ""),
            row.get("method", ""),
            row.get("cells_generated", ""),
            row.get("candidate_count", ""),
            row.get("direct_cell_count", ""),
            row.get("uncertain_cell_count", ""),
            row.get("uncertain_cell_ids", ""),
            row.get("square_side", ""),
            row.get("median_grid_error", ""),
            row.get("max_grid_error", ""),
            row.get("review_required", ""),
            row.get("overlay_path", ""),
            row.get("status", ""),
            row.get("notes", ""),
        ]
        for row in summary_rows
    ]

    style_sheet(
        summary_sheet,
        summary_headers,
        summary_values,
    )

    for row_number in range(2, len(summary_values) + 2):
        review_cell = summary_sheet.cell(
            row=row_number,
            column=16,
        )

        status_cell = summary_sheet.cell(
            row=row_number,
            column=18,
        )

        if review_cell.value == "Yes":
            review_cell.fill = PatternFill(
                "solid",
                fgColor="FFF2CC",
            )

        if status_cell.value in {
            "PASS_AUTO",
            "PASS_MANUAL",
        }:
            status_cell.fill = PatternFill(
                "solid",
                fgColor="C6EFCE",
            )
        elif status_cell.value == "CHECK_AUTO":
            status_cell.fill = PatternFill(
                "solid",
                fgColor="FFF2CC",
            )
        elif status_cell.value == "FAIL":
            status_cell.fill = PatternFill(
                "solid",
                fgColor="F4CCCC",
            )

    coordinate_sheet = workbook.create_sheet(
        "Cell Coordinates"
    )

    coordinate_headers = [
        "Day Order",
        "Day",
        "Tray",
        "Tray No",
        "Capture ID",
        "Cell ID",
        "Row",
        "Column",
        "Centre X",
        "Centre Y",
        "Square X0",
        "Square Y0",
        "Square X1",
        "Square Y1",
        "Square Side",
        "Grid Error",
        "Coordinate Source",
        "Needs Review",
    ]

    coordinate_values = [
        [
            row.get("day_order", ""),
            row.get("day", ""),
            row.get("tray", ""),
            row.get("tray_no", ""),
            row.get("capture_id", ""),
            row.get("cell_id", ""),
            row.get("row", ""),
            row.get("column", ""),
            row.get("x", ""),
            row.get("y", ""),
            row.get("square_x0", ""),
            row.get("square_y0", ""),
            row.get("square_x1", ""),
            row.get("square_y1", ""),
            row.get("square_side", ""),
            row.get("grid_error", ""),
            row.get("coordinate_source", ""),
            row.get("needs_review", ""),
        ]
        for row in coordinate_rows
    ]

    style_sheet(
        coordinate_sheet,
        coordinate_headers,
        coordinate_values,
    )

    readme = workbook.create_sheet("Read Me")

    readme["A1"] = "Script 03 — Square 70-Cell Grid Detection"
    readme["A1"].font = Font(
        bold=True,
        size=14,
    )

    notes = [
        "This script detects cell locations only. It does not count seedlings or calculate germination.",
        "Each tray is modelled as a fixed 7 rows × 10 columns = 70-cell lattice.",
        "Every cell is represented by a square ownership zone, not a circle.",
        "The square zone is deliberately larger than the physical cup so seedlings that grow beyond the cup rim can still be assigned to their original planting cell.",
        "Green squares are directly supported by local cup/soil evidence.",
        "Yellow squares are fitted from the global lattice and should be reviewed only when the report says Review Required = Yes.",
        "Blue squares represent manual four-corner cell grids.",
        "Script 04 will use these square ownership zones for visible-emergence and seedling-growth analysis.",
        "Do not copy these D/RGB square coordinates to multispectral images. MS imagery requires an independent MS grid later.",
    ]

    for row_number, note in enumerate(notes, start=3):
        readme.cell(
            row=row_number,
            column=1,
            value=note,
        )

    readme.column_dimensions["A"].width = 120

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(output_path)


# ============================================================
# 13) MAIN WORKFLOW
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Script 03: Detect fixed 7 x 10 square-cell zones "
            "from cropped Second Trial D/RGB images."
        )
    )

    parser.add_argument(
        "--days",
        help='Example: --days "Day 1,Day 9"',
    )

    parser.add_argument(
        "--trays",
        help='Example: --trays "Tray 1,Tray 7"',
    )

    parser.add_argument(
        "--mode",
        choices=["auto", "manual"],
        default="auto",
        help=(
            "auto = robust automatic D/RGB grid; "
            "manual = click Cell 1, Cell 10, Cell 70 and Cell 61."
        ),
    )

    parser.add_argument(
        "--max-dimension",
        type=int,
        default=DEFAULT_MAX_DIMENSION,
        help=(
            "Temporary maximum size for faster auto detection. "
            "Final coordinates are always saved at full crop resolution."
        ),
    )

    parser.add_argument(
        "--square-zone-ratio",
        type=float,
        default=DEFAULT_SQUARE_ZONE_RATIO,
        help=(
            "Square side as a fraction of grid spacing. "
            "Default: 0.90."
        ),
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing Script 03 outputs for selected trays.",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="List cropped D/RGB files without analysing them.",
    )

    args = parser.parse_args()

    if (
        args.square_zone_ratio <= 0
        or args.square_zone_ratio >= 1.0
    ):
        print(
            "ERROR: --square-zone-ratio must be greater than 0 "
            "and less than 1.0."
        )
        return 1

    if not INPUT_ROOT.exists():
        print(
            "ERROR: Script 01 output folder not found:\n"
            f"{INPUT_ROOT}"
        )
        return 1

    selected_days = parse_filter_list(args.days)
    selected_trays = parse_filter_list(args.trays)

    day_folders = sorted(
        [
            folder
            for folder in INPUT_ROOT.iterdir()
            if folder.is_dir()
            and folder.name.casefold() in DAY_NAME_TO_ORDER
        ],
        key=day_sort_key,
    )

    jobs = []

    for day_folder in day_folders:
        if (
            selected_days
            and day_folder.name.casefold()
            not in selected_days
        ):
            continue

        tray_folders = sorted(
            [
                folder
                for folder in day_folder.iterdir()
                if folder.is_dir()
            ],
            key=lambda folder: natural_key(folder.name),
        )

        for tray_folder in tray_folders:
            if (
                selected_trays
                and tray_folder.name.casefold()
                not in selected_trays
            ):
                continue

            d_images = find_d_images(tray_folder)

            if not d_images:
                jobs.append(
                    {
                        "day_folder": day_folder,
                        "tray_folder": tray_folder,
                        "d_image": None,
                    }
                )
            else:
                for d_image in d_images:
                    jobs.append(
                        {
                            "day_folder": day_folder,
                            "tray_folder": tray_folder,
                            "d_image": d_image,
                        }
                    )

    if not jobs:
        print("No cropped D/RGB images were found.")
        return 1

    print("\nSCRIPT 03 — SQUARE 70-CELL GRID DETECTION")
    print("=" * 70)
    print(f"Input crops:\n{INPUT_ROOT}")
    print(f"\nOutput folder:\n{OUTPUT_ROOT}")
    print(f"\nMode: {args.mode.upper()}")
    print(f"Square-zone ratio: {args.square_zone_ratio}")

    for job in jobs:
        if job["d_image"] is None:
            print(
                f"MISSING D IMAGE: "
                f"{job['day_folder'].name} > "
                f"{job['tray_folder'].name}"
            )
        else:
            print(
                f"READY: "
                f"{job['day_folder'].name} > "
                f"{job['tray_folder'].name} > "
                f"{job['d_image']['path'].name}"
            )

    if args.dry_run:
        print("\nDry run complete. No outputs created.")
        return 0

    report_folder = OUTPUT_ROOT / "_reports"

    manifest_path = report_folder / "cell_grid_manifest.csv"
    coordinates_path = report_folder / "cell_coordinates.csv"
    workbook_path = report_folder / "cell_grid_report.xlsx"

    config_path = (
        OUTPUT_ROOT
        / "_config"
        / "manual_grid_configuration.json"
    )

    existing_summary = load_existing_csv(manifest_path)
    existing_coordinates = load_existing_csv(coordinates_path)

    summary_by_key = {
        record_key(
            row.get("day", ""),
            row.get("tray", ""),
            row.get("capture_id", ""),
        ): row
        for row in existing_summary
    }

    coordinates_by_key: dict[str, list[dict]] = {}

    for row in existing_coordinates:
        key = record_key(
            row.get("day", ""),
            row.get("tray", ""),
            row.get("capture_id", ""),
        )

        coordinates_by_key.setdefault(key, [])
        coordinates_by_key[key].append(row)

    manual_config = load_json(config_path)

    for job in jobs:
        day_name = job["day_folder"].name
        tray_name = job["tray_folder"].name

        day_order = DAY_NAME_TO_ORDER.get(
            day_name.casefold(),
            999,
        )

        tray_no = tray_number_from_name(tray_name)

        if job["d_image"] is None:
            capture_id = ""
            key = record_key(
                day_name,
                tray_name,
                capture_id,
            )

            summary_by_key[key] = {
                "day_order": day_order,
                "day": day_name,
                "tray": tray_name,
                "tray_no": tray_no,
                "capture_id": capture_id,
                "source_d": "",
                "method": args.mode,
                "cells_generated": 0,
                "candidate_count": "",
                "direct_cell_count": "",
                "uncertain_cell_count": "",
                "uncertain_cell_ids": "",
                "square_side": "",
                "median_grid_error": "",
                "max_grid_error": "",
                "review_required": "Yes",
                "overlay_path": "",
                "status": "FAIL",
                "notes": "No cropped D/RGB image found.",
            }

            coordinates_by_key[key] = []
            continue

        capture_id = job["d_image"]["capture_id"]
        d_path = job["d_image"]["path"]

        key = record_key(
            day_name,
            tray_name,
            capture_id,
        )

        output_folder = (
            OUTPUT_ROOT
            / day_name
            / tray_name
        )

        overlay_path = (
            output_folder
            / f"{safe_name(capture_id)}_70_square_grid_overlay.jpg"
        )

        if (
            not args.overwrite
            and key in summary_by_key
            and overlay_path.exists()
        ):
            print(
                f"SKIPPED EXISTING: "
                f"{day_name} > {tray_name} > {capture_id}"
            )
            continue

        try:
            rgb = read_rgb(d_path)

            if args.mode == "auto":
                cells, metrics = auto_detect_square_grid(
                    rgb,
                    args.max_dimension,
                    args.square_zone_ratio,
                )

                status = decide_auto_status(metrics)

                method = "automatic_global_square_grid"

                review_required = (
                    "No"
                    if status == "PASS_AUTO"
                    else "Yes"
                )

                if status == "PASS_AUTO":
                    notes = (
                        "Global square grid fitted successfully. "
                        "Only routine spot checks are recommended."
                    )
                elif status == "CHECK_AUTO":
                    notes = (
                        "Global square grid generated, but some cells "
                        "are fitted rather than directly supported. "
                        "Review the listed yellow squares."
                    )
                else:
                    notes = (
                        "Automatic grid confidence was insufficient. "
                        "Rerun this tray in manual mode."
                    )

            else:
                corner_points = get_manual_corner_centres(
                    rgb,
                    (
                        f"Manual square grid | {day_name} | "
                        f"{tray_name} | {capture_id}"
                    ),
                )

                if corner_points is None:
                    raise ValueError(
                        "Manual cell-centre selection was cancelled."
                    )

                cells, metrics = build_manual_square_grid(
                    corner_points,
                    args.square_zone_ratio,
                )

                status = "PASS_MANUAL"
                method = "manual_four_corner_square_grid"
                review_required = "No"

                notes = (
                    "Grid generated from manually clicked centres of "
                    "Cell 1, Cell 10, Cell 70 and Cell 61."
                )

                manual_config["records"][key] = {
                    "day": day_name,
                    "tray": tray_name,
                    "capture_id": capture_id,
                    "method": method,
                    "square_zone_ratio": args.square_zone_ratio,
                    "manual_corner_cell_centres": (
                        corner_points.tolist()
                    ),
                }

                save_json(
                    config_path,
                    manual_config,
                )

            overlay_title = (
                f"{day_name} | {tray_name} | "
                f"{capture_id} | {status} | 70 square cells"
            )

            save_square_grid_overlay(
                rgb,
                cells,
                overlay_path,
                overlay_title,
            )

            summary_by_key[key] = {
                "day_order": day_order,
                "day": day_name,
                "tray": tray_name,
                "tray_no": tray_no,
                "capture_id": capture_id,
                "source_d": relative_path(
                    d_path,
                    INPUT_ROOT,
                ),
                "method": method,
                "cells_generated": len(cells),
                "candidate_count": metrics["candidate_count"],
                "direct_cell_count": metrics[
                    "direct_cell_count"
                ],
                "uncertain_cell_count": metrics[
                    "uncertain_cell_count"
                ],
                "uncertain_cell_ids": ", ".join(
                    str(value)
                    for value in metrics["uncertain_cell_ids"]
                ),
                "square_side": round(
                    float(metrics["square_side"]),
                    3,
                ),
                "median_grid_error": round(
                    float(metrics["median_grid_error"]),
                    4,
                ),
                "max_grid_error": round(
                    float(metrics["max_grid_error"]),
                    4,
                ),
                "review_required": review_required,
                "overlay_path": relative_path(
                    overlay_path,
                    OUTPUT_ROOT,
                ),
                "status": status,
                "notes": notes,
            }

            coordinates_by_key[key] = []

            for cell in cells:
                coordinates_by_key[key].append(
                    {
                        "day_order": day_order,
                        "day": day_name,
                        "tray": tray_name,
                        "tray_no": tray_no,
                        "capture_id": capture_id,
                        "cell_id": cell["cell_id"],
                        "row": cell["row"],
                        "column": cell["column"],
                        "x": round(float(cell["x"]), 3),
                        "y": round(float(cell["y"]), 3),
                        "square_x0": round(
                            float(cell["square_x0"]),
                            3,
                        ),
                        "square_y0": round(
                            float(cell["square_y0"]),
                            3,
                        ),
                        "square_x1": round(
                            float(cell["square_x1"]),
                            3,
                        ),
                        "square_y1": round(
                            float(cell["square_y1"]),
                            3,
                        ),
                        "square_side": round(
                            float(cell["square_side"]),
                            3,
                        ),
                        "grid_error": round(
                            float(cell["grid_error"]),
                            4,
                        ),
                        "coordinate_source": cell[
                            "coordinate_source"
                        ],
                        "needs_review": cell["needs_review"],
                    }
                )

            print(
                f"{status}: {day_name} > {tray_name} > "
                f"{capture_id} | direct="
                f"{metrics['direct_cell_count']}/70 | uncertain="
                f"{metrics['uncertain_cell_count']}"
            )

        except Exception as error:
            summary_by_key[key] = {
                "day_order": day_order,
                "day": day_name,
                "tray": tray_name,
                "tray_no": tray_no,
                "capture_id": capture_id,
                "source_d": relative_path(
                    d_path,
                    INPUT_ROOT,
                ),
                "method": args.mode,
                "cells_generated": 0,
                "candidate_count": "",
                "direct_cell_count": "",
                "uncertain_cell_count": "",
                "uncertain_cell_ids": "",
                "square_side": "",
                "median_grid_error": "",
                "max_grid_error": "",
                "review_required": "Yes",
                "overlay_path": "",
                "status": "FAIL",
                "notes": str(error),
            }

            coordinates_by_key[key] = []

            print(
                f"FAIL: {day_name} > {tray_name} > "
                f"{capture_id} | {error}"
            )

    summary_rows = sorted(
        summary_by_key.values(),
        key=lambda row: (
            numeric_or_default(row.get("day_order")),
            natural_key(row.get("tray", "")),
            natural_key(row.get("capture_id", "")),
        ),
    )

    coordinate_rows = []

    for rows in coordinates_by_key.values():
        coordinate_rows.extend(rows)

    coordinate_rows.sort(
        key=lambda row: (
            numeric_or_default(row.get("day_order")),
            natural_key(row.get("tray", "")),
            natural_key(row.get("capture_id", "")),
            numeric_or_default(row.get("cell_id")),
        ),
    )

    summary_fields = [
        "day_order",
        "day",
        "tray",
        "tray_no",
        "capture_id",
        "source_d",
        "method",
        "cells_generated",
        "candidate_count",
        "direct_cell_count",
        "uncertain_cell_count",
        "uncertain_cell_ids",
        "square_side",
        "median_grid_error",
        "max_grid_error",
        "review_required",
        "overlay_path",
        "status",
        "notes",
    ]

    coordinate_fields = [
        "day_order",
        "day",
        "tray",
        "tray_no",
        "capture_id",
        "cell_id",
        "row",
        "column",
        "x",
        "y",
        "square_x0",
        "square_y0",
        "square_x1",
        "square_y1",
        "square_side",
        "grid_error",
        "coordinate_source",
        "needs_review",
    ]

    write_csv(
        manifest_path,
        summary_fields,
        summary_rows,
    )

    write_csv(
        coordinates_path,
        coordinate_fields,
        coordinate_rows,
    )

    create_excel_report(
        workbook_path,
        summary_rows,
        coordinate_rows,
    )

    pass_auto = sum(
        row.get("status") == "PASS_AUTO"
        for row in summary_rows
    )

    pass_manual = sum(
        row.get("status") == "PASS_MANUAL"
        for row in summary_rows
    )

    check_auto = sum(
        row.get("status") == "CHECK_AUTO"
        for row in summary_rows
    )

    failed = sum(
        row.get("status") == "FAIL"
        for row in summary_rows
    )

    print("\n" + "=" * 70)
    print("SCRIPT 03 FINISHED")
    print("=" * 70)
    print(f"PASS_AUTO: {pass_auto}")
    print(f"PASS_MANUAL: {pass_manual}")
    print(f"CHECK_AUTO: {check_auto}")
    print(f"FAIL: {failed}")
    print(f"\nExcel report:\n{workbook_path}")
    print(f"\nCell coordinates:\n{coordinates_path}")
    print(f"\nSquare overlays:\n{OUTPUT_ROOT}")

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())