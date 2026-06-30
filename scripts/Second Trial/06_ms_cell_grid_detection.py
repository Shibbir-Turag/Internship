from __future__ import annotations

"""
SCRIPT 06 — ROBUST INDEPENDENT MS CELL GRID DETECTION

This script detects the fixed 7 x 10 = 70-cell grid independently on MS_NIR
images. It does not use D/RGB coordinates.

Key correction:
- Each Hough-circle configuration is evaluated separately.
- The script selects the single most regular near-70-cell detection result.
- It does NOT merge circle detections from several thresholds.
- Final cell positions always come from one fitted global 7 x 10 lattice.

Outputs:
outputs/Second Trial/06_MS_Cell_Grid_Detection/
    Day X/Tray Y/
        <capture>_70_ms_square_grid_overlay.jpg

    _reports/
        ms_grid_manifest.csv
        ms_cell_coordinates.csv
        ms_cell_grid_report.xlsx

    _config/
        ms_grid_detection_settings.json
        manual_ms_grid_points.json
"""

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

import cv2
import matplotlib

matplotlib.use("TkAgg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageOps


# ============================================================
# 1) PATHS — CHANGE PROJECT_ROOT ONLY WHEN REUSING
# ============================================================

PROJECT_ROOT = Path(
    r"C:\Users\tshib\OneDrive\Desktop\Internship"
)

INPUT_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "01_Crop_Dual_Reference"
)

OUTPUT_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "06_MS_Cell_Grid_Detection"
)


# ============================================================
# 2) TRAY / DETECTION SETTINGS
# ============================================================

ROWS = 7
COLS = 10
EXPECTED_CELLS = ROWS * COLS

MS_BANDS = (
    "MS_G",
    "MS_R",
    "MS_RE",
    "MS_NIR",
)

SUPPORTED_EXTENSIONS = {
    ".jpg",
    ".jpeg",
    ".png",
    ".tif",
    ".tiff",
}

DAY_NAME_TO_ORDER = {
    "day 1": 1,
    "day 2": 2,
    "day 3": 3,
    "day 4": 4,
    "day 5": 5,
    "day 9": 9,
    "first day": 1,
    "second day": 2,
    "third day": 3,
    "fourth day": 4,
    "fifth day": 5,
    "ninth day": 9,
}

DEFAULT_MAX_DIMENSION = 1800
DEFAULT_SQUARE_ZONE_RATIO = 0.90

# Only near-70 candidate configurations are considered.
MIN_CANDIDATES_FOR_AUTO = 55
MAX_CANDIDATES_FOR_AUTO = 90

# PASS_AUTO requirements.
PASS_MIN_SUPPORTED_CELLS = 68
PASS_MAX_MEDIAN_ERROR = 0.12
PASS_MAX_SPACING_CV = 0.12

# CHECK_AUTO requirements.
CHECK_MIN_SUPPORTED_CELLS = 60
CHECK_MAX_MEDIAN_ERROR = 0.22
CHECK_MAX_SPACING_CV = 0.20


# ============================================================
# 3) GENERAL HELPERS
# ============================================================

def natural_key(value: object):
    return [
        int(part) if part.isdigit() else part.casefold()
        for part in re.split(r"(\d+)", str(value))
    ]


def day_sort_key(folder: Path):
    return (
        DAY_NAME_TO_ORDER.get(
            folder.name.casefold(),
            999,
        ),
        natural_key(folder.name),
    )


def parse_filter_list(value: str | None):
    if not value:
        return None

    return {
        item.strip().casefold()
        for item in value.split(",")
        if item.strip()
    }


def tray_number(tray_name: str):
    match = re.search(r"(\d+)", tray_name)
    return int(match.group(1)) if match else ""


def safe_name(value: object):
    return re.sub(
        r"[^A-Za-z0-9_.-]+",
        "_",
        str(value),
    )


def relative_path(path: Path | None, root: Path):
    if path is None:
        return ""

    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


def record_key(
    day: str,
    tray: str,
    capture_id: str,
):
    return f"{day}|{tray}|{capture_id}"


def int_or_default(value: object, default=999):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


# ============================================================
# 4) FIND COMPLETE MS BAND SETS
# ============================================================

def parse_ms_filename(path: Path):
    """
    Example:
        DJI_20260618124331_0005_MS_NIR.TIF
    """

    match = re.match(
        r"^(?P<capture>.+)_"
        r"(?P<band>MS_G|MS_R|MS_RE|MS_NIR)$",
        path.stem.upper(),
    )

    if match is None:
        return None

    return (
        match.group("capture"),
        match.group("band"),
    )


def find_ms_sets(tray_folder: Path):
    grouped: dict[str, dict[str, Path]] = defaultdict(dict)

    for file in tray_folder.iterdir():
        if not file.is_file():
            continue

        if file.suffix.casefold() not in SUPPORTED_EXTENSIONS:
            continue

        parsed = parse_ms_filename(file)

        if parsed is None:
            continue

        capture_id, band = parsed
        grouped[capture_id][band] = file

    return [
        (capture_id, grouped[capture_id])
        for capture_id in sorted(
            grouped,
            key=natural_key,
        )
    ]


# ============================================================
# 5) IMAGE READING
# ============================================================

def read_band_array(path: Path):
    with Image.open(path) as image:
        image = ImageOps.exif_transpose(image)
        return np.asarray(image)


def read_band_shape(path: Path):
    image = read_band_array(path)

    if image.ndim < 2:
        raise ValueError(
            f"Unsupported band image shape: {image.shape}"
        )

    return image.shape[1], image.shape[0]


def to_gray_8bit(image: np.ndarray):
    """
    Converts the raw MS_NIR image into an 8-bit display/detection image.

    The original NIR values are not changed or saved over.
    """

    if image.ndim == 3:
        gray = image[:, :, :3].astype(np.float32).mean(axis=2)
    else:
        gray = image.astype(np.float32)

    low, high = np.percentile(
        gray,
        [1, 99],
    )

    if high <= low:
        low = float(gray.min())
        high = float(gray.max())

    if high <= low:
        return np.zeros(
            gray.shape,
            dtype=np.uint8,
        )

    normalised = (
        (gray - low)
        * 255.0
        / (high - low)
    )

    return np.clip(
        normalised,
        0,
        255,
    ).astype(np.uint8)


def resize_for_detection(
    gray: np.ndarray,
    max_dimension: int,
):
    height, width = gray.shape[:2]
    largest_dimension = max(height, width)

    if (
        max_dimension <= 0
        or largest_dimension <= max_dimension
    ):
        return gray.copy(), 1.0

    scale = max_dimension / largest_dimension

    resized = cv2.resize(
        gray,
        (
            int(round(width * scale)),
            int(round(height * scale)),
        ),
        interpolation=cv2.INTER_AREA,
    )

    return resized, scale


# ============================================================
# 6) GRID FITTING HELPERS
# ============================================================

def kmeans_1d(
    values: np.ndarray,
    cluster_count: int,
):
    values = np.asarray(
        values,
        dtype=np.float32,
    ).reshape(-1, 1)

    criteria = (
        cv2.TERM_CRITERIA_EPS
        + cv2.TERM_CRITERIA_MAX_ITER,
        100,
        0.1,
    )

    cv2.setRNGSeed(42)

    _, _, centres = cv2.kmeans(
        values,
        cluster_count,
        None,
        criteria,
        20,
        cv2.KMEANS_PP_CENTERS,
    )

    return np.sort(
        centres.reshape(-1)
    )


def build_nodes(
    x_centres: np.ndarray,
    y_centres: np.ndarray,
):
    nodes = []

    for row, y_value in enumerate(
        y_centres,
        start=1,
    ):
        for column, x_value in enumerate(
            x_centres,
            start=1,
        ):
            nodes.append(
                {
                    "cell_id": (
                        (row - 1) * COLS
                        + column
                    ),
                    "row": row,
                    "column": column,
                    "x": float(x_value),
                    "y": float(y_value),
                }
            )

    return nodes


def greedy_matches(
    nodes: list[dict],
    circles: np.ndarray,
    pitch: float,
    maximum_normalised_distance: float,
):
    """
    Match one Hough circle to one expected lattice node.
    """

    pairs = []

    for node_index, node in enumerate(nodes):
        for circle_index, circle in enumerate(circles):
            distance = np.hypot(
                node["x"] - circle[0],
                node["y"] - circle[1],
            )

            normalised_distance = (
                distance / max(pitch, 1e-6)
            )

            if (
                normalised_distance
                <= maximum_normalised_distance
            ):
                pairs.append(
                    (
                        normalised_distance,
                        node_index,
                        circle_index,
                    )
                )

    pairs.sort()

    matches = {}
    used_nodes = set()
    used_circles = set()

    for distance, node_index, circle_index in pairs:
        if node_index in used_nodes:
            continue

        if circle_index in used_circles:
            continue

        matches[node_index] = (
            circle_index,
            float(distance),
        )

        used_nodes.add(node_index)
        used_circles.add(circle_index)

    return matches


def fit_affine_grid(
    nodes: list[dict],
    circles: np.ndarray,
    matches: dict,
):
    """
    Fit:
        x = a0 + a1 * column + a2 * row
        y = b0 + b1 * column + b2 * row

    The final 70 cell centres are always generated from this full lattice,
    not from individual raw Hough circle coordinates.
    """

    if len(matches) < 20:
        raise ValueError(
            "Too few reliable circles to fit the MS lattice."
        )

    features = []
    positions = []

    for node_index, (
        circle_index,
        _error,
    ) in matches.items():
        node = nodes[node_index]
        circle = circles[circle_index]

        features.append(
            [
                1.0,
                node["column"] - 1,
                node["row"] - 1,
            ]
        )

        positions.append(
            [
                circle[0],
                circle[1],
            ]
        )

    coefficients, _, _, _ = np.linalg.lstsq(
        np.asarray(
            features,
            dtype=np.float64,
        ),
        np.asarray(
            positions,
            dtype=np.float64,
        ),
        rcond=None,
    )

    fitted_nodes = []

    for row in range(1, ROWS + 1):
        for column in range(1, COLS + 1):
            x_value, y_value = np.asarray(
                [
                    1.0,
                    column - 1,
                    row - 1,
                ],
                dtype=np.float64,
            ) @ coefficients

            fitted_nodes.append(
                {
                    "cell_id": (
                        (row - 1) * COLS
                        + column
                    ),
                    "row": row,
                    "column": column,
                    "x": float(x_value),
                    "y": float(y_value),
                }
            )

    x_pitch = float(
        np.linalg.norm(coefficients[1])
    )

    y_pitch = float(
        np.linalg.norm(coefficients[2])
    )

    return fitted_nodes, x_pitch, y_pitch


def evaluate_circle_configuration(
    circles: np.ndarray,
):
    """
    Evaluate one Hough detection configuration independently.

    This is the key correction: circle sets are never merged.
    """

    candidate_count = len(circles)

    if (
        candidate_count < MIN_CANDIDATES_FOR_AUTO
        or candidate_count > MAX_CANDIDATES_FOR_AUTO
    ):
        return None

    x_centres = kmeans_1d(
        circles[:, 0],
        COLS,
    )

    y_centres = kmeans_1d(
        circles[:, 1],
        ROWS,
    )

    x_spacing = np.diff(x_centres)
    y_spacing = np.diff(y_centres)

    if (
        len(x_spacing) != COLS - 1
        or len(y_spacing) != ROWS - 1
    ):
        return None

    x_pitch_initial = float(
        np.median(x_spacing)
    )

    y_pitch_initial = float(
        np.median(y_spacing)
    )

    initial_pitch = min(
        x_pitch_initial,
        y_pitch_initial,
    )

    if initial_pitch <= 0:
        return None

    spacing_cv = max(
        float(
            np.std(x_spacing)
            / max(np.mean(x_spacing), 1e-6)
        ),
        float(
            np.std(y_spacing)
            / max(np.mean(y_spacing), 1e-6)
        ),
    )

    nodes = build_nodes(
        x_centres,
        y_centres,
    )

    matches = greedy_matches(
        nodes,
        circles,
        initial_pitch,
        maximum_normalised_distance=0.55,
    )

    if len(matches) < 20:
        return None

    nodes, x_pitch, y_pitch = fit_affine_grid(
        nodes,
        circles,
        matches,
    )

    refined_pitch = min(
        x_pitch,
        y_pitch,
    )

    matches = greedy_matches(
        nodes,
        circles,
        refined_pitch,
        maximum_normalised_distance=0.32,
    )

    inlier_matches = {
        node_index: match
        for node_index, match in matches.items()
        if match[1] <= 0.24
    }

    if len(inlier_matches) >= 20:
        nodes, x_pitch, y_pitch = fit_affine_grid(
            nodes,
            circles,
            inlier_matches,
        )

        refined_pitch = min(
            x_pitch,
            y_pitch,
        )

        matches = greedy_matches(
            nodes,
            circles,
            refined_pitch,
            maximum_normalised_distance=0.30,
        )

    errors = []

    supported_node_indexes = set()

    for node_index in range(EXPECTED_CELLS):
        if node_index in matches:
            error = matches[node_index][1]
        else:
            error = 1.0

        errors.append(float(error))

        if error <= 0.24:
            supported_node_indexes.add(node_index)

    supported_cell_count = len(
        supported_node_indexes
    )

    median_error = float(
        np.median(errors)
    )

    max_error = float(max(errors))

    # Prefer circle sets close to 70 with a regular, low-error lattice.
    quality_score = (
        supported_cell_count * 30.0
        - abs(candidate_count - EXPECTED_CELLS) * 25.0
        - median_error * 600.0
        - spacing_cv * 500.0
    )

    return {
        "nodes": nodes,
        "matches": matches,
        "candidate_count": candidate_count,
        "supported_cell_count": supported_cell_count,
        "median_grid_error": median_error,
        "max_grid_error": max_error,
        "spacing_cv": spacing_cv,
        "x_pitch": x_pitch,
        "y_pitch": y_pitch,
        "quality_score": quality_score,
    }


# ============================================================
# 7) AUTOMATIC MS_NIR GRID DETECTION
# ============================================================

def select_best_ms_grid(
    gray: np.ndarray,
):
    """
    Test multiple individual Hough configurations and select the most regular
    7 x 10 lattice. No configurations are merged.
    """

    enhanced = cv2.createCLAHE(
        clipLimit=2.0,
        tileGridSize=(8, 8),
    ).apply(gray)

    blurred = cv2.GaussianBlur(
        enhanced,
        (5, 5),
        0,
    )

    height, width = gray.shape[:2]

    estimated_pitch = min(
        width / COLS,
        height / ROWS,
    )

    min_radius = max(
        7,
        int(round(estimated_pitch * 0.16)),
    )

    max_radius = max(
        min_radius + 6,
        int(round(estimated_pitch * 0.54)),
    )

    candidates = []

    # Strict thresholds are intentionally tested first.
    # They usually detect one circle per real planting recess.
    for min_distance_ratio in (
        0.50,
        0.60,
        0.70,
    ):
        for param2 in (
            68,
            64,
            60,
            56,
            52,
            48,
        ):
            min_distance = max(
                15,
                int(
                    round(
                        estimated_pitch
                        * min_distance_ratio
                    )
                ),
            )

            circles = cv2.HoughCircles(
                blurred,
                cv2.HOUGH_GRADIENT,
                dp=1.2,
                minDist=min_distance,
                param1=100,
                param2=param2,
                minRadius=min_radius,
                maxRadius=max_radius,
            )

            if circles is None:
                continue

            circle_array = np.round(
                circles[0]
            ).astype(np.float32)

            result = evaluate_circle_configuration(
                circle_array
            )

            if result is None:
                continue

            result["param2"] = param2
            result["min_distance_ratio"] = (
                min_distance_ratio
            )
            result["min_distance"] = min_distance
            result["min_radius"] = min_radius
            result["max_radius"] = max_radius

            candidates.append(result)

    if not candidates:
        raise ValueError(
            "No reliable near-70-cell Hough configuration "
            "was found in MS_NIR."
        )

    best = max(
        candidates,
        key=lambda item: item["quality_score"],
    )

    return best


def auto_detect_grid(
    nir_image: np.ndarray,
    max_dimension: int,
    square_zone_ratio: float,
):
    """
    Generate 70 square MS cell zones from MS_NIR.
    """

    full_gray = to_gray_8bit(nir_image)

    detection_gray, scale = resize_for_detection(
        full_gray,
        max_dimension,
    )

    best = select_best_ms_grid(
        detection_gray
    )

    pitch = min(
        best["x_pitch"],
        best["y_pitch"],
    )

    square_side_detection = (
        pitch * square_zone_ratio
    )

    half_side_detection = (
        square_side_detection / 2.0
    )

    cells = []

    uncertain_cell_ids = []

    for node_index, node in enumerate(
        best["nodes"]
    ):
        if (
            node_index in best["matches"]
            and best["matches"][node_index][1] <= 0.24
        ):
            coordinate_source = (
                "ms_nir_circle_supported_lattice"
            )
            needs_review = "No"
            grid_error = best["matches"][node_index][1]
        else:
            coordinate_source = (
                "ms_nir_fitted_lattice"
            )
            needs_review = "Yes"
            grid_error = 1.0
            uncertain_cell_ids.append(
                node["cell_id"]
            )

        x_value = node["x"] / scale
        y_value = node["y"] / scale
        square_side = (
            square_side_detection / scale
        )
        half_side = (
            half_side_detection / scale
        )

        cells.append(
            {
                "cell_id": node["cell_id"],
                "row": node["row"],
                "column": node["column"],
                "x": float(x_value),
                "y": float(y_value),
                "square_side": float(square_side),
                "square_x0": float(
                    x_value - half_side
                ),
                "square_y0": float(
                    y_value - half_side
                ),
                "square_x1": float(
                    x_value + half_side
                ),
                "square_y1": float(
                    y_value + half_side
                ),
                "grid_error": float(grid_error),
                "coordinate_source": coordinate_source,
                "needs_review": needs_review,
            }
        )

    metadata = {
        "candidate_count": best["candidate_count"],
        "supported_cell_count": best[
            "supported_cell_count"
        ],
        "uncertain_cell_count": len(
            uncertain_cell_ids
        ),
        "uncertain_cell_ids": uncertain_cell_ids,
        "square_side": square_side_detection / scale,
        "x_pitch": best["x_pitch"] / scale,
        "y_pitch": best["y_pitch"] / scale,
        "median_grid_error": best[
            "median_grid_error"
        ],
        "max_grid_error": best[
            "max_grid_error"
        ],
        "spacing_cv": best["spacing_cv"],
        "selected_hough_param2": best["param2"],
        "selected_min_distance_ratio": best[
            "min_distance_ratio"
        ],
        "selected_min_distance": best[
            "min_distance"
        ],
        "selected_min_radius": best[
            "min_radius"
        ],
        "selected_max_radius": best[
            "max_radius"
        ],
        "detection_scale": scale,
    }

    return full_gray, cells, metadata


# ============================================================
# 8) MANUAL FOUR-CORNER FALLBACK
# ============================================================

def get_manual_corner_points(
    gray: np.ndarray,
    title: str,
):
    figure, axis = plt.subplots(
        figsize=(15, 10)
    )

    axis.imshow(
        gray,
        cmap="gray",
    )

    axis.set_title(
        title
        + "\nClick: Cell 1 → Cell 10 → Cell 70 → Cell 61"
        + "\nRight-click removes the latest point.",
        fontsize=12,
    )

    axis.axis("off")

    plt.tight_layout()
    plt.show(block=False)

    points = plt.ginput(
        n=4,
        timeout=0,
        show_clicks=True,
        mouse_add=1,
        mouse_pop=3,
        mouse_stop=2,
    )

    plt.close(figure)

    if len(points) != 4:
        return None

    return np.asarray(
        points,
        dtype=np.float32,
    )


def build_manual_grid(
    corner_points: np.ndarray,
    square_zone_ratio: float,
):
    source_points = np.asarray(
        [
            [0, 0],
            [COLS - 1, 0],
            [COLS - 1, ROWS - 1],
            [0, ROWS - 1],
        ],
        dtype=np.float32,
    )

    transform = cv2.getPerspectiveTransform(
        source_points,
        corner_points.astype(np.float32),
    )

    lattice_points = np.asarray(
        [
            [[column, row]]
            for row in range(ROWS)
            for column in range(COLS)
        ],
        dtype=np.float32,
    )

    positions = cv2.perspectiveTransform(
        lattice_points,
        transform,
    ).reshape(-1, 2)

    horizontal_distances = []

    for row in range(ROWS):
        for column in range(COLS - 1):
            left = positions[
                row * COLS + column
            ]

            right = positions[
                row * COLS + column + 1
            ]

            horizontal_distances.append(
                float(np.linalg.norm(right - left))
            )

    vertical_distances = []

    for row in range(ROWS - 1):
        for column in range(COLS):
            top = positions[
                row * COLS + column
            ]

            bottom = positions[
                (row + 1) * COLS + column
            ]

            vertical_distances.append(
                float(np.linalg.norm(bottom - top))
            )

    x_pitch = float(
        np.median(horizontal_distances)
    )

    y_pitch = float(
        np.median(vertical_distances)
    )

    square_side = (
        min(x_pitch, y_pitch)
        * square_zone_ratio
    )

    half_side = square_side / 2.0

    cells = []

    for index, (x_value, y_value) in enumerate(
        positions,
        start=1,
    ):
        row = ((index - 1) // COLS) + 1
        column = ((index - 1) % COLS) + 1

        cells.append(
            {
                "cell_id": index,
                "row": row,
                "column": column,
                "x": float(x_value),
                "y": float(y_value),
                "square_side": float(square_side),
                "square_x0": float(
                    x_value - half_side
                ),
                "square_y0": float(
                    y_value - half_side
                ),
                "square_x1": float(
                    x_value + half_side
                ),
                "square_y1": float(
                    y_value + half_side
                ),
                "grid_error": 0.0,
                "coordinate_source": (
                    "manual_four_corner_ms_grid"
                ),
                "needs_review": "No",
            }
        )

    metadata = {
        "candidate_count": "",
        "supported_cell_count": EXPECTED_CELLS,
        "uncertain_cell_count": 0,
        "uncertain_cell_ids": [],
        "square_side": square_side,
        "x_pitch": x_pitch,
        "y_pitch": y_pitch,
        "median_grid_error": 0.0,
        "max_grid_error": 0.0,
        "spacing_cv": 0.0,
        "selected_hough_param2": "",
        "selected_min_distance_ratio": "",
        "selected_min_distance": "",
        "selected_min_radius": "",
        "selected_max_radius": "",
        "detection_scale": 1.0,
    }

    return cells, metadata


# ============================================================
# 9) STATUS / OVERLAY
# ============================================================

def determine_status(metadata: dict):
    if (
        metadata["supported_cell_count"]
        >= PASS_MIN_SUPPORTED_CELLS
        and metadata["median_grid_error"]
        <= PASS_MAX_MEDIAN_ERROR
        and metadata["spacing_cv"]
        <= PASS_MAX_SPACING_CV
    ):
        return "PASS_AUTO"

    if (
        metadata["supported_cell_count"]
        >= CHECK_MIN_SUPPORTED_CELLS
        and metadata["median_grid_error"]
        <= CHECK_MAX_MEDIAN_ERROR
        and metadata["spacing_cv"]
        <= CHECK_MAX_SPACING_CV
    ):
        return "CHECK_AUTO"

    return "FAIL"


def save_overlay(
    gray: np.ndarray,
    cells: list[dict],
    output_path: Path,
    title: str,
):
    overlay = cv2.cvtColor(
        gray,
        cv2.COLOR_GRAY2BGR,
    )

    image_height, image_width = overlay.shape[:2]

    median_side = float(
        np.median(
            [
                cell["square_side"]
                for cell in cells
            ]
        )
    )

    line_width = max(
        2,
        int(round(median_side / 32)),
    )

    font_scale = max(
        0.35,
        min(1.0, median_side / 88),
    )

    font_thickness = max(
        1,
        int(round(font_scale * 2)),
    )

    for cell in cells:
        x0 = max(
            0,
            int(round(cell["square_x0"])),
        )

        y0 = max(
            0,
            int(round(cell["square_y0"])),
        )

        x1 = min(
            image_width - 1,
            int(round(cell["square_x1"])),
        )

        y1 = min(
            image_height - 1,
            int(round(cell["square_y1"])),
        )

        if (
            cell["coordinate_source"]
            == "manual_four_corner_ms_grid"
        ):
            colour = (255, 150, 0)  # Blue

        elif cell["needs_review"] == "Yes":
            colour = (0, 180, 255)  # Yellow

        else:
            colour = (0, 180, 0)  # Green

        cv2.rectangle(
            overlay,
            (x0, y0),
            (x1, y1),
            colour,
            line_width,
        )

        label_x = max(
            1,
            int(
                round(
                    cell["x"]
                    - median_side * 0.10
                )
            ),
        )

        label_y = min(
            image_height - 5,
            int(
                round(
                    cell["y"]
                    + median_side * 0.08
                )
            ),
        )

        cv2.putText(
            overlay,
            str(cell["cell_id"]),
            (label_x, label_y),
            cv2.FONT_HERSHEY_SIMPLEX,
            font_scale,
            colour,
            font_thickness,
            cv2.LINE_AA,
        )

    header_height = max(
        42,
        int(round(median_side * 0.30)),
    )

    cv2.rectangle(
        overlay,
        (0, 0),
        (image_width, header_height),
        (255, 255, 255),
        thickness=-1,
    )

    cv2.putText(
        overlay,
        title,
        (12, int(header_height * 0.68)),
        cv2.FONT_HERSHEY_SIMPLEX,
        max(
            0.42,
            min(0.85, font_scale),
        ),
        (0, 0, 0),
        2,
        cv2.LINE_AA,
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    cv2.imwrite(
        str(output_path),
        overlay,
        [
            cv2.IMWRITE_JPEG_QUALITY,
            95,
        ],
    )


# ============================================================
# 10) REPORT HELPERS
# ============================================================

def load_csv_records(path: Path):
    if not path.exists():
        return []

    try:
        return pd.read_csv(path).to_dict(
            orient="records"
        )
    except Exception:
        return []


def save_excel_report(
    path: Path,
    manifest_rows: list[dict],
    coordinate_rows: list[dict],
):
    manifest_frame = pd.DataFrame(manifest_rows)
    coordinate_frame = pd.DataFrame(coordinate_rows)

    readme_frame = pd.DataFrame(
        {
            "Notes": [
                "This script detects the 7 × 10 cell layout independently from MS_NIR images.",
                "D/RGB coordinates are not used or copied into this MS workflow.",
                "MS_G, MS_R, MS_RE and MS_NIR are treated as aligned only within the multispectral image set.",
                "Green squares are supported by the selected MS_NIR circle configuration.",
                "Yellow squares are fitted grid positions that need review.",
                "Blue squares are manually generated from Cell 1, Cell 10, Cell 70 and Cell 61.",
                "Only PASS_AUTO and PASS_MANUAL grid results should be used in Script 07.",
            ]
        }
    )

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with pd.ExcelWriter(
        path,
        engine="openpyxl",
    ) as writer:
        manifest_frame.to_excel(
            writer,
            sheet_name="MS Grid Summary",
            index=False,
        )

        coordinate_frame.to_excel(
            writer,
            sheet_name="MS Cell Coordinates",
            index=False,
        )

        readme_frame.to_excel(
            writer,
            sheet_name="Read Me",
            index=False,
        )

    workbook = load_workbook(path)

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 34

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(
                color="FFFFFF",
                bold=True,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in worksheet.columns:
            letter = column_cells[0].column_letter

            max_length = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )

            worksheet.column_dimensions[letter].width = min(
                max(max_length + 2, 12),
                58,
            )

    workbook.save(path)


# ============================================================
# 11) MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Script 06: robust independent 7 x 10 MS cell "
            "grid detection from cropped MS_NIR images."
        )
    )

    parser.add_argument(
        "--days",
        help='Example: --days "Day 1,Day 9"',
    )

    parser.add_argument(
        "--trays",
        help='Example: --trays "Tray 1,Tray 7"',
    )

    parser.add_argument(
        "--mode",
        choices=("auto", "manual"),
        default="auto",
    )

    parser.add_argument(
        "--max-dimension",
        type=int,
        default=DEFAULT_MAX_DIMENSION,
    )

    parser.add_argument(
        "--square-zone-ratio",
        type=float,
        default=DEFAULT_SQUARE_ZONE_RATIO,
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
    )

    args = parser.parse_args()

    if not 0 < args.square_zone_ratio < 1:
        print(
            "ERROR: --square-zone-ratio must be "
            "greater than 0 and less than 1."
        )
        return 1

    if not INPUT_ROOT.exists():
        print(
            "ERROR: Script 01 crop folder not found:\n"
            f"{INPUT_ROOT}"
        )
        return 1

    selected_days = parse_filter_list(args.days)
    selected_trays = parse_filter_list(args.trays)

    jobs = []

    day_folders = sorted(
        [
            folder
            for folder in INPUT_ROOT.iterdir()
            if (
                folder.is_dir()
                and folder.name.casefold()
                in DAY_NAME_TO_ORDER
            )
        ],
        key=day_sort_key,
    )

    for day_folder in day_folders:
        if (
            selected_days
            and day_folder.name.casefold()
            not in selected_days
        ):
            continue

        tray_folders = sorted(
            [
                folder
                for folder in day_folder.iterdir()
                if folder.is_dir()
            ],
            key=lambda folder: natural_key(
                folder.name
            ),
        )

        for tray_folder in tray_folders:
            if (
                selected_trays
                and tray_folder.name.casefold()
                not in selected_trays
            ):
                continue

            ms_sets = find_ms_sets(tray_folder)

            if not ms_sets:
                jobs.append(
                    {
                        "day_folder": day_folder,
                        "tray_folder": tray_folder,
                        "capture_id": "",
                        "bands": {},
                    }
                )

            for capture_id, bands in ms_sets:
                jobs.append(
                    {
                        "day_folder": day_folder,
                        "tray_folder": tray_folder,
                        "capture_id": capture_id,
                        "bands": bands,
                    }
                )

    if not jobs:
        print(
            "No cropped multispectral band sets were found."
        )
        return 1

    print("\nSCRIPT 06 — ROBUST MS CELL GRID DETECTION")
    print("=" * 70)
    print(f"Input crops:\n{INPUT_ROOT}")
    print(f"\nOutput folder:\n{OUTPUT_ROOT}")
    print(f"\nMode: {args.mode.upper()}")

    for job in jobs:
        missing_bands = [
            band
            for band in MS_BANDS
            if band not in job["bands"]
        ]

        status = (
            "READY"
            if not missing_bands
            else (
                "INCOMPLETE: "
                + ", ".join(missing_bands)
            )
        )

        print(
            f"{status}: "
            f"{job['day_folder'].name} > "
            f"{job['tray_folder'].name} > "
            f"{job['capture_id']}"
        )

    if args.dry_run:
        print(
            "\nDry run complete. No outputs created."
        )
        return 0

    reports_root = OUTPUT_ROOT / "_reports"

    manifest_path = (
        reports_root
        / "ms_grid_manifest.csv"
    )

    coordinates_path = (
        reports_root
        / "ms_cell_coordinates.csv"
    )

    workbook_path = (
        reports_root
        / "ms_cell_grid_report.xlsx"
    )

    settings_path = (
        OUTPUT_ROOT
        / "_config"
        / "ms_grid_detection_settings.json"
    )

    manual_points_path = (
        OUTPUT_ROOT
        / "_config"
        / "manual_ms_grid_points.json"
    )

    existing_manifest = load_csv_records(
        manifest_path
    )

    existing_coordinates = load_csv_records(
        coordinates_path
    )

    manifest_by_key = {
        record_key(
            str(row.get("day", "")),
            str(row.get("tray", "")),
            str(row.get("capture_id", "")),
        ): row
        for row in existing_manifest
    }

    coordinates_by_key: dict[str, list[dict]] = defaultdict(list)

    for row in existing_coordinates:
        key = record_key(
            str(row.get("day", "")),
            str(row.get("tray", "")),
            str(row.get("capture_id", "")),
        )

        coordinates_by_key[key].append(row)

    if manual_points_path.exists():
        try:
            manual_points_config = json.loads(
                manual_points_path.read_text(
                    encoding="utf-8"
                )
            )
        except Exception:
            manual_points_config = {"records": {}}
    else:
        manual_points_config = {"records": {}}

    for job in jobs:
        day = job["day_folder"].name
        tray = job["tray_folder"].name
        capture_id = job["capture_id"]
        bands = job["bands"]

        key = record_key(
            day,
            tray,
            capture_id,
        )

        day_order = DAY_NAME_TO_ORDER.get(
            day.casefold(),
            999,
        )

        tray_no = tray_number(tray)

        band_paths = {
            band: relative_path(
                bands.get(band),
                INPUT_ROOT,
            )
            if band in bands
            else ""
            for band in MS_BANDS
        }

        missing_bands = [
            band
            for band in MS_BANDS
            if band not in bands
        ]

        output_folder = (
            OUTPUT_ROOT
            / day
            / tray
        )

        overlay_path = (
            output_folder
            / (
                f"{safe_name(capture_id)}"
                "_70_ms_square_grid_overlay.jpg"
            )
        )

        if (
            not args.overwrite
            and key in manifest_by_key
            and overlay_path.exists()
        ):
            print(
                f"SKIPPED EXISTING: "
                f"{day} > {tray} > {capture_id}"
            )
            continue

        if missing_bands:
            manifest_by_key[key] = {
                "day_order": day_order,
                "day": day,
                "tray": tray,
                "tray_no": tray_no,
                "capture_id": capture_id,
                "ms_g_path": band_paths["MS_G"],
                "ms_r_path": band_paths["MS_R"],
                "ms_re_path": band_paths["MS_RE"],
                "ms_nir_path": band_paths["MS_NIR"],
                "band_dimensions_match": "No",
                "method": args.mode,
                "cells_generated": 0,
                "candidate_count": "",
                "supported_cell_count": "",
                "uncertain_cell_count": "",
                "uncertain_cell_ids": "",
                "square_side": "",
                "median_grid_error": "",
                "max_grid_error": "",
                "spacing_cv": "",
                "selected_hough_param2": "",
                "selected_min_distance_ratio": "",
                "review_required": "Yes",
                "overlay_path": "",
                "status": "FAIL",
                "notes": (
                    "Missing MS band(s): "
                    + ", ".join(missing_bands)
                ),
            }

            coordinates_by_key[key] = []
            continue

        try:
            band_shapes = {
                band: read_band_shape(path)
                for band, path in bands.items()
                if band in MS_BANDS
            }

            dimensions_match = (
                len(set(band_shapes.values())) == 1
            )

            if not dimensions_match:
                raise ValueError(
                    "MS band dimensions do not match: "
                    + str(band_shapes)
                )

            nir_image = read_band_array(
                bands["MS_NIR"]
            )

            if args.mode == "auto":
                gray, cells, metadata = auto_detect_grid(
                    nir_image,
                    args.max_dimension,
                    args.square_zone_ratio,
                )

                status = determine_status(metadata)

                method = (
                    "robust_auto_ms_nir_global_lattice"
                )

                notes = (
                    "Selected one independent near-70-cell "
                    "MS_NIR Hough configuration and fitted a "
                    "global 7 x 10 lattice."
                )

            else:
                gray = to_gray_8bit(nir_image)

                manual_points = get_manual_corner_points(
                    gray,
                    (
                        f"Manual MS grid | {day} | "
                        f"{tray} | {capture_id}"
                    ),
                )

                if manual_points is None:
                    raise ValueError(
                        "Manual corner selection was cancelled."
                    )

                cells, metadata = build_manual_grid(
                    manual_points,
                    args.square_zone_ratio,
                )

                status = "PASS_MANUAL"

                method = (
                    "manual_four_corner_ms_square_grid"
                )

                notes = (
                    "Grid created from manually selected "
                    "Cell 1, Cell 10, Cell 70 and Cell 61 centres."
                )

                manual_points_config.setdefault(
                    "records",
                    {},
                )[key] = {
                    "day": day,
                    "tray": tray,
                    "capture_id": capture_id,
                    "method": method,
                    "square_zone_ratio": (
                        args.square_zone_ratio
                    ),
                    "manual_corner_cell_centres": (
                        manual_points.tolist()
                    ),
                }

                manual_points_path.parent.mkdir(
                    parents=True,
                    exist_ok=True,
                )

                manual_points_path.write_text(
                    json.dumps(
                        manual_points_config,
                        indent=2,
                    ),
                    encoding="utf-8",
                )

            overlay_title = (
                f"{day} | {tray} | {capture_id} | "
                f"{status} | 70 MS square cells"
            )

            save_overlay(
                gray,
                cells,
                overlay_path,
                overlay_title,
            )

            manifest_by_key[key] = {
                "day_order": day_order,
                "day": day,
                "tray": tray,
                "tray_no": tray_no,
                "capture_id": capture_id,
                "ms_g_path": band_paths["MS_G"],
                "ms_r_path": band_paths["MS_R"],
                "ms_re_path": band_paths["MS_RE"],
                "ms_nir_path": band_paths["MS_NIR"],
                "band_dimensions_match": "Yes",
                "method": method,
                "cells_generated": len(cells),
                "candidate_count": metadata[
                    "candidate_count"
                ],
                "supported_cell_count": metadata[
                    "supported_cell_count"
                ],
                "uncertain_cell_count": metadata[
                    "uncertain_cell_count"
                ],
                "uncertain_cell_ids": ", ".join(
                    str(value)
                    for value in metadata[
                        "uncertain_cell_ids"
                    ]
                ),
                "square_side": round(
                    float(metadata["square_side"]),
                    3,
                ),
                "median_grid_error": round(
                    float(
                        metadata["median_grid_error"]
                    ),
                    4,
                ),
                "max_grid_error": round(
                    float(
                        metadata["max_grid_error"]
                    ),
                    4,
                ),
                "spacing_cv": round(
                    float(metadata["spacing_cv"]),
                    4,
                ),
                "selected_hough_param2": metadata[
                    "selected_hough_param2"
                ],
                "selected_min_distance_ratio": metadata[
                    "selected_min_distance_ratio"
                ],
                "review_required": (
                    "No"
                    if status in {
                        "PASS_AUTO",
                        "PASS_MANUAL",
                    }
                    else "Yes"
                ),
                "overlay_path": relative_path(
                    overlay_path,
                    OUTPUT_ROOT,
                ),
                "status": status,
                "notes": notes,
            }

            coordinates_by_key[key] = [
                {
                    "day_order": day_order,
                    "day": day,
                    "tray": tray,
                    "tray_no": tray_no,
                    "capture_id": capture_id,
                    "cell_id": cell["cell_id"],
                    "row": cell["row"],
                    "column": cell["column"],
                    "x": round(
                        float(cell["x"]),
                        3,
                    ),
                    "y": round(
                        float(cell["y"]),
                        3,
                    ),
                    "square_x0": round(
                        float(cell["square_x0"]),
                        3,
                    ),
                    "square_y0": round(
                        float(cell["square_y0"]),
                        3,
                    ),
                    "square_x1": round(
                        float(cell["square_x1"]),
                        3,
                    ),
                    "square_y1": round(
                        float(cell["square_y1"]),
                        3,
                    ),
                    "square_side": round(
                        float(cell["square_side"]),
                        3,
                    ),
                    "grid_error": round(
                        float(cell["grid_error"]),
                        4,
                    ),
                    "coordinate_source": cell[
                        "coordinate_source"
                    ],
                    "needs_review": cell["needs_review"],
                }
                for cell in cells
            ]

            print(
                f"{status}: {day} > {tray} > {capture_id} | "
                f"supported={metadata['supported_cell_count']}/70 | "
                f"candidates={metadata['candidate_count']} | "
                f"uncertain={metadata['uncertain_cell_count']}"
            )

        except Exception as error:
            manifest_by_key[key] = {
                "day_order": day_order,
                "day": day,
                "tray": tray,
                "tray_no": tray_no,
                "capture_id": capture_id,
                "ms_g_path": band_paths["MS_G"],
                "ms_r_path": band_paths["MS_R"],
                "ms_re_path": band_paths["MS_RE"],
                "ms_nir_path": band_paths["MS_NIR"],
                "band_dimensions_match": "Unknown",
                "method": args.mode,
                "cells_generated": 0,
                "candidate_count": "",
                "supported_cell_count": "",
                "uncertain_cell_count": "",
                "uncertain_cell_ids": "",
                "square_side": "",
                "median_grid_error": "",
                "max_grid_error": "",
                "spacing_cv": "",
                "selected_hough_param2": "",
                "selected_min_distance_ratio": "",
                "review_required": "Yes",
                "overlay_path": "",
                "status": "FAIL",
                "notes": str(error),
            }

            coordinates_by_key[key] = []

            print(
                f"FAIL: {day} > {tray} > {capture_id} | {error}"
            )

    manifest_rows = sorted(
        manifest_by_key.values(),
        key=lambda row: (
            int_or_default(row.get("day_order")),
            natural_key(row.get("tray", "")),
            natural_key(row.get("capture_id", "")),
        ),
    )

    coordinate_rows = [
        row
        for rows in coordinates_by_key.values()
        for row in rows
    ]

    coordinate_rows.sort(
        key=lambda row: (
            int_or_default(row.get("day_order")),
            natural_key(row.get("tray", "")),
            natural_key(row.get("capture_id", "")),
            int_or_default(row.get("cell_id")),
        )
    )

    manifest_frame = pd.DataFrame(manifest_rows)
    coordinate_frame = pd.DataFrame(coordinate_rows)

    reports_root.mkdir(
        parents=True,
        exist_ok=True,
    )

    manifest_frame.to_csv(
        manifest_path,
        index=False,
    )

    coordinate_frame.to_csv(
        coordinates_path,
        index=False,
    )

    save_excel_report(
        workbook_path,
        manifest_rows,
        coordinate_rows,
    )

    settings = {
        "purpose": (
            "Independent MS_NIR 7 x 10 cell grid "
            "detection for aligned MS_G, MS_R, MS_RE and MS_NIR bands."
        ),
        "critical_fix": (
            "Hough circle configurations are evaluated separately. "
            "Circle candidates from different thresholds are never merged."
        ),
        "rows": ROWS,
        "columns": COLS,
        "expected_cells": EXPECTED_CELLS,
        "square_zone_ratio": args.square_zone_ratio,
        "hough_param2_values": [
            68,
            64,
            60,
            56,
            52,
            48,
        ],
        "min_distance_ratio_values": [
            0.50,
            0.60,
            0.70,
        ],
    }

    settings_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    settings_path.write_text(
        json.dumps(
            settings,
            indent=2,
        ),
        encoding="utf-8",
    )

    status_counts = {
        status: sum(
            row.get("status") == status
            for row in manifest_rows
        )
        for status in (
            "PASS_AUTO",
            "PASS_MANUAL",
            "CHECK_AUTO",
            "FAIL",
        )
    }

    print("\n" + "=" * 70)
    print("SCRIPT 06 FINISHED")
    print("=" * 70)

    for status, count in status_counts.items():
        print(f"{status}: {count}")

    print(f"\nExcel report:\n{workbook_path}")
    print(f"\nMS grid overlays:\n{OUTPUT_ROOT}")

    return 0 if status_counts["FAIL"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())