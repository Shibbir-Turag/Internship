from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Any

import numpy as np
import tifffile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageOps


# ============================================================
# 1) EDIT THESE PATHS ONLY WHEN REUSING THE SCRIPT
# ============================================================

PROJECT_ROOT = Path(
    r"C:\Users\tshib\OneDrive\Desktop\Internship"
)

INPUT_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "01_Crop_Dual_Reference"
)

OUTPUT_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "02_Crop_Quality_Check"
)


# ============================================================
# 2) PROJECT SETTINGS
# ============================================================

DAY_NAME_TO_ORDER = {
    "day 1": 1,
    "day 2": 2,
    "day 3": 3,
    "day 4": 4,
    "day 5": 5,
    "day 9": 9,
    "first day": 1,
    "second day": 2,
    "third day": 3,
    "fourth day": 4,
    "fifth day": 5,
    "ninth day": 9,
}

REQUIRED_BANDS = [
    "D",
    "MS_G",
    "MS_R",
    "MS_RE",
    "MS_NIR",
]

MS_BANDS = [
    "MS_G",
    "MS_R",
    "MS_RE",
    "MS_NIR",
]

# A crop with less than this percentage of non-zero pixels
# is not automatically failed, but is marked CHECK.
MIN_COVERAGE_PERCENT = 75.0

PREVIEW_PANEL_WIDTH = 360
PREVIEW_PANEL_HEIGHT = 280


# ============================================================
# 3) SORTING / PATH HELPERS
# ============================================================

def natural_key(text: str):
    return [
        int(part) if part.isdigit() else part.casefold()
        for part in re.split(r"(\d+)", text)
    ]


def day_sort_key(folder: Path):
    return (
        DAY_NAME_TO_ORDER.get(folder.name.casefold(), 999),
        natural_key(folder.name),
    )


def parse_filter_list(value: str | None):
    if not value:
        return None

    return {
        item.strip().casefold()
        for item in value.split(",")
        if item.strip()
    }


def tray_number_from_name(tray_name: str):
    match = re.search(r"(\d+)", tray_name)

    return int(match.group(1)) if match else ""


def safe_filename(text: str):
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", text)


def relative_path(path: Path | None, root: Path):
    if path is None:
        return ""

    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


# ============================================================
# 4) OUTPUT FILE GROUPING
# ============================================================

def parse_output_band(path: Path):
    """
    Examples:
        DJI_20260623124403_0008_D.JPG
        DJI_20260623124403_0008_MS_NIR.TIF
    """

    stem = path.stem.upper()

    for band in ["MS_NIR", "MS_RE", "MS_R", "MS_G", "D"]:
        match = re.match(
            rf"^(?P<capture>.+)_{re.escape(band)}$",
            stem,
        )

        if match:
            return match.group("capture"), band

    return stem, "UNKNOWN"


def find_crop_sets(tray_folder: Path):
    """
    Returns:
    {
        capture_id: {
            "D": [Path(...)],
            "MS_G": [Path(...)],
            ...
        }
    }
    """

    crop_sets: dict[str, dict[str, list[Path]]] = {}

    files = sorted(
        [
            file
            for file in tray_folder.iterdir()
            if file.is_file()
            and file.suffix.casefold() in {
                ".jpg",
                ".jpeg",
                ".tif",
                ".tiff",
            }
        ],
        key=lambda file: natural_key(file.name),
    )

    for file in files:
        capture_id, band = parse_output_band(file)

        if band == "UNKNOWN":
            continue

        crop_sets.setdefault(capture_id, {})
        crop_sets[capture_id].setdefault(band, [])
        crop_sets[capture_id][band].append(file)

    return crop_sets


# ============================================================
# 5) FILE INSPECTION
# ============================================================

def inspect_d_image(path: Path):
    """Read D/RGB crop and calculate basic coverage."""

    try:
        with Image.open(path) as image:
            image = ImageOps.exif_transpose(image)
            image = image.convert("RGB")

            array = np.asarray(image)

        height, width = array.shape[:2]

        brightness = array.max(axis=2)
        coverage = float(np.count_nonzero(brightness > 5) / brightness.size * 100)

        return {
            "readable": True,
            "width": width,
            "height": height,
            "shape": str(array.shape),
            "dtype": str(array.dtype),
            "coverage_percent": coverage,
            "error": "",
        }

    except Exception as error:
        return {
            "readable": False,
            "width": "",
            "height": "",
            "shape": "",
            "dtype": "",
            "coverage_percent": "",
            "error": str(error),
        }


def inspect_ms_band(path: Path):
    """Read multispectral TIFF and calculate dimensions and coverage."""

    try:
        array = tifffile.imread(path)

        original_shape = array.shape

        if array.ndim == 3:
            if array.shape[0] == 1:
                array = array[0]
            elif array.shape[-1] == 1:
                array = array[:, :, 0]
            else:
                raise ValueError(
                    f"Expected single-band TIFF, found shape {original_shape}."
                )

        if array.ndim != 2:
            raise ValueError(
                f"Expected 2D TIFF, found shape {original_shape}."
            )

        height, width = array.shape

        coverage = float(
            np.count_nonzero(array > 0)
            / array.size
            * 100
        )

        return {
            "readable": True,
            "width": width,
            "height": height,
            "shape": str(original_shape),
            "dtype": str(array.dtype),
            "coverage_percent": coverage,
            "error": "",
        }

    except Exception as error:
        return {
            "readable": False,
            "width": "",
            "height": "",
            "shape": "",
            "dtype": "",
            "coverage_percent": "",
            "error": str(error),
        }


# ============================================================
# 6) PREVIEW GENERATION
# ============================================================

def normalise_ms_for_preview(array: np.ndarray):
    """
    Display-only conversion of raw band values to grayscale.
    It does not alter the source TIFF.
    """

    values = array.astype(np.float32)

    low, high = np.percentile(values, [1, 99])

    if high <= low:
        high = low + 1

    scaled = np.clip(
        (values - low) * 255 / (high - low),
        0,
        255,
    ).astype(np.uint8)

    return Image.fromarray(scaled).convert("RGB")


def load_preview(path: Path | None, band: str):
    """Load D image or a normalised MS preview."""

    if path is None or not path.exists():
        return None

    try:
        if band == "D":
            with Image.open(path) as image:
                image = ImageOps.exif_transpose(image)
                return image.convert("RGB")

        array = tifffile.imread(path)

        if array.ndim == 3:
            if array.shape[0] == 1:
                array = array[0]
            elif array.shape[-1] == 1:
                array = array[:, :, 0]

        if array.ndim != 2:
            return None

        return normalise_ms_for_preview(array)

    except Exception:
        return None


def make_panel(
    image: Image.Image | None,
    label: str,
):
    """Create one labelled visual QA panel."""

    panel = Image.new(
        "RGB",
        (PREVIEW_PANEL_WIDTH, PREVIEW_PANEL_HEIGHT),
        "white",
    )

    draw = ImageDraw.Draw(panel)

    draw.text(
        (12, 10),
        label,
        fill="black",
    )

    if image is None:
        draw.text(
            (12, 45),
            "Missing or unreadable",
            fill="red",
        )
        return panel

    image_copy = image.copy()

    image_copy.thumbnail(
        (
            PREVIEW_PANEL_WIDTH - 16,
            PREVIEW_PANEL_HEIGHT - 55,
        )
    )

    x = (PREVIEW_PANEL_WIDTH - image_copy.width) // 2
    y = 40 + (
        PREVIEW_PANEL_HEIGHT
        - 40
        - image_copy.height
    ) // 2

    panel.paste(image_copy, (x, y))

    return panel


def create_qa_preview(
    output_path: Path,
    day_name: str,
    tray_name: str,
    capture_id: str,
    selected_files: dict[str, Path],
):
    """
    Create one five-panel QA preview:
    D | MS_G | MS_R | MS_RE | MS_NIR
    """

    ordered_bands = [
        "D",
        "MS_G",
        "MS_R",
        "MS_RE",
        "MS_NIR",
    ]

    header_height = 42

    canvas = Image.new(
        "RGB",
        (
            PREVIEW_PANEL_WIDTH * len(ordered_bands),
            PREVIEW_PANEL_HEIGHT + header_height,
        ),
        "#EAEAEA",
    )

    draw = ImageDraw.Draw(canvas)

    draw.text(
        (12, 12),
        f"{day_name} | {tray_name} | {capture_id}",
        fill="black",
    )

    for index, band in enumerate(ordered_bands):
        preview = load_preview(
            selected_files.get(band),
            band,
        )

        panel = make_panel(
            preview,
            band,
        )

        canvas.paste(
            panel,
            (index * PREVIEW_PANEL_WIDTH, header_height),
        )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    canvas.save(output_path)


# ============================================================
# 7) REPORT CREATION
# ============================================================

def style_worksheet(
    worksheet,
    headers: list[str],
    rows: list[list[Any]],
):
    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    thin = Side(
        style="thin",
        color="D9E2F3",
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(
            row=1,
            column=column,
            value=header,
        )

        cell.fill = header_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = border

    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            cell = worksheet.cell(
                row=row_index,
                column=column,
                value=value,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

            cell.border = border

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 34

    for column in worksheet.columns:
        letter = column[0].column_letter

        longest = max(
            len(str(cell.value))
            if cell.value is not None
            else 0
            for cell in column
        )

        worksheet.column_dimensions[letter].width = min(
            max(12, longest + 2),
            50,
        )


def create_excel_report(
    output_path: Path,
    summary_rows: list[dict],
    file_rows: list[dict],
):
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Crop QA Summary"

    summary_headers = [
        "Day Order",
        "Day",
        "Tray",
        "Tray No",
        "Capture ID",
        "Required Bands Present",
        "Duplicate Bands",
        "Unreadable Bands",
        "MS Dimensions Match",
        "MS Dimensions",
        "Lowest Coverage %",
        "Preview Path",
        "Status",
        "Notes",
    ]

    summary_values = [
        [
            row["day_order"],
            row["day"],
            row["tray"],
            row["tray_no"],
            row["capture_id"],
            row["required_bands_present"],
            row["duplicate_bands"],
            row["unreadable_bands"],
            row["ms_dimensions_match"],
            row["ms_dimensions"],
            row["lowest_coverage_percent"],
            row["preview_path"],
            row["status"],
            row["notes"],
        ]
        for row in summary_rows
    ]

    style_worksheet(
        summary_sheet,
        summary_headers,
        summary_values,
    )

    for row_number in range(2, len(summary_values) + 2):
        status_cell = summary_sheet.cell(
            row=row_number,
            column=13,
        )

        status = status_cell.value

        if status == "PASS":
            status_cell.fill = PatternFill(
                "solid",
                fgColor="C6EFCE",
            )
        elif status == "CHECK":
            status_cell.fill = PatternFill(
                "solid",
                fgColor="FFF2CC",
            )
        elif status == "FAIL":
            status_cell.fill = PatternFill(
                "solid",
                fgColor="F4CCCC",
            )

    file_sheet = workbook.create_sheet("File Details")

    file_headers = [
        "Day Order",
        "Day",
        "Tray",
        "Tray No",
        "Capture ID",
        "Band",
        "File",
        "Relative Path",
        "Width",
        "Height",
        "Shape",
        "Data Type / Mode",
        "Coverage %",
        "Readable",
        "Error",
    ]

    file_values = [
        [
            row["day_order"],
            row["day"],
            row["tray"],
            row["tray_no"],
            row["capture_id"],
            row["band"],
            row["file_name"],
            row["relative_path"],
            row["width"],
            row["height"],
            row["shape"],
            row["dtype"],
            row["coverage_percent"],
            row["readable"],
            row["error"],
        ]
        for row in file_rows
    ]

    style_worksheet(
        file_sheet,
        file_headers,
        file_values,
    )

    readme_sheet = workbook.create_sheet("Read Me")

    readme_sheet["A1"] = "Script 02 — Crop Quality Check"
    readme_sheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    notes = [
        "This script does not alter any crop files. It only inspects Script 01 outputs.",
        "D/RGB dimensions are intentionally allowed to differ from multispectral dimensions.",
        "MS_G, MS_R, MS_RE and MS_NIR must have identical dimensions before NDVI/NDRE analysis.",
        "PASS means all required bands were found, readable, and MS dimensions matched.",
        "CHECK means visual review is recommended, such as lower image coverage or duplicate files.",
        "FAIL means missing required bands, unreadable files or mismatched multispectral dimensions.",
        "Open several QA previews from each day before continuing to cell detection.",
        "The preview panels are only for human quality checking and are not analytical inputs.",
    ]

    for row_number, note in enumerate(notes, start=3):
        readme_sheet.cell(
            row=row_number,
            column=1,
            value=note,
        )

    readme_sheet.column_dimensions["A"].width = 120

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(output_path)


def write_csv(
    output_path: Path,
    summary_rows: list[dict],
):
    fields = [
        "day_order",
        "day",
        "tray",
        "tray_no",
        "capture_id",
        "required_bands_present",
        "duplicate_bands",
        "unreadable_bands",
        "ms_dimensions_match",
        "ms_dimensions",
        "lowest_coverage_percent",
        "preview_path",
        "status",
        "notes",
    ]

    with output_path.open(
        "w",
        newline="",
        encoding="utf-8",
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=fields,
        )

        writer.writeheader()
        writer.writerows(summary_rows)


# ============================================================
# 8) MAIN WORKFLOW
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Script 02: Validate crop outputs from "
            "01_Crop_Dual_Reference."
        )
    )

    parser.add_argument(
        "--days",
        help='Example: --days "Day 1,Day 9"',
    )

    parser.add_argument(
        "--trays",
        help='Example: --trays "Tray 1,Tray 2"',
    )

    parser.add_argument(
        "--no-previews",
        action="store_true",
        help="Do not create visual QA preview PNGs.",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="List crop sets that would be checked.",
    )

    args = parser.parse_args()

    if not INPUT_ROOT.exists():
        print(
            "ERROR: Script 01 output folder was not found:\n"
            f"{INPUT_ROOT}"
        )
        return 1

    selected_days = parse_filter_list(args.days)
    selected_trays = parse_filter_list(args.trays)

    day_folders = sorted(
        [
            folder
            for folder in INPUT_ROOT.iterdir()
            if folder.is_dir()
            and folder.name.casefold() in DAY_NAME_TO_ORDER
        ],
        key=day_sort_key,
    )

    jobs = []

    for day_folder in day_folders:
        if (
            selected_days
            and day_folder.name.casefold() not in selected_days
        ):
            continue

        tray_folders = sorted(
            [
                folder
                for folder in day_folder.iterdir()
                if folder.is_dir()
            ],
            key=lambda folder: natural_key(folder.name),
        )

        for tray_folder in tray_folders:
            if (
                selected_trays
                and tray_folder.name.casefold() not in selected_trays
            ):
                continue

            crop_sets = find_crop_sets(tray_folder)

            for capture_id, band_files in crop_sets.items():
                jobs.append(
                    {
                        "day_folder": day_folder,
                        "tray_folder": tray_folder,
                        "capture_id": capture_id,
                        "band_files": band_files,
                    }
                )

    if not jobs:
        print("No Script 01 crop outputs were found.")
        return 1

    print("\nSCRIPT 02 — CROP QUALITY CHECK")
    print("=" * 70)
    print(f"Crop input:\n{INPUT_ROOT}")
    print(f"\nQA output:\n{OUTPUT_ROOT}\n")

    for job in jobs:
        print(
            f"READY: {job['day_folder'].name} > "
            f"{job['tray_folder'].name} > "
            f"{job['capture_id']}"
        )

    if args.dry_run:
        print("\nDry run complete. No reports or previews created.")
        return 0

    summary_rows = []
    file_rows = []

    for job in jobs:
        day_name = job["day_folder"].name
        tray_name = job["tray_folder"].name
        tray_no = tray_number_from_name(tray_name)

        capture_id = job["capture_id"]
        band_files = job["band_files"]

        day_order = DAY_NAME_TO_ORDER.get(
            day_name.casefold(),
            999,
        )

        selected_files = {}
        duplicate_bands = []

        for band, paths in band_files.items():
            if len(paths) > 1:
                duplicate_bands.append(band)

            selected_files[band] = paths[0]

        missing_bands = [
            band
            for band in REQUIRED_BANDS
            if band not in selected_files
        ]

        inspection = {}

        for band, path in selected_files.items():
            if band == "D":
                info = inspect_d_image(path)
            else:
                info = inspect_ms_band(path)

            inspection[band] = info

            file_rows.append(
                {
                    "day_order": day_order,
                    "day": day_name,
                    "tray": tray_name,
                    "tray_no": tray_no,
                    "capture_id": capture_id,
                    "band": band,
                    "file_name": path.name,
                    "relative_path": relative_path(path, INPUT_ROOT),
                    "width": info["width"],
                    "height": info["height"],
                    "shape": info["shape"],
                    "dtype": info["dtype"],
                    "coverage_percent": info["coverage_percent"],
                    "readable": "Yes" if info["readable"] else "No",
                    "error": info["error"],
                }
            )

        unreadable_bands = [
            band
            for band, info in inspection.items()
            if not info["readable"]
        ]

        ms_sizes = []

        for band in MS_BANDS:
            if (
                band in inspection
                and inspection[band]["readable"]
            ):
                ms_sizes.append(
                    (
                        inspection[band]["width"],
                        inspection[band]["height"],
                    )
                )

        unique_ms_sizes = sorted(set(ms_sizes))

        ms_dimensions_match = (
            len(ms_sizes) == 4
            and len(unique_ms_sizes) == 1
        )

        coverage_values = [
            float(info["coverage_percent"])
            for info in inspection.values()
            if info["readable"]
            and info["coverage_percent"] != ""
        ]

        lowest_coverage = (
            round(min(coverage_values), 3)
            if coverage_values
            else ""
        )

        notes = []

        if missing_bands:
            notes.append(
                "Missing: " + ", ".join(missing_bands)
            )

        if duplicate_bands:
            notes.append(
                "Duplicates: " + ", ".join(duplicate_bands)
            )

        if unreadable_bands:
            notes.append(
                "Unreadable: " + ", ".join(unreadable_bands)
            )

        if not ms_dimensions_match:
            notes.append(
                "MS dimensions missing or inconsistent."
            )

        if (
            lowest_coverage != ""
            and lowest_coverage < MIN_COVERAGE_PERCENT
        ):
            notes.append(
                f"Low crop coverage: {lowest_coverage}%."
            )

        if (
            missing_bands
            or unreadable_bands
            or not ms_dimensions_match
        ):
            status = "FAIL"

        elif (
            duplicate_bands
            or (
                lowest_coverage != ""
                and lowest_coverage < MIN_COVERAGE_PERCENT
            )
        ):
            status = "CHECK"

        else:
            status = "PASS"

        preview_path = ""

        if not args.no_previews:
            safe_capture_id = safe_filename(capture_id)

            preview_file = (
                OUTPUT_ROOT
                / day_name
                / tray_name
                / f"{safe_capture_id}_crop_qa_preview.png"
            )

            create_qa_preview(
                preview_file,
                day_name,
                tray_name,
                capture_id,
                selected_files,
            )

            preview_path = relative_path(
                preview_file,
                OUTPUT_ROOT,
            )

        summary_rows.append(
            {
                "day_order": day_order,
                "day": day_name,
                "tray": tray_name,
                "tray_no": tray_no,
                "capture_id": capture_id,
                "required_bands_present": (
                    "Yes"
                    if not missing_bands
                    else "No"
                ),
                "duplicate_bands": (
                    ", ".join(duplicate_bands)
                    if duplicate_bands
                    else ""
                ),
                "unreadable_bands": (
                    ", ".join(unreadable_bands)
                    if unreadable_bands
                    else ""
                ),
                "ms_dimensions_match": (
                    "Yes"
                    if ms_dimensions_match
                    else "No"
                ),
                "ms_dimensions": (
                    str(unique_ms_sizes)
                    if unique_ms_sizes
                    else ""
                ),
                "lowest_coverage_percent": lowest_coverage,
                "preview_path": preview_path,
                "status": status,
                "notes": (
                    "All automatic checks passed."
                    if not notes
                    else " | ".join(notes)
                ),
            }
        )

        print(
            f"{status}: {day_name} > {tray_name} > "
            f"{capture_id}"
        )

    summary_rows.sort(
        key=lambda row: (
            row["day_order"],
            natural_key(row["tray"]),
            natural_key(row["capture_id"]),
        )
    )

    file_rows.sort(
        key=lambda row: (
            row["day_order"],
            natural_key(row["tray"]),
            natural_key(row["capture_id"]),
            REQUIRED_BANDS.index(row["band"])
            if row["band"] in REQUIRED_BANDS
            else 99,
        )
    )

    report_path = (
        OUTPUT_ROOT
        / "_reports"
        / "crop_quality_report.xlsx"
    )

    csv_path = (
        OUTPUT_ROOT
        / "_reports"
        / "crop_quality_manifest.csv"
    )

    create_excel_report(
        report_path,
        summary_rows,
        file_rows,
    )

    write_csv(
        csv_path,
        summary_rows,
    )

    pass_count = sum(
        row["status"] == "PASS"
        for row in summary_rows
    )

    check_count = sum(
        row["status"] == "CHECK"
        for row in summary_rows
    )

    fail_count = sum(
        row["status"] == "FAIL"
        for row in summary_rows
    )

    print("\n" + "=" * 70)
    print("SCRIPT 02 FINISHED")
    print("=" * 70)
    print(f"PASS: {pass_count}")
    print(f"CHECK: {check_count}")
    print(f"FAIL: {fail_count}")
    print(f"\nExcel report:\n{report_path}")
    print(f"\nCSV report:\n{csv_path}")
    print(f"\nQA previews:\n{OUTPUT_ROOT}")

    return 0 if fail_count == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())