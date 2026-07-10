from __future__ import annotations

"""
SCRIPT 07 — MULTISPECTRAL VEGETATION INDICES

Purpose
-------
Calculate relative NDVI and NDRE from cropped multispectral bands using the
independent 7 x 10 square grid produced by Script 06.

Inputs
------
- Script 01 cropped MS_G, MS_R, MS_RE, MS_NIR bands
- Script 06 MS grid manifest and cell coordinates

Outputs
-------
outputs/Second Trial/07_MS_Vegetation_Indices/
    Day X/Tray Y/
        <capture>_relative_indices.npz
        <capture>_ndvi_preview.png
        <capture>_ndre_preview.png
        <capture>_ndvi_square_overlay.png
        <capture>_ndre_square_overlay.png

    _reports/
        ms_index_tray_summary.csv
        ms_index_cell_results.csv
        ms_vegetation_index_report.xlsx

    _config/
        index_settings.json

Important interpretation rule
-----------------------------
If the original MS bands are not calibrated reflectance products, the results
must be described as relative image-derived NDVI and NDRE values. They are
appropriate for comparing trays captured with the same workflow.
"""

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageOps


# ============================================================
# 1) PATHS — CHANGE PROJECT_ROOT ONLY WHEN REUSING
# ============================================================

PROJECT_ROOT = Path(
    r"C:\Users\tshib\OneDrive\Desktop\Internship"
)

CROP_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "01_Crop_Dual_Reference"
)

GRID_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "06_MS_Cell_Grid_Detection"
)

OUTPUT_ROOT = (
    PROJECT_ROOT
    / "outputs"
    / "Second Trial"
    / "07_MS_Vegetation_Indices"
)

GRID_MANIFEST_CSV = (
    GRID_ROOT
    / "_reports"
    / "ms_grid_manifest.csv"
)

GRID_COORDINATES_CSV = (
    GRID_ROOT
    / "_reports"
    / "ms_cell_coordinates.csv"
)


# ============================================================
# 2) SETTINGS
# ============================================================

EXPECTED_CELLS = 70

ACCEPTED_GRID_STATUSES = {
    "PASS_AUTO",
    "PASS_MANUAL",
}

MS_BANDS = (
    "MS_G",
    "MS_R",
    "MS_RE",
    "MS_NIR",
)

DAY_NAME_TO_ORDER = {
    "day 1": 1,
    "day 2": 2,
    "day 3": 3,
    "day 4": 4,
    "day 5": 5,
    "day 9": 9,
    "first day": 1,
    "second day": 2,
    "third day": 3,
    "fourth day": 4,
    "fifth day": 5,
    "ninth day": 9,
}

EPSILON = 1e-6
MIN_VALID_PIXELS_PER_CELL = 20


# ============================================================
# 3) GENERAL HELPERS
# ============================================================

def natural_key(value: object):
    return [
        int(part) if part.isdigit() else part.casefold()
        for part in re.split(r"(\d+)", str(value))
    ]


def parse_filter_list(value: str | None):
    if not value:
        return None

    return {
        item.strip().casefold()
        for item in value.split(",")
        if item.strip()
    }


def safe_name(value: object):
    return re.sub(
        r"[^A-Za-z0-9_.-]+",
        "_",
        str(value),
    )


def record_key(
    day: str,
    tray: str,
    capture_id: str,
):
    return f"{day}|{tray}|{capture_id}"


def tray_number(tray_name: str):
    match = re.search(r"(\d+)", tray_name)
    return int(match.group(1)) if match else ""


def int_or_default(value: object, default=999):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def relative_path(path: Path | None, root: Path):
    if path is None:
        return ""

    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


def resolve_band_path(value: object):
    relative_value = str(value).strip()

    if not relative_value:
        return None

    path = Path(
        relative_value.replace("\\", "/")
    )

    if path.is_absolute():
        return path

    return CROP_ROOT / path


# ============================================================
# 4) LOAD INPUT REPORTS
# ============================================================

def read_csv_required(
    path: Path,
    required_columns: list[str],
):
    if not path.exists():
        raise FileNotFoundError(
            f"Required input file not found:\n{path}"
        )

    dataframe = pd.read_csv(path)

    missing = [
        column
        for column in required_columns
        if column not in dataframe.columns
    ]

    if missing:
        raise ValueError(
            f"{path.name} is missing required column(s): "
            + ", ".join(missing)
        )

    return dataframe


def load_grid_inputs():
    manifest = read_csv_required(
        GRID_MANIFEST_CSV,
        [
            "day_order",
            "day",
            "tray",
            "tray_no",
            "capture_id",
            "ms_g_path",
            "ms_r_path",
            "ms_re_path",
            "ms_nir_path",
            "status",
        ],
    )

    coordinates = read_csv_required(
        GRID_COORDINATES_CSV,
        [
            "day_order",
            "day",
            "tray",
            "tray_no",
            "capture_id",
            "cell_id",
            "row",
            "column",
            "x",
            "y",
            "square_x0",
            "square_y0",
            "square_x1",
            "square_y1",
            "square_side",
        ],
    )

    manifest["day_order"] = pd.to_numeric(
        manifest["day_order"],
        errors="coerce",
    )

    manifest["tray_no"] = pd.to_numeric(
        manifest["tray_no"],
        errors="coerce",
    )

    manifest = manifest.dropna(
        subset=["day_order", "tray_no"]
    ).copy()

    manifest["day_order"] = (
        manifest["day_order"].astype(int)
    )

    manifest["tray_no"] = (
        manifest["tray_no"].astype(int)
    )

    manifest = manifest.loc[
        manifest["status"]
        .astype(str)
        .str.upper()
        .isin(ACCEPTED_GRID_STATUSES)
    ].copy()

    coordinates["day_order"] = pd.to_numeric(
        coordinates["day_order"],
        errors="coerce",
    )

    coordinates["tray_no"] = pd.to_numeric(
        coordinates["tray_no"],
        errors="coerce",
    )

    coordinates["cell_id"] = pd.to_numeric(
        coordinates["cell_id"],
        errors="coerce",
    )

    coordinates = coordinates.dropna(
        subset=[
            "day_order",
            "tray_no",
            "cell_id",
        ]
    ).copy()

    coordinates["day_order"] = (
        coordinates["day_order"].astype(int)
    )

    coordinates["tray_no"] = (
        coordinates["tray_no"].astype(int)
    )

    coordinates["cell_id"] = (
        coordinates["cell_id"].astype(int)
    )

    for column in [
        "x",
        "y",
        "square_x0",
        "square_y0",
        "square_x1",
        "square_y1",
        "square_side",
    ]:
        coordinates[column] = pd.to_numeric(
            coordinates[column],
            errors="coerce",
        )

    coordinates = coordinates.dropna(
        subset=[
            "x",
            "y",
            "square_x0",
            "square_y0",
            "square_x1",
            "square_y1",
            "square_side",
        ]
    ).copy()

    manifest["grid_key"] = manifest.apply(
        lambda row: record_key(
            str(row["day"]),
            str(row["tray"]),
            str(row["capture_id"]),
        ),
        axis=1,
    )

    coordinates["grid_key"] = coordinates.apply(
        lambda row: record_key(
            str(row["day"]),
            str(row["tray"]),
            str(row["capture_id"]),
        ),
        axis=1,
    )

    return manifest, coordinates


# ============================================================
# 5) IMAGE READING AND INDEX CALCULATION
# ============================================================

def read_raw_band(path: Path):
    """
    Reads a MS band without changing its original pixel values.

    If a band unexpectedly has three channels, the channels are averaged into
    one intensity plane. Standard DJI MS bands are normally single-channel.
    """

    if not path.exists():
        raise FileNotFoundError(
            f"MS band image not found:\n{path}"
        )

    with Image.open(path) as image:
        image = ImageOps.exif_transpose(image)
        array = np.asarray(image)

    if array.ndim == 2:
        return array.astype(np.float32)

    if array.ndim == 3:
        return array[:, :, :3].astype(np.float32).mean(axis=2)

    raise ValueError(
        f"Unsupported image dimensions for {path.name}: {array.shape}"
    )


def calculate_indices(
    red: np.ndarray,
    red_edge: np.ndarray,
    nir: np.ndarray,
):
    """
    NDVI  = (NIR - Red) / (NIR + Red)
    NDRE  = (NIR - Red Edge) / (NIR + Red Edge)
    """

    if (
        red.shape != nir.shape
        or red_edge.shape != nir.shape
    ):
        raise ValueError(
            "MS_R, MS_RE and MS_NIR dimensions do not match: "
            f"MS_R={red.shape}, MS_RE={red_edge.shape}, "
            f"MS_NIR={nir.shape}"
        )

    ndvi_denominator = nir + red
    ndre_denominator = nir + red_edge

    ndvi_valid = (
        np.isfinite(nir)
        & np.isfinite(red)
        & (ndvi_denominator > EPSILON)
    )

    ndre_valid = (
        np.isfinite(nir)
        & np.isfinite(red_edge)
        & (ndre_denominator > EPSILON)
    )

    ndvi = np.full(
        nir.shape,
        np.nan,
        dtype=np.float32,
    )

    ndre = np.full(
        nir.shape,
        np.nan,
        dtype=np.float32,
    )

    ndvi[ndvi_valid] = (
        nir[ndvi_valid]
        - red[ndvi_valid]
    ) / ndvi_denominator[ndvi_valid]

    ndre[ndre_valid] = (
        nir[ndre_valid]
        - red_edge[ndre_valid]
    ) / ndre_denominator[ndre_valid]

    return ndvi, ndre, ndvi_valid, ndre_valid


# ============================================================
# 6) CELL-LEVEL INDEX SUMMARY
# ============================================================

def clipped_square_bounds(
    cell: dict,
    width: int,
    height: int,
):
    x0 = max(
        0,
        int(round(cell["square_x0"])),
    )

    y0 = max(
        0,
        int(round(cell["square_y0"])),
    )

    x1 = min(
        width,
        int(round(cell["square_x1"])),
    )

    y1 = min(
        height,
        int(round(cell["square_y1"])),
    )

    return x0, y0, x1, y1


def index_statistics(values: np.ndarray):
    values = values[np.isfinite(values)]

    if len(values) < MIN_VALID_PIXELS_PER_CELL:
        return {
            "valid_pixels": int(len(values)),
            "mean": np.nan,
            "median": np.nan,
            "std": np.nan,
            "p10": np.nan,
            "p90": np.nan,
        }

    return {
        "valid_pixels": int(len(values)),
        "mean": float(np.mean(values)),
        "median": float(np.median(values)),
        "std": float(np.std(values)),
        "p10": float(np.percentile(values, 10)),
        "p90": float(np.percentile(values, 90)),
    }


def evaluate_cells(
    ndvi: np.ndarray,
    ndre: np.ndarray,
    cell_rows: pd.DataFrame,
):
    height, width = ndvi.shape[:2]

    results = []

    for row in cell_rows.sort_values("cell_id").itertuples():
        cell = {
            "cell_id": int(row.cell_id),
            "row": int(row.row),
            "column": int(row.column),
            "x": float(row.x),
            "y": float(row.y),
            "square_x0": float(row.square_x0),
            "square_y0": float(row.square_y0),
            "square_x1": float(row.square_x1),
            "square_y1": float(row.square_y1),
            "square_side": float(row.square_side),
            "coordinate_source": getattr(
                row,
                "coordinate_source",
                "",
            ),
            "needs_review": getattr(
                row,
                "needs_review",
                "",
            ),
        }

        x0, y0, x1, y1 = clipped_square_bounds(
            cell,
            width,
            height,
        )

        if x1 <= x0 or y1 <= y0:
            ndvi_stats = index_statistics(
                np.asarray([], dtype=np.float32)
            )

            ndre_stats = index_statistics(
                np.asarray([], dtype=np.float32)
            )
        else:
            ndvi_stats = index_statistics(
                ndvi[y0:y1, x0:x1]
            )

            ndre_stats = index_statistics(
                ndre[y0:y1, x0:x1]
            )

        results.append(
            {
                **cell,
                "ndvi_valid_pixels": ndvi_stats["valid_pixels"],
                "ndvi_mean": ndvi_stats["mean"],
                "ndvi_median": ndvi_stats["median"],
                "ndvi_std": ndvi_stats["std"],
                "ndvi_p10": ndvi_stats["p10"],
                "ndvi_p90": ndvi_stats["p90"],
                "ndre_valid_pixels": ndre_stats["valid_pixels"],
                "ndre_mean": ndre_stats["mean"],
                "ndre_median": ndre_stats["median"],
                "ndre_std": ndre_stats["std"],
                "ndre_p10": ndre_stats["p10"],
                "ndre_p90": ndre_stats["p90"],
            }
        )

    return results


# ============================================================
# 7) VISUAL OUTPUTS
# ============================================================

def index_colour_image(
    index_array: np.ndarray,
):
    """
    Use a fixed -1 to +1 scale so all tray previews are visually comparable.
    Invalid pixels are shown as black.
    """

    valid = np.isfinite(index_array)

    scaled = np.zeros(
        index_array.shape,
        dtype=np.uint8,
    )

    clipped = np.clip(
        index_array,
        -1.0,
        1.0,
    )

    scaled[valid] = np.round(
        (clipped[valid] + 1.0)
        * 127.5
    ).astype(np.uint8)

    colour = cv2.applyColorMap(
        scaled,
        cv2.COLORMAP_TURBO,
    )

    colour[~valid] = (0, 0, 0)

    return colour


def add_header(
    image: np.ndarray,
    text: str,
    font_scale: float = 0.65,
):
    height, width = image.shape[:2]

    header_height = 42

    output = np.zeros(
        (height + header_height, width, 3),
        dtype=np.uint8,
    )

    output[:header_height] = (255, 255, 255)
    output[header_height:] = image

    cv2.putText(
        output,
        text,
        (12, 28),
        cv2.FONT_HERSHEY_SIMPLEX,
        font_scale,
        (0, 0, 0),
        2,
        cv2.LINE_AA,
    )

    return output


def save_index_preview(
    index_array: np.ndarray,
    output_path: Path,
    title: str,
):
    image = index_colour_image(index_array)
    image = add_header(image, title)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    cv2.imwrite(
        str(output_path),
        image,
        [
            cv2.IMWRITE_PNG_COMPRESSION,
            3,
        ],
    )


def save_square_overlay(
    index_array: np.ndarray,
    cells: list[dict],
    output_path: Path,
    title: str,
):
    image = index_colour_image(index_array)

    height, width = image.shape[:2]

    median_side = float(
        np.median(
            [
                cell["square_side"]
                for cell in cells
            ]
        )
    )

    line_width = max(
        2,
        int(round(median_side / 32)),
    )

    font_scale = max(
        0.35,
        min(1.0, median_side / 88),
    )

    font_thickness = max(
        1,
        int(round(font_scale * 2)),
    )

    for cell in cells:
        x0 = max(
            0,
            int(round(cell["square_x0"])),
        )

        y0 = max(
            0,
            int(round(cell["square_y0"])),
        )

        x1 = min(
            width - 1,
            int(round(cell["square_x1"])),
        )

        y1 = min(
            height - 1,
            int(round(cell["square_y1"])),
        )

        if str(cell["needs_review"]).casefold() == "yes":
            colour = (0, 180, 255)  # Yellow
        else:
            colour = (0, 180, 0)  # Green

        cv2.rectangle(
            image,
            (x0, y0),
            (x1, y1),
            colour,
            line_width,
        )

        label_x = max(
            1,
            int(
                round(
                    cell["x"]
                    - median_side * 0.10
                )
            ),
        )

        label_y = min(
            height - 5,
            int(
                round(
                    cell["y"]
                    + median_side * 0.08
                )
            ),
        )

        cv2.putText(
            image,
            str(cell["cell_id"]),
            (label_x, label_y),
            cv2.FONT_HERSHEY_SIMPLEX,
            font_scale,
            colour,
            font_thickness,
            cv2.LINE_AA,
        )

    image = add_header(image, title)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    cv2.imwrite(
        str(output_path),
        image,
        [
            cv2.IMWRITE_PNG_COMPRESSION,
            3,
        ],
    )


# ============================================================
# 8) REPORT OUTPUTS
# ============================================================

def load_existing_csv(path: Path):
    if not path.exists():
        return []

    try:
        return pd.read_csv(path).to_dict(
            orient="records"
        )
    except Exception:
        return []


def style_excel_report(path: Path):
    workbook = load_workbook(path)

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 34

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(
                color="FFFFFF",
                bold=True,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in worksheet.columns:
            letter = column_cells[0].column_letter

            longest = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )

            worksheet.column_dimensions[letter].width = min(
                max(12, longest + 2),
                58,
            )

    workbook.save(path)


def create_excel_report(
    output_path: Path,
    tray_rows: list[dict],
    cell_rows: list[dict],
):
    tray_frame = pd.DataFrame(tray_rows)
    cell_frame = pd.DataFrame(cell_rows)

    readme_frame = pd.DataFrame(
        {
            "Notes": [
                "NDVI = (NIR - Red) / (NIR + Red).",
                "NDRE = (NIR - Red Edge) / (NIR + Red Edge).",
                "These are relative image-derived indices unless source bands are calibrated reflectance data.",
                "The square cell zones were independently detected in MS_NIR by Script 06.",
                "D/RGB grid coordinates are not used in this multispectral workflow.",
                "Day 1 to Day 5 will later be used for continuous trend comparisons. Day 9 is a later follow-up observation.",
                "Script 08 will compare treatment and environment groups using these tray- and cell-level results.",
            ]
        }
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        tray_frame.to_excel(
            writer,
            sheet_name="Tray Summary",
            index=False,
        )

        cell_frame.to_excel(
            writer,
            sheet_name="Cell Results",
            index=False,
        )

        readme_frame.to_excel(
            writer,
            sheet_name="Read Me",
            index=False,
        )

    style_excel_report(output_path)


# ============================================================
# 9) MAIN WORKFLOW
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Script 07: calculate relative NDVI and NDRE "
            "from MS bands using Script 06 square zones."
        )
    )

    parser.add_argument(
        "--days",
        help='Example: --days "Day 1,Day 9"',
    )

    parser.add_argument(
        "--trays",
        help='Example: --trays "Tray 1,Tray 7"',
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help=(
            "Overwrite output files for selected "
            "day/tray/capture combinations."
        ),
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate eligible Script 06 grid records only.",
    )

    args = parser.parse_args()

    if not CROP_ROOT.exists():
        print(
            "ERROR: Script 01 crop folder was not found:\n"
            f"{CROP_ROOT}"
        )
        return 1

    if not GRID_ROOT.exists():
        print(
            "ERROR: Script 06 output folder was not found:\n"
            f"{GRID_ROOT}"
        )
        return 1

    manifest, coordinates = load_grid_inputs()

    if manifest.empty:
        print(
            "ERROR: No PASS_AUTO or PASS_MANUAL Script 06 grid records were found."
        )
        return 1

    selected_days = parse_filter_list(args.days)
    selected_trays = parse_filter_list(args.trays)

    if selected_days:
        manifest = manifest.loc[
            manifest["day"]
            .astype(str)
            .str.casefold()
            .isin(selected_days)
        ].copy()

    if selected_trays:
        manifest = manifest.loc[
            manifest["tray"]
            .astype(str)
            .str.casefold()
            .isin(selected_trays)
        ].copy()

    if manifest.empty:
        print(
            "No eligible Script 06 records matched the selected filters."
        )
        return 1

    coordinates_by_key = {
        key: group.copy()
        for key, group in coordinates.groupby(
            "grid_key"
        )
    }

    print("\nSCRIPT 07 — MS VEGETATION INDICES")
    print("=" * 70)
    print(f"MS crop input:\n{CROP_ROOT}")
    print(f"\nScript 06 grid input:\n{GRID_ROOT}")
    print(f"\nOutput folder:\n{OUTPUT_ROOT}")

    jobs = []

    for row in manifest.sort_values(
        ["day_order", "tray_no", "capture_id"]
    ).itertuples():
        key = str(row.grid_key)

        cell_count = len(
            coordinates_by_key.get(
                key,
                pd.DataFrame(),
            )
        )

        state = (
            "READY"
            if cell_count == EXPECTED_CELLS
            else f"INVALID GRID: {cell_count}/70 cells"
        )

        print(
            f"{state}: "
            f"{row.day} > {row.tray} > {row.capture_id}"
        )

        jobs.append(row)

    if args.dry_run:
        print(
            "\nDry run complete. No index files or reports created."
        )
        return 0

    reports_root = OUTPUT_ROOT / "_reports"

    tray_summary_path = (
        reports_root
        / "ms_index_tray_summary.csv"
    )

    cell_results_path = (
        reports_root
        / "ms_index_cell_results.csv"
    )

    excel_report_path = (
        reports_root
        / "ms_vegetation_index_report.xlsx"
    )

    settings_path = (
        OUTPUT_ROOT
        / "_config"
        / "index_settings.json"
    )

    existing_tray_rows = load_existing_csv(
        tray_summary_path
    )

    existing_cell_rows = load_existing_csv(
        cell_results_path
    )

    tray_by_key = {
        record_key(
            str(row.get("day", "")),
            str(row.get("tray", "")),
            str(row.get("capture_id", "")),
        ): row
        for row in existing_tray_rows
    }

    cells_by_key: dict[str, list[dict]] = defaultdict(list)

    for row in existing_cell_rows:
        key = record_key(
            str(row.get("day", "")),
            str(row.get("tray", "")),
            str(row.get("capture_id", "")),
        )

        cells_by_key[key].append(row)

    for job in jobs:
        day = str(job.day)
        tray = str(job.tray)
        capture_id = str(job.capture_id)
        key = record_key(day, tray, capture_id)

        output_folder = (
            OUTPUT_ROOT
            / day
            / tray
        )

        safe_capture = safe_name(capture_id)

        npz_path = (
            output_folder
            / f"{safe_capture}_relative_indices.npz"
        )

        ndvi_preview_path = (
            output_folder
            / f"{safe_capture}_ndvi_preview.png"
        )

        ndre_preview_path = (
            output_folder
            / f"{safe_capture}_ndre_preview.png"
        )

        ndvi_overlay_path = (
            output_folder
            / f"{safe_capture}_ndvi_square_overlay.png"
        )

        ndre_overlay_path = (
            output_folder
            / f"{safe_capture}_ndre_square_overlay.png"
        )

        if (
            not args.overwrite
            and key in tray_by_key
            and npz_path.exists()
            and ndvi_overlay_path.exists()
            and ndre_overlay_path.exists()
        ):
            print(
                f"SKIPPED EXISTING: "
                f"{day} > {tray} > {capture_id}"
            )
            continue

        cell_grid = coordinates_by_key.get(key)

        if cell_grid is None or len(cell_grid) != EXPECTED_CELLS:
            tray_by_key[key] = {
                "day_order": int(job.day_order),
                "day": day,
                "tray": tray,
                "tray_no": int(job.tray_no),
                "capture_id": capture_id,
                "grid_status": str(job.status),
                "cells_processed": 0,
                "cells_with_valid_ndvi": 0,
                "cells_with_valid_ndre": 0,
                "mean_cell_ndvi": np.nan,
                "median_cell_ndvi": np.nan,
                "mean_cell_ndre": np.nan,
                "median_cell_ndre": np.nan,
                "npz_path": "",
                "ndvi_preview_path": "",
                "ndre_preview_path": "",
                "ndvi_overlay_path": "",
                "ndre_overlay_path": "",
                "status": "FAIL",
                "notes": (
                    "Script 06 grid did not contain "
                    f"exactly {EXPECTED_CELLS} cells."
                ),
            }

            cells_by_key[key] = []
            continue

        try:
            band_paths = {
                "MS_G": resolve_band_path(
                    job.ms_g_path
                ),
                "MS_R": resolve_band_path(
                    job.ms_r_path
                ),
                "MS_RE": resolve_band_path(
                    job.ms_re_path
                ),
                "MS_NIR": resolve_band_path(
                    job.ms_nir_path
                ),
            }

            missing_band_paths = [
                band
                for band, path in band_paths.items()
                if path is None or not path.exists()
            ]

            if missing_band_paths:
                raise FileNotFoundError(
                    "Missing MS band file(s): "
                    + ", ".join(missing_band_paths)
                )

            bands = {
                band: read_raw_band(path)
                for band, path in band_paths.items()
            }

            dimensions = {
                band: array.shape
                for band, array in bands.items()
            }

            if len(set(dimensions.values())) != 1:
                raise ValueError(
                    "MS band dimensions do not match: "
                    + str(dimensions)
                )

            ndvi, ndre, ndvi_valid, ndre_valid = (
                calculate_indices(
                    bands["MS_R"],
                    bands["MS_RE"],
                    bands["MS_NIR"],
                )
            )

            output_folder.mkdir(
                parents=True,
                exist_ok=True,
            )

            np.savez_compressed(
                npz_path,
                ndvi=ndvi,
                ndre=ndre,
                ndvi_valid_mask=ndvi_valid,
                ndre_valid_mask=ndre_valid,
            )

            cells = evaluate_cells(
                ndvi,
                ndre,
                cell_grid,
            )

            title_base = (
                f"{day} | {tray} | {capture_id}"
            )

            save_index_preview(
                ndvi,
                ndvi_preview_path,
                f"{title_base} | Relative NDVI",
            )

            save_index_preview(
                ndre,
                ndre_preview_path,
                f"{title_base} | Relative NDRE",
            )

            save_square_overlay(
                ndvi,
                cells,
                ndvi_overlay_path,
                f"{title_base} | Relative NDVI | 70 MS square cells",
            )

            save_square_overlay(
                ndre,
                cells,
                ndre_overlay_path,
                f"{title_base} | Relative NDRE | 70 MS square cells",
            )

            cell_frame = pd.DataFrame(cells)

            valid_ndvi_cells = int(
                cell_frame["ndvi_mean"]
                .notna()
                .sum()
            )

            valid_ndre_cells = int(
                cell_frame["ndre_mean"]
                .notna()
                .sum()
            )

            status = (
                "PASS"
                if (
                    valid_ndvi_cells == EXPECTED_CELLS
                    and valid_ndre_cells == EXPECTED_CELLS
                )
                else "CHECK"
            )

            tray_by_key[key] = {
                "day_order": int(job.day_order),
                "day": day,
                "tray": tray,
                "tray_no": int(job.tray_no),
                "capture_id": capture_id,
                "grid_status": str(job.status),
                "cells_processed": len(cells),
                "cells_with_valid_ndvi": valid_ndvi_cells,
                "cells_with_valid_ndre": valid_ndre_cells,
                "mean_cell_ndvi": float(
                    cell_frame["ndvi_mean"].mean()
                ),
                "median_cell_ndvi": float(
                    cell_frame["ndvi_median"].median()
                ),
                "mean_cell_ndre": float(
                    cell_frame["ndre_mean"].mean()
                ),
                "median_cell_ndre": float(
                    cell_frame["ndre_median"].median()
                ),
                "npz_path": relative_path(
                    npz_path,
                    OUTPUT_ROOT,
                ),
                "ndvi_preview_path": relative_path(
                    ndvi_preview_path,
                    OUTPUT_ROOT,
                ),
                "ndre_preview_path": relative_path(
                    ndre_preview_path,
                    OUTPUT_ROOT,
                ),
                "ndvi_overlay_path": relative_path(
                    ndvi_overlay_path,
                    OUTPUT_ROOT,
                ),
                "ndre_overlay_path": relative_path(
                    ndre_overlay_path,
                    OUTPUT_ROOT,
                ),
                "status": status,
                "notes": (
                    "Relative NDVI and NDRE calculated "
                    "from aligned MS bands inside Script 06 square zones."
                ),
            }

            cells_by_key[key] = []

            for cell in cells:
                cells_by_key[key].append(
                    {
                        "day_order": int(job.day_order),
                        "day": day,
                        "tray": tray,
                        "tray_no": int(job.tray_no),
                        "capture_id": capture_id,
                        "cell_id": cell["cell_id"],
                        "row": cell["row"],
                        "column": cell["column"],
                        "centre_x": round(
                            cell["x"],
                            3,
                        ),
                        "centre_y": round(
                            cell["y"],
                            3,
                        ),
                        "square_x0": round(
                            cell["square_x0"],
                            3,
                        ),
                        "square_y0": round(
                            cell["square_y0"],
                            3,
                        ),
                        "square_x1": round(
                            cell["square_x1"],
                            3,
                        ),
                        "square_y1": round(
                            cell["square_y1"],
                            3,
                        ),
                        "square_side": round(
                            cell["square_side"],
                            3,
                        ),
                        "ndvi_valid_pixels": cell[
                            "ndvi_valid_pixels"
                        ],
                        "ndvi_mean": cell["ndvi_mean"],
                        "ndvi_median": cell[
                            "ndvi_median"
                        ],
                        "ndvi_std": cell["ndvi_std"],
                        "ndvi_p10": cell["ndvi_p10"],
                        "ndvi_p90": cell["ndvi_p90"],
                        "ndre_valid_pixels": cell[
                            "ndre_valid_pixels"
                        ],
                        "ndre_mean": cell["ndre_mean"],
                        "ndre_median": cell[
                            "ndre_median"
                        ],
                        "ndre_std": cell["ndre_std"],
                        "ndre_p10": cell["ndre_p10"],
                        "ndre_p90": cell["ndre_p90"],
                        "coordinate_source": cell[
                            "coordinate_source"
                        ],
                        "needs_review": cell[
                            "needs_review"
                        ],
                    }
                )

            print(
                f"{status}: {day} > {tray} > {capture_id} | "
                f"valid NDVI={valid_ndvi_cells}/70 | "
                f"valid NDRE={valid_ndre_cells}/70"
            )

        except Exception as error:
            tray_by_key[key] = {
                "day_order": int(job.day_order),
                "day": day,
                "tray": tray,
                "tray_no": int(job.tray_no),
                "capture_id": capture_id,
                "grid_status": str(job.status),
                "cells_processed": 0,
                "cells_with_valid_ndvi": 0,
                "cells_with_valid_ndre": 0,
                "mean_cell_ndvi": np.nan,
                "median_cell_ndvi": np.nan,
                "mean_cell_ndre": np.nan,
                "median_cell_ndre": np.nan,
                "npz_path": "",
                "ndvi_preview_path": "",
                "ndre_preview_path": "",
                "ndvi_overlay_path": "",
                "ndre_overlay_path": "",
                "status": "FAIL",
                "notes": str(error),
            }

            cells_by_key[key] = []

            print(
                f"FAIL: {day} > {tray} > {capture_id} | {error}"
            )

    tray_rows = sorted(
        tray_by_key.values(),
        key=lambda row: (
            int_or_default(row.get("day_order")),
            natural_key(row.get("tray", "")),
            natural_key(row.get("capture_id", "")),
        ),
    )

    cell_rows = [
        row
        for rows in cells_by_key.values()
        for row in rows
    ]

    cell_rows.sort(
        key=lambda row: (
            int_or_default(row.get("day_order")),
            natural_key(row.get("tray", "")),
            natural_key(row.get("capture_id", "")),
            int_or_default(row.get("cell_id")),
        )
    )

    reports_root.mkdir(
        parents=True,
        exist_ok=True,
    )

    pd.DataFrame(tray_rows).to_csv(
        tray_summary_path,
        index=False,
    )

    pd.DataFrame(cell_rows).to_csv(
        cell_results_path,
        index=False,
    )

    create_excel_report(
        excel_report_path,
        tray_rows,
        cell_rows,
    )

    settings = {
        "purpose": (
            "Relative NDVI and NDRE calculation using independent "
            "Script 06 multispectral square cell zones."
        ),
        "ndvi_formula": "(MS_NIR - MS_R) / (MS_NIR + MS_R)",
        "ndre_formula": "(MS_NIR - MS_RE) / (MS_NIR + MS_RE)",
        "accepted_grid_statuses": sorted(
            ACCEPTED_GRID_STATUSES
        ),
        "minimum_valid_pixels_per_cell": (
            MIN_VALID_PIXELS_PER_CELL
        ),
        "important_interpretation": (
            "NDVI and NDRE are relative image-derived indices unless "
            "the source MS bands are calibrated reflectance products."
        ),
    }

    settings_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    settings_path.write_text(
        json.dumps(
            settings,
            indent=2,
        ),
        encoding="utf-8",
    )

    status_counts = {
        status: sum(
            row.get("status") == status
            for row in tray_rows
        )
        for status in ("PASS", "CHECK", "FAIL")
    }

    print("\n" + "=" * 70)
    print("SCRIPT 07 FINISHED")
    print("=" * 70)

    for status, count in status_counts.items():
        print(f"{status}: {count}")

    print(f"\nExcel report:\n{excel_report_path}")
    print(f"\nTray summary:\n{tray_summary_path}")
    print(f"\nCell results:\n{cell_results_path}")
    print(f"\nNDVI/NDRE outputs:\n{OUTPUT_ROOT}")

    return 0 if status_counts["FAIL"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())